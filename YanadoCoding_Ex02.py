########################################
# 2차시 수업 학습 목표
# 1. 콜렉션(Collection)
#    - 리스트(List)
#    - 튜플(Tuple)
#    - 집합(Set)
#    - 사전(Dict)
# 2. 엑셀 라이브러리
########################################

########################################
# 과제 - 1 : 구구단 출력
# 1) 1개단의 구구단 출력   : 반복문 사용
# 2) 전체단의 구구단 출력  : 중첩 반복문 사용
# 3) 함수 이용 구구단 출력 : 1)을 함수로 구현 후 반복 호출

for i in range(1, 10):
    print(3, 'X', i, '=', 3 * i) # 3단 출력

def gugudan(dan):
    for i in range(1, 10):
        print(dan, 'X', i, '=', dan * i)

for dan in range(2, 10):
    gugudan(dan)


########################################
# 1. 콜렉션(Collection)

A = [3, 1, 4, 2] # 리스트 생성
print(A)
print(type(A)) # 데이터형 확인
print(A[1], A[-1]) # 인덱스로 참조, 인덱스는 0부터 시작, 음수의 인덱스는 -1부터 뒤에서부터 시작

# 파이썬 기본 함수
print(len(A), sum(A), min(A), max(A), sorted(A))
print(A) # sorted() 함수는 정렬된 결과만 보여주며, 리스트는 동일

A.sort() # 리스트 내장 함수 : 리스트 요소가 정렬됨.
print(A)

# 객체가 어떤 변수와 메소드(method)를 가지고 있는지 나열
print(dir(A))

# 객체의 이름, 매개변수, 독스트링(doc 속성)을 도움말 형태로 출력
# print(help(A))

A.reverse() # 리스트 내장 함수 : 역순
print(A)

# 리스트 슬라이싱
print(A[1:3], A[:3], A[2:], A[::2], A[::-1])

# 문자열 인덱싱
seasonName = ['Winter', 'Spring', 'Summer', 'Autumn', 'None'] #문자열 리스트
print(seasonName[1], seasonName[1][1], seasonName[0][:2:-1])

## 예제 1)
# 에라토스테네스의 체를 리스트로 구현해 보자.
MAX = 100 # 최대값 상수, 0 ~ 100
def eratosthenes(size):
    '''에라토스테네스의 체 : Boolean 리스트를 생성'''
    a = [False] * 2 + [True] * (size - 1) # 인덱스 0, 1은 False, 그 외는 True로 초기화
    for k in range(2, int(size**.5) + 1): # 2 ~ 10 반복
        a[k * 2::k] = [False] * ((size - k) // k) # k를 제외한 모든 k의 배수를 False로 변경
    return a

primeMask = eratosthenes(MAX) # 불리언 리스트 반환
print(primeMask)

# 사용자 입력으로 소수 여부 출력
# print(primeMask[int(input('is Prime? (0 ~ 100) : '))]) # 소수 불리언 리스트의 인덱스로 값 참조

## 예제 2)
# 숫자를 한글로 읽어보자.
UNIT = {0:"영", 1:"일", 2:"이", 3:"삼", 4:"사", 5:"오", 6:"육", 7:"칠", 8:"팔", 9:"구"}
numStr = '2021'
for d in numStr:
    print(UNIT[int(d)], end=' ')
print()

########################################
# 2. 엑셀 라이브러리

# import openpyxl
from openpyxl import Workbook
import pandas as pd

# 엑셀 파일 쓰기
write_wb = Workbook()

# 이름이 있는 시트를 생성
gugudan_ws = write_wb.create_sheet('구구단')

# 구구단 시트에 입력
write_ws = write_wb.active
# write_ws['A1'] = '2단'
# 행 단위로 입력, A1 ~ H1
write_ws.append(['2단', '3단', '4단', '5단', '6단', '7단', '8단', '9단'])

# 구구단을 중첩 반복문을 통해 셀 단위로 추가
for i in range(1, 10): # 행 반복 1 ~ 9
    for dan in range(2, 10): # 열 반복 2 ~ 9
        str = f'{dan} X {i} = {dan * i}' # 구구단 포맷으로 문자열 저장
        write_ws.cell(i + 1, dan - 1, str) # (행, 열, 값)으로 셀에 입력

# 엑셀 저장
write_wb.save("구구단.xlsx")

df_gugudan = pd.read_excel("구구단.xlsx")
print(df_gugudan.head())

print(df_gugudan['7단'])

