# -*- coding: utf-8 -*-

import re, os, shutil

import urllib 
import importlib


import xlsx_util
import list_util 
import utf_util
import gen_util


import xlrd 

import time_table as t 
import semester as sem 
import datetime, math

import openpyxl
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.worksheet.pagebreak import Break


import subprocess, time

dropbox_dir = "/cygdrive/c/Users/SSHS/Dropbox/수강신청/2022_1/수강신청계획서"
stats_dir = "/cygdrive/c/Users/SSHS/Dropbox/수강신청/2022_1/수강신청현황"

if not os.path.exists(dropbox_dir) :
    dropbox_dir = "/home/cape/tmp/Dropbox/수강신청/2022_1/수강신청계획서"
    stats_dir = "/home/cape/tmp/Dropbox/수강신청/2022_1/수강신청현황"


incl_excl_xlsx = "2022_2_수강포함배제목록.xlsx"
target_semesters = [ "2022_2" ]

#on_leave = [ "1620김온겸", "1720오유찬"]
on_leave = [ "2820반딧불" ]

students = t.students + on_leave


#prefixing_filename = "prefixed_enrollments.xlsx"

exclusive_subject_dict = dict( list_util.partition( 
re.split( r'\s+', 
'''경제학 커뮤니케이션
커뮤니케이션 경제학 
한국사 철학
철학 한국사''') ,2))


subject_credit_dict = {
"R&EI": 4,
"R&EII": 4,
"객체지향프로그래밍": 3,
"건강과체육I": 1,
"건강과체육II": 1,
"경제학": 2,
"고급물리학I": 3,
"고급물리학II": 3,
"고급생명과학I": 3,
"고급생명과학II": 3,
"고급지구과학": 3,
"고급커뮤니케이션": 2,
"고급화학I": 3,
"고급화학II": 3,
"고전문학": 2,
"과제연구I": 4,
"과제연구II": 4,
"과제연구III": 2,
"과제연구IV": 2,
"과학사": 2,
"국어I": 2,
"국어II": 2,
"기초통계학": 3,
"독서IV": 1,
"독서I": 1,
"독서II": 1,
"독서III": 1,
"디자인": 2,
"매체언어비평": 2,
"문법": 2,
"물리학I": 3,
"물리학II": 3,
"물리학III": 4,
"물리학IV": 3,
"물리학실험I": 1,
"물리학실험II": 1,
"미술I": 1,
"미술II": 1,
"미적분학I": 4,
"미적분학II": 4,
"생명과학III": 3,
"생명과학IV": 3,
"생명과학I": 3,
"생명과학II": 3,
"생명과학실험I": 1,
"생명과학실험II": 1,
"생활미술": 2,
"생활음악": 2,
"생활체육": 2,
"선형대수학": 3,
"세계문화지리": 3,
"세계사": 3,
"수리정보탐구": 2,
"수학I": 4,
"수학II": 4,
"수학III": 3,
"수학IV": 4,
"시사영어": 3,
"여가와체육I": 1,
"여가와체육II": 1,
"영미문화탐구": 2,
"영어I": 3,
"영어II": 3,
"영어III": 3,
"영어독해": 3,
"영어회화I": 1,
"영어회화II": 1,
"영작문": 3,
"예술사": 2,
"위탁교육": 2,
"융합과학": 3,
"융합과학탐구": 2,
"음악I": 1,
"음악II": 1,
"이공계체험학습": 1,
"자료구조": 3,
"자연탐사": 2,
"작문": 2,
"정수론": 3,
"정치경제": 3,
"졸업논문I": 2,
"졸업논문II": 2,
"중국어I": 2,
"중국어II": 2,
"지구과학I": 3,
"지구과학II": 3,
"지구과학III": 3,
"철학": 3,
"커뮤니케이션": 2,
"컴퓨터과학I": 2,
"컴퓨터과학II": 2,
"한국사": 3,
"현대문학": 2,
"화학I": 3,
"화학II": 3,
"화학III": 4,
"화학IV": 3,
"화학실험I": 1,
"화학실험II": 1,
"창의융합특강I": 2,
"창의융합특강II": 2,
"창의융합특강III": 2,
"창의융합특강IV": 2,
"창의융합특강V": 2,
"창의융합특강VI": 2,
"창의융합특강VII": 2,
"창의융합특강VIII": 2,
"창의융합특강IX": 2,
"창의융합특강X": 2,
"창의융합특강XI": 2,
"창의융합특강XII": 2,
"창의융합특강XIII": 2,
"창의융합특강XIV": 2,
"창의융합특강XV": 2,
"창의융합특강XVI": 2,
"창의융합특강XVII": 2,
"창의융합특강XVIII": 2,
"창의융합특강XIX": 2,
"창의융합특강XX": 2,
"창의융합특강XXI": 2,
"창의융합특강XXII": 2,
"창의융합특강XXIII": 2,
"창의융합특강XXIV": 2,
"창의융합특강XXV": 2,
"창의융합특강XXVI": 2,
"창의융합특강XXVII": 2,
"창의융합특강XXVIII": 2,
"창의융합특강XXIX": 2,
"창의융합특강XXX": 2,
"창의융합특강XXXI": 2,
"창의융합특강XXXII": 2,
"창의융합특강XXXIII": 2,
"창의융합특강XXXIV": 2,
"창의융합특강XXXV": 2,
"창의융합특강XXXVI": 2 }


isypyo_subjects = re.split(r'\s+',
'''R&EI
R&EII
객체지향프로그래밍
건강과체육I
건강과체육II
경제학
고급물리학I
고급물리학II
고급생명과학I
고급생명과학II
고급지구과학
고급커뮤니케이션
고급화학I
고급화학II
고전문학
과제연구I
과제연구II
과제연구III
과제연구IV
과학사
국어I
국어II
기초통계학
독서IV
독서I
독서II
독서III
디자인
매체언어비평
문법
물리학I
물리학II
물리학III
물리학IV
물리학실험I
물리학실험II
미술I
미술II
미적분학I
미적분학II
생명과학III
생명과학IV
생명과학I
생명과학II
생명과학실험I
생명과학실험II
생활미술
생활음악
생활체육
선형대수학
세계문화지리
세계사
수리정보탐구
수학I
수학II
수학III
수학IV
시사영어
여가와체육I
여가와체육II
영미문화탐구
영어I
영어II
영어III
영어독해
영어회화I
영어회화II
영작문
예술사
위탁교육
융합과학
융합과학탐구
음악I
음악II
이공계체험학습
자료구조
자연탐사
작문
정수론
정치경제
졸업논문I
졸업논문II
중국어I
중국어II
지구과학I
지구과학II
지구과학III
철학
커뮤니케이션
컴퓨터과학I
컴퓨터과학II
한국사
현대문학
화학I
화학II
화학III
화학IV
화학실험I
화학실험II
창의융합특강I
창의융합특강II
창의융합특강III
창의융합특강IV
창의융합특강V
창의융합특강VI
창의융합특강VII
창의융합특강VIII
창의융합특강IX
창의융합특강X
창의융합특강XI
창의융합특강XII''')



updated_hours_dict = dict(sem.hours_dict)

updated_hours_dict.update( {
"고급커뮤니케이션": 3,
"국어I": 3,
"국어II": 3,
"컴퓨터과학I": 3,
"물리학I": 4,
"화학II": 4,
"창의융합특강XVI/유기분광학과":2
})


data = { } 




def arabic_to_roman(x) : 
    ''' 1 -> I, 2 -> II, 3 -> III, 4 -> IV '''
    return x.replace("1", "I").replace("2", "II").replace("3", "III").replace("4", "IV")



def generate_classes( subj, num_of_classes ) : 
    ''' return a list of  subj+"_"+str(k) '''
    num = int(num_of_classes) 
    return [ subj + "_" + str(k) for k in range(1, num + 1) ] 
    
def generate_class_times( clss, num_of_hours ) : 
    ''' return a list of  clss+"$"+str(k) '''
    hours = int(num_of_hours) 
    return [ clss + "$" + str(k) for k in range(1, hours + 1) ] 

def subject_name(clss) : 
    ind = clss.find("_")
    if ind < 0 : 
        return clss 
    else : 
        return clss[:ind]



                
def read_year_advance_into_dict(xl_filename) : 
    name_convs = xlsx_util.read_xlsx_sheet_into_list(xl_filename ,2,1)
    output = { }
    for c in name_convs :
            #print "Read %s -> %s " % (c[0], c[1])
            print("Read {} -> {} ".format(c[0], c[1]))
            output[ c[0] ] = c[1]
    return output

def export_year_advance_name_subjects(sugang_csv, advance_xlsx, out_filename):
    old_name_subjects = read_sugang_csv(sugang_csv)
    advance_dict = read_year_advance_into_dict(advance_xlsx)
    output = [ ] 
    for old_name, subj in old_name_subjects : 
        output.append( ( advance_dict[old_name], old_name, subj ) )
        print("Read ({}, {}, {}) ".format(* output[-1]))
    output.sort()
    xlsx_util.write_tuples_into_xlsx(out_filename, output, 
        in_column_names = "신학번명 구학번명 수강신청과목".split() )
    


def read_xls_rows(xls_name, with_filename=False) : 
    if with_filename : 
        dir_name, fname = os.path.split(xls_name)
        output = [ [fname]] 
    else : 
        output = [ ] 
    sheet = xlrd.open_workbook(xls_name).sheet_by_index(0)
    for i in range(sheet.nrows) : 
        output.append(list(map(xlsx_util.strip_typeinfo,sheet.row_values(i))))
    return output


def read_xls_files_in_directory(dir_name, with_filename = False) : 
    fnames = [dir_name + "/" + c for c in os.listdir(dir_name) 
        if c.endswith(".xls")]
    output = [ ]
    for fn in fnames : 
        print("Reading <{}>..".format(fn))
        output.extend(read_xls_rows(fn, with_filename=with_filename))
    return output


def read_student_name_from_row(row) : 
    my_str = row[1]
    if my_str.find("성명 :")>0 and my_str.find("제 ")>=0 : 
        result = re.search(r'(\d)학년\s+(\d)반\s+(\d+)번\s+성명\s*:\s*(\S+)', 
                    my_str)
        if result is None : 
            return None
        else : 
            return str(int(float( result.group(1) ))*1000 + 
                        int(float( result.group(2) ))*100 + 
                            int(float( result.group(3) )) ) + \
                                result.group(4) 
    else : 
        return None

def contains_year_subject(row) : 
    if re.search(r'\d{4}', row[1]) and len(row[2])>0 and len(row[3])>0 : 
        return True 
    else : 
        return False

def convert_semester(sem_str) : 
    if re.search(r'\d', sem_str) : 
        return str(int(float(sem_str)))
    elif sem_str.find("여름")>=0 : 
        return "summer"
    elif sem_str.find("겨울")>=0 : 
        return "winter"
    else : 
        return None
        

def read_ban_prefix(row): 
    my_str = row[0]
    if my_str.find("학년")>0 and my_str.find("반")>=0 : 
        result = re.search(r'제\s*(\d)학년\s*(\d)반', my_str)
        if result is None : 
            return None
        else : 
            return result.group(1)+result.group(2) 
    else : 
        return None
        

def read_coursework_history(data) : 
    output = [ ] 
    name_subj_pairs = t.name_subjects[:]
    curr_st_name = None
    curr_ban_prefix = None
    for s_row in data : 
        new_name = read_student_name_from_row(s_row)
        if new_name : 
            curr_st_name = new_name
            continue
        row = s_row[0:]
        if contains_year_subject(row) : 
            year = str(int(float(row[1])))
            semester = convert_semester(row[2])
            subj = utf_util.convert_romans(row[4])
            if row[6] == "재수강 이전" or row[6] == "F" : 
                print("Found {} {} {} {}, Discarding it".format(curr_st_name,
                           year + "_" + semester, subj, row[6] ))
            else : 
                output.append( (curr_st_name, "{}_{}".format(
                    year, semester), subj))
                name_subj_pairs.append( (curr_st_name, subj) )

    sem_str = sem.semester_prefix[:-1]
    next_sem_str = sem_str[:-1] + "2" 

    print("Adding 졸업논문I 졸업논문II..")
    students = list_util.union( [c[0] for c in output])
    output.extend( [ (st_name, next_sem_str, "졸업논문II") 
        for st_name in students])
    output.extend( [ (st_name, sem_str, "졸업논문I") 
            for st_name in students])
    

    output.sort()
    return output
        
def save_coursework_history_into_xlsx(filename, data) : 
    xlsx_util.write_tuples_into_xlsx(filename, data)


def sg_name( substr ) :
    mystr = str(substr)
    output = [ ]
    for c in sorted(name_code_dict.keys()) :
        if c.find(mystr) >= 0 :
            output.append(c)
    if len(output) == 0 :
        return None
    else :
        return ' '.join(output)


def plan_name_check(filename):
    rows = xlsx_util.read_xlsx_sheet_into_list(filename,sheet_name="수강신청") 
    st_num = rows[2][0]
    st_name = rows[2][1]
    result = re.search(r'^(\d{4})', os.path.split(filename)[1])
    correct_num = result.group(1)
    if len(st_num)>0 and len(st_name) > 0 : 
        if st_num != correct_num : 
            print("{} != {} Wrong number in <{}>".format(st_num, correct_num,
                        filename))
            return False
        if sg_name(st_num) == st_num + st_name : 
            return True 
        elif (st_num + st_name) in name_code_dict : 
            return True 
        else : 
            print("Wrong name in <{}>".format(filename))
            print(st_num + st_name, "!=",  sg_name(correct_num) )
            return False 
    else : 
        print("Empty name in <{}>".format(filename))
        print("|{}|{}|".format(st_num ,  st_name))
        return False 

def check_names( dir_name = dropbox_dir, 
        verbose = False, check_function=plan_name_check):    
    fnames = [dir_name + "/" + c for c in os.listdir(dir_name) 
        if c.endswith(".xlsx") and (not c.startswith("~")) ]
    output = []
    print("Reading <{}>..".format( ' '.join(fnames)))
    for fn in fnames : 
        if verbose : print(fn)
        try :
            if not check_function(fn):
                new_name = fname_to_st_name(fn)
                if new_name is not None : 
                    list_util.check_append(output, new_name)
                else : 
                    print("No student name for <{}>".format(fn))
        except Exception as e : 
            print(e )
            print(fn)
    return output
    
wrong_name_message = "제출된 수강신청계획서에 잘못된 학번 이름이 " + \
        "기재되어 있습니다. 확인 정정하여 다시 제출하기 바랍니다." 

wrong_version_message = "잘못된 수강신청계획서 양식이 사용되었습니다. " + \
        "2020-11-29 버전을 사용하여 다시 제출하기 바랍니다." 

due_date_message = "오늘 중에 수강신청계획서를 제출하기 바랍니다."

def missing_submissions( dir_name = dropbox_dir, verbose = False):    
    fnames = [dir_name + "/" + c for c in os.listdir(dir_name) 
        if c.endswith(".xlsx") and (not c.startswith("~")) ]
    submitted = []
    print("Reading <{}>..".format( ' '.join(fnames)))
    for fn in fnames : 
        if verbose : print(fn)
        st_name = fname_to_st_name(fn)
        if st_name is None : 
            print("No student name for <{}>".format(fn))
        else :
            submitted.append(st_name)
    if sem.is_next_spring_semester() :  
        all_students = [c for c in students if c[0] in "1 2".split()]    
    else : 
        all_students = [c for c in students if c[0] in "2 3".split()]    
    return sorted(list_util.complement(all_students,submitted))
    


def version_check(filename, check_str = sem.version_check_pattern ) : 
    rows = xlsx_util.read_xlsx_sheet_into_list(filename,sheet_name="수강신청") 
    v_str = rows[0][0]
    st_num = rows[2][0]
    st_name = rows[2][1]
    if re.search(check_str, v_str) is None:
        print("Wrong version of <{}>, {}{}".format(filename, st_num, st_name))
        print(v_str, "!=", sem.version_check_pattern)
        return False 
    elif v_str[0]!=st_num[0] : 
        print("Version year mismatch for <{}>, {}{}".format(filename,
                st_num, st_name))
        print(v_str, "<->",st_num + st_name)
        return False 
    else : 
        return True 

def fname_to_st_name(filename): 
    sp = re.search(r'^(\d{4})', os.path.split(filename)[1])
    if sp is not None : 
        st_name = sg_name( sp.group(1) ) 
        if st_name is None : 
            return sp.group(1)
        else : 
            return st_name
    else : 
        return str(None)

def check_versions(dir_name = dropbox_dir, verbose = False):    
    fnames = [dir_name + "/" + c for c in os.listdir(dir_name) 
        if c.endswith(".xlsx") and (not c.startswith("~"))]
    output = []
    print("Reading <{}>..".format( ' '.join(fnames)))
    for fn in fnames : 
        if verbose is True : 
            print(fn)
        try :
            if not version_check(fn):
                new_name = fname_to_st_name(fn)
                if new_name is not None : 
                    list_util.check_append(output, new_name)
                else : 
                    #print "No student name for <%s>" % fn
                    print("No student name for <{}>".format(fn))
        except Exception as e : 
            print(e )
            print(fn)
    return output


def check_plan_integrity(filename, check_str = sem.version_check_pattern, 
        verbose = False ): 
    rows = xlsx_util.read_xlsx_sheet_into_list(filename) 
    v_str = rows[0][0]
    output = { "version": False, "name": False }
    if re.search(check_str, v_str) is not None:
        output["version"] = True
    else : 
        if verbose : 
            print("Wrong version of <{}>, {}{}".format(filename, rows[1][14],
                                                    rows[1][25]))
            
            print(v_str, "!=", sem.version_check_pattern)
    st_num = rows[1][14]
    st_name = rows[1][25]
    result = re.search(r'^(\d{4})', os.path.split(filename)[1])
    correct_num = result.group(1)
    if len(st_num)>0 and len(st_name) > 0 : 
        if st_num != correct_num : 
            if verbose:
                print("Wrong number in <{}>".format(filename))
                print(st_num , "!=",  correct_num)
        elif sg_name(st_num) == st_num + st_name : 
            output["name"] = True
        else : 
            if verbose : 
                print("Wrong name in <{}>".format(filename))
                print(st_num + st_name, "!=",  sg_name(correct_num) )
    else : 
        if verbose: 
            print("Empty name in <{}>".format(filename))
            print("|{}|{}|".format(st_num ,  st_name))
    return output

def check_name_versions( dir_name = dropbox_dir, 
        verbose = True, check_function=check_plan_integrity):    
    fnames = [dir_name + "/" + c for c in os.listdir(dir_name) 
        if c.endswith(".xlsx") and (not c.startswith("~")) ]
    output = { "name": [ ], "version": [ ], "xlsx":[]}
    #print "Reading <%s>.." % ' '.join(fnames)
    print("Reading <{}>..".format( ' '.join(fnames)))
    for fn in fnames : 
        if verbose : print(fn)
        try :
            check_result = check_function(fn, verbose=verbose)
            if not check_result["name"] : 
                output["name"].append( fname_to_st_name(fn) )
            if not check_result["version"] : 
                output["version"].append( fname_to_st_name(fn) )
        except Exception as e : 
            print(e )
            print(fn)
            output["xlsx"].append(fname_to_st_name(fn))
    return output
    


            
    




def override_history_philosophy(st_name, subj_sem_pairs):
    hp_subjs = [sem.one_four_spring_soph, sem.five_eight_spring_soph]

    if st_name[0] == "1" : 
        new_pairs = [c for c in subj_sem_pairs 
                    if not((c[0] in hp_subjs) and c[1][0]=="2") ]
        if st_name[1] in "1 2 3 4".split() : 
            new_pairs.append( (sem.one_four_spring_soph, "2학년1학기") )
            new_pairs.append( (sem.five_eight_spring_soph, "2학년2학기") )
        else : 
            new_pairs.append( (sem.one_four_spring_soph, "2학년2학기") )
            new_pairs.append( (sem.five_eight_spring_soph, "2학년1학기") )
        return new_pairs

    elif st_name[0] == "2" : 
        new_pairs = [c for c in subj_sem_pairs 
                    if not((c[0] in hp_subjs) and c[1][0]=="2") ]
        if st_name in t.members(hp_subjs[0]):
            new_pairs.append( (sem.one_four_spring_soph, "2학년1학기") )
            new_pairs.append( (sem.five_eight_spring_soph, "2학년2학기") )
        else : 
            new_pairs.append( (sem.one_four_spring_soph, "2학년2학기") )
            new_pairs.append( (sem.five_eight_spring_soph, "2학년1학기") )
        return new_pairs



def read_name_subject_relsems(plan_filename, use_filename=False, 
    with_cell_coords = False  ) : 
    rows = xlsx_util.read_xlsx_sheet_into_list(plan_filename, 
            sheet_name="수강신청") 
        
    st_name = rows[2][0] + rows[2][1]
    if st_name not in students : 
        try_st_name = sg_name(st_name)
    else :
        try_st_name = st_name
    if try_st_name is None and use_filename is True: 
        sr = re.search(r'(\d{4})\D+_', os.path.split(plan_filename)[1])
        if sr : 
            try_st_name = sg_name(sr.group(1))

    if try_st_name is None: 
        print("Found a wrong name {} in {}".format(st_name, plan_filename))
        return [ ]
    else : 
        st_name = try_st_name
    col_pairs = [ (3,5), (7,9), (11,13), (15,17)]
    end_row_num = 90
    subject_semester_pairs = [ ] 
    subject_semester_coords = [ ] 
    for i in range(end_row_num):
        for pair in col_pairs : 
            subj = rows[i][ pair[0] ]
            my_sem = rows[i][ pair[1] ]
            if len(my_sem)>0 and (not my_sem.startswith("=")) \
              and subj in sem.all_subjects:
                subject_semester_pairs.append( (subj, my_sem))
                subject_semester_coords.append( (subj, my_sem, i, pair[1]))
    if ("한국사", "2학년1학기") in subject_semester_pairs : 
        subject_semester_pairs.append( ("철학", "2학년2학기") )
        subject_semester_coords.append(("철학", "2학년2학기", 19 ,9) )
    elif ("한국사", "2학년2학기") in subject_semester_pairs :  
        subject_semester_pairs.append( ("철학", "2학년1학기") )
        subject_semester_coords.append(("철학", "2학년1학기", 19 ,9) )
    else : 
        print("Something's wrong. Couldn't find <한국사> in {}.".format(
            plan_filename))
        print("Continuing anyway." )

    if (st_name.startswith("1") and sem.is_next_spring_semester())  \
      or (st_name.startswith("2") and (not sem.is_next_spring_semester()) ):
        subject_semester_pairs = override_history_philosophy(st_name, 
                            subject_semester_pairs)
    


    #col_pairs = [ (3,5), (7,9), (11,13), (15,17)]

    reen_start_rownum = 32
    max_reen = 11
    reen_start_colnum = 21
    reen_pairs = [ ] 
    for i in range(reen_start_rownum, reen_start_rownum + max_reen):
        subj = rows[i][reen_start_colnum + 1]
        if len(subj)>0:
            tmp_relsem = rows[i][reen_start_colnum][:-2]
            subject_semester_pairs.append( (subj, tmp_relsem) )
            subject_semester_coords.append( (subj, tmp_relsem, 
                i, reen_start_colnum + 1) )

    name_subj_relsems = [ (st_name, subj, relsem) 
            for subj, relsem in subject_semester_pairs]
    name_subj_relsem_coords = [ (st_name, subj, relsem, i, j)
            for subj, relsem, i, j in subject_semester_coords]
    if with_cell_coords : 
        return name_subj_relsem_coords
    else : 
        return name_subj_relsems
            


relsem_list = re.split(r'\s+',
'''1학년1학기 1학년여름학기 1학년2학기 1학년겨울학기
2학년1학기 2학년여름학기 2학년2학기 2학년겨울학기
3학년1학기 3학년여름학기 3학년2학기''')

freshmen_semesters = re.split(r'\s+',
'''2022_1 2022_summer 2022_2 2022_winter 
2023_1 2023_summer 2023_2 2023_winter 
2024_1 2024_summer 2024_2''') 

sophomore_semesters = re.split(r'\s+',
'''2021_1 2021_summer 2021_2 2021_winter 
2022_1 2022_summer 2022_2 2022_winter 
2023_1 2023_summer 2023_2''') 

senior_semesters = re.split(r'\s+',
'''2020_1 2020_summer 2020_2 2020_winter 
2021_1 2021_summer 2021_2 2021_winter 
2022_1 2022_summer 2022_2''')


fs_semesters = re.split(r'\s+',
'''2019_1 2019_summer 2019_2 2019_winter 
2020_1 2020_summer 2020_2 2020_winter 
2021_1 2021_summer 2021_2 2021_winter 
2022_1 2022_summer 2022_2''') 


fresh_relsem_dict = dict(list_util.transposed( 
        [ relsem_list, freshmen_semesters]))
soph_relsem_dict = dict(list_util.transposed( 
        [ relsem_list, sophomore_semesters]))
sen_relsem_dict = dict(list_util.transposed( 
        [ relsem_list, senior_semesters]))

def read_name_subject_sems(filename, use_filename=False):
    triples = read_name_subject_relsems(filename, use_filename=use_filename)
    if len(triples)==0:
        return [ ]
    st_name = triples[0][0]
    if st_name.startswith("1"): 
        my_dict = fresh_relsem_dict 
    elif st_name.startswith("2"): 
        my_dict = soph_relsem_dict 
    else: 
        my_dict = sen_relsem_dict 
    before_conv = [ (c[0], c[1], my_dict[c[2]]) for c in triples ]

    output = [ (c[0], gen_util.apply_dict(sem.special_lecture_formal_dict, 
                 utf_util.convert_romans(c[1])) , c[2]) for c in before_conv]
    

    return output

def target_semester(off_set):
    return sem.semester_list[ sem.current_semester_index + off_set ]




def read_sugang_plan(filename, 
        target_semesters = target_semesters, 
        conv_func = sem.convert_function, use_filename=False ) : 
    triples = read_name_subject_sems(filename, use_filename=use_filename)
    if len(triples)==0:
        return [ ]
    output = [ conv_func( (c[0], c[1], c[2]) ) for c in triples 
        if c[2] in target_semesters ] 
    return output
    


def deal_with_exceptions(ns_pairs):
    output = ns_pairs[:]
    print(  "1816김재성 - 정치와법 -> 1816김재성 - 세계문화지리" )
    output.remove( ("1816김재성", "정치와법") )
    output.append( ("1816김재성", "세계문화지리") )
    return output

def modify_by_include_exclude_xlsx( include_exclude_xlsx , name_subj_sems, 
        target_semesters = target_semesters ):

    current_triples = name_subj_sems[:]

    data = xlsx_util.read_xlsx_sheets_into_list(include_exclude_xlsx)
    new_triples = [ ] 

    for sem_bytes in data.keys():
        sem_str = sem_bytes.decode("utf-8")
        if not sem_str.find("수강포함명단") >= 0 : 
            continue
        for tg_sem in target_semesters : 
            if not sem_str.find(tg_sem ) >= 0 : 
                continue
            subj_namelists = list_util.transposed( data[sem_bytes] )
            for line in subj_namelists: 
                subj = line[0]
                for st_name in line[1:] : 
                    if len(st_name) > 0 : 
                        new_triples.append( (st_name, subj, tg_sem) )


    for st_name, subj, sem_str in new_triples : 
        list_util.check_append(current_triples, (st_name, subj, sem_str), 
            silent=True )

    excl_triples = [ ] 
    for sem_bytes in data.keys():
        sem_str = sem_bytes.decode("utf-8")
        if not sem_str.find("수강배제명단") >= 0 : 
            continue
        for tg_sem in target_semesters : 
            if not sem_str.find(tg_sem ) >= 0 : 
                continue
            subj_namelists = list_util.transposed( data[sem_bytes] )
            for line in subj_namelists: 
                subj = line[0]
                for st_name in line[1:] : 
                    if len(st_name) > 0 : 
                        excl_triples.append( (st_name, subj, tg_sem) )

    for st_name, subj, sem_str in excl_triples : 
        list_util.check_remove(current_triples, (st_name, subj, sem_str), 
            silent=True )

    return current_triples


            
        
    
include_exclude_xlsx =  stats_dir + "/2022_2_수강포함배제목록.xlsx"


def read_sugang_plans_from_directory(dir_name = dropbox_dir, 
        verbose = False, target_semesters = target_semesters, 
            use_filename=False, 
        all_semesters=False, include_exclude = include_exclude_xlsx ):    
    fnames = [dir_name + "/" + c for c in os.listdir(dir_name) 
        if c.endswith(".xlsx") and not c.startswith("~")]
    special_ns_pairs = [ ] 
    output = [ ]
    print("Reading <{}>..".format(' '.join(fnames)))
    for fn in fnames : 
        try :
            if verbose : 
                print("<{}>..".format(fn))
            if all_semesters is False : 
                new_pairs = read_sugang_plan(fn, 
                    target_semesters = target_semesters, 
                                  use_filename=use_filename)
                if verbose : 
                    if len(new_pairs)>0 : 
                        print(new_pairs[0][0], end=" ")
                        print(' '.join([c[1] for c in new_pairs]))
                for c in new_pairs : 
                    if c in output : 
                        print(fn, 
                 "({},{}) already has been read!, adding another..".format(*c))
                    output.append(c)
            else : 
                new_triples = read_name_subject_sems(fn, 
                                use_filename=use_filename)
                if verbose : 
                    if len(new_triples)>0 : 
                        print(new_triples[0][0], end=" ")
                        print(' '.join([c[1]+c[2] for c in new_triples]))
                for c in new_triples : 
                    if c in output : 
                        print(fn, 
              "({},{},{}) already has been read!, adding another..".format(*c))
                    output.append(c)
        except Exception as e : 
            print(e )
            print(fn)
    if all_semesters is False : 
        js_p = junior_subject_pairs(target_semesters )
        #js_p = [ (c[0], c[1], sem.next_semester_prefix[:-1]) for c in js_p]
        #if off_set == 1 and (not sem.is_next_spring_semester()) : 
            #js_p = [ ] 
        output.extend(js_p)

    #modified_output = deal_with_exceptions(output)
    if include_exclude is not None : 
        modified_output = modify_by_include_exclude_xlsx( include_exclude, 
            output, target_semesters = target_semesters )
    else : 
        modified_output = output

    return modified_output
    
def read_isupyo(xlsx_filename) : 
    rows = xlsx_util.read_xlsx_sheet_into_list(xlsx_filename, 
                sheet_name="과목명")
    st_name = rows[0][5] + rows[0][7]
    output = [ ] 
    for row in rows[1:] : 
        check = row[2]
        if len(check)==0 or check == "0" or check == 0 : 
            continue 
        else : 
            subj = utf_util.convert_romans(row[0])
            output.append( (st_name, subj) )
    return output


def read_isupyo_from_directory(dir_name, verbose = False) : 
    fnames = [dir_name + "/" + c for c in os.listdir(dir_name) 
        if c.endswith(".xlsx")]
    output = [ ]
    #print "Reading <%s>.." % ' '.join(fnames)
    print("Reading <{}>..".format( ' '.join(fnames)))
    for fn in fnames : 
        try :
            new_pairs = read_isupyo(fn)
        except Exception as e : 
            print(e )
            print(fn)
        if verbose : 
            print(new_pairs[0][0],)
            print(' '.join([c[1] for c in new_pairs]))
        for c in new_pairs : 
            if c in output : 
                #print fn, "(%s,%s) already has been read!, adding another.." % c
                print(fn, 
                  "({},{}) already has been read!, adding another..".format(c))
        output.extend(new_pairs) 
    return output
    


def isupyo_name_check(filename):
    rows = xlsx_util.read_xlsx_sheet_into_list(filename, 
                sheet_name="과목명")
    st_num = rows[0][5] 
    st_name = rows[0][7]
    result = re.search(r'(\d{4})', filename)
    correct_num = result.group(1)
    if len(st_num)>0 and len(st_name) > 0 : 
        if st_num != correct_num : 
            #print "Wrong number in <%s>" % filename
            print("Wrong number in <{}>".format(filename))
            return False
        if sg_name(st_num) == st_num + st_name : 
            return True 
        else : 
            #print "Wrong name in <%s>" % filename
            print("Wrong name in <{}>".format(filename))
            print(st_num + st_name, "!=",  sg_name(correct_num) )
            return False 
    else : 
        #print "Empty name in <%s>" % filename
        #print "|%s|%s|" % (st_num ,  st_name)
        print("Empty name in <{}>".format(filename))
        print("|{}|{}|".format(st_num ,  st_name))
        return False 


def pair_differences(submitted_data, neis_data) : 
    left = list_util.collect_on_first(list_util.complement( 
                submitted_data, neis_data))
    right = list_util.collect_on_first(list_util.complement( 
                neis_data, submitted_data ))

    left_dict = { }
    for row in left : 
        left_dict[row[0]] = row 
    right_dict = { }
    for row in right : 
        right_dict[row[0]] = row 

    st_names = list_util.union( [c[0] for c in left], [d[0] for d in right])
    st_names.sort()
    output = [ ] 
    if len(left) > 0 : 
        left_maxlen = max( [len(row) for row in left])
    else : 
        left_maxlen = 1
    for st_name in st_names : 
        if st_name in left_dict : 
            left_part = list_util.pad( left_dict[st_name], left_maxlen ) 
        else : 
            left_part = list_util.pad( [ ],  left_maxlen ) 
        if st_name in right_dict : 
            right_part =  right_dict[st_name]
        else : 
            right_part =  [ ]
        output.append( left_part +  right_part)
    return output
        


def read_pass_records(filename) : 
    rows = xlsx_util.read_xlsx_sheet_into_list(filename)
    #for row in rows : 
        #print '|'.join(row)
        #print 
    output = [ ]
    for row in rows : 
        if len(row[1]) > 0 and row[1].startswith("3") : 
            st_name = row[1] + row[2] 
            if len(row[6]) > 0 : 
                pass_subj = row[6]
            else : 
                pass_subj = "통과" 
            output.append( (st_name, pass_subj) )
    return output
            

def read_isupyo_pass(xlsx_filename) : 
    rows = xlsx_util.read_xlsx_sheet_into_list(xlsx_filename, 
                sheet_name="과목명")
    st_name = rows[0][5] + rows[0][7]
    pass_str = rows[0][9]
    subs_subj = rows[0][11]
    if pass_str == "통과" :
        return (st_name, pass_str) 
    else : 
        return (st_name, subs_subj) 



def read_isupyo_pass_from_directory(dir_name, verbose = False) : 
    fnames = [dir_name + "/" + c for c in os.listdir(dir_name) 
        if c.endswith(".xlsx") and (not c.startswith("~")) ]
    output = [ ]
    #print "Reading <%s>.." % ' '.join(fnames)
    print("Reading <{}>.." % ' '.join(fnames))
    for fn in fnames : 
        try :
            new_pair = read_isupyo_pass(fn)
        except Exception as e : 
            print(e )
            print(fn)
        if verbose : 
            print(fn , end=" ")
            #print "(%s, %s)" % new_pair 
            print("({}, {})".format(*new_pair))
        if c in output : 
            #print fn, "(%s,%s) already has been read!, adding another.." % c
            print(fn, 
                "({},{}) already has been read!, adding another..".format(c))
        output.append(new_pair) 
    return output
    

        
def save_pass_records_differences(output_filename, 
                pass_records_fname, dir_name) : 
    recs = read_pass_records(pass_records_fname ) 
    st_data = read_isupyo_pass_from_directory(dir_name) 
    diff_data = pair_differences(st_data, recs) 
    if len(diff_data) == 0 : 
        print("All coincide.")
    else : 
        xlsx_util.write_tuples_into_xlsx(output_filename, diff_data)



def read_summer_enroll(filename) : 
    dict_data = xlsx_util.read_xlsx_sheet_into_dict( filename, 5, 1 )
    pairs = [ k for k in dict_data if len(dict_data[k]) == 1 ]
    output = [ (  st_num + dict_data[ ( st_num, "이름" ) ], subj ) 
                for st_num, subj in pairs ] 
    return output



def save_isupyo_differences(output_filename, summer_enroll_filename, 
             neis_dir, isupyo_dir) : 
    summer_pairs = [ c for c in read_summer_enroll(summer_enroll_filename) 
                        if c[0].startswith("3")]
    st_data = read_isupyo_from_directory(isupyo_dir)
    neis_rows = read_xls_files_in_directory(neis_dir) 
    neis_triples = read_coursework_history(neis_rows)  
    neis_pairs = [ (c[0],c[2]) for c in neis_triples ]
    neis_pairs.extend( summer_pairs )
    diff_data = pair_differences(st_data, neis_pairs)
    xlsx_util.write_tuples_into_xlsx(output_filename, diff_data)



def ban_name_name(ban_str, in_name) : 
    #print ban_str, in_name
    for c in t.names : 
        if c.find(in_name) >= 0 and c.startswith( ban_str ) : 
            return c 
    else : 
        return None
    
def read_ban_from_table_row(row) : 
    my_str = row[0]
    if my_str.find(".xls")>0 :  
        return my_str.replace(".xls", "").replace("-","")
    else :
        return None


def read_student_name_from_table_row(row, ban_str) : 
    my_str = row[0]
    if my_str.find("학년도")>0 and my_str.find("반")>=0 : 
        result = re.search(r'(\d{4})학년도 (\d)학기\s+(\d)반\s*(\S+)', 
                    my_str)
        if result is None : 
            return None
        else : 
            return ban_name_name( ban_str, result.group(4)) 
    else : 
        return None

def unpara_classname(clss) :
    return re.sub( r'\((\d{1,2})\)', r'_\1', clss )


def read_classname_slot_from_table_row(row) : 
    i_col = { 0:1, 1:3, 2:4, 3:5, 4:11 }
    my_str = row[0]
    if my_str.find("교시")>0 :  
        output = [ ]
        slot_num_str = my_str[0]
        for i in range(5) : 
            if len( row[ i_col[i] ] ) > 0 : 
                c_name = unpara_classname(
                            utf_util.convert_romans(row[i_col[i]]))
                #if c_name.find("12")>=0 and c_name.find("실험")>=0 : 
                    #print row[i_col[i]], c_name
                output.append( (c_name, t.days[i] + slot_num_str) )
        return output
    else : 
        return None


def read_neis_table_data(data) : 
    output = [ ] 
    curr_st_name = None
    curr_ban = None
    for row in data : 
        new_ban = read_ban_from_table_row(row) 
        if new_ban : 
            curr_ban = new_ban 
            continue
        new_name = read_student_name_from_table_row(row, curr_ban)
        if new_name : 
            curr_st_name = new_name
            continue
        class_slots = read_classname_slot_from_table_row(row) 
        if class_slots : 
            output.extend( [(curr_st_name, clss, sl) 
                    for clss, sl in class_slots] )
            continue
    return output 

def remove_ABCD(my_str) : 
    output = my_str.replace("A", "")
    output = output.replace("B", "")
    output = output.replace("C", "")
    output = output.replace("D", "")
    return output


def name_class_slots() : 
    output = [ ] 
    for st_name, pt in t.name_classparts : 
        if not st_name.endswith("t") :
            output.append( ( st_name, t.class_name_of(pt), 
                t.classpart_slot[pt]) )
    return output 

def compute_number_of_classes( subj, st_count) : 
    if st_count < 6 : 
        return 0 
    elif subj in sem.junior_switch_subjects + sem.sophomore_switch_subjects : 
        return 4 
    elif subj in sem.exp_subjects : 
        return math.ceil( st_count / 12.0)
    elif subj in sem.junior_subjects["2022_1"] \
            + sem.junior_subjects["2022_2"] : 
        return 8
    else : 
        if st_count >= 90 :
            return math.floor(0.5 + (st_count / 16.0) )
        elif st_count >= 17 and st_count <= 32 : 
            return 2
        elif (st_count % 15) >= 6 : 
            return math.ceil( st_count / 15.0)
        else : 
            return math.floor( st_count / 15.0)
        



def export_enrollment_stats_xlsx( name_subj_sems , 
        out_filename = None , target_semesters = target_semesters ):
    title_str = sem.next_semester_str \
        + " 수강신청 현황 (%s)" % datetime.date.today().strftime("%Y-%m-%d")
    if out_filename is None : 
        filename = \
            "수강신청현황-%s.xlsx" % datetime.date.today().strftime("%Y%m%d")
    else : 
        filename = out_filename
    print("Reading style from <enrollment-stats-template.xlsx>.")
    shutil.copyfile( "enrollment-stats-template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    ws =  wb[ "Sheet1"]
    ws.cell(row=1, column=3).value = title_str
    target_sem = sem.next_semester_prefix[:-1]

    #if target_sem[-1] == "1" :  
        #j_subjs = sem.junior_subjects_spring 
    #else : 
        #j_subjs = sem.junior_subjects_fall 
    j_subjs = sem.junior_subjects[target_sem]


    #print(name_subj_sems[:20])
    #print(list_util.union([p[1] for p in name_subj_sems if p[2]==target_sem ]))

    subject_list = list_util.union( [p[1] for p in name_subj_sems 
            if p[2]==target_sem ], 
                    "한국사 철학".split(), j_subjs )
    subject_list.sort()

    subj_rows = [ ] 
    for subj in subject_list : 
        st_count = len( [p for p in name_subj_sems if p[1] == subj 
            and p[2] == target_sem])

        #if sem.is_next_spring_semester():
            #if subj in sem.sophomore_switch_subjects : 
                #st_count = int(sem.num_sophomores/2) 
            
        s_group = sem.subject_group_dict[subj]
        num_classes = compute_number_of_classes(subj, st_count)
        subj_rows.append(   (s_group, subj, st_count, num_classes )  )

    subj_rows.sort(key = 
        (lambda x: ( sem.ordered_group.index(x[0]), x[1]))  )
    xlsx_util.fill_in_sheet_by_tuples(ws, subj_rows, 3,2)
    credits = [ ]
    for tup in subj_rows : 
        hours = updated_hours_dict[tup[1]]
        if tup[1] in sem.exp_subjects : 
            cr = hours-1
        else : 
            cr = hours
        credits.append( (cr, hours) )
    xlsx_util.fill_in_sheet_by_tuples(ws, credits, 3,8)

    for sem_str in target_semesters : 
        output = [ ] 
        subject_list = list_util.union( [ c[1] for c in name_subj_sems 
            if c[2] == sem_str])
        subject_list.sort()
        for subj in subject_list :
            output.append( [ subj ] 
                + sorted(list_util.union( [ p[0] for p in name_subj_sems 
                    if p[1] == subj and p[2] == sem_str])) )
        ws =  wb[ sem_str ]
        xlsx_util.fill_in_sheet_by_tuples(ws, list_util.transposed(output), 1,1)

    print("Saving to <{}>..".format(filename))
    wb.save( filename )

def junior_skip_subject(st_name, sem_str ) : 
    if sem_str.endswith("1") : 
        if st_name[1] in "1 2 3 4".split() : 
            return sem.five_eight_spring_subj
        else : 
            return sem.one_four_spring_subj
    elif sem_str.endswith("2") : 
        if st_name[1] in "1 2 3 4".split() : 
            return sem.one_four_spring_subj
        else : 
            return sem.five_eight_spring_subj
    else : 
        return None


def junior_subject_pairs(target_semesters = target_semesters) : 
    output = [ ] 

    for sem_str in target_semesters : 
        if sem_str.endswith("1") : 
            j_names = sem.pre_junior_names
            subj_list = sem.junior_subjects[sem_str]
        elif sem_str.endswith("2") : 
            j_names = sem.junior_names 
            subj_list = sem.junior_subjects[sem_str]
        else : 
            j_names = [ ] 
            subj_list = [ ] 

        for subj in subj_list : 
            for st_name in j_names : 
                if not subj == junior_skip_subject(st_name, sem_str ) : 
                    output.append( (st_name, subj, sem_str) )
    return output
    
def export_missing_submissions(dir_name = dropbox_dir, verbose = False):    
    missing_names = missing_submissions(dir_name = dir_name, verbose = verbose)
    wrong_name_inputs = check_names(dir_name = dir_name, verbose = verbose)
    wrong_versions = check_versions(dir_name = dir_name, verbose = verbose)
    output = [ ] 
    output.append( ["학번이름오류"] + wrong_name_inputs )
    output.append( ["양식오류"] + wrong_versions )
    output.append( ["미제출"] + missing_names )
    filename = "미제출현황-{}.xlsx".format( 
        datetime.date.today().strftime("%Y%m%d"))
    out_filename = stats_dir + "/" + filename
    xlsx_util.write_tuples_into_xlsx(out_filename,
        list_util.transposed(output), in_column_names = "first row" )
    


def read_and_save_coursework_plans() : 
    filename = "수강신청현황-{}.xlsx".format( 
                datetime.date.today().strftime("%Y%m%d"))
    nss_triples = read_sugang_plans_from_directory( dropbox_dir, 
        verbose=True, use_filename=True, 
        include_exclude = include_exclude_xlsx )
    export_enrollment_stats_xlsx(  nss_triples , 
        out_filename = stats_dir + "/" + filename)


def file_list_str() : 
    #xlsx_flist, err_out = subprocess.Popen( [ 'ls', '-l', dropbox_dir ], 
                #stdout=subprocess.PIPE, stderr=subprocess.PIPE).communicate()
    #data = [ c for c in xlsx_flist.split("\n") if c.find(".xlsx") >= 0]
    xlsx_flist = [ dropbox_dir + "/" + fname 
                    for fname in os.listdir(dropbox_dir) 
                    if fname.find(".xlsx")>= 0]
    data = [  fname + " " + str(os.path.getmtime(fname)) 
                for fname in xlsx_flist]
    return '\n'.join(data)
    

def update_flist() : 
    output = file_list_str()
    print("Writing <plan_filelist.txt>..")
    with open("plan_filelist.txt", "w") as f:
        f.write(output)

def check_for_changes() : 
    xlsx_flist = file_list_str()
    with open("plan_filelist.txt", "r") as f:
        curr_data = f.read()
    if curr_data != xlsx_flist : 
        return True
    else : 
        return False


def export_differences_into_xlsx(filename, diff_tuples) : 
    template_filename = "difference-list-template.xlsx"
    #print "Reading style from <%s>" % template_filename
    print("Reading style from <{}>".format(template_filename))
    shutil.copyfile( template_filename, filename)
    #print "Copied %s to %s" %( template_filename, filename)
    print("Copied {} to {}".format( template_filename, filename))
    wb = openpyxl.load_workbook(filename)
    ws = wb[  "변경목록" ]
    xlsx_util.fill_in_sheet_by_tuples(ws, diff_tuples, row_ind=2)
    #print "Saving to <%s>" % filename
    print("Saving to <{}>".format(filename))
    wb.save(filename)


def export_isupyo(isu_data, num_name, directory="coursework-check"): 
    isu_list = [ ] 
    for triple in isu_data : 
        if num_name == triple[0] : 
            isu_list.append(triple[2])
    #print "Found for %s, %s" % ( num_name, ' '.join(isu_list)) 
    print("Found for {}, {}".format( num_name, ' '.join(isu_list)))
    tmp_name = re.sub(r'\d+', '', num_name) 
    tmp_num = num_name.replace(tmp_name, '')
    #filename = directory + "/" + "교육과정_이수현황표_%s_%s_배포용.xlsx" % ( tmp_num, tmp_name)
    filename = directory + "/" \
        + "교육과정_이수현황표_{}_{}_배포용.xlsx".format( tmp_num, tmp_name)
    shutil.copyfile( "isupyo-template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    ws =  wb[ unicode( "과목명", "utf-8") ]
    print("Writing")
    ws.cell(row=1, column=6).value = int(tmp_num)
    ws.cell(row=1, column=8).value = tmp_name
    for i in range(2, 119) : 
        subj = utf_util.convert_romans(ws.cell(row=i, 
                column = 1).value.encode(encoding="utf-8"))
        if subj in isu_list : 
            ws.cell(row=i, column = 3).value = 1
            print( ws.cell(row=i, column = 1).value.encode(encoding="utf-8"), 
                end=" " )
            print( "-",  end=" ")
            print(ws.cell(row=i, column = 3).value , end=" " )
    sheet_one =  wb[ unicode( "졸업요건", "utf-8") ]
    #patch_border.fix_borders_on_sheet(sheet_one)
    #print "Saving to <%s>.." % filename
    print("Saving to <{}>..".format(filename))
    wb.save( filename )

def export_isupyo_of_all(isu_data) : 
    senior_names = sorted( [ c for c in t.names if c.startswith("3") ])
    #print "Working on %s .." % ' '.join(senior_names)
    print("Working on {} .." % ' '.join(senior_names))
    for st_name in senior_names : 
        export_isupyo(isu_data, st_name) 


def export_credit_lists(isu_data, filename, class_list = sem.sen_ban_list, 
        directory="coursework-check", st_list = None, date = None, 
            salt = None) :
    f_name = directory + "/"  + filename 

    print("Reading style from <credit-list-template.xlsx>.")
    shutil.copyfile( "credit-list-template.xlsx", f_name)
    wb = openpyxl.load_workbook(f_name)

    template_ws = wb["Template"]

    if salt is None :
        my_salt = datetime.datetime.today().strftime("%Y-%m-%d")
    else :
        my_salt = salt



    for clss in class_list :
        prefix = clss.replace("-", "")
        name_list = [nm for nm in students if nm.startswith(prefix)]
        if st_list is not None :
            name_list = list_util.intersect(name_list, st_list)
        name_list.sort()
        if(len(name_list)==0) :
            continue

        #print "Writing weekly table for %s." % clss
        print("Writing weekly table for {}.".format(clss))
        new_worksheet = wb.create_sheet(  clss )
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        for k in range( 38, 801, 40  ):
            new_worksheet.row_dimensions[k+2].hidden = True
        for k in range(40,801,40):
            #new_worksheet.page_breaks.append( Break(id=k) )
            new_worksheet.row_breaks.append( Break(id=k) )

        for i, st_name in enumerate(name_list) :
            tmp_name = re.sub(r'\d+', '', st_name) 
            tmp_num = st_name.replace(tmp_name, '')

            new_worksheet.cell( row = 1 + 40*i, column = 2 ).value = \
                tmp_num + " " + tmp_name 

            if date is None :
                date_str = datetime.datetime.today().strftime("%Y-%m-%d")
            else :
                date_str = date
            new_worksheet.cell( row = 1 + 40*i, column = 8 ).value = date_str



            isu_list = [ ] 
            fall_isu_list = [ ] 
            fall_str = sem.next_semester_prefix[:-1]
            for triple in isu_data : 
                if st_name == triple[0] : 
                    isu_list.append(triple[2])
                    if triple[1] == fall_str : 
                        fall_isu_list.append(triple[2])
            #fill_in_time_table(new_worksheet, i*20, st_name, t_name=t_name)
            isu_output = [ ] 
            for j, subj in enumerate(isypyo_subjects) : 
                #subj_credit = ("%s(%d)" % (subj, subject_credit_dict[subj]))
                subj_credit = ("{}({})".format(subj,subject_credit_dict[subj]))
                if subj in fall_isu_list : 
                    isu_output.append( ( j+1, subj_credit, "E" ))
                elif subj in isu_list : 
                    isu_output.append( ( j+1, subj_credit, "O" ))
                else : 
                    isu_output.append( ( j+1, subj_credit, "" ))
            xlsx_util.fill_in_sheet_by_tuples(new_worksheet, 
                list_util.column_pack(isu_output, 37),  3+i*40, 1) 

        for k in range( (i+1)*40 + 1, 801  ):
            new_worksheet.row_dimensions[k].hidden = True

        if len(my_salt)>0:
            new_worksheet.protection.password = \
                gen_util.generate_password(my_salt,clss)


    wb.remove(template_ws)
    #print "Saving to <%s>."%f_name
    print("Saving to <{}>.".format(f_name))
    wb.save(f_name)

def year_advance_coursework_plan(xl_filename, adv_dict, 
        directory="coursework-plan", output_directory = ".", 
        indices=[]) : 
    #print xl_filename
    old_num =  re.search(  r'([1-3][1-8][01]\d)_' , xl_filename ).group(1)
    old_st_name = None
    for old_name in adv_dict : 
        if old_name.startswith(old_num) : 
            old_st_name = old_name 
            break 
    else : 
        #print "Number <%s> not found among old_names" % old_num
        print("Number <{}> not found among old_names".format(old_num))
        return False

    new_st_name = adv_dict[old_st_name]
    new_num, new_name = gen_util.num_name_split(new_st_name)
    input_filename = directory + "/" + xl_filename
    #new_filename = (output_directory +"/" + "%s_%s_수강신청계획서_%s_배포용.xlsx" % (new_num, new_name, sem.next_semester_prefix[:-1]))
    new_filename = (output_directory +"/" \
        + "{}_{}_수강신청계획서_{}_배포용.xlsx".format(new_num, 
            new_name, sem.next_semester_prefix[:-1]))
    

    #print ("Copying %s to %s" % ( input_filename, new_filename))
    print(("Copying {} to {}".format( input_filename, new_filename)))
    shutil.copyfile( input_filename, new_filename )
    
    wb = openpyxl.load_workbook(new_filename)
    ws =  wb[ unicode( "계획서", "utf-8") ]
    ws.cell(row=1, column=1).value = "수강신청 계획서(2019학년도 2학기)"
    ws.cell(row=42, column=1).value = "5. 엑셀2007 이상 버전을 사용할 것." +\
        " 제출시 파일명은 배포용->제출 으로 변경할 것."
    ws.cell(row=2, column=15).value = new_num 
    #ws.cell(row=2, column=26).value = new_name
    #patch_border.fix_borders_on_sheet(ws, style="medium", indices = indices )
    #patch_border.fix_plan_borders(ws)
    #print "Saving %s.." % new_filename
    print("Saving {}..".format(new_filename))
    wb.save( new_filename )

def plan_year_advance_from_directory(adv_dict, from_dir = "coursework-plan", 
        output_dir = "." ):
    fnames = [ c for c in os.listdir(from_dir) 
        if c.endswith(".xlsx") and not c.startswith("~")]
    for fname in fnames : 
        year_advance_coursework_plan(fname,  adv_dict, 
            directory=from_dir, output_directory = output_dir)

def read_english_tuple(xl_filename):
    rows = xlsx_util.read_xlsx_sheet_into_list(xl_filename, 
            sheet_name="수강신청") 
    output = [ ] 
    if len(rows[4][21])>0: 
        output.append( ('', rows[4][21], 4, 21))
    return output


def plan_form_input_data(xl_filename):
    reen_start_colnum = 21
    data = read_name_subject_relsems(xl_filename, use_filename=True, 
            with_cell_coords=True)
    st_name = data[0][0]
    new_num, new_name = gen_util.num_name_split(sg_name(st_name))
    name_inputs = [ ] 
    name_inputs.append( ('', new_num, 2,0) )
    name_inputs.append( ('', new_name, 2,1) )

    english_inputs = read_english_tuple(xl_filename)

    main_inputs = [ (c[1], c[2], c[3], c[4]) for c in data 
        if c[4]!=reen_start_colnum+1 and (c[1] not in sem.sem_fix_subjects)]
    reen_inputs = [ ('', c[1], c[3], c[4]) for c in data 
        if c[4]==reen_start_colnum+1 ]
    return (name_inputs + main_inputs + reen_inputs + english_inputs)


def copy_coursework_plan_into_new_form(xl_filename, 
        directory="plan-files", output_directory = "." ): 
    sr = re.search(  r'([1-3][1-8][012]\d)([^_]+)_' , 
            xl_filename )
    new_num, new_name = gen_util.num_name_split(sg_name(sr.group(1)))
    input_filename = directory + "/" + xl_filename
    #new_filename = (output_directory +"/" + "%s%s_%s_수강신청계획서.xlsx" % (new_num, new_name, name_code_dict[new_num + new_name]) )
    new_filename = (output_directory +"/"  \
        + "{}{}_{}_수강신청계획서.xlsx".format(new_num, 
            new_name, name_code_dict[new_num + new_name]) )


    input_data = plan_form_input_data(input_filename)

    if new_num.startswith("2") : 
        template_fname = "sophomore-plan-template.xlsx"
    elif new_num.startswith("3") : 
        template_fname = "senior-plan-template.xlsx"
    elif new_num.startswith("1") : 
        template_fname = "freshmen-plan-template.xlsx"
    else : 
        print(new_num, new_name , "not senior nor sophomore")
        return False

    #print ("Copying %s to %s" % ( template_fname, new_filename))
    print(("Copying {} to {}".format( template_fname, new_filename)))
    shutil.copyfile( template_fname, new_filename )
    
    wb = openpyxl.load_workbook(new_filename)
    ws =  wb[ unicode( "수강신청", "utf-8") ]
    for subj, val, i, j in input_data : 
        ws.cell(row=i+1, column=j+1).value = val

    #patch_border.fix_plan_borders(ws)
    #print "Saving %s.." % new_filename
    print("Saving {}..".format(new_filename))
    wb.save( new_filename )

    
def copy_coursework_plans_from_directory( from_dir = "tmp", output_dir = "." ):
    fnames = [ c for c in os.listdir(from_dir) 
        if c.endswith(".xlsx") and not c.startswith("~")]
    for fname in fnames : 
        copy_coursework_plan_into_new_form( fname, directory=from_dir, 
           output_directory = output_dir )  

def get_name_url_pairs(filename, salt = sem.salt ) : 
    output = [ ] 
    with open(filename, "rb") as f: 
        links = [ line.strip() for line in f ]
    for i,url in enumerate(links) :         
        print(i, end=" ")
        #print "Downloading from %s.." % url
        print("Downloading from {}..".format(url))
        response = urllib.urlopen(url)
        down_data = response.read().decode('utf-8') 
        #print "Downloaded %s.." % down_data[:40]
        print("Downloaded {}..".format(down_data[:40]))
        sr = re.search( r'([1-3][1-8][01]\d)([^_]+)_.*\.xlsx', down_data)
        st_name = sg_name( sr.group(1).encode("utf-8") )
        #print "Found %s (%s).." % (st_name, sr.group(2).encode("utf-8") )
        print("Found {} ({})..".format(st_name, sr.group(2).encode("utf-8") ))
        output.append( (st_name + "_" + 
          gen_util.password_of(st_name, salt = salt), url))
    return output

def add_ABCD_to_neis_data(isu_data): 
    output = [ ] 
    for c in isu_data: 
        r_name = sg_name( c[0] )
        if r_name is None or len(r_name) == 0  : 
            print("Somethings wrong..", end=" " )
            print(' '.join(c))
            return None
        elif r_name[-1] in "A B C D E".split() :
            #print "Fixing for %s" % ' '.join(c)
            print("Fixing for {}".format(' '.join(c)))
            output.append(  (r_name, ) + tuple( c[1:]))
        else : 
            output.append(c)
    return output

def change_subject_names( ns_data, change_dict = sem.special_lecture_dict): 
    output = [ ]
    for st_name, subj in ns_data : 
        if subj in change_dict : 
            output.append( (st_name, change_dict[subj]) )
        else : 
            output.append( (st_name, subj) )
    return output

            
def read_name_subjects_from_xlsx(filename, 
        change_dict = sem.special_lecture_formal_dict):
    data = list_util.transposed( xlsx_util.read_xlsx_sheet_into_list(filename))
    output = [ ] 
    for row in data : 
        subj = row[0]
        if len(subj)>0: 
            for st_name in row[1:] : 
                if len(st_name) > 0 : 
                    list_util.check_append(output, (st_name, subj), 
                        lambda x: ' '.join(x) ) 
    return change_subject_names(output, change_dict = change_dict)


def subjects_of(st_name, ns_pairs) : 
    return sorted( [c[1] for c in ns_pairs if c[0] == st_name] )

def members(subj, ns_pairs) : 
    return sorted( [c[0] for c in ns_pairs if c[1] == subj] )


def xlsx_semester_str(target_sem):
    year, sem_num = target_sem.split("_")
    output = f"{year}학년 {sem_num}학기"
    output = output.replace("winter", "겨울 계절")
    output = output.replace("summer", "여름 계절")
    return output


def fill_in_sugang_checks(new_worksheet, offset, st_name, ns_pairs, 
        date=None, hist_slash_phil=False, 
            target_semester=target_semesters[-1] ) :
    if date is None :
        date_str = datetime.datetime.today().strftime("%Y-%m-%d")
    else :
        date_str = date
    name_in_xlsx = gen_util.convert_name_for_xlsx(st_name)
    print("Writing sugang check for {}.".format(st_name))
    new_worksheet.cell( row = 3 + offset, column = 5 ).value = \
            xlsx_semester_str(target_semester) 
    new_worksheet.cell( row = 2 + offset, column = 5 ).value = name_in_xlsx
    new_worksheet.cell( row = 2 + offset, column = 2 ).value = date_str
    new_worksheet.cell( row = 19 + offset, column = 2 ).value = \
        xlsx_semester_str(target_semester) \
            + " 수강신청을 위와 같이 하였음을 확인합니다."
    temp_list = sorted( subjects_of(st_name, ns_pairs) )
    st_hour_list = [  updated_hours_dict[subj]  for subj in temp_list]
    st_size_list  = [ ]
    st_subj_list = [ ]
    for subj in temp_list :
        if hist_slash_phil is True and (subj in "철학 한국사".split()) and \
                    sem.next_semester_prefix[-2]=="1" :
            st_subj_list.append( "한국사/철학")
            if sem.is_spring_semester() :
                st_size_list.append( "({}명)".format(int(sem.num_juniors/2.0)))
            else :
                st_size_list.append( "({}명)".format(
                    int(sem.num_sophomores/2.0) ))
        else :
            st_subj_list.append( subj)
            #st_size_list.append( "(%d명)" % len(members(subj, ns_pairs)) )
            st_size_list.append( "({}명)".format(len(members(subj, ns_pairs))))

    xlsx_util.fill_in_sheet_by_list(new_worksheet, st_subj_list,
                5 + offset, 2, direction="column")
    xlsx_util.fill_in_sheet_by_list(new_worksheet, st_size_list,
                5 + offset, 5, direction="column")
    xlsx_util.fill_in_sheet_by_list(new_worksheet, st_hour_list,
                5 + offset, 6, direction="column")
    new_worksheet.cell( row = 19 + offset, column = 6 ).value = \
            ("합계: {} 시간".format(sum(st_hour_list) ))





def export_sugang_checks_of_all_classes( filename, nss_triples, 
        class_list = sem.ban_list, date = None, salt = None, 
            st_list = students, target_semester=target_semesters[-1] ) :
    print("Reading style from <sugang-check-template.xlsx>.")
    shutil.copyfile( "sugang-check-template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb["Template"]

    if salt is None :
        my_salt = datetime.datetime.today().strftime("%Y-%m-%d")
    else :
        my_salt = salt

    ns_pairs = [ (c[0], c[1]) for c in nss_triples if c[2] == target_semester]

    for clss in class_list :
        print("Writing subjects lists for {}.".format(clss))
        new_worksheet = wb.create_sheet(  clss)
        new_worksheet.print_area = "A1:E400"
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        prefix = clss.replace("-", "")
        name_list = [nm for nm in st_list if nm.startswith(prefix)]
        name_list.sort()
        for i, st_name in enumerate(name_list) :
            fill_in_sugang_checks(new_worksheet, i*20, st_name, 
                ns_pairs, date=date, target_semester=target_semester )
        for k in range( (i+1)*20 + 1, 401  ):
            new_worksheet.row_dimensions[k].hidden = True
        if len(my_salt)>0:
            new_worksheet.protection.password = \
                    gen_util.generate_password(my_salt,clss)
    wb.remove(template_ws)
    print("Saving to <{}>.".format(filename))
    wb.save(filename)


def change_into_isu_data( ns_pairs, sem_str, 
        change_dict = sem.special_lecture_isupyo_dict  ) : 
    tmp_pairs = change_subject_names(ns_pairs, change_dict = change_dict)
    output = [ (c[0], sem_str, c[1]) for c in tmp_pairs]
    return output

def export_moodle_users_csv(filename, st_list = None, salt = sem.salt): 
    cohort_name = "enrollment_" + sem.semester_prefix[:-1]
    if st_list is None : 
        in_st_list = [ st_name for st_name in t.students 
            if st_name[0] in "1 2".split() ]
    else : 
        in_st_list = st_list[:]
    first_row = "username password firstname lastname email cohort1".split()
    output = [ first_row ] 
    for st_name in in_st_list : 
        num, name = gen_util.num_name_split(st_name)
        username = english_id(st_name)
        firstname = num 
        lastname = name 
        email = username + "@lms.savano.org"
        password = gen_util.password_of(username, salt = salt) 
        output.append( [username,password, firstname, lastname, 
            email, cohort_name])
    with open(filename, "wb") as f: 
        #print "Saving %s" % filename
        print("Saving {}".format(filename))
        for rec in output : 
            f.write( ",".join(rec) + "\n")
        
        

def get_filename_url_pairs(filename, salt = sem.salt ) : 
    output = [ ] 
    with open(filename, "rb") as f: 
        links = [ line.strip() for line in f ]
    for i,url in enumerate(links) :         
        print(i, end=" ")
        #print "Downloading from %s.." % url
        print("Downloading from {}..".format(url))
        response = urllib.urlopen(url)
        down_data = response.read().decode('utf-8') 
        #print "Downloaded %s.." % down_data[:40]
        print("Downloaded {}..".format(down_data[:40]))
        sr = re.search( r'<title>(.*?\.xlsx)', down_data)
        try : 
            f_name =  sr.group(1).encode("utf-8") 
            #print "Found %s .." % f_name
            print("Found {} ..".format(f_name))
            output.append( (f_name + 
                gen_util.password_of(f_name, salt = salt), url))
        except Exception as e: 
            print(e)
    return output


def english_id(st_name): 
    num, name = gen_util.num_name_split(st_name)
    syllables = gen_util.korean_syllables(re.sub(r"[ABCD]+","",name))
    return (num + ''.join( map( gen_util.dict_to_fn( sem.given_name_conv), 
        syllables[1:] )) +  sem.last_name_conv[syllables[0]])

def print_duplicate_name_subjects(name_sem_subjs):
    output = [] 
    for name, sem, subj in name_sem_subjs: 
        search_result = [rec for rec in output 
                if rec[0]==name and rec[2]==subj]
        if len(search_result) > 0 : 
            print(name, sem, subj)
            for rec in search_result:
                print(' '.join(rec))
        output.append( (name, sem, subj) )

def make_code_message_dict(st_names, salt = sem.salt):
    output = { }
    for st_name in st_names:
        code = gen_util.generate_password(salt, st_name, decimal=True)
        num_str, name_str = gen_util.num_name_split(st_name)
        #msg = ("%s 학생의 본인확인코드는 %s 입니다. " + "수강신청을 할 때 꼭 이 코드를 사용해야 합니다.") % (name_str, code)
        msg = ("{} 학생의 본인확인코드는 {} 입니다. " + \
          "수강신청계획서 파일명을 {}_{}_수강신청계획서.xlsx 와 같이 하여 제출하세요.").format(name_str, 
                        code, st_name, code)
        output[st_name] = msg 
    return output

def is_correct_filename(filename, salt = sem.salt, strict=False):
    sr = re.search( r'^(\d{4}).+(\d{6}).*', filename)
    if sr is None : 
        return False
    st_name = sg_name(sr.group(1))
    if st_name is None or len(st_name) == 0 : 
        return False
    if gen_util.generate_password(salt, st_name, decimal=True) == sr.group(2):
        if strict is False : 
            return True
        else : 
            if filename ==  st_name + "_" + sr.group(2) \
                    + "_수강신청계획서.xlsx" : 
                return True
            else : 
                return False
    else :
        return False

code_name_dict = dict( [ (gen_util.generate_password(sem.salt, 
            st_name, decimal=True), st_name) for st_name in students])
name_code_dict = dict( [ ( st_name, gen_util.generate_password(sem.salt, 
            st_name, decimal=True)) for st_name in students])
fd_name_dict = dict( [ (c[2], c[0]) for c in sem.st_info_tuples ])
#code_name_dict["999852"] = "1420김종우" 
#name_code_dict["1420김종우" ] = "999852"
#fd_name_dict["19023"] = "1420김종우" 

def fname_correction_pair(filename, isupyo=False):
    if isupyo is False : 
        postfix_str = "_수강신청계획서.xlsx"
    else : 
        postfix_str = "_교육과정_이수현황표.xlsx"
    sr = re.search(r'(\d{6})',filename)
    if sr and (sr.group(1) in code_name_dict ) :
        code = sr.group(1)
        new_name = code_name_dict[code] + "_" + code + postfix_str 
        return (filename, new_name)
    else :
        dsr = re.search(r'([12]\d{4})',filename)
        if dsr and ( dsr.group(1) in fd_name_dict ) :
            st_name = fd_name_dict[ dsr.group(1) ]
            code = name_code_dict[st_name] 
            new_name = st_name  + "_" + code + postfix_str 
            return (filename, new_name)
        #print "Now using st_num for %s" % filename
        print("Now using st_num for {}".format(filename))
        dsr = re.search(r'([123][1-8][012]\d{1})',filename)
        if dsr : 
            candis = list_util.prefix_filter(dsr.group(1), 
                                        name_code_dict.keys())
        if dsr and candis != []:
            st_name = candis[0]
            code = name_code_dict[st_name] 
            new_name = st_name  + "_" + code + postfix_str 
            return (filename, new_name)
        else :
            print("Wrong filename " + filename)
            return (filename, None)

def rename_by_pairs(fname_pairs, directory="plan-files"):
    for fname, new_name in fname_pairs : 
        full_name = directory + "/" + fname
        full_new_name = directory + "/" + new_name
        #print "Renaming <%s> -> <%s>" % ( full_name, full_new_name)
        print("Renaming <{}> -> <{}>".format( full_name, full_new_name))
        shutil.move(full_name, full_new_name)




def copy_plan_files(from_dir = "plan-files", to_dir = dropbox_dir ):
    xlsx_filenames = [c for c in os.listdir(from_dir) if c.endswith(".xlsx")]
    output = [ ] 
    for fname in xlsx_filenames : 
        new_name = re.sub(r'_\d{6}_', '_', fname)
        output.append( (from_dir + "/" + fname, 
            to_dir + "/" + new_name) )
    for pair in output : 
        #print "Copying %s to %s" % pair 
        print("Copying {} to {}".format(pair[0], pair[1]))
        shutil.copyfile(pair[0], pair[1])
        
def ns_pairs_from_name_subjects(my_tuples):
    output = [ ] 
    for rec in my_tuples : 
        st_name = rec[0]
        for subj in rec[1:] : 
            if len(subj)>0 : 
                output.append( (st_name, subj) )
    return output

def rename_pairs(dir_name, osx = True ):
    f_names = [c for c in os.listdir(dir_name) 
                if c.endswith(".xlsx") and (not c.startswith("~")) ]
    output_pairs = [fname_correction_pair(c) for c in f_names]
    if osx : 
        conv_func = gen_util.nfd_to_nfc 
    else : 
        conv_func = (lambda x : x)
    convd_pairs = [ (conv_func(c[0]), c[1]) for c in output_pairs]
    filtered_pairs = [ c for c in convd_pairs if c[0] != c[1] ]
    return filtered_pairs
        

def export_subject_selection(nss_triples, ban_str=None, 
        target_semesters=target_semesters):
    for sem_str in target_semesters : 
        ns_pairs = [ (c[0], c[1]) for c in nss_triples if  c[2] == sem_str]

        if ban_str is None : 
            filename = "{}-수강신청목록-{}.xlsx".format( sem_str, 
                datetime.date.today().strftime( "%Y%m%d"))
            restrict_pairs = list(
                filter(lambda pair: (not pair[0].startswith("0")), ns_pairs))
            tmp_tuples = list_util.collect_on_first(restrict_pairs)
            tmp_tuples.sort()

        else : 
            filename = "{}-{}-{}-수강신청목록.xlsx".format(ban_str[0], 
                ban_str[1], sem_str )
            restrict_pairs = list(filter(lambda pair: pair[0].startswith(
                    ban_str), ns_pairs))
            tmp_tuples = list_util.collect_on_first(restrict_pairs)
            tmp_tuples.sort()

        data = [ [b[0],  sum( [updated_hours_dict[c] for c in b[1:]]) ] + 
                [d + "(" + str(updated_hours_dict[d]) + ")" for d in b[1:] ] 
                for b in tmp_tuples]
        

        template_filename = "subject_selection_template.xlsx"
        print("Reading style from <{}>".format(template_filename))
        shutil.copyfile( template_filename, filename)
        print("Copied {} to {}".format( template_filename, filename))
        wb = openpyxl.load_workbook(filename)
        ws = wb[  "신청목록" ]
        xlsx_util.fill_in_sheet_by_tuples(ws, data, 2, 1)
        print("Saving to <{}>..".format(filename))
        wb.save( filename )

    


def read_xlsx_files_in_directory(dir_name, with_filename = False) : 
    fnames = [ os.path.join(dir_name, c) for c in os.listdir(dir_name) 
        if c.endswith(".xlsx") and c[0]!="~"]
    output = [ ]
    for fn in fnames : 
        print("Reading <{}>..".format(fn))
        if with_filename : 
            output.extend( [ [fn]]  )
        output.extend(xlsx_util.read_xlsx_sheet_into_list(fn, 
            convert_fn= (lambda x: '' if x is None else str(x))))
    return output


def export_name_english_supplements(dir_name = dropbox_dir, verbose = True):    
    fnames = [dir_name + "/" + c for c in os.listdir(dir_name) 
        if c.endswith(".xlsx") and (not c.startswith("~")) and c[0] == "3" ]
    output = []
    for filename in fnames : 
        short_fname = os.path.split(filename)[1]
        bar_position = short_fname.find("_")
        st_name = short_fname[:bar_position]
        data = xlsx_util.read_xlsx_sheet_into_list(filename)
        english_subj = data[4][21]
        if english_subj is None : 
            english_subj = ''
        if verbose : 
            print( f"Appending ({st_name}, {english_subj})..")
        output.append( (st_name, english_subj))

    xlsx_util.write_tuples_into_xlsx("영어대체과목_신청목록.xlsx", output)
    

def read_subject_names(hyeonhwang_filename, 
        target_semester = target_semesters[-1] ):
    data = xlsx_util.read_xlsx_sheet_into_list(hyeonhwang_filename, 
        sheet_name= target_semester )
    subj_names = list_util.transposed(data)
    output = [ ] 
    for rec in subj_names : 
        subj = rec[0]
        for st_name in rec[1:]:
            if len(st_name)>0 : 
                output.append( (st_name, subj) )
    return output

def export_sugang_update_records(filename_one, filename_two):
    diff_filename = "diff-" + filename_one.replace(".xlsx",'') + "-" + \
                        filename_two
    data_one = read_subject_names(filename_one)
    data_two = read_subject_names(filename_two)
    output = pair_differences(data_one, data_two)
    export_differences_into_xlsx(diff_filename, output)


if __name__ == "__main__" : 
    t.load()
    importlib.reload(t)
    students = t.students + on_leave

    code_name_dict = dict( [ (gen_util.generate_password(sem.salt, 
            st_name, decimal=True), st_name) for st_name in students])
    name_code_dict = dict( [ ( st_name, gen_util.generate_password(sem.salt, 
            st_name, decimal=True)) for st_name in students])
    fd_name_dict = dict( [ (c[2], c[0]) for c in sem.st_info_tuples ])

    for k in range(10000) : 
        if check_for_changes() : 
            try: 
                read_and_save_coursework_plans() 
                update_flist()
            except Exception as e : 
                print(e )
            print("Waiting for another update..")
        time.sleep(300)
        
        



