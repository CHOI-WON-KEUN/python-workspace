
import openpyxl, re 
import shutil
import list_util as l 

def strip_typeinfo(c, utf_input=True ) :
    if type(c) in [float, int ] :
        return str(c)
    else : 
        #if type(c) is unicode :
            #utf_c = c.replace(u'\ufeff', '').encode("utf-8")
        #elif type(c) is str : 
            #utf_c = c
        if utf_input : 
            if c is None : 
                return ''
            if type(c) is str : 
                uni_c = c
            else :
                uni_c = c.decode()
            if uni_c.startswith("=") : 
                return uni_c
            else : 
                return re.sub(r'^[^\:]+:', '',  uni_c)
        else : 
            if type(c) is str : 
                utf_c = c.encode()
            else : 
                return '' 
            if utf_c.decode().startswith("=") : 
                return utf_c
            else : 
                return re.sub(r'^[^\:]+:'.encode(), '',  utf_c)




def read_xlsx_sheet_into_dict( filename, 
            head_row_ind, head_column_ind, sheet_name=None, 
            convert_fn = strip_typeinfo ) : 
    ''' read xlsx file <filename>, read entries of head_row th row, 
        read entries of head_column th column, 
        read non-empty cells in str format and save them in 
        a dict in { (row_name, col_name): val } format.'''
    wb = openpyxl.load_workbook( filename ) 
    if sheet_name is not None : 
        #ws = wb.get_sheet_by_name( unicode(sheet_name , 'utf-8') )
        ws = wb[ sheet_name  ]
    else :
        ws = wb.active 
    main_row = [ ( convert_fn(c.value), j+1 ) 
        for j, c in  enumerate(list(ws.rows)[ head_row_ind  - 1 ] ) 
        if j+1 > head_column_ind ]
    main_column = [ (convert_fn(row[head_column_ind - 1].value), i+1) 
                        for i, row in enumerate(ws.rows) 
                        if i+1 > head_row_ind ]
    
    output = { }

    for c_name, j in main_row : 
        for r_name, i in main_column : 
            val = convert_fn( ws.cell(row=i, column=j).value)
            if len(val) > 0 : 
                output[ (r_name, c_name) ] = val

    return output
    

def write_dict_into_xlsx(filename, array_dict, in_column_names = None, 
                   in_row_names = None , sheetname = None, conv_func = None ) : 
    if conv_func is None : 
        conv_func = ( lambda x: x )
    wb = openpyxl.Workbook()
    ws = wb.active 
    if sheetname is not None : 
        ws.title =  sheetname 
    else : 
        ws.title =  "Sheet1"


    if in_column_names is not None : 
        column_names = in_column_names 
    else : 
        column_names = sorted( l.union( [ c[1] for c in array_dict ] ) )

    if in_row_names is not None : 
        row_names = in_row_names 
    else : 
        row_names = sorted( l.union( [ c[0] for c in array_dict ] ) )
        

    main_row = [ (c_name, i+2) for i,c_name in enumerate(column_names) ] 
    main_column = [ (r_name, j+2) for j,r_name in enumerate(row_names) ] 
    for c_name, j in main_row : 
        ws.cell(row = 1, column = j).value = c_name
    for r_name, i in main_column : 
        ws.cell(row = i, column = 1).value = r_name
    for c_name, j in main_row : 
        for r_name, i in main_column : 
            if (r_name, c_name) in array_dict : 
                cell_val = array_dict[(r_name, c_name)]
                if len( str( cell_val ) ) > 0 : 
                    _ = ws.cell(row = i, column = j, 
                                value = conv_func( cell_val  )  )
    #print "Saving to <%s>.." % filename 
    print("Saving to <{}>..".format(filename))
    wb.save( filename )



def fill_in_xlsx_by_dict(output_filename, template_filename, array_dict, 
    row_ind=1, column_ind=1, sheet_name = None, convert_fn = strip_typeinfo ) : 
    #print "Reading style from <%s>" % template_filename
    print("Reading style from <{}>".format(template_filename))
    shutil.copyfile( template_filename, output_filename)
    #print "Copied %s to %s" %( template_filename, output_filename)
    print("Copied {} to {}".format( template_filename, output_filename))
    wb = openpyxl.load_workbook(output_filename)
    if sheet_name is not None : 
        ws = wb[ sheet_name  ]
    else :
        ws = wb.active 
    main_row = [ ] 
    main_column = [ ] 
    for i, row in enumerate(ws.rows) : 
        if i == row_ind - 1 : 
            for j, entry in enumerate(row) : 
                main_row.append( (j+1, convert_fn(entry.value)) )
        main_column.append( (i+1, convert_fn(row[ column_ind - 1].value)))
        
    for j, n in main_row : 
        for i, sl in main_column : 
            if (sl, n) in array_dict : 
                ws.cell( row=i, column=j ).value = array_dict[ (sl, n) ]
                #print "Writing (%s, %s) -> %s" % (sl, n, array_dict[ (sl,n)])
                print("Writing ({}, {}) -> {}".format(sl, 
                    n, array_dict[ (sl,n)]))
    #print "Saving to <%s>.." % output_filename
    print("Saving to <{}>..".format(output_filename))
    wb.save( output_filename )
    

def fill_in_sheet_by_list( work_sheet, input_list, 
        row_ind=1, column_ind=1, direction = "column" ) : 
        if direction == "column" : 
            for i, item  in enumerate(input_list) : 
                work_sheet.cell( row=i+row_ind, column=column_ind  ).value = item
                #print "Writing (%d, %d) -> %s" % (i+row_ind,column_ind, item)  
                print("Writing ({}, {}) -> {}".format(i+row_ind, 
                    column_ind, item))
        else : 
            for j, item  in enumerate(input_list) : 
                work_sheet.cell( row=row_ind, column=column_ind + j  ).value = item
                #print "Writing (%d, %d) -> %s" % (row_ind, j + column_ind, item)  
                print("Writing ({}, {}) -> {}".format(row_ind, 
                    j + column_ind, item))


def fill_in_sheet_by_tuples( work_sheet, input_tuples, row_ind=1, column_ind=1 ) : 
    for i, rec  in enumerate(input_tuples) : 
        for j, item in enumerate(rec) : 
            if len(str(item))>0 and item is not None: 
                try:
                    #work_sheet.cell( row=i+row_ind, 
                        #column=j+column_ind  ).value = item
                    #print("Writing ({}, {}) -> {}".format(i+row_ind, 
                        #j+column_ind, str(item)))
                #except AttributeError as e : 
                    #print( "cell({},{})".format(i+row_ind, j+column_ind))
                    #print(e)
                work_sheet.cell( row=i+row_ind, 
                    column=j+column_ind  ).value = item
                print("Writing ({}, {}) -> {}".format(i+row_ind, 
                    j+column_ind, str(item)))


def fill_in_xlsx_by_tuples(output_filename,  template_filename, tuple_list, 
        row_ind=1, column_ind=1) : 
    #print "Reading style from <%s>" % template_filename
    print("Reading style from <{}>".format(template_filename))
    shutil.copyfile( template_filename, output_filename)
    #print "Copied %s to %s" %( template_filename, output_filename)
    print("Copied {} to {}".format( template_filename, output_filename))
    wb = openpyxl.load_workbook(output_filename)
    ws = wb.active 
    fill_in_sheet_by_tuples(ws, tuple_list, row_ind=row_ind, 
        column_ind=column_ind)
    #print "Saving to <%s>.." % output_filename
    print("Saving to <{}>..".format(output_filename))
    wb.save( output_filename )




def fill_in_sheet_by_dict(ws,  array_dict, row_ind=1, column_ind=1, 
        sheet_name = None) : 
    main_row = [ ] 
    main_column = [ ] 
    for i, row in enumerate(ws.rows) : 
        if i == row_ind - 1 : 
            for j, entry in enumerate(row) : 
                main_row.append( (j+1, strip_typeinfo(entry.value)) )
        main_column.append( (i+1, strip_typeinfo(row[ column_ind - 1].value)))
    for j, n in main_row : 
        for i, sl in main_column : 
            if (sl, n) in array_dict and len(str(array_dict[(sl,n)]))>0 : 
                ws.cell( row=i, column=j ).value = array_dict[ (sl, n) ]
                print("Writing ({}, {}) -> {}".format(sl, n, 
                    array_dict[ (sl,n)]))
    

def fill_in_sheet_by_list( work_sheet, input_list, 
        row_ind=1, column_ind=1, direction = "column" ) : 
        if direction == "column" : 
            for i, item  in enumerate(input_list) : 
                if item is not None and len(item)>0 : 
                    work_sheet.cell( row=i+row_ind, column=column_ind  ).value = item
                    print("Writing ({}, {}) -> {}".format(i+row_ind, 
                        column_ind, item))
        else : 
            for j, item  in enumerate(input_list) : 
                if item is not None and len(item)>0 : 
                    work_sheet.cell( row=row_ind, column=column_ind + j  ).value = item
                    print("Writing ({}, {}) -> {}".format(row_ind, 
                        j + column_ind, item))




def write_list_into_xlsx(filename, my_list, 
                column_name = None, sheetname = None, conv_func = None ) : 
    if conv_func is None : 
        conv_func = ( lambda x: x )
    wb = openpyxl.Workbook()
    ws = wb.active 
    if sheetname is not None : 
        ws.title =  sheetname
    else : 
        ws.title =  "Sheet1"
    if column_name is not None : 
        ws.cell(row = 1, column = 1).value = column_name 
    for i, entry in enumerate(my_list) : 
        ws.cell(row = i+2 , column = 1).value = entry
    #print "Saving to <%s>.." % filename 
    print("Saving to <{}>..".format(filename))
    wb.save( filename )



def write_tuples_into_xlsx(filename, my_tuples, 
                in_column_names = None, sheetname = None, conv_func = None ) : 
    if conv_func is None : 
        conv_func = ( lambda x: x )
    wb = openpyxl.Workbook()
    ws = wb.active 
    if sheetname is not None : 
        ws.title =  sheetname
    else : 
        ws.title =  "Sheet1"
    if in_column_names is not None : 
        column_names = in_column_names 
    elif in_column_names == "first row" : 
        column_names = my_tuples[0]
    else : 
        column_names = [ "Column" +  str(j)  
                for j in range(1, max( map(len, my_tuples)) + 1)]
    if in_column_names != "first row" : 
        for i, c_name in enumerate(column_names) : 
            ws.cell(row = 1, column = i+1).value = c_name
        for i, entries in enumerate(my_tuples) : 
            for j, entry in enumerate(entries) : 
                ws.cell(row = i+2 , column = j+1).value = entry
    else : 
        for i, entries in enumerate(my_tuples) : 
            for j, entry in enumerate(entries) : 
                ws.cell(row = i+1 , column = j+1).value = entry
    #print "Saving to <%s>.." % filename 
    print("Saving to <{}>..".format(filename))
    wb.save( filename )




def column_names_in( my_dict ) : 
    return l.union( [ c[1] for c in my_dict ] )
                        
def row_names_in( my_dict ) :                     
    return l.union( [ c[0] for c in my_dict ] )



def read_xlsx_sheet_into_list( filename, 
            head_row_ind = 1, head_column_ind = 1, sheet_name=None, convert_fn=strip_typeinfo ) : 
    ''' read xlsx file <filename>, get a list of lists including the 
        head row and the head column entries'''
    wb = openpyxl.load_workbook( filename ) 
    if sheet_name is not None : 
        ws = wb[  sheet_name  ]
    else :
        ws = wb.active 
    output = [ ]
    for j, c in  enumerate(list(ws.rows) ) :
        if j >= head_row_ind - 1 : 
            c_ind = head_column_ind - 1 
            output.append( [convert_fn(entry.value) 
                                        for entry in c[ c_ind : ]])
    return output




def read_xlsx_partial_table_into_dict( worksheet, 
            head_row_ind, head_row_len, head_column_ind, head_column_len) : 
    ''' read part of worksheet to save it into a dict and return the dict 
            of the form { (row_name, col_name): val } 
        empty head_row, head_columns are ignored ''' 
    ws = worksheet
    main_row = [ ( strip_typeinfo(c.value), j+1 ) 
        for j, c in  enumerate(list(ws.rows)[ head_row_ind  - 1 ] ) 
        if j+1 > head_column_ind and j + 1 <= head_column_ind + head_row_len ]
    main_column = [ (strip_typeinfo(row[head_column_ind - 1].value), i+1) 
            for i, row in enumerate(ws.rows) 
            if i+1 > head_row_ind and i+1 <= head_row_ind + head_column_len ]
    output = { }
    for c_name, j in main_row : 
        for r_name, i in main_column : 
            val = strip_typeinfo( ws.cell(row=i, column=j).value)
            if len(val) > 0 and len(c_name) > 0 and len(r_name) > 0 : 
                output[ (r_name, c_name) ] = val
    return output
    
def read_txt_into_list(filename) : 
    output = [ ] 
    with open(filename) as f : 
        for line in f : 
            output.append(line.strip())
    return output


def export_txt_into_xlsx(filename, xl_filename) : 
    data = read_txt_into_list(filename)
    write_list_into_xlsx(xl_filename, data)




def read_xlsx_sheets_into_list( filename, 
            head_row_ind = 1, head_column_ind = 1, convert_fn = strip_typeinfo ) : 
    ''' read xlsx file <filename>, get a dict sheetname -> lists including the 
        head row and the head column entries'''
    wb = openpyxl.load_workbook( filename ) 
    sheet_list_dict = { }
    for ws in wb : 
        #print sheet_name.replace("\\u", "\u") 
        #ws = wb[sheet_name]
        output = [ ]
        for j, c in  enumerate(list(ws.rows) ) :
            if j >= head_row_ind - 1 : 
                c_ind = head_column_ind - 1 
                output.append( [convert_fn(entry.value) 
                                        for entry in c[ c_ind : ]])
        sheet_list_dict[ws.title.encode("utf-8")] = output 
    return sheet_list_dict

