# -*- coding: utf-8 -*-

from __future__ import print_function

import openpyxl, re, os, pickle, shutil, random

from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.worksheet.pagebreak import Break

import hashlib
import functools
import itertools

import xlsx_util, list_util,  utf_util 
import gen_util


import math, datetime

import semester as sem 






def bujang_str(t_name) : 
    output = [ ] 
    for k in sem.bujang_dict : 
        if t_name == sem.bujang_dict[k] : 
            output.append(k)
    output.sort()
    output.reverse()
    if len(output) == 0 : 
        return ''
    else : 
        return '\n'.join(output)










            










try : 
    name_classparts
except NameError : 
    name_classparts = [ ] 
    name_subjects = [ ] 
    juniors = [ ] 
    classparts = [ ] 
    classpart_slot = { }
    teachers = [ ] 
    neis_subject_name_pairs = [ ] 
    classrooms = { }
    students = [ ] 

import sugang


def fill_merged_cells( tuples, fill_idx, ref_idx ) : 
    output = [ ] 
    for i, rec in enumerate(tuples) : 
        if i == 0 : 
            output.append(rec) 
        else : 
            if len(rec[fill_idx]) == 0 and len( output[-1][fill_idx] ) > 0 \
               and len( output[-1][ref_idx] ) > 0 and len( rec[ref_idx] ) > 0: 
                new_rec = rec[:]
                new_rec[fill_idx] = output[-1][fill_idx]
                output.append(new_rec)
            else : 
                output.append(rec)
    return output
                    

def select_columns(tuples, indices) : 
    output = [ ] 
    for rec in tuples : 
        new_rec = [ ] 
        for idx in indices : 
            new_rec.append( rec[idx] )
        output.append(new_rec)
    return output

def read_teacher_classes(filename): 
    data = xlsx_util.read_xlsx_sheet_into_list(filename)
    filled_data = fill_merged_cells(data,2,4) 
    trimmed_data = select_columns(filled_data, [2,4] + range(5,20) )
    grep_data = filter( lambda x: len(x[0])>0 and len(x[1])>0 
        and (not x[1].startswith("연구활동")) 
        and (not x[2].startswith("분반별")), trimmed_data)
    return grep_data


def save_teacher_parts(filename, nsh_tuples): 
    name_parts = [ ] 
    for rec in nsh_tuples : 
        t_name = rec[0]
        subj = re.sub( r'\(.+\)','',  utf_util.convert_romans(rec[1]))
        class_nums = [ i - 1 for i in range(2,17) if len(rec[i])>0 ] 
        for cnum in class_nums :
            for pnum in range(1, hours(subj)+1): 
                #name_parts.append( (t_name, "%s_%d$%d" % (subj, cnum, pnum)))
                name_parts.append((t_name, "{}_{}${}".format(subj, cnum, pnum)))
    teacher_list = list_util.union( [rec[0] for rec in nsh_tuples])
    output = list_util.collect_on_first(map(tuple,map( reversed, name_parts )))
    output.sort( key = ( lambda x: (teacher_list.index(x[1]), x[0])) ) 
    xlsx_util.write_tuples_into_xlsx(filename, output)
            


def add_teacher_to_part( t_name, pt ) : 
    #print "Adding %s to %s .." %(t_name, pt)
    print("Adding {} to {} ..".format(t_name, pt))
    check_append(name_classparts, (t_name, pt))

def add_teacher_to_class(t_name, clss) : 
    #print "%s -> %s .. " % (t_name, clss)
    print("{} -> {} .. ".format(t_name, clss))
    for pt in parts_of(clss) : 
        add_teacher_to_part(t_name, pt)

def remove_teacher_from_part(t_name, pt):
    #print "Removing %s from %s .." %(t_name, pt)
    print("Removing {} from {} ..".format(t_name, pt))
    check_remove(name_classparts, (t_name, pt), pair_str)

def remove_teacher_from_class(t_name, clss) : 
    #print "Removing %s from %s .. " % (t_name, clss)
    print("Removing {} from {} .. ".format(t_name, clss))
    for pt in parts_of(clss) : 
        remove_teacher_from_part(t_name, pt)


def move_teacher(t_name, from_clss, to_clss):
    remove_teacher_from_class(t_name, from_clss) 
    add_teacher_to_class(t_name, to_clss)


def read_baedang_data(filename) :
    ''' read [statistic.xlsx] "수업배당" and return a list of 
         (  teachername+"t", subj_clnum$time) pairs '''
    baedang_dict = xlsx_util.read_xlsx_sheet_into_dict(
                        filename, 1, 1, "수업배당")
    output = [ ]
    for clss_time, col_name in baedang_dict :
        if col_name.startswith("담당교사") :
            t_name  = baedang_dict[ (clss_time,col_name) ]
            if len(t_name) > 0 :
                if t_name.endswith("조교") :
                    output.append( (t_name + "a", clss_time ) )
                else :
                    output.append( (t_name + "t", clss_time ) )
                print(output[-1][0], output[-1][1])
    return output



def read_teacher_assignment( filename ) : 
    print("Removing teacher assignment data.." )
    t_data = [ c for c in name_classparts 
                if c[0].endswith("t") or  c[0].endswith("a") ] 
    for c in t_data : 
        name_classparts.remove(c)
        #print "%s %s is removed." % (c[0], c[1]) 
        print("{} {} is removed.".format(c[0], c[1]))
    print() 
    #print "Loading <%s>.." % filename
    print("Loading <{}>..".format(filename))
    name_classparts.extend( read_baedang_data(filename) )
    

def read_enrollments( filename ) : 
    global name_subjects 
    name_subjects = [ ] 
    enroll_dict = sugang.read_enrollment_data( filename ) 
    counter = 0
    for i, subj in enroll_dict : 
        name = enroll_dict[ (i, subj) ]
        name_subjects.append( (name, subj) )
        counter += 1 
    #print "%d many (name, subj) are read." % counter 
    print("{} many (name, subj) are read." % counter)



def classparts_of(name) : 
    return t_sorted([c[1] for c in name_classparts if c[0] == name ]) 

def members(classpart_or_class) : 
    input_str = classpart_or_class
    cat = category(input_str)
    if cat is "part" : 
        return list_util.collect_items( name_classparts, 
                    fn = lambda x: x == input_str)
    elif cat is "class" : 
        return sorted(list_util.union(list_util.collect_items( name_classparts, 
                    fn = lambda x: x.startswith(input_str + "$"))))
    elif cat is "subject" : 
        return list_util.collect_items( name_subjects, 
                    fn = (lambda x: x == input_str))
    else : 
        return None




def belongs_to(part_or_clss, clss_or_subj) : 
    if part_or_clss.startswith( clss_or_subj + "$" ) or \
            part_or_clss.startswith( clss_or_subj + "_" ) or \
            part_or_clss == clss_or_subj : 
        return True 
    else : 
        return False


def st_members_of(classpart_or_clss) : 
    return sorted(list_util.union([c[0] for c in name_classparts 
        if belongs_to(c[1],  classpart_or_clss )
                      and (not (c[0].endswith("t") or c[0].endswith("a")))] ))

def t_sorted(t_names) : 
    return sorted( t_names, key = (lambda x: (len(x), x)))


def t_members_of(classpart_or_clss) : 
    if classpart_or_clss.find("_") < 0 : 
        output = [ ] 
        subj = classpart_or_clss
        for c in classes_of(subj) : 
            output.extend( [ n for n in members(c) if n.endswith("t")])
        return t_sorted(list_util.union(output))
    else : 
        return t_sorted( [c for c in members(classpart_or_clss) 
                    if c.endswith("t") ] )


def subjects_of(a_name) : 
    if a_name in teachers : 
        return list_util.union( 
            [subject_name_of(c) for c in current_classes(a_name)]) 
    else : 
        return [subj for n, subj in name_subjects if n == a_name]


def used_slots_of(name, skip = [ ] ) : 
    pts = [pt for pt in classparts_of(name) if subject_name_of(pt) not in skip]
    return [classpart_slot[pt] for pt in pts ]


def available_slots_of(classpart) : 
    t_st_mems = members(classpart) 
    tmp_slots = list_util.union( list_util.flatten( [ used_slots_of(name) 
                    for name in t_st_mems] ), [ classpart_slot[classpart] ]  )
    return list_util.complement( slots, tmp_slots )

def used_slot_str(my_slots) : 
    output = ' '.join( map( lambda x: x if x in my_slots else "빔_" , 
                        slots ))
    return output

def teacher_slot_trans(t_name) : 
    my_slots = used_slots_of(t_name)
    if t_name in sem.teacher_avoid_slots_dict : 
        a_slots = sem.teacher_avoid_slots_dict[t_name] 
    else : 
        a_slots = [ ] 
    output_fn = (lambda x: x if x in my_slots 
                    else ( "___" if x in a_slots else "빔_"))
    return output_fn

def teacher_used_slot_str(t_name) : 
    output = ' '.join( map(  teacher_slot_trans(t_name) , slots ))
    return output

def slots_in_tabbing(non_used_slots, cl_slots = [ ], count_dict = { } ) : 
    output = [ ] 
    for sl in slots : 
        if sl in non_used_slots : 
            output.append( sl )
        elif sl in cl_slots : 
            output.append( "_익" )
        elif sl in count_dict : 
            #output.append( "%03d" % count_dict[sl] )
            output.append( "{:03d}".format(count_dict[sl]) )
        else : 
            output.append( "___" )
    return ' '.join(output)




def remaining_shorthands(st_name) : 
    output = remaining_subjects(st_name)
    output.sort(key = sem.ordered_subjects.index)
    return ''.join( [sem.shorthands[c] for c in output])


def class_shorthands(x):
    return  (sem.shorthands[subject_name_of(x)]+ class_number_of(x) )

def subjects_in_tabbing( st_name , cl_slots = [ ], c_num=False  ) : 
    output = [ ] 
    st_slots = used_slots_of(st_name)
    for sl in slots : 
        if sl in st_slots : 
            if c_num is True:
                output.append( class_shorthands( class_name_of(
                    name_slot_part(st_name, sl))))
            else: 
                output.append( sem.shorthands[ subject_name_of(
                    name_slot_part(st_name, sl))])
        else : 
            if c_num is True:
                output.append( "----/") 
            else: 
                output.append( "---/") 
    return ''.join(output)




def korean_ljust(my_str, n) : 
    curr_len = len( re.sub( r'\W{3}', 'aa', my_str ) ) 
    padding = ' ' *  max( n - curr_len, 0 ) 
    return (my_str + padding) 
    
    


def show_possible_slots(name_or_subj, t_num = None, slot_list = None, 
        exchange=False) : 
    if name_or_subj in names : 
        t_name = name_or_subj 
        print(korean_ljust(t_name, 24), ":",end=" " )
        print(teacher_used_slot_str( t_name))
        print()
        part_list = classparts_of(t_name)
    else : 
        if type(name_or_subj) is not list : 
            subj = name_or_subj 
            part_list = parts_of(subj)
        else : 
            part_list = functools.reduce( lambda x,y:x+y, 
                          [ sorted(parts_of(subj)) for subj in name_or_subj])  
        t_list = t_sorted(filter( lambda x:x.endswith("t"), 
                list_util.union(list_util.flatten( 
                        map( members, part_list )))))
        if t_num is not None : 
            if slot_list is None : 
                return show_possible_slots( t_list[ t_num - 1 ] )
            else: 
                if exchange is True : 
                    assign_fn = exchange_slots
                else : 
                    assign_fn = assign_slots
                if type(slot_list) is str : 
                    my_slot_list = slot_list.split()
                else : 
                    my_slot_list = slot_list
                return assign_fn(t_list[ t_num - 1 ], my_slot_list)
        list_util.jpr( [ '"' + c + '"' for c in t_list])

        a_list = sorted(filter( lambda x:x.endswith("a"), 
                list_util.union(list_util.flatten( 
                        map( members, part_list )))))
        list_util.jpr( [ '"' + c + '"' for c in a_list])

    for c in part_list  :
        print(korean_ljust(c + " " + str(len(st_members_of(c))), 20),end=" " )
        if classpart_slot[c] is not None : 
            print(":"+ classpart_slot[c] + ":",end=" ")
        else :
            print(":"+ "   " + ":",end=" ")
        print(slots_in_tabbing( available_slots_of(c), [classpart_slot[c]] ))


sps = show_possible_slots






def show_unused_slots_of_class(clss) : 
    for t_name in t_members_of(clss) : 
        print(korean_ljust(t_name, 24), ":",end=" " )
        print(teacher_used_slot_str( t_name))
    slot_list = [ ] 
    for st_name in st_members_of(clss) : 
        slot_list.append( (korean_ljust(st_name, 24), 
                    used_slot_str( used_slots_of(st_name)) ) )
    slot_list.sort(key = (lambda x:x[1]))
    for sl in slot_list : 
        print(sl[0], ":", sl[1])
    print()
    for c in parts_of(clss)  :
        print(korean_ljust(c, 20),end=" " )
        if classpart_slot[c] is not None : 
            print(":"+ classpart_slot[c] + ":",end=" ")
        else :
            print(":"+ "   " + ":",end=" ")
        print(slots_in_tabbing( available_slots_of(c), [classpart_slot[c]] ))


def show_slots_of_subject(subj, avoid_slot_list = [ ], c_num = None, 
        st_list = [ ] ) : 
    for t_name in t_members_of(subj) : 
        print(korean_ljust(t_name, 24), ":",end=" " )
        print(teacher_used_slot_str( t_name))
    slot_list = [ ] 
    slot_counter = { } 
    double_slot_counter = { }
    for sl in slots : 
        slot_counter[sl] = 0
        double_slot_counter[sl] = 0
    if c_num is None : 
        if len(st_list) == 0 : 
            disp_names = remaining_names_of(subj) 
        else : 
            disp_names = list_util.intersect( members(subj), st_list )
    else : 
        disp_names = st_members_of(subj + "_" + str(c_num)) 
    for st_name in disp_names : 
        st_slots = used_slots_of(st_name)
        if len( list_util.intersect( avoid_slot_list, st_slots ) ) == 0  : 
            slot_list.append( (st_name, used_slot_str( st_slots ) ) )
            for sl in slot_counter : 
                if sl not in st_slots : slot_counter[sl] += 1
            for i, sl in enumerate(slots[:-1]) : 
                if (sl not in st_slots) and (slots[i+1] not in st_slots) : 
                    double_slot_counter[sl] += 1

    slot_list.sort(key = (lambda x:[ len(used_slots_of(x[0])), x[1]]))

    for i, sl in enumerate(slot_list) : 
        print(str(i).rjust(2), korean_ljust(sl[0],20), ":", sl[1])

    print("  ".rjust(2), korean_ljust(" ",20), ":",end=" " )
    print(' '.join( [str( slot_counter[sl] ).ljust(3) for sl in slots ]))
    print("  ".rjust(2), korean_ljust(" ",20), ":",end=" " )
    print(' '.join( [str( double_slot_counter[sl] ).ljust(3) for sl in slots ]))
    print()

    for c in parts_of(subj)  :
        print(korean_ljust(c, 20),end=" " )
        if classpart_slot[c] is not None : 
            print(":"+ classpart_slot[c] + ":",end=" ")
        else :
            print(":"+ "   " + ":",end=" ")
        print(slots_in_tabbing( available_slots_of(c), [classpart_slot[c]] ))
    return [c[0] for c in slot_list]

sss = show_slots_of_subject


def assign_available_slots(clss, indices = None) : 
    my_slots = available_slots_of(clss + "$1")
    if len(my_slots) == 0 : 
        #print "No available slots for %s, Skipping.." % clss
        print("No available slots for {}, Skipping..".format(clss))
        return None
    if indices is None : 
        free_slots = my_slots 
    else : 
        free_slots = [ my_slots[k] for k in indices ] 
    
    for i, pt in enumerate(sorted(parts_of(clss))) :
        if free_slots[i] in slots : 
            if classpart_slot[pt] is not None : 
                #d = raw_input( "Going to move %s %s to %s %s ? (y/n/q)" % (pt, classpart_slot[pt], pt, free_slots[i]) )
                d = raw_input( "Going to move {} {} to {} {} ? (y/n/q)".format(
                        pt, classpart_slot[pt], pt, free_slots[i]) )
                if d == "q" : 
                    #print "Quitting slot sequence"
                    print("Quitting slot sequence")
                    break 
                elif d != "y" : 
                    print("Continuing to the next slot")
                    continue 
            classpart_slot[pt] = free_slots[i]
            print(pt, "->", classpart_slot[pt])
        else : 
            #print "Skipping %s .. %s -> %s unchanged" % (pt,pt, classpart_slot[pt])
            print("Skipping {} .. {} -> {} unchanged".format(pt,pt, 
                classpart_slot[pt]) )


days = ["월", "화", "수", "목", "금" ]
slots = [ ]
for d in days :
    if d == "수" :
        slots.extend( [ d + str(i) for i in range(1,5) ] )
    elif d == "금" :
        slots.extend( [ d + str(i) for i in range(1,8) ] )
    else :
        slots.extend( [ d + str(i) for i in range(1,8) ] )
times = "1 2 3 4 5 6 7".split()
slot_names = { }
for i, my_time in enumerate( times ): 
    for j, my_day in enumerate(days) : 
        if (my_day + my_time) in slots : 
            slot_names[ (i,j) ] = my_day+my_time


    
slot_dict =  dict( list_util.transposed(
    [ ("m1 m2 m3 m4 m5 m6 m7 t1 t2 t3 t4 t5 t6 t7 w1 w2 w3 w4 " + 
    "h1 h2 h3 h4 h5 h6 h7 f1 f2 f3 f4 f5 f6").split(), slots ]))

slot_dict.update( {"None": None, None: None} )


def assign_part_slot(part, slot_input) : 
    if slot_input in slot_dict : 
        slot  = slot_dict[slot_input]
    elif slot_input in slots : 
        slot = slot_input
    else : 
        slot = None

    if part not in classparts : 
        #print "No such classpart <%s>, doing nothing" % part
        print("No such classpart <{}>, doing nothing".format(part))
        return False
    elif slot is None : 
        classpart_slot[part] = None
        #print part, "->", classpart_slot[part]
        print(part, "->", classpart_slot[part])
        return True
    elif slot not in slots : 
        #print "No such slot <%s>, doing nothing" % slot
        print("No such slot <{}>, doing nothing".format(slot))
        return False
    else : 
        classpart_slot[part] = slot
        #print part, "->", classpart_slot[part]
        print(part, "->", classpart_slot[part])
        return True

def assign_class_slots(clss, input_slots) : 
    print(clss)
    my_slots = ["a" for k in parts_of(clss)]
    for i, sl in enumerate(input_slots) : 
        if sl in slot_dict : 
            my_slots[i] = slot_dict[sl]
        else : 
            my_slots[i] = sl
    for i, c in enumerate(parts_of(clss)) :
        if my_slots[i] in slots : 
            if classpart_slot[c] is not None : 
                #d = raw_input( "Going to move %s %s to %s %s ? (y/n/q)" % (c, classpart_slot[c], c, my_slots[i]) )
                d = raw_input( "Going to move {} {} to {} {} ? (y/n/q)".format(
                    c, classpart_slot[c], c, my_slots[i]) )
                if d == "q" : 
                    print("Quitting slot sequence")
                    break 
                elif d != "y" : 
                    print("Continuing to the next slot" )
                    continue 
            classpart_slot[c] = my_slots[i]
            print(c, "->", classpart_slot[c])
        elif str(my_slots[i]) == "None" : 
            classpart_slot[c] = None
            print(c, "->", "None" )
        else : 
            #print "Skipping %s .. %s -> %s unchanged" % (c,c, classpart_slot[c])
            print("Skipping {} .. {} -> {} unchanged".format(c,c, 
                classpart_slot[c]))
    
def assign_c_slots(subj, c_num, slot_list) : 
    clss = subj + "_" + str(c_num)
    assign_class_slots(clss, slot_list)
        


def assign_slots(name, input_slots): 
    print(name)
    my_slots = ["a" for k in classparts_of(name)]
    for i, sl in enumerate(input_slots) : 
        if sl in slot_dict : 
            my_slots[i] = slot_dict[sl]
        else : 
            my_slots[i] = sl
        
    for i, c in enumerate(classparts_of(name)) :
        if (my_slots[i] in slots) or (my_slots[i] in slot_dict) : 
            if classpart_slot[c] is not None : 
                #d = raw_input( "Going to move %s %s to %s %s ? (y/n/q)" % (c, classpart_slot[c], c, my_slots[i]) )
                d = raw_input( "Going to move {} {} to {} {} ? (y/n/q)".format(c, classpart_slot[c], c, my_slots[i]) )
                if d == "q" : 
                    print("Quitting slot sequence")
                    break 
                elif d != "y" : 
                    print("Continuing to the next slot" )
                    continue 
            assign_part_slot(c, my_slots[i])
        else : 
            #print "Skipping %s .. %s -> %s unchanged" % (c,c, classpart_slot[c])
            print("Skipping {} .. {} -> {} unchanged".format(c,c, 
                classpart_slot[c]))

ass = assign_slots 


def subject_name_of(classpart) : 
    ind = classpart.find("_")
    if ind < 0 :
        return classpart
    else :
        return classpart[:ind]

def class_name_of(classpart) : 
    ind = classpart.find("$")
    if ind < 0 :
        return classpart
    else :
        return classpart[:ind]



def check_append( my_list, new_item, str_fn = str, verbose=True ) : 
    if new_item in my_list : 
        if verbose:
            #print "%s is already in the list. Doing nothing." % str_fn(new_item)
            print("{} is already in the list. Doing nothing.".format(
                str_fn(new_item)))
        return False
    else : 
        my_list.append(new_item)
        if verbose:
            #print "%s is appended." % str_fn(new_item)
            print("{} is appended.".format(str_fn(new_item)))
        return True

def is_assistant_name( t_name ) : 
    if t_name.find("조교t") >= 0 : 
        return True 
    else : 
        return False 



def add_assistants() : 
    for cp in classparts : 
        if cp in sem.default_classroom_assignments_part : 
            a_names = [ n + "조교t" 
                for n in sem.default_classroom_assignments_part[cp].split("/")] 
            for a_name in a_names : 
                if check_append(name_classparts, (a_name, cp ),verbose=False): 
                    print(name_classparts[-1][0], "->", name_classparts[-1][1] )

        clss = class_name_of(cp)
        if clss in sem.default_classroom_assignments_class : 
            a_names = [ n + "조교t" for n 
                in sem.default_classroom_assignments_class[clss].split("/")] 
            for a_name in a_names : 
                if check_append(name_classparts, (a_name, cp ),verbose=False): 
                    print(name_classparts[-1][0], "->", name_classparts[-1][1] )

        subj = subject_name_of(cp) 
        if subj in sem.default_classroom_assignments_subject : 
            a_names = [ n + "조교t" for n 
                in sem.default_classroom_assignments_subject[subj].split("/")] 
            for a_name in a_names : 
                if check_append(name_classparts, (a_name, cp ),verbose=False): 
                    print(name_classparts[-1][0], "->", name_classparts[-1][1] )

    for t_name in teachers : 
        if t_name in sem.default_name_classroom_dict : 
            c_room = sem.default_name_classroom_dict[t_name] 
            a_name = c_room + "조교t" 
            for cp in classparts_of(t_name): 
                if subject_name_of(cp) in subjects:
                    if check_append(name_classparts, ( a_name , cp ), 
                            verbose=False ) : 
                        print(name_classparts[-1][0], "->",end=" " )
                        print(name_classparts[-1][1] )



def remove_assistants() : 
    pairs = [ (nm, pt) for (nm,pt) in name_classparts 
                    if nm.endswith("조교t")]
    for pair in pairs : 
        check_remove(name_classparts, pair, pair_str)

def update_assistants() : 
    reload(sem)
    remove_assistants()
    add_assistants()


def change_name( curr_name, new_name, pair_list = name_classparts, 
    classpart_list = None ) : 
    if classpart_list is None : 
        c_list = [c for c in pair_list if c[0] == curr_name ]
    else : 
        c_list = [c for c in pair_list 
                    if c[0] == curr_name and c[1] in classpart_list ]
    for c in c_list : 
            #print "(%s, %s) -> (%s, %s)" % (c[0], c[1], new_name, c[1]) 
            print("({}, {}) -> ({}, {})".format(c[0], c[1], new_name, c[1]))
            pair_list.remove(c)
            check_append( pair_list, (new_name, c[1] ) ) 


def change_name_in_all(curr_name, new_name) : 
    change_name(curr_name, new_name, name_subjects)
    change_name(curr_name, new_name, name_classparts)



def replace_subject( curr_subj, new_subj, pair_list = name_classparts ) : 
    replace_subject_in_name_subjects(curr_subj, new_subj)

    c_list = [c for c in pair_list if c[1].startswith( curr_subj + "_" ) ]
    for c in c_list : 
            curr_part = c[1]
            new_part = curr_part.replace( curr_subj + "_", new_subj + "_") 
            #print "(%s, %s) -> (%s, %s)" % (c[0], curr_part, c[0], new_part )
            print("({}, {}) -> ({}, {})".format(c[0], curr_part, 
                c[0], new_part ))
            pair_list.remove(c)
            check_append( pair_list, (c[0], new_part ) ) 
            if curr_part in classpart_slot : 
                classpart_slot[ new_part ] = classpart_slot[curr_part]
                #print "(%s, %s) -> (%s, %s)" % (curr_part, classpart_slot[curr_part], new_part, classpart_slot[new_part])
                print("({}, {}) -> ({}, {})".format(curr_part, 
                 classpart_slot[curr_part], new_part, classpart_slot[new_part]))
                del classpart_slot[curr_part]
            if curr_part in classrooms : 
                classrooms[ new_part ] = classrooms[curr_part]
                #print "(%s, %s) -> (%s, %s)" % (curr_part, classrooms[curr_part], new_part, classrooms[new_part])
                print("({}, {}) -> ({}, {})".format(curr_part,
                classrooms[curr_part], new_part, classrooms[new_part]))
                del classrooms[curr_part]


def replace_subject_in_name_subjects( curr_subj, new_subj, 
                    pair_list = name_subjects ):
    c_list = [c for c in pair_list if c[1] == curr_subj ]
    for c in c_list : 
            #print "(%s, %s) -> (%s, %s)" % (c[0], c[1], c[0], new_subj) 
            print("({}, {}) -> ({}, {})".format(c[0], c[1], c[0], new_subj))
            pair_list.remove(c)
            check_append( pair_list, (c[0], new_subj ) ) 

def exchange_name( name_one, name_two ) : 
    tmp_name = "TemporaryName"
    change_name( name_one, tmp_name, name_classparts )
    change_name( name_one, tmp_name, name_subjects )

    change_name( name_two, name_one, name_classparts )
    change_name( name_two, name_one, name_subjects )

    change_name( tmp_name,  name_two,  name_classparts )
    change_name( tmp_name, name_two,  name_subjects )


def apply_name_changes(name_name_dict) : 
    tmp_prefix = "TemporaryHead"
    for st_name in name_name_dict : 
        tmp_name = tmp_prefix + st_name 
        change_name( st_name, tmp_name, name_classparts)
        change_name( st_name, tmp_name, name_subjects)
    for st_name in name_name_dict : 
        tmp_name = tmp_prefix + st_name 
        new_name = name_name_dict[ st_name ]
        change_name(tmp_name, new_name, name_classparts)
        change_name(tmp_name, new_name, name_subjects)


def save_slots_xlsx(filename) : 
    teacher_names = sorted( list_util.union( 
            [ c[0] for c in name_classparts if c[0].endswith("t")]))
    output = { } 
    print("Building cp-t_name-slot dictionary")
    for name in teacher_names : 
        for cp in classparts_of(name) : 
            if classpart_slot[  cp ] is not None : 
                output[ (cp, name) ] = classpart_slot[  cp ]
            else : 
                output[ (cp, name) ] = "TBF" 

    print("Saving .. ")

    xlsx_util.write_dict_into_xlsx( filename, output,  teacher_names, 
            t_sorted(classparts))
            
    

def loop_check( name_list = None, return_list = False, slot_sharing=True ): 
    if name_list is None : 
        name_list = list_util.union( [c[0] for c in name_classparts] )
    name_pt_slots = [ ] 
    for name in name_list : 
        for cp in classparts_of(name) : 
            if classpart_slot[cp] is not None : 
                name_pt_slots.append( (name, cp,  classpart_slot[cp]) )
        if name in sem.teacher_avoid_slots_dict : 
            for sl in sem.teacher_avoid_slots_dict[name] : 
                name_pt_slots.append( (name, name+sl+"_1$1" ,  sl) )
    name_slots = [ ] 
    if slot_sharing : 
        mod_nps = [ ] 
        for name, pt, slot in name_pt_slots :
            if pt in sem.slot_sharing_part_dict and \
                (name, sem.slot_sharing_part_dict[pt],slot) in mod_nps : 
                continue
            else : 
                mod_nps.append( (name, pt, slot) )
    else : 
        mod_nps = name_pt_slots

    for name, pt, slot in mod_nps : 
        name_slots.append( (name, slot) )
        #if slot_sharing and (pt in sem.slot_sharing_part_dict) : 
            #if (name,  sem.slot_sharing_part_dict[pt], slot) in name_pt_slots : 
                #continue
            #else : 
                #name_slots.append( (name, slot) )
        #else : 
            #name_slots.append( (name, slot) )
        
    dup_list = [ item for item in name_slots if name_slots.count(item) > 1 ]

    #list_util.join_npr(dup_list)
    if len(dup_list) == 0 : 
        return True 
    else : 
        for name in sorted(list_util.union( [c[0] for c in dup_list] )) : 
            for cp in classparts_of(name) : 
                if  (name, classpart_slot[cp])  in dup_list : 
                    print(name, cp, classpart_slot[cp])
        if return_list is False : 
            return False 
        else : 
            return sorted(list_util.union( [c[0] for c in dup_list] )) 

lc = loop_check 
    

def clear_slot( classpart_or_clss ) : 
    if is_classname( classpart_or_clss ) : 
        for pt in parts_of(classpart_or_clss) :  
            clear_slot(pt)
    else :     
        classpart = classpart_or_clss 
        if classpart in classpart_slot : 
            #print "Deleting %s %s to %s None" % (classpart, classpart_slot[classpart], classpart )
            print("Deleting {} {} to {} None".format(classpart,
                    classpart_slot[classpart], classpart ))
            classpart_slot[classpart] = None 
        else : 
            #print "No %s in classpart_slot. Doing nothing." % classpart 
            print("No {} in classpart_slot. Doing nothing.".format(classpart))
        
def clear_slots_names(subj, c_num_str) : 
    cl_list = [ subj + "_" + c for c in c_num_str.split() ]
    for cl in cl_list : 
        remove_all_names_from_class(cl)
        clear_slot(cl)

        
def clear_names(subj, c_num_str) : 
    cl_list = [ subj + "_" + c for c in c_num_str.split() ]
    for cl in cl_list : 
        remove_all_names_from_class(cl)


def remove_from_subject_at_slot_names(subj, slot_short, name_list) : 
    if slot_short in slot_dict : 
        slot = slot_dict[slot_short]
    else : 
        slot = slot_short
    for st_name in name_list : 
        pt = name_slot_part(st_name, slot)
        if subj == subject_name_of(pt) : 
            remove_name_from_class(st_name, class_name_of(pt))
        



def classes_of( subj ) : 
    return sorted( list_util.union( [ class_name_of(cp) for cp in classparts 
                    if subject_name_of(cp) == subj ] ), 
                key = (lambda x: (len(x), x)) ) 

    
def parts_of( clss_or_subj ) : 
    return sorted([ cp for cp in classparts if belongs_to(cp, clss_or_subj)],
                key = (lambda x: (len(x), x)) ) 


def initialize_classpart_slot_classrooms() : 
    global classpart_slot, classrooms 
    print("Initializing classpart_slot")
    classpart_slot = { }
    classrooms = { }
    for cp in classparts : 
        classpart_slot[cp] = None
        classrooms[cp] = None
        print(cp,end=" ")
    print()
    print("All set to be None")
        
    

def clear_slots_of(subj_or_clss) : 
    for cp in parts_of(subj_or_clss) :  
        clear_slot(cp)



teachers = sorted( list_util.union( 
            [ c[0] for c in name_classparts if c[0].endswith("t") 
                    and (not c[0].endswith("조교t"))]))




trashbinname = "trash-bin"
def leave_a_backup( fname ) :
    backup_fname = os.path.expanduser( '~/' + trashbinname +  \
            os.sep + os.path.split(fname)[1] )
    if fname[-1] == '/' : fname = fname[:-1]
    for i in range(1,10000) :
        j = backup_fname.rfind('.')
        if j >= 0 :
            tmpname = backup_fname[0:j] + '-' + str(i) + backup_fname[j:]
        else :
            tmpname = backup_fname + '-' + str(i)
        if os.path.exists(   tmpname ) :
            continue
        else :
            print("Executing os command: ",end=" ")
            print('cp ' + fname + ' '+   tmpname)
            os.system('cp ' + fname + ' '+   tmpname )
            break



    



save_filename = "classpart_info.p"
def save(save_filename = save_filename) :
    leave_a_backup( save_filename )
    with open( save_filename, "wb" ) as f :
        #print "Writing class_info to <%s> " % save_filename
        print("Writing class_info to <{}> ".format(save_filename))
        pickle.dump( 
            [name_classparts, name_subjects, juniors , 
              classparts, classpart_slot, teachers, 
                 classrooms, 
                neis_subject_name_pairs 
] , f )

def load( filename = "classpart_info.p" ) :
    global name_classparts, name_subjects, juniors 
    global classparts, classpart_slot, teachers 
    global   classrooms
    global neis_subject_name_pairs 
    with open( filename ) as f :
        print("Loading " + filename)

        name_classparts, name_subjects, juniors, classparts, classpart_slot, \
            teachers,   classrooms, \
               neis_subject_name_pairs \
                = pickle.load( open(filename, "rb") )
        #print "%s has %d items" % ("name_classparts", len(name_classparts))
        #print "%s has %d items" % ("name_subjects", len(name_subjects))
        #print "%s has %d items" % ("juniors", len(juniors))
        #print "%s has %d items" % ("classparts", len(classparts))
        #print "%s has %d items" % ("classpart_slot", len(classpart_slot))
        #print "%s has %d items" % ("teachers", len(teachers))
        #print "%s has %d items" % ("classrooms", len(classrooms))
        #print "%s has %d items" % ("neis_subject_name_pairs",len(neis_subject_name_pairs))
        print("{} has {} items".format("name_classparts", len(name_classparts)))
        print("{} has {} items".format("name_subjects", len(name_subjects)))
        print("{} has {} items".format("juniors", len(juniors)))
        print("{} has {} items".format("classparts", len(classparts)))
        print("{} has {} items".format("classpart_slot", len(classpart_slot)))
        print("{} has {} items".format("teachers", len(teachers)))
        print("{} has {} items".format("classrooms", len(classrooms)))
        print("{} has {} items".format("neis_subject_name_pairs",
                        len(neis_subject_name_pairs)))


names = list(set( [c[0] for c in name_classparts] + 
            [c[0] for c in name_subjects] ))
names.sort()

subjects = list(set( [c[1] for c in name_subjects]))
subjects.sort()

students = list(set( [c[0] for c in name_subjects]))
students.sort()


classparts = list_util.union( [ c[1] for c in name_classparts ] )


neis_subject_name = dict( neis_subject_name_pairs )


def name( substr ) :
    mystr = str(substr)
    output = [ ] 
    for c in names :
        if c.find(mystr) >= 0 :
            output.append(c)
    if len(output) == 0 : 
        return None 
    else :
        return ' '.join(output)



def slot_list_of( name ) : 
    return [ classpart_slot[c] for c in classparts_of(name)] 


def put_name_into_class(name, clss ) : 
    subj = subject_name_of(clss)
    if (not name.endswith("t")) : 
        if ( (name, subj) not in name_subjects) : 
            #print "%s is not enrolled in %s" % (name, subj)
            print("{} is not enrolled in {}".format(name, subj))
            return False 
        else : 
            for pt in parts_of(subj) : 
                if (name, pt) in name_classparts : 
                    #print "Removing %s from %s" % (name, pt) 
                    print("Removing {} from {}".format(name, pt))
                    name_classparts.remove( (name,pt) ) 
    for pt in parts_of(clss) : 
        #print "Trying %s -> %s" % (name, pt)
        print("Trying {} -> {}".format(name, pt))
        check_append( name_classparts, (name,pt) ) 
    return True

def remove_name_from_class(name, clss) : 
    for pt in parts_of(clss) : 
        if (name, pt) in name_classparts : 
            if not name.endswith("t") : 
                #print "Removing %s from %s" % (name, pt) 
                print("Removing {} from {}".format(name, pt))
                name_classparts.remove( (name,pt) ) 
            else : 
                #print "Keeping %s in %s." % (name, pt)
                print("Keeping {} in {}.".format(name, pt))
        else : 
            #print "%s not in %s, Doing nothing." % (name, pt)
            print("{} not in {}, Doing nothing.".format(name, pt))



def remove_names_from_class(n_list, clss) : 
    for n in n_list : 
        remove_name_from_class(n, clss)


def remove_all_names_from_class(clss, count = None, fix_names = [ ] ) : 
    if count is None : 
        st_list = list_util.complement(st_members_of(clss), fix_names)
    else : 
        st_list = list_util.complement(st_members_of(clss), fix_names)
    for n in st_list : 
        remove_name_from_class(n, clss)

def remove_all_names_from_parts_of(subj) : 
    for c in classes_of(subj)  : 
        remove_all_names_from_class(c)


def remove_name_from_classes(name, classes) : 
    for c in classes : 
        remove_name_from_class(name, c)

def remove_name_from_all_classes(name) :
    remove_name_from_classes(name, current_classes(name))
    
def remove_names_from_class_of_subject(name_list, subj) : 
    for st_name in name_list : 
        clss = name_subject_class(st_name, subj) 
        if len(clss)>0 : 
            remove_name_from_class(st_name, clss)
        else : 
            #print "%s not in any class of %s" % (st_name, subj)
            print("{} not in any class of {}".format(st_name, subj))

        
    




def put_names_into_class( names, clss) : 
    for name in names : 
        put_name_into_class(name, clss) 

def put_name_into_classes( name, classes, c_nums = None  ) : 
    if c_nums is None : 
        for c in classes :
            put_name_into_class(name, c)
    else : 
        subjs = classes
        for i, c_num in enumerate(c_nums) : 
            put_name_into_class(name, subjs[i] + "_" + str(c_num) ) 
            

def remaining_names_of(subj) : 
    my_members = [c[0] for c in name_subjects if c[1] == subj ] 
    enrolled = list_util.union( 
            list_util.flatten( [members(cp) for cp in parts_of(subj)]))
    return list_util.complement( my_members, enrolled )
    
    
def parts_with_none_classroom() : 
    return [ pt for pt in classparts if classrooms[pt] is None ]

def classes_with_none_classroom() : 
    output = list_util.union( [class_name_of(pt) for pt 
                in parts_with_none_classroom() 
                    if subject_name_of(pt) not in "담임업무".split() ] )
    output.sort(key=(lambda x: (sem.ordered_group.index( 
                    sem.subject_group_dict[subject_name_of(x)]),x) )  )
    return output


def classes_with_empty_slots() : 
    output = [ ] 
    for subj in subjects : 
        for clss in classes_of(subj) : 
            if None in [classpart_slot[c] for c in parts_of(subj)] : 
                output.append(clss)
    output.sort()
    return output

def subjects_with_empty_slots() : 
    output = [ ] 
    for subj in subjects : 
        if None in [classpart_slot[c] for c in parts_of(subj)] : 
            output.append(subj)
    output.sort()
    return output


def first_part(clss) : 
    return ( clss + "$1" )


def classes_with_no_students() : 
    output = [ ] 
    for subj in subjects : 
        for clss in classes_of(subj) : 
            if len(st_members_of(clss)) == 0 : 
                output.append(clss)
    output.sort()
    return output

def total_hours(name) : 
    return sum( [ hours(subj) for n, subj in name_subjects if n == name] )


def put_name_into_smallest_classes( name, keep_current = True ) : 
    if not keep_current : 
        subjs = subjects_of(name)
    else : 
        subjs = list_util.complement(subjects_of(name), 
                    map(subject_name_of, current_classes(name)))
    for subj in subjs : 
        tmp_clss = min( classes_of(subj), 
             key =  (lambda x: len(st_members_of(first_part(x)))))
        put_name_into_class(name, tmp_clss)



def assign_random_ones() : 
    for c in classes_with_no_students() : 
        if len(st_members_of(c)) > 0 : 
            #print "%s is non-empty. Skipping it" % ( c,)
            print("{} is non-empty. Skipping it".format( c))
        else : 
            tmp_list = sorted( remaining_names_of(subj) , key = total_hours )
            if len(tmp_list) == 0 : 
                print("Everyone is already enrolled. Doing nothing..")
            else : 
                put_name_into_first_empty_classes( tmp_list[-1] )




    
def is_classname( mystr ) : 
    if mystr.find("_") >= 0 and mystr.find("$") < 0 : 
        return True 
    else : 
        return False 
    
def is_classpart_name( mystr ) : 
    if mystr.find("_") >= 0 and mystr.find("$") >= 0 : 
        return True 
    else : 
        return False 

def category(part_str):
    if re.search( r'_\d+\$\d+$' , part_str) : 
        return "part"
    elif re.search( r'_\d+$' , part_str) : 
        return "class"
    elif not re.search( r'[_\$]+' , part_str) : 
        return "subject"
    else : 
        return None



def classes_with_none_slots() :  
    pts = [ pt for pt in classparts if classpart_slot[pt] is None ] 
    return sorted(list_util.union( [ class_name_of(pt) for pt in pts ] ))

def has_none_slot(c_class) :  
    for pt in parts_of(c_class) : 
        if classpart_slot[pt] is None : 
            return True 
    else : 
        return False 

def has_all_none_slots(c_class) :  
    for pt in parts_of(c_class) : 
        if classpart_slot[pt] is not None : 
            return False
    else : 
        return True


def remove_duplicates() : 
    for nc in name_classparts : 
        nc_count = name_classparts.count(nc)
        if nc_count > 1 : 
            #print "%d instances of (%s, %s), removing %d of them" % (nc_count, nc[0], nc[1], nc_count - 1)
            print("{} instances of ({}, {}), removing {} of them".format(
                        nc_count, nc[0], nc[1], nc_count - 1))
            for k in range(nc_count - 1) : 
                name_classparts.remove(nc)
                #print "(%s, %s) is removed." % ( nc[0], nc[1])
                print("({}, {}) is removed.".format( nc[0], nc[1]))


def clear_all_st_names( classname ) :
    for pt in parts_of(classname): 
        st_names = st_members_of(pt)
        for n in st_names : 
            #print "Removing %s from %s." % (n, pt) 
            print("Removing {} from {}.".format(n, pt))
            name_classparts.remove( (n,pt) )
        
def slots_of(c_class) : 
    return [ classpart_slot[pt] for pt in parts_of(c_class) 
                if classpart_slot[pt] is not None ] 



def enumerate_assignments( class_list, curr_slots,  subject_list, 
       size_bound = 0, excludes = [ ], includes = [ ],  randomize = False  ) : 
    if len( subject_list ) == 0 : 
        yield class_list 
    else : 
        classes = [c for c in  classes_of( subject_list[0] ) 
                    if (not has_all_none_slots(c)) 
           and len(list_util.intersect(slots_of(c), curr_slots)) == 0 
                       and c not in excludes 
                    and not ( size_bound > 0 and size_bound < 
  len( list_util.complement( st_members_of(c), [ name ] )) + 1) ] 
        curr_includes = [c for c in includes 
                            if subject_list[0] == subject_name_of(c) ]
        if len(curr_includes) > 0 : 
            intersected_classes = list_util.intersect(classes, curr_includes )
        else : 
            intersected_classes = classes
        if len(intersected_classes ) > 0 : 
            classes = intersected_classes 
            if randomize is True : 
                random.shuffle(classes)
            else : 
                classes.sort( key = (lambda x: len(st_members_of(x))))
            for c in classes : 
                new_list = class_list[:] + [c]
                new_slots = curr_slots[:] +  slots_of(c)
                for c_list in enumerate_assignments( new_list, new_slots, 
                        subject_list[1:], size_bound, excludes, includes, 
                            randomize ) : 
                    yield c_list
        else : 
            yield class_list 



def has_class_without_none_slot(subject) : 
    for c in classes_of(subject) : 
        if not has_none_slot(c) : 
            return True
    else :
        return False
        

def has_class_with_assigned_slots(subject) : 
    for c in classes_of(subject) : 
        if not has_all_none_slots(c) : 
            return True
    else :
        return False
        


def current_classes( st_name ) : 
    return list_util.union( [class_name_of(p) for p in classparts_of(st_name) ] ) 


def name_subject_class(st_name, subject) : 
    c_list = [c for c in current_classes(st_name) 
                if subject_name_of(c) == subject ]
    return '|'.join(c_list)
    


def average_size( subject ) : 
    return float( len( [c for c in name_subjects 
        if c[1] == subject ] )) / len( classes_of(subject))

def num_of_assignable_classes(st_name, subj, excludes = [], bound=40, 
        includes = []) : 
    if subj in map(subject_name_of, includes):
        return 1
    output = [c for c in list_util.complement( classes_of(subj), excludes ) 
                if not overfulled_with(c, st_name, n = bound, exp_n = 28)]  
    return len(output)

def search_class_assignments( name, count = 20, excludes = [ ], 
            input_includes = [ ], bound = 0 , verbose = False, 
                randomize = False, yield_partial = False, 
   sort_key =  None 
 ) : 
    all_subs = [s for s in subjects_of(name) if has_class_with_assigned_slots(s)]

    if sort_key is None : 
        sort_fn = (lambda s: num_of_assignable_classes(name, s, excludes, 
                        bound ) )
    else : 
        sort_fn = sort_key

    all_subs.sort( key = sort_fn  )
    includes = [ ] 
    for c in input_includes : 
        if is_classname(c) : 
            includes.append(c)
        else : 
            includes.append( name_subject_class(name,c) )
    for c in includes : 
        all_subs.remove( subject_name_of(c) ) 
        all_subs.insert(0, subject_name_of(c) ) 
    i = 0
    n = len(all_subs)
    print(name)
    for s in all_subs : 
        print(s, average_size(s), ' ',end=" " )
    print
    output = [ ] 
    for classes in  enumerate_assignments( [], [], all_subs, 
            bound, excludes, includes, randomize  ) : 
        if verbose is True : 
            print(' '.join( [ c + ' '.join(current_slots(c)) for c in classes]))
        if yield_partial is True : 
            if i > count : 
                break 
            if len(output) == 0 : 
                output.append(classes)
            elif  sum( map( lambda x: hours(subject_name_of(x)), 
                    output[-1])) < sum( 
                        map( lambda x: hours(subject_name_of(x)), classes )) :
                output[-1] = classes[:] 
                print(' '.join(classes))
                continue
            else : 
                continue
        elif len(classes) == n : 
            print(i,end=" ")
            for c in classes : 
                st_num = len( list_util.complement( st_members_of(c), [ name ] )) + 1 
                #print ("<%s %d>" % (c, st_num)).ljust(18), 
                print(("<{} {}>".format(c, st_num)).ljust(18), end=" ")
            print()
            output.append(classes)
            i += 1 
            if i >= count : 
                break 
    return(output)




def search_with_fixes(name, fixes = [ ], count = 20, excludes = [ ], 
        includes = [ ], bound = 0, randomize=False, sort_key = None ) : 
    my_includes = [ name_subject_class( name, s ) for s in fixes ] + includes
    return search_class_assignments (name, count, excludes, my_includes, 
            bound, randomize=randomize, sort_key = sort_key )
    




def is_assignable( name, verbose = False, num_output = False ) : 
    all_subs = [s for s in subjects_of(name) if has_class_with_assigned_slots(s)]
    all_subs.sort( key = (lambda x : len( classes_of(x))))
    includes = [ ] 
    excludes = [ ] 
    i = 0
    n = len(all_subs)
    print(name)
    for s in all_subs : 
        print(s,end=" "  )
    print
    for cnt, classes in  enumerate(enumerate_assignments( [], [], all_subs, 
            40, excludes, includes  )) : 
        if cnt % 10 == 0 : 
            print(cnt,end=" ")
        if verbose : 
            print(' '.join(classes))
        if len(classes) == n : 
            print
            for c in classes : 
                st_num = len( list_util.complement( st_members_of(c), [ name ] )) + 1 
                #print ("<%s %d>" % (c, st_num)).ljust(18), 
                print(("<{} {}>".format(c, st_num)).ljust(18), end=" ")
            print
            if num_output is True : 
                return cnt
            else : 
                return True 
    else : 
        return False 
            

    
def remaining_names( subject ) : 
    tmp = [ members(c) for c in classes_of(subject) ]
    return list_util.complement( members(subject), *tmp )

def remaining_subjects( name ) : 
    finished_classes = [clss for clss in current_classes(name) 
                if None not in [classpart_slot[cp] for cp in parts_of(clss)]]
    return sorted(list_util.complement(subjects_of(name), 
                map( subject_name_of, finished_classes), 
                    sem.skip_subject_list), 
                key = sem.ordered_subjects.index)


def expected_class_size(st_name, clss, removing=False ) : 
    if removing is False : 
        return len( list_util.complement( 
                    st_members_of(clss), [ st_name ] )) + 1 
    else : 
        return len( list_util.complement(st_members_of(clss), [ st_name ] )) 

def overfulled_with(clss, st_name, n = 20, exp_n = 16) : 
    if n == 0 : 
        return False
    ecs = expected_class_size(st_name, clss) 
    if (ecs > n) or ( clss.find("실험") >= 0 and ecs > exp_n ) : 
        return True 
    else : 
        return False 


def score_name_class_assignment( name, classes ) : 
    output = 0.0
    for c in classes : 
        ecs = expected_class_size(name, c)
        if ecs > 20 : 
            output += 1000 * (ecs - 20)
        elif c.find("실험")>= 0 and ecs > 17 : 
            output += 900* (ecs - 17)
        elif ecs < 7 : 
            output -= 400 * (7 - ecs)

        a_size = average_size(subject_name_of(c))
        output -= a_size - ecs
    return output


def print_class_sizes(name, classes) : 
    for c in classes : 
        st_num = len( list_util.complement( st_members_of(c), [ name ] )) + 1 
        #print korean_ljust( ("<%s %d>" % (c, st_num)) , 18), 
        print(korean_ljust( ("<{} {}>".format(c, st_num)) , 18), end=" ")
    print



def choose_best_assignment(name, count = 30, excludes = [ ], includes = [ ], 
        fix_subjects = [ ], bound = 0, randomize=False, sort_key = None ) : 
    candis = [ ] 
    for i in range(count) : 
        cnd = search_with_fixes(name, fixes = fix_subjects , 
                count = 1, excludes = excludes, 
                includes = includes, bound = bound, randomize=randomize, 
                sort_key = sort_key )[0] 
        candis.append(cnd)
    output = min( candis, key = (lambda x: score_name_class_assignment(name,x)))
    print("\n\n")
    print_class_sizes(name, output)
    return output



def choose_and_assign(st_name, count = 1, includes = [ ], excludes=[], 
                        bound = 0, fix_subjects = [ ], randomize=False, 
                        across = False, sort_key = None) : 
    #sem_excludes = sem.exclude_classes_of(st_name, across = across)
    #my_excludes = list_util.union( excludes, sem_excludes)
    my_excludes = excludes
    my_includes = includes
    
    classes = choose_best_assignment(st_name, count, my_excludes , my_includes, 
                    bound=bound, fix_subjects = fix_subjects, 
                    randomize=randomize, sort_key = sort_key) 
    put_name_into_classes(st_name, classes)



#def put_random_one_into_class( clss, candidates = None, n=19, exp_n=16, 
        #count=1, randomize=False, lower_bound = 13, fix_names = []) : 
    #subj = subject_name_of(clss)
    #if candidates is None : 
        #candis = members(subj)
    #else : 
        #candis = list_util.intersect(candidates, members(subj))
    #st_name = random.choice( list_util.complement(candis, 
                                #st_members_of(clss), fix_names))
    #subjs = [s for s in subjects_of(st_name) if len(classes_of(s)) > 2 ]
    #small_cs = list_util.union(* [ [c for c in classes_of(subj) 
         #if expected_class_size(st_name, c) < lower_bound] for subj in subjs 
            #if has_class_with_assigned_slots(subj) ])
    #big_cs = list_util.union(* [ [c for c in classes_of(subj) 
         #if overfulled_with(c, st_name, n, exp_n) ] 
            #for subj in subjects_of(st_name) ])

    #ex_classes = list_util.union( big_cs ,  small_cs )
    #inc_classes = list_util.union([clss], small_cs) 
    #try: 
        #classes = choose_best_assignment(st_name, count, ex_classes, 
                #includes = inc_classes , randomize=randomize ) 
    #except IndexError : 
        #return False
    #put_name_into_classes(st_name, classes)
    #return True


def move_random_one_from( subj, c_num, exclude_nums = [ ],  
    fixed_names = [ ], excl_classes = [ ], n = 19, exp_n = 16, count = 5, 
        randomize=False) : 
    if type(c_num) is list : 
        cl_nums = c_num 
    else : 
        cl_nums = [ c_num ]
    st_list = list_util.union( *[st_members_of( subj + "_" + str(k)) 
                        for k in cl_nums])
    st_name = random.choice( list_util.complement( st_list, fixed_names  ) )
    move_one_from_class(st_name, subj, exclude_nums + [c_num], excl_classes, 
        n, exp_n, count, randomize=randomize)


def try_move_random_ones_from_to( subj, c_num, dest_c_num, exclude_nums = [ ],  
    fix_names = [ ], excl_classes = [ ], n = 19, exp_n = 12, count = 1, 
    lower_bound = 9, iter_num = 1, stop_count=0, randomize=False ) : 
    candidates = st_members_of(subj + "_" + str(c_num))
    clss = subj + "_" + str(dest_c_num)
    for k in range(iter_num) : 
        try: 
            put_random_one_into_class( clss, candidates = candidates, 
                fix_names = fix_names, n = n, exp_n = exp_n, count = count,  
                lower_bound = lower_bound)
            if stop_count > 0 and k >= (stop_count-1) : 
                show_stats([subj])
                break
        except (IndexError, KeyboardInterrupt): 
            pass


def try_move_random_one_to( subj,  dest_c_num, exclude_nums = [ ],  
    fix_names = [ ], excl_classes = [ ], n = 30, exp_n = 16, count = 1, 
    lower_bound = 1, iter_num = 1, stop_count=0, randomize=False ) : 
    clss = subj + "_" + str(dest_c_num)
    candidates = list_util.complement(members(subj), 
                    st_members_of(clss))
    for k in range(iter_num) : 
        try: 
            put_random_one_into_class( clss, candidates = candidates, 
                fix_names = fix_names, n = n, exp_n = exp_n, count = count,  
                lower_bound = lower_bound)
            if stop_count > 0 and k >= (stop_count-1) : 
                show_stats([subj])
                break
        except (IndexError, KeyboardInterrupt): 
            pass




def try_put_random_ones_into_class(subj, c_num, excludes = [ ], bound = 20, 
                      count=1, fix_subjects = [ ], fix_names = [], 
                        randomize=True, num = 1):
    output = [ ] 
    for k in range(num) : 
        try:
            st_name = put_random_one_into_class(subj, 
                        c_num, excludes = excludes, bound = bound, count=count, 
                            fix_subjects = [ ], fix_names = [], 
                            randomize=randomize)  
            output.append(st_name)
        except IndexError:
            pass
    return output
    
def move_one_from_class(st_name, subj , ex_nums = [ ], excl_classes = [ ], 
         n=20, exp_n = 16, count = 5, randomize=False ) :  
    clss = name_subject_class(st_name, subj)
    excludes = [ subj + "_" + str(k) for k in ex_nums ] + [clss] + excl_classes
    all_classes = list_util.flatten( [ classes_of(s) 
                        for s in subjects_of(st_name)])
    ex_classes = list_util.union([ c for c in all_classes 
            if overfulled_with(c, st_name, n, exp_n)] ,  excludes )
    classes = choose_best_assignment(st_name, count, ex_classes, 
                randomize=randomize ) 
    put_name_into_classes(st_name, classes)
    
    
def assign_classes(st_name,   excl_classes = [ ], 
         n=20, exp_n = 16, count = 5, randomize=False ) :  
    excludes =  excl_classes
    all_classes = list_util.flatten( [ classes_of(s) 
                        for s in subjects_of(st_name)])
    ex_classes = list_util.union([ c for c in all_classes 
            if overfulled_with(c, st_name, n, exp_n)] ,  excludes )
    classes = choose_best_assignment(st_name, count, ex_classes, 
                randomize=randomize ) 
    put_name_into_classes(st_name, classes)
    



def show_subject_info(subject):
    c_names = classes_of(subject)
    list_util.jpr(c_names + [str( average_size(subject)) ])
    print()
    for c in c_names : 
        t_names = [n for n in members(c) if n.endswith("t") ]
        t_names.sort(key = (lambda x: (len(x), x)))
        if len(t_names) > 0 : 
            print(c, "("+str(hours(subject))+")", t_names[0])
        else :
            print(c, "("+str(hours(subject))+")" )
        tmp = ' '.join( [ str(classpart_slot[pt]) for pt in parts_of(c) ] ) 
        #print "<%s>"% tmp, 
        print("<{}>".format(tmp),  end=" ")
        print(len( st_members_of(c) ),end=" ")
        print(' '.join( list_util.union( 
                [ str( classrooms[pt] ) for pt in parts_of(c)]) ))
        for c in parts_of(c) :
            print(korean_ljust(c, 20),end=" " )
            if classpart_slot[c] is not None : 
                print(":"+ classpart_slot[c] + ":",end=" ")
            else :
                print(":"+ "   " + ":",end=" ")
            print(slots_in_tabbing( available_slots_of(c), [classpart_slot[c]] ))

        print(' '.join( members(c)))
        print()
    print("Remaining names",end=" " )
    print(' '.join(remaining_names(subject)))

ssi = show_subject_info



def subjects_to_balance( n = 20 ) : 
    output = [ ] 
    if n > 0 : 
        size_test = (lambda c: (len(st_members_of(c)) > n) )
    else : 
        size_test = (lambda c: (len(st_members_of(c)) < -n) 
                and (len(st_members_of(c))>1)  )
    for s in subjects : 
        for c in classes_of(s) : 
            if size_test(c) : 
                #print s, "%.1f" % average_size(s), 
                #print c, len(st_members_of(c))
                print(s, "{:.1f}".format(average_size(s)),  end=" ")
                print(c, len(st_members_of(c)))
                output.append( subject_name_of(c))
    return list_util.union( output )


                
def full_classes(n=19, exp_n = 12, gap = 3, small = False) : 
    output = [ ] 
    small_output = [ ] 
    k = 0
    for s in subjects : 
        av_size = average_size(s)
        max_c_num = len(classes_of(s))
        s_ind = subjects.index(s)
        for c in classes_of(s) : 
            c_size = len(st_members_of(c))
            if c_size == n and (c_size - av_size <= 2) : 
                #print "%02d %s %d %.1f" % (s_ind, s, max_c_num, av_size),
                print("{:02d} {} {} {:.1f}".format(s_ind, s, max_c_num, av_size),
                      end=" ")
                print(c, c_size)
            elif c_size >= n : 
                output.append(c)
                #print "%02d %s %d %.1f" % (s_ind, s, max_c_num, av_size),
                print("{:02d} {} {} {:.1f}".format(s_ind, s, max_c_num, 
                    av_size),  end=" ")
                print(c, c_size)
                k += 1
            elif c.find("실험") >= 0 and c_size >= exp_n : 
                output.append(c)
                #print "%02d %s %d %.1f" % (s_ind, s, max_c_num, av_size),
                print("{:02d} {} {} {:.1f}".format(s_ind, s, max_c_num, 
                    av_size),  end=" ")
                print(c, c_size)
                k += 1
            elif ( av_size - c_size >= gap ) : 
                small_output.append(c)
                #print "%02d %s %d %.1f" % (s_ind, s, max_c_num, av_size),
                print("{:02d} {} {} {:.1f}".format(s_ind, s, max_c_num, 
                    av_size),  end=" ")
                print(c, c_size)
    if small :
        return small_output
    else : 
        return output


def gap_of_subject(subj) : 
    sizes = [size_of(c) for c in classes_of(subj)]
    my_gap = max(sizes) - min(sizes)
    return my_gap

def show_gap_subjects(subj_list = subjects, gap = 1 ):
    tmp_list = [ [subj, map(size_of, classes_of(subj))] for subj in subj_list]
    sec_list = [ [c[0], max(c[1]) - min(c[1]), max(c[1]), min(c[1]), c[1]] 
                    for c in tmp_list ]
    thr_list = [ c for c in sec_list if c[1] > gap ]
    thr_list.sort(key = lambda x: (x[1], -len(classes_of(x[0]))))
    for i,x in enumerate(thr_list) : 
        #print i, "%s |%d=%d-%d|" % (x[0], x[1], x[2], x[3] ), 
        print(i, "{} |{}={}-{}|".format(x[0], x[1], x[2], x[3] ), end=' ')
        for num in x[4]:
            #print "<%d>" % num, 
            print("<{}>".format(num), end=' ')
        print
    return [c[0] for c in thr_list]


def check_class_sizes(exp_n=12) : 
    output = True
    for subj in subjects : 
        for c in classes_of(subj) : 
            sz = size_of(c)
            if sz < 6 or sz > 20 : 
                output = False 
                #print "%s <%s>" % (c, sz)
                print("{} <{}>".format(c, sz))
            if subj.find("실험")>=0 and sz > exp_n : 
                output = False 
                #print "%s <%s>" % (c, sz)
                print("{} <{}>".format(c, sz))

    my_parts = [ ] 
    for room in sem.classroom_size_bound:
        for sl in slots : 
            curr_part = room_slot_part(room, sl)
            if len(curr_part)>0 : 
                if size_of(curr_part) > sem.classroom_size_bound[room]:
                    print(curr_part, len(st_members_of(curr_part)))
                    output = False
    return output
    

def check_day_condition():
    output = True
    c_list = list_util.flatten( [classes_of(c) for c 
                in list_util.intersect(sem.three_day_subjects, subjects)] )
    for clss in c_list : 
        if len( list_util.union( [ c[:-1] for c in slots_of(clss)])) != 3 : 
            print(clss, ' '.join(slots_of(clss)))
            output = False
    chinese_subjects = "중국어I 중국어II".split()
    c_list = list_util.flatten( [classes_of(c) for c 
                in list_util.intersect(chinese_subjects, subjects)] )
    for clss in c_list : 
        if len(list_util.intersect(sem.chinese_days, 
            [c[:-1] for c in slots_of(clss)])) != 1 : 
            print(clss, ' '.join(slots_of(clss)))
            output = False
    for clss in list_util.flatten( [classes_of(subj) for subj in subjects]):
        if not has_two_days(clss):
            print(clss, ' '.join(slots_of(clss)))
            output = False

    
    return output

    
def check_validity() :     
    print("Checking students enrolled to classes")
    for st_name in students : 
        if len( remaining_subjects(st_name) ) > 0 : 
            print(st_name, ' '.join( remaining_subjects(st_name) ))
            return False 

    print("Checking no time slots conflict" )
    if not loop_check() : 
        return False 
  
    print("Checking double booking" )
    if not double_booking_check() : 
        return False

    print("Checking class sizes" )
    if not check_class_sizes() : 
        return False

    return True



all_classrooms = \
    ( "융합공통강의실 물리강의실1 사회강의실1 사회강의실2 사회강의실3" + \
" 생물실험실2 생물실험실1 화학강의실1 화학강의실2 수학강의실3" + \
" 정보강의실 물리실험실1 외국어강의실5 외국어강의실4 외국어강의실3" + \
" 외국어강의실2 외국어강의실1 화학강의실3 수학강의실6 수학강의실4" + \
" 수학강의실5 수학강의실2 음악실 수학강의실1 생물강의실1" + \
" 컴퓨터실2 국어강의실3 국어강의실2 국어강의실1 미술실" + \
" 생물강의실2 컴퓨터실1 우암공통강의실2 우암공통강의실3 우암공통강의실1" + \
" 화학실험실1 상상토의실 우암공통강의실4 물리강의실2 지구과학강의실2" + \
" 수학강의실7 물리실험실3 물리실험실2 지구과학실험실1 지구과학실험실2 " + \
" 생물실험실3 화학실험실2 화학RnE실" + \
" 융합과학실" + \
" 지구과학강의실1 소회의실 체육관" ).split()


all_classrooms.sort()




default_classroom_assignments_class = { 
"기초통계학_6": "수학강의실3",
"기초통계학_7": "수학강의실3",
"기초통계학_8": "수학강의실3",
"미적분학II_1": "수학강의실1",
"미적분학II_2": "수학강의실1",
"미적분학I_1": "수학강의실2",
#"선형대수학_1": "수학강의실7",
#"선형대수학_2": "수학강의실7",
"선형대수학_3": "수학강의실7",
"선형대수학_4": "수학강의실7",
"선형대수학_5": "수학강의실7",
"선형대수학_6": "수학강의실1",
"수리정보탐구_1": "정보강의실",
"수리정보탐구_2": "정보강의실",
"수리정보탐구_3": "정보강의실",
"수학III_1": "수학강의실4",
"수학III_2": "수학강의실4",
"수학III_3": "수학강의실6",
"수학III_4": "수학강의실6",
"수학III_5": "수학강의실6",
"수학III_6": "수학강의실4",
"수학III_7": "수학강의실4",
"수학III_8": "수학강의실4",
#"수학I_1": "수학강의실3",
"수학I_2": "수학강의실3",
"수학I_3": "수학강의실3",
#"수학I_4": "수학강의실3",
"수학I_5": "수학강의실5",
"수학I_6": "수학강의실5",
"수학I_7": "수학강의실5",
"수학I_8": "수학강의실5",
"정수론_1": "수학강의실1"}




def assign_default_classrooms() : 
    for subj in sem.default_classroom_assignments_subject : 
        for pt in parts_of(subj) : 
            classrooms[pt] = sem.default_classroom_assignments_subject[subj] 
            #print "%s %s -> %s" % (subj, classpart_slot[pt], classrooms[pt])
            print("{} {} -> {}".format(subj, classpart_slot[pt], 
                classrooms[pt]))
    for clss in sem.default_classroom_assignments_class : 
        for pt in parts_of(clss) : 
            classrooms[pt] = sem.default_classroom_assignments_class[clss] 
            #print "%s %s -> %s" % (clss, classpart_slot[pt], classrooms[pt])
            print("{} {} -> {}".format(clss, classpart_slot[pt], 
                classrooms[pt]))
    for pt in sem.default_classroom_assignments_part : 
        classrooms[pt] = sem.default_classroom_assignments_part[pt]
        #print "%s %s -> %s" % (pt, classpart_slot[pt], classrooms[pt])
        print("{} {} -> {}".format(pt, classpart_slot[pt], classrooms[pt]))

def assign_classroom( clss, room ) : 
    for pt in parts_of(clss) : 
        classrooms[pt] = room
        #print "%s %s -> %s" % (clss, classpart_slot[pt], classrooms[pt])
        print("{} {} -> {}".format(clss, classpart_slot[pt], classrooms[pt]))

def occupied_classrooms(slot) : 
    output = [ ] 
    for pt in classparts : 
        if classpart_slot[pt] == slot : 
            if classrooms[pt] is not None : 
                output.append( classrooms[pt] )
    return output 
    
def free_classrooms(slot) : 
    occ_rooms = occupied_classrooms(slot) 
    for room in occ_rooms[:] : 
        if room.find("/") > 0 : 
            occ_rooms.extend( room.split("/"))
    output =  sorted( list_util.complement( sem.all_classrooms, occ_rooms ) )
    return output
            
def print_class_info(clss):
    subj = subject_name_of(clss)
    #print "%s, %.1f" % (subj, average_size(subj)), 
    print("{}, {:.1f}".format(subj, average_size(subj)), end=' ')
    #print ' '.join( classes_of(subj) ), 
    print(' '.join( classes_of(subj) ), end=' ')
    t_names = [n for n in members(clss) if n.endswith("t") ]
    if len(t_names) > 0 : 
        print("("+str(hours(subj))+")", t_names[0])
    else :
        print("("+str(hours(subj))+")" )
    tmp = ' '.join( [ str(classpart_slot[pt]) for pt in parts_of(clss) ] ) 
    #print "<%s>"% tmp, 
    print("<{}>".format(tmp), end=' ')
    #print len( st_members_of(clss) ),
    print(len( st_members_of(clss) ), end=' ')
    print(' '.join( list_util.union( 
            [ str( classrooms[pt] ) for pt in parts_of(clss)]) ))
    for pt in parts_of(clss) :
        print(korean_ljust(pt, 20),end=" " )
        if classpart_slot[pt] is not None : 
            print(":"+ classpart_slot[pt] + ":",end=" ")
        else :
            print(":"+ "   " + ":",end=" ")
        print(slots_in_tabbing( available_slots_of(pt), [classpart_slot[pt]] ))

    print(' '.join( members(clss)))
    print()


    
def free_classrooms_for_class(clss) : 
    output = list_util.intersect( * 
            [ free_classrooms( classpart_slot[pt] ) for pt in parts_of(clss) ])
    list_util.npr(output) 
    print_class_info(clss)
    return output




def double_booking_check( allow = ["생활체육", "융합과학탐구"] ) :  
    slot_rooms = [ ] 
    for pt in classparts : 
        if subject_name_of(pt) in allow : 
            continue
        if classpart_slot[pt] is not None : 
            if classrooms[pt] is not None : 
                for room in classrooms[pt].split("/") :
                    slot_rooms.append( ( classpart_slot[pt], room ) )

    dup_list = [ item for item in slot_rooms if slot_rooms.count(item) > 1 ]

    if len(dup_list) == 0 : 
        return True 
    else : 
        for sl_rm in sorted(list_util.union( dup_list ) ) : 
            for cp in classparts : 
                if classrooms[cp] is None : 
                    continue
                if  sl_rm[0] == classpart_slot[cp] \
                        and sl_rm[1] in  classrooms[cp].split("/"):
                    print(' '.join(t_members_of(cp)),end=" " )
                    print(cp, sl_rm[0], classrooms[cp] )
        return False 


#def assign_classroom_to_classes(subj, c_nums, room ) : 
    #for num in c_nums : 
        #assign_classroom( subj + "_" + str(num), room)


def assign_classroom_to_part(pt, room) : 
    classrooms[pt] = room
    #print "%s %s -> %s" % (pt, classpart_slot[pt], classrooms[pt])
    print("{} {} -> {}".format(pt, classpart_slot[pt], classrooms[pt]))


def part_using_slot_classroom( slot, room ) : 
    output = [ ] 
    for pt in classparts : 
        if classpart_slot[pt] == slot and classrooms[pt] == room : 
            output.append(pt)
    list_util.jpr(output)
    return ' '.join(output)



def convert_name_for_xlsx(name) : 
    if name.endswith("t") : 
        return name[:-1]
    else : 
        return re.sub( r'[ABCD]$', '',  
                re.sub( r'(\d{4})', r'\1 ', name))





def name_slot_part(name, myslot) : 
    pt_list = [pt for n, pt in name_classparts 
                    if classpart_slot[pt] == myslot and n == name  ]
    return ' '.join(sorted(pt_list))
    

def class_slot_part(clss, myslot) : 
    pt_list = [pt for pt in parts_of(clss) 
                    if classpart_slot[pt] == myslot ]
    return ' '.join(pt_list)
    



def export_weekly_tables_xlsx( filename, in_name_list = None ) : 
    if in_name_list is None : 
        name_list = sorted( [n for n in teachers if n.find("조교") < 0])
    else : 
        name_list = in_name_list
    print("Reading style from <weekly-table-template.xlsx>." )
    shutil.copyfile( "weekly-table-template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb[ "Template" ]

    for name in name_list : 
        name_in_xlsx = convert_name_for_xlsx(name)
        #print "Writing weekly table for %s." % name
        print("Writing weekly table for {}.".format(name))
        new_worksheet = wb.create_sheet(  name_in_xlsx)
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        new_worksheet.page_setup.orientation='landscape'
        fill_in_time_table(new_worksheet, 0 , name )
    wb.remove(template_ws)
    #print "Saving to <%s>."%filename
    print("Saving to <{}>.".format(filename))
    wb.save(filename)




def show_stats( subj_list = [ ], exclude_juniors = True ) : 
    if len(subj_list) > 0 : 
        sublist = subj_list 
    elif exclude_juniors is True : 
        sublist = list_util.complement(subjects, sem.junior_subjects) 
    else : 
        sublist = subjects
    for subj in sublist : 
        print(subjects.index(subj), subj,end=" " )
        #print "("+ str(hours(subj))  +")", "%.1f"% average_size(subj)
        print("("+ str(hours(subj))  +")", "{:.1f}".format(average_size(subj)))
        for c in classes_of(subj) : 
            #print c[-2:], "<%d>" % (len(st_members_of(c)),end=" " ), 
            print(c[-2:], "<{}>".format(len(st_members_of(c))), end=' ')
        print("\n")
    


def read_subject_names(filename) : 
    hyeonhwang = xlsx_util.read_xlsx_sheet_into_dict( filename, 3, 1, 
                    "수강신청현황")
    subject_dept = { } 
    row_names = list_util.union( [ c[0] for c in hyeonhwang ] )
    for r in row_names : 
        dept = hyeonhwang[ (r, "교과") ] 
        subj  = hyeonhwang[ (r, "과목명") ] 
        subject_dept[subj] = dept 
        #print "%s -> %s" % ( subj, subject_dept[subj] ) 
        print("{} -> {}".format( subj, subject_dept[subj] ))
    return subject_dept

sub_group_names = ["국어", "사회", "수학", "공통", "물리", "화학",
                    "생물", "지구과학", "정보", "영어", "중국어", "음악",
                        "미술", "예술", "체육", "보건상담"]


def sorted_by_days( my_list ) : 
    return sorted( my_list, key = (lambda x : (days.index(x[:-1]), x[-1]) ) )






#def strip_typeinfo(c) :
    #if type(c) is str :
        #return re.sub(r'^[^\:]+:', '',  c)
    #elif type(c) is unicode :
        #return re.sub(r'^[^\:]+:', '',  c).replace(u'\ufeff',
            #'').encode('utf-8')
    #elif type(c) in [float, int, long] :
        #return str(c)
    #else :
        #return ''

def convert_romans( mystr ) :
    conv_pairs = [("\xe2\x85\xa0", "I"), ("\xe2\x85\xa1", "II"),
            ("\xe2\x85\xa2", "III"), ("\xe2\x85\xa3", "IV"), (" ", "")]
    output = mystr
    for c in conv_pairs :
        output = output.replace( c[0], c[1])
    return output


def name_subject_pairs_from_neis_xlsx(filename) : 
    records = xlsx_util.read_xlsx_sheet_into_list(filename) 
    output = [ ] 
    curr_subject = None 
    grades = map(str, range(1,4))
    c_nums = map(str, range(1,9))
    for r in records : 
        if r[1].startswith("2017") : 
            curr_subject = \
                utf_util.convert_romans(re.sub( 
              r'\d{4}학년도\s*\d학년\s*\d학기\s*', '', r[1])).replace(" ", "")
            print(curr_subject)
        elif curr_subject is not None \
                and (r[2] in grades) and (r[4] in c_nums) : 
            st_name = str(int(r[2])*1000 + int(r[4])*100 + int(r[5])) + r[6] 
            output.append( (st_name, curr_subject) )
            print(output[-1][0], "->", output[-1][1])
    return output
            


classroom_shorthand = {
"국어강의실1":"국강1", 
"국어강의실2":"국강2", 
"국어강의실3":"국강3", 
"물리강의실1":"물강1", 
"물리강의실2":"물강2", 
"물리실험실1":"물실1", 
"물리실험실2":"물실2", 
"물리실험실3":"물실3", 
"미술실":"미술실", 
"사회강의실1":"사강1", 
"사회강의실2":"사강2", 
"사회강의실3":"사강3", 
"생물강의실1":"생강1", 
"생물강의실2":"생강2", 
"생물실험실1":"생실1", 
"생물실험실2":"생실2", 
"생물실험실3":"생실3", 
"소회의실":"소회의실", 
"수학강의실1":"수강1", 
"수학강의실2":"수강2", 
"수학강의실3":"수강3", 
"수학강의실4":"수강4", 
"수학강의실5":"수강5", 
"수학강의실6":"수강6", 
"수학강의실7":"수강7", 
"외국어강의실1":"외강1", 
"외국어강의실2":"외강2", 
"외국어강의실3":"외강3", 
"외국어강의실4":"외강4", 
"외국어강의실5":"외강5", 
"우암공통강의실1":"우공1", 
"우암공통강의실2":"우공2", 
"우암공통강의실3":"우공3", 
"우암공통강의실4":"우공4", 
"융합공통강의실":"융공강", 
"음악실":"음악실", 
"정보강의실":"정보강", 
"지구과학강의실1":"지강1", 
"지구과학강의실2":"지강2", 
"지구과학실험실1":"지실1", 
"지구과학실험실2":"지실2", 
"체육관":"체육관", 
"컴퓨터실1":"컴실1", 
"컴퓨터실2":"컴실2", 
"정보강의실/컴퓨터실2":"정보/컴실2", 
"상상토의실":"상상실", 
"화학RnE실":"화학RnE", 
"화학강의실1":"화강1", 
"화학강의실2":"화강2", 
"화학강의실3":"화강3", 
"화학실험실1":"화실1", 
"화학실험실2":"화실2", 
"미술실/음악실":"미/음",
"예지융합실험실":"예융실", 
"예지공통실험실":"예공실", 
"융합과학실":"융과실", 
"예지공통실험실/생물실험실3/물리실험실2/화학실험실2":"해당실험실", 
"예지공통실험실/생물실험실3":"예공실/생실3",
"융합과학실/생물실험실3/물리실험실2/화학실험실2":"해당실험실", 
"융합과학실/생물실험실3":"융과실/생실3"
, "물리실험실2/우암공통강의실2":"물실2/우공2"
}


def join_clss_room(clss, room) : 
    if len(clss)>=19 : 
        my_joiner = " "
    else :
        my_joiner = "\n"
    #return "%s%s%s" % (clss, my_joiner, room) 
    return "{}{}{}".format(clss, my_joiner, room) 


def export_teacher_table_xlsx(filename, skip = sem.skip_subject_list ) :  
    column_names = [c.replace("t", "") for c in teachers if c.find('조교')<0 ]
    row_names = slots
    slot_teacher_clss = { }
    for n in column_names : 
        slot_teacher_clss[ ("과목", n) ] = '\n'.join( [subj 
            for subj in sorted(subjects_of(n + "t")) 
                if subj not in skip ])
        slot_teacher_clss[ ("부장/담임", n) ] = bujang_str(n) 
        for curr_slot in row_names : 
            curr_part = name_slot_part(n + "t", curr_slot)
            if len(curr_part) > 0 and subject_name_of(curr_part) not in skip : 
                curr_room = sem.classroom_shorthand[classrooms[ 
                    curr_part.split()[0]  ] ]
                class_and_room = join_clss_room(re.sub( r'_(\d+)', r'(\1)', 
                    class_name_of(curr_part)),  curr_room )
                slot_teacher_clss[ (curr_slot, n) ] = class_and_room
                #print "Found (%s,%s) -> %s" % (curr_slot, n, class_and_room.replace("\n", " "))
                print("Found ({},{}) -> {}".format(curr_slot, n,
                                        class_and_room.replace("\n", " ")))
    t_list = teachers[:]
    t_list.sort( key = lambda x : 
            (sem.ordered_departments.index(sem.name_department_dict[x]), 
               x ))

    nums = range(1,len(t_list)+1)
    slot_num_clss = { }
    for i,t_name in enumerate(t_list) : 
        xl_name = t_name.replace("t","")
        ind = str(i + 1)
        slot_num_clss[ ("이름", ind) ] = xl_name 
        slot_num_clss[ ("교과", ind) ] = sem.name_department_dict[t_name] 
        slot_num_clss[ ("시수", ind) ] = sem.hour_sum_formula(ind)
        for (curr_slot, key_name ) in slot_teacher_clss : 
            if key_name == xl_name : 
                slot_num_clss[ (curr_slot, ind) ] = \
                        slot_teacher_clss[ (curr_slot, key_name) ] 

    xlsx_util.fill_in_xlsx_by_dict(filename, "big_time_table_template.xlsx", 
                 slot_num_clss , 1, 2  )




def show_info( a_name ) : 
    print(a_name)
    curr_classes = current_classes(a_name)
    curr_subjs = subjects_of(a_name) 
    curr_classes.sort()
    curr_subjs.sort()
    for s in curr_subjs : 
        #print "(%d) %s" % ( hours(s), s),
        print("({}) {}".format( hours(s), s), end=' ')
    #print "%d hours in total." % total_hours(a_name), "\n" 
    print("{} hours in total.".format(total_hours(a_name)),"\n")
    for c in curr_classes : 
        print(c)
        print_class_info(c)
    #print "%d hours in total." % sum( [ hours(c) for c in curr_classes])
    print("{} hours in total.".format(sum( [ 
                        hours(c) for c in curr_classes])))
    show_possible_slots(a_name)
    list_util.jpr( remaining_subjects(a_name) )




def export_student_table_xlsx(filename, teacher_names = False) :  
    row_names = students
    column_names = slots
    student_slot_clss = { }
    for n in row_names : 
        for curr_slot in column_names : 
            curr_part = name_slot_part(n, curr_slot)
            if len(curr_part) > 0 : 
                if not teacher_names : 
                    curr_room = sem.classroom_shorthand[classrooms[curr_part] ]
                    class_and_room = join_clss_room(re.sub( r'_(\d+)', r'(\1)', 
                        class_name_of(curr_part)),  curr_room )
                    student_slot_clss[ (n, curr_slot) ] = class_and_room
                else : 
                    t_names = ' '.join( [ convert_name_for_xlsx(c) for c 
                      in  t_members_of(curr_part)  if not is_assistant_name(c)])
                    student_slot_clss[ (n, curr_slot) ] = '\n'.join( [ 
                        re.sub( r'_(\d+)', r'(\1)', 
                            class_name_of(curr_part)),  t_names] )
                #print "Found (%s,%s) -> %s" % (n, curr_slot, student_slot_clss[ (n, curr_slot) ])
                print("Found ({},{}) -> {}".format(n, curr_slot,
                                        student_slot_clss[ (n, curr_slot) ]))

    template_filename = "students_big_time_table_template.xlsx"
    print("Reading style from <%s>" % template_filename)
    shutil.copyfile( template_filename, filename)
    print("Copied %s to %s" %( template_filename, filename))
    wb = openpyxl.load_workbook(filename)
    ws = wb[  "학생시간표" ]
    xlsx_util.fill_in_sheet_by_list(ws, row_names, 2, 1, direction="column")
    xlsx_util.fill_in_sheet_by_list(ws, column_names, 1, 2, direction="row")

    xlsx_util.fill_in_sheet_by_dict(ws, student_slot_clss, 1, 1  )
    #print "Saving to <%s>.." % filename
    print("Saving to <{}>..".format(filename))
    wb.save( filename )


def export_baedangpyo_xlsx(filename) : 
    row_names = list_util.flatten( [ classes_of(subj) for subj in subjects ] )
    column_names = "시간1 시간2 시간3 시간4 담당자1 강의실1".split()
    baedang_info = { } 
    for clss in row_names : 
        for i, pt in enumerate( sorted( parts_of(clss) )) : 
            teacher_list = [ n.replace("t","")  
                for n in t_members_of(pt) if n.find("조교")<0 ]
            baedang_info[ (clss, "담당자" + str(i+1))] = " ".join(teacher_list)
            baedang_info[ (clss, "시간"+str(i+1)) ] = str(classpart_slot[pt]) 
            baedang_info[ (clss, "강의실"+str(i+1)) ] = str(classrooms[pt]) 
        for i, pt in enumerate( reversed(sorted( parts_of(clss) ))) : 
            c_num = len(parts_of(clss)) - i 
            if c_num > 1 : 
                for key_str in "담당자 강의실".split() : 
                    if baedang_info[ (clss, key_str + str(c_num))] == \
                        baedang_info[ (clss, key_str + str(c_num-1))] : 
                            baedang_info[ (clss, key_str + str(c_num))] = "-"
        
        #print "%s -> %s %s %s" % (clss, " ".join(slots_of(clss)), " ".join(t_members_of(clss)), "/".join( [classrooms[pt] for pt in parts_of(clss)]))
        print("{} -> {} {} {}".format(clss, " ".join(slots_of(clss)),
                          " ".join(t_members_of(clss)),
                    "/".join( [classrooms[pt] for pt in parts_of(clss)])))
    xlsx_util.fill_in_xlsx_by_dict(filename, 
        "baedangpyo_template.xlsx", baedang_info , 1, 3  )


def room_slot_class(room, sl) : 
    pts = [ pt for pt in classparts 
         if (room in str(classrooms[pt]).split("/")) 
                and classpart_slot[pt] == sl ]
    return " ".join( list_util.union( [class_name_of(pt) for pt in pts]))


def room_slot_part(room, sl) : 
    pts = [ pt for pt in classparts 
         if (room in str(classrooms[pt]).split("/")) 
                    and classpart_slot[pt] == sl ]
    return " ".join( list_util.union( pts ))



def export_room_table_xlsx(filename) :  
    row_names = sem.all_classrooms
    column_names = slots
    room_slot_clss = { }
    for room in row_names : 
        for curr_slot in column_names : 
            curr_class = room_slot_class(room, curr_slot)
            if len(curr_class) > 0 : 
                para_cname = para_classname( curr_class)
                room_slot_clss[ (room, curr_slot) ] = para_cname
                #print "Found (%s,%s) -> %s" % (room, curr_slot, room_slot_clss[ (room, curr_slot) ] )
                print("Found ({},{}) -> {}".format(room, curr_slot,
                                room_slot_clss[ (room, curr_slot) ] ))
    xlsx_util.fill_in_xlsx_by_dict(filename, 
        "rooms_big_time_table_template.xlsx", room_slot_clss , 1, 1  )




def export_bunban_table_xlsx(filename) :  
    template_filename = "bunban_baejeong_template.xlsx"

    #print "Reading style from <%s>" % template_filename
    print("Reading style from <{}>".format(template_filename))
    shutil.copyfile( template_filename, filename)
    #print "Copied %s to %s" %( template_filename, filename)
    print("Copied {} to {}".format( template_filename, filename))
    wb = openpyxl.load_workbook(filename)
    row_names = students
    column_names = subjects

    ws = wb[  "분반배정표" ]

    for i, row in enumerate(ws.rows) : 
        if i > 0 and row[1].value.startswith("=") : 
            start_row_num = i + 1
            break 

    xlsx_util.fill_in_sheet_by_list(ws, row_names, start_row_num, 
                1, direction="column")
    xlsx_util.fill_in_sheet_by_list(ws, column_names, 1, 3, direction="row")
    student_subject_class = { }
    for n in row_names : 
        curr_classes = current_classes(n)
        for clss in curr_classes : 
            subj = subject_name_of(clss)
            c_num = int(class_number_of( clss ))
            student_subject_class[ (n, subj) ] = c_num 
            #print "Found (%s,%s) -> %d" % (n, subj, student_subject_class[ (n, subj) ] )  
            print("Found ({},{}) -> {}".format(n, subj,
                        student_subject_class[ (n, subj) ] ))
    xlsx_util.fill_in_sheet_by_dict(ws, student_subject_class , 1, 1  )
    ws = wb[ "학생별신청목록" ]
    xlsx_util.fill_in_sheet_by_list(ws, row_names, 2, 1, direction="column")
    xlsx_util.fill_in_sheet_by_list(ws, column_names, 1, 3, direction="row")
    student_subject_hours = { }
    for n in row_names : 
        subjs = subjects_of(n) 
        for subj in subjs : 
            student_subject_hours[ (n, subj) ] = hours(subj) 
            #print "Found (%s,%s) -> %d" % (n, subj, student_subject_hours[ (n, subj) ] )  
            print("Found ({},{}) -> {}".format(n, subj,
                        student_subject_hours[ (n, subj) ] ))
    xlsx_util.fill_in_sheet_by_dict(ws, student_subject_hours , 1, 1  )
    #print "Saving to <%s>.." % filename
    print("Saving to <{}>..".format(filename))
    wb.save( filename )

def shorten_slot_list( slot_list ) : 
    curr_day = None
    output = [ ] 
    for sl in slot_list : 
        if sl[:-1] != curr_day : 
            curr_day = sl[:-1]
            output.append(" ")
            output.append( sl ) 
        else : 
            output.append( sl[-1:] )
    return (''.join(output))[1:]

def export_teacher_schedule_txt(filename) : 
    t_list = [n for n in teachers if n.find("조교t") < 0 ];
    with open(filename, "wb") as f : 
        for n in t_list : 
            f.write( (n[:-1] + " " +  shorten_slot_list( 
                sorted(used_slots_of(n, list_util.complement( 
   sem.skip_subject_list, "부장회의 담임업무".split())  ), 
        key = slots.index))).encode())
            f.write("\n".encode())
        #print "Saving to <%s>" % filename
        print("Saving to <{}>".format(filename))


def apply_shortening(my_str, shortening_dict) : 
    output = my_str
    for k in shortening_dict : 
        output = output.replace(k, shortening_dict[k])
    return output


def export_weekly_tables_of_classes( filename, class_list ) : 
    print("Reading style from <weekly-table-template.xlsx>." )
    shutil.copyfile( "weekly-table-template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb[ "Template" ]
    for clss in class_list : 
        clss_in_xlsx = re.sub( r'_(\d+)', r'(\1)', clss ).replace("/", "_")
        #print "Writing weekly table for %s." % clss
        print("Writing weekly table for {}.".format(clss))
        new_worksheet = wb.create_sheet( clss_in_xlsx)
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        new_worksheet.cell( row = 1, column = 3 ).value = clss_in_xlsx 
        for i in range(7) : 
            for j in range(5) : 
                if (i,j) in slot_names : 
                    curr_slot = slot_names[ (i,j) ] 
                    curr_part = class_slot_part(clss, curr_slot)
                    if len(curr_part) > 0 : 
                        curr_teachers = ' '.join( t_members_of(curr_part))
                        class_and_teachers = re.sub( r'_(\d+)', r'(\1)', 
                             class_name_of(curr_part))  + "\n" + curr_teachers
                        new_worksheet.cell( row = i + 3, 
                                            column = j + 2 ).value = \
             apply_shortening(class_and_teachers, sem.weekly_table_shortening)
    wb.remove(template_ws)
    #print "Saving to <%s>."%filename
    print("Saving to <{}>.".format(filename))
    wb.save(filename)


def fill_in_time_table(new_worksheet, offset, st_name, t_name = False ) : 
    name_in_xlsx = convert_name_for_xlsx(st_name)
    #print "Writing weekly table for %s." % name_in_xlsx
    print("Writing weekly table for {}.".format(name_in_xlsx))
    new_worksheet.cell( row = 1 + offset, column = 2 ).value = sem.semester_str
    new_worksheet.cell( row = 1 + offset, column = 3 ).value = name_in_xlsx 
    if sem.is_in_sci_dept(st_name) : 
        for i in range(4,7) : 
            new_worksheet.cell( row = i + 3 + offset, column = 2+2).value = \
                    "연구활동지도" 
    for i in range(7) : 
        for j in range(5) : 
            if (i,j) in slot_names : 
                curr_slot = slot_names[ (i,j) ] 
                curr_part = name_slot_part(st_name, curr_slot)
                if len(curr_part) > 0 and \
                  subject_name_of(curr_part) not in sem.skip_subject_list: 
                    curr_room = str(classrooms[ curr_part.split()[0]  ] )
                    t_names = ' '.join( [ n[:-1] 
                      for n in t_members_of(curr_part) if n.find("조교") < 0])
                    if t_name is True : 
                        class_and_room = re.sub( r'_(\d+)', r'(\1)', 
                            class_name_of(curr_part))  + "\n" + t_names
                    else : 
                        class_and_room = re.sub( r'_([,\d]+)', r'(\1)', 
                            class_name_of(sem.part_merge(curr_part))) \
                                + "\n" + curr_room
                        #class_and_room = re.sub( r'_([,\d]+)', r'(\1)', 
                            #class_name_of(curr_part)) \
                                #+ "\n" + curr_room
                    new_worksheet.cell( row = i + 3 + offset, 
                                        column = j + 2 ).value = \
                  apply_shortening(class_and_room, sem.weekly_table_shortening)





ban_list =  re.split(r'\s+', 
'''1-1 1-2 1-3 1-4 1-5 1-6 1-7 1-8
2-1 2-2 2-3 2-4 2-5 2-6 2-7 2-8
3-1 3-2 3-3 3-4 3-5 3-6 3-7 3-8''')


def export_weekly_tables_of_all_classes( filename, class_list = sem.ban_list, 
                                         t_name = False, st_list = None ) : 
    print("Reading style from <weekly-table-template-for-class.xlsx>." )
    shutil.copyfile( "weekly-table-template-for-class.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb["Template"]
    for clss in class_list : 
        prefix = clss.replace("-", "")
        name_list = [nm for nm in names if nm.startswith(prefix)]
        if st_list is not None : 
            name_list = list_util.intersect(name_list, st_list)
        name_list.sort()
        if(len(name_list)==0) : 
            continue

        #print "Writing weekly table for %s." % clss
        print("Writing weekly table for {}.".format(clss))
        new_worksheet = wb.create_sheet(  clss)
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        new_worksheet.page_setup.orientation='landscape'

        for k in range( 15, 401, 20  ): 
            new_worksheet.row_dimensions[k].hidden = True
        for k in range(20,401,20):
            #new_worksheet.page_breaks.append( Break(id=k) )
            new_worksheet.row_breaks.append( Break(id=k) )

        for i, st_name in enumerate(name_list) : 
            fill_in_time_table(new_worksheet, i*20, st_name, t_name=t_name)
        for k in range( (i+1)*20 + 1, 401  ): 
            new_worksheet.row_dimensions[k].hidden = True
    wb.remove(template_ws)
    #print "Saving to <%s>."%filename
    print("Saving to <{}>.".format(filename))
    wb.save(filename)






def para_classname(clss) : 
    return re.sub( r'_(\d+)', r'(\1)', clss )

def export_weekly_table_for_neis( clss ) : 
    #filename = "%s.xlsx" % para_classname(clss)
    filename = "{}.xlsx".format(para_classname(clss))
    output = { }
    neis_subj = neis_subject_name[subject_name_of(clss)]
    for pt in parts_of(clss) : 
        my_slot = classpart_slot[pt]
        my_day = my_slot[:-1] + "요일"
        my_time = my_slot[-1:]  + "교시"
        output[ ( my_time, my_day ) ] = neis_subj

    xlsx_util.fill_in_xlsx_by_dict(filename, "NEIS_timetable_form.xlsx", 
                 output , 1, 1  )



def neis_input_show_students( subj, c_num, ex_cls = None ) : 
    if ex_cls is None : 
        ex_classes = [ subj + "_" + str(k) for k in range(1,c_num) ] 
    else : 
        ex_classes = [ subj + "_" + str(k) for k in  ex_cls ] 
    clss = subj + "_" + str(c_num)
    my_list = list_util.complement( members(subj), 
                    * [st_members_of(c) for c in ex_classes] )
    my_list.sort()
    my_members = st_members_of(clss)
    counter = 0
    for n in my_list : 
        if n in my_members : 
            counter += 1
            print(str(counter).rjust(2), n, " ", n )
        else  : 
            print("  ", "          ", " ", n)
    


def export_classes_on_weekly_table(filename) : 
    output = { }
    for sl in slots : 
        curr_classes = sorted( [ 
          class_name_of(pt) for pt in classparts if sl == classpart_slot[pt]])
        available_classes = [c for c in curr_classes 
                                if len(st_members_of(c)) < 20 and 
                  not ( c.find("실험") >= 0 and len(st_members_of(c)) >=16) ]
        if len(available_classes) > 0 : 
            output[(sl[-1:] + "교시", sl[:-1]  )] = '\n'.join( [  
                para_classname(clss) + "<" + ' '.join( [ 
                  str(classpart_slot[pt]) for pt in parts_of(clss) ] ) + ">" 
                    for clss in available_classes] )

    xlsx_util.fill_in_xlsx_by_dict(filename, 
            "classes-weekly-table-template.xlsx", output , 2, 1  )


def current_slots( classname ) : 
    my_slots = list_util.complement( 
                [ classpart_slot[pt] for pt in parts_of(classname) ], 
                [None])
    my_slots.sort(key = slots.index)
    return my_slots


absent_students = "".split()





def export_chulseokbu( class_name ) : 
    filename = "출석부_" + para_classname(class_name) + ".xlsx"
    name_list = sorted( list_util.complement(st_members_of(class_name), 
                                sem.absent_students) )
    print("Reading style from <chulseokbu-template.xlsx>." )
    shutil.copyfile( "chulseokbu_template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb["Template"]
    new_worksheet = wb.create_sheet( para_classname(class_name) )
    instance = WorksheetCopy(template_ws, new_worksheet)
    WorksheetCopy.copy_worksheet(instance)
    new_worksheet.cell( row = 1, column = 3 ).value =para_classname(class_name)
    new_worksheet.cell( row = 1, column = 4 ).value = \
                ' '.join(current_slots(class_name)) + "\n" \
                + '/'.join( list_util.union( 
            [ classrooms[pt] for pt in parts_of(class_name) ] ))
    new_worksheet.cell( row = 1, column = 8 ).value = \
        '\n'.join([ convert_name_for_xlsx(n) 
                for n in t_members_of(class_name) if n.find("조교")<0])
    for i, name in enumerate(name_list) : 
        name_in_xlsx = convert_name_for_xlsx(name)
        #print "Adding <%s> to name list." % name
        print("Adding <{}> to name list.".format(name))

        new_worksheet.cell( row = i+3, column = 2 ).value = i + 1 
        new_worksheet.cell( row = i+3, column = 3 ).value = name_in_xlsx
    wb.remove(template_ws)
    #print "Saving to <%s>."%filename
    print("Saving to <{}>.".format(filename))
    wb.save(filename)





def export_chulseokbu_of_subject( subj ) : 
    filename = "분반별명단_" + sem.subject_group_dict[subj]  + "_" \
                    + subj.replace("/", "_")  + ".xlsx"

    print("Reading style from <chulseokbu-template.xlsx>." )
    shutil.copyfile( "chulseokbu_template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb["Template"]
    for class_name in classes_of(subj) : 
        name_list = sorted( list_util.complement(st_members_of(class_name), 
                                sem.absent_students) )
        new_worksheet = wb.create_sheet(
                       para_classname(class_name).replace("/", "_"))
        #print("Creating Sheet <%s>.." % (para_classname(class_name), ))
        print("Creating Sheet <{}>..".format(para_classname(class_name)))
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        new_worksheet.cell( row = 1, column = 3 ).value = \
                    para_classname(class_name)
        new_worksheet.cell( row = 1, column = 4 ).value = \
                 ' '.join(current_slots(class_name)) + "\n" \
                        + '/'.join( list_util.union( 
                    [ classrooms[pt] for pt in parts_of(class_name) ] ))
        new_worksheet.cell( row = 1, column = 8 ).value = \
            '\n'.join([ convert_name_for_xlsx(n) 
                for n in t_members_of(class_name) if n.find("조교")<0])
        for i, name in enumerate(name_list) : 
            name_in_xlsx = convert_name_for_xlsx(name)
            #print "Adding <%s> to name list." % name
            print("Adding <{}> to name list.".format(name))
            new_worksheet.cell( row = i+3, column = 2 ).value = i + 1 
            new_worksheet.cell( row = i+3, column = 3 ).value = name_in_xlsx

    wb.remove(template_ws)
    #print "Saving to <%s>."%filename
    print("Saving to <{}>.".format(filename))
    wb.save(filename)




def export_chulseokbu_of_all_subjects_xlsx( filename, phone_number = False ) : 
    if phone_number is True:
        name_id_dict = dict( [ (c[0],  c[2] + "@sshs.hs.kr") 
                for c in sem.st_info_tuples ] )
    print("Reading style from <chulseokbu_template_compact.xlsx>." )
    shutil.copyfile( "chulseokbu_template_compact.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb["Template"]
    s_list = subjects[:]
    s_list.sort(key = (lambda x: sem.subject_group_dict[x] + x) ) 
    for subj in s_list : 
        c_list = classes_of(subj)
        #c_list.sort( key = lambda x:("%02d"% int(class_number_of(x))) )
        c_list.sort( key = lambda x:("{:02d}".format(int(class_number_of(x)))))
        new_worksheet = wb.create_sheet( subj.replace("/", "_"))
        #print "Creating Sheet <%s>.." % subj  
        print("Creating Sheet <{}>..".format(subj))
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        for k, class_name in enumerate(c_list) : 
            name_list = sorted( list_util.complement(st_members_of(class_name), 
                                sem.absent_students) )
            new_worksheet.cell( row = 1, column = 3 + k*11 ).value = \
                    para_classname(class_name)
            new_worksheet.cell( row = 1, column = 4  + k*11).value = \
                 ' '.join(current_slots(class_name)) + "\n" \
                        + '/'.join( list_util.union( 
                    [ classrooms[pt] for pt in parts_of(class_name) ] ))
            new_worksheet.cell( row = 1, column = 8  + k*11).value = \
                '\n'.join([ convert_name_for_xlsx(n) 
                    for n in t_members_of(class_name) if n.find("조교")<0])
            for i, st_name in enumerate(name_list) : 
                name_in_xlsx = convert_name_for_xlsx(st_name)
                #print "Adding <%s> to name list." % st_name
                print("Adding <{}> to name list.".format(st_name))
                new_worksheet.cell( row = i+3, column = 2+k*11 ).value = i + 1 
                new_worksheet.cell( row = i+3, 
                    column = 3+k*11).value = name_in_xlsx
                if phone_number is True : 
                    new_worksheet.cell( row = i+3, column = 4+ k*11 ).value = \
                        sem.phone_dict[st_name]
                    new_worksheet.cell( row = i+3, column = 7+ k*11 ).value = \
                        name_id_dict[st_name]
        for j in range( (k+1)*11 + 1, 165 ): 
            new_worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(j)].hidden = True
                
    wb.remove(template_ws)
    #print "Saving to <%s>."%filename
    print("Saving to <{}>.".format(filename))
    wb.save(filename)



def export_chulseokbu_of_classes(filename,  in_classes ) : 
    print("Reading style from <chulseokbu-template.xlsx>." )
    shutil.copyfile( "chulseokbu_template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb["Template"]
    for class_name in in_classes:
        name_list = sorted( list_util.complement(st_members_of(class_name), 
                                sem.absent_students) )
        new_worksheet = wb.create_sheet(
                       para_classname(class_name).replace("/", "_"))
        #print "Creating Sheet <%s>.." % (para_classname(class_name), )
        print("Creating Sheet <{}>..".format(para_classname(class_name)))
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        new_worksheet.cell( row = 1, column = 3 ).value = \
                    para_classname(class_name)
        new_worksheet.cell( row = 1, column = 4 ).value = \
                 ' '.join(current_slots(class_name)) + "\n" \
                        + '/'.join( list_util.union( 
                    [ classrooms[pt] for pt in parts_of(class_name) ] ))
        new_worksheet.cell( row = 1, column = 8 ).value = \
            '\n'.join([ convert_name_for_xlsx(n) 
                for n in t_members_of(class_name) if n.find("조교")<0])
        for i, name in enumerate(name_list) : 
            name_in_xlsx = convert_name_for_xlsx(name)
            #print "Adding <%s> to name list." % name
            print("Adding <{}> to name list.".format(name))
            new_worksheet.cell( row = i+3, column = 2 ).value = i + 1 
            new_worksheet.cell( row = i+3, column = 3 ).value = name_in_xlsx
    wb.remove(template_ws)
    #print "Saving to <%s>."%filename
    print("Saving to <{}>.".format(filename))
    wb.save(filename)


def class_list_of_teachers( t_names ) : 
    output =  list_util.union(* 
                [current_classes( t_n ) for t_n in t_names])
    output.sort( key = lambda x: (subject_name_of(x), int(class_number_of(x))))
    return [c for c in output if subject_name_of(c) not in sem.skip_subject_list]


def export_chulseokbu_of_department( dept ) : 
    t_list = [c for c in sem.name_department_dict 
                if dept == sem.name_department_dict[c]] 
    export_chulseokbu_of_classes(dept + "_" +sem.semester_prefix 
                    +"분반별_명단.xlsx", class_list_of_teachers(t_list) )

def export_chulseokbu_of_all_departments() : 
    d_list = list_util.union( [sem.name_department_dict[c] for c 
                  in sem.name_department_dict ])
    for dept in d_list : 
        export_chulseokbu_of_department( dept ) 
        


def export_slot_t_names_xlsx( filename ) : 
    t_names = [c for c in teachers if c.find("조교") < 0 ] 

    my_slots = list_util.flatten( [ [ d + str(k) for k in range(1,8) ] 
                                        for d in days] )

    output = [ ] 

    for curr_slot in my_slots : 
        curr_t_list = [n[:-1] for n in t_names 
                                    if curr_slot in used_slots_of(n)]
        output.append( [curr_slot] + curr_t_list )

    xlsx_util.write_tuples_into_xlsx( filename, list_util.transposed(output))







def name_class_pairs_from_neis_xlsx(filename) : 
    wb = openpyxl.load_workbook( filename )
    ws = wb.active
    result = re.search( r'(\d)\D(\d)\.', filename )
    grade_class = result.group(1) + result.group(2)
    output = [ ] 
    for i, row in enumerate(ws.rows) : 
        if xlsx_util.strip_typeinfo(row[0].value).startswith("1교시") : 
            name_str = xlsx_util.strip_typeinfo( 
                ws.cell( row = i-2, column = 1 ).value)
            my_name = ''.join([n for n in students 
                            if n.startswith(grade_class) and 
                                n.find(name_str.split()[-1])>= 0])
            tmp = xlsx_util.read_xlsx_partial_table_into_dict(ws, i, 12, 1, 14)
            print(my_name, end=" ")
            for pair in tmp : 
                clss = utf_util.convert_romans(tmp[pair]).replace("(", "_")
                clss = clss.replace(")", "")
                print(clss, end=" ")
                if (my_name, clss) not in output : 
                    output.append( (my_name, clss) )
            print()
    return output
            
def name_class_pairs_from_neis_files(filenames) : 
    output = [ ] 
    for fname in filenames : 
        nc_pairs = name_class_pairs_from_neis_xlsx(fname)
        output.extend(nc_pairs)
    return output

def name_class_pairs_for_neis_diff():
    pairs = list_util.union([ (c[0], class_name_of(c[1])) 
        for c in name_classparts if c[0][0]!="1" and c[0][-1] !="t"] )
    sp_subjects = list_util.prefix_filter("창의융합특강", subjects)
    sp_classes = list_util.flatten([classes_of(subj) for subj in sp_subjects])
    sp_dict = dict( [  (c, re.sub( r'/[^_]+_', '_', c)) for c in sp_classes])
    #list_util.show_dict(sp_dict)
    conv_f = gen_util.dict_to_fn(sp_dict)
    output = [ (p[0], conv_f(p[1])) for p in pairs]
    return output

avoid_series = [ "월1 월2 월3",
"월2 월3 월4",
"월5 월6 월7",
"화1 화2 화3",
"화2 화3 화4",
"화5 화6 화7",
"수1 수2 수3",
"수2 수3 수4",
"수5 수6 수7",
"목1 목2 목3",
"목2 목3 목4",
"목5 목6 목7",
"금1 금2 금3",
"금2 금3 금4",
"금5 금6 금7"]

sec_avoid_series = [ "월1 월2 월3 월4",
"월4 월5 월6 월7",
"화1 화2 화3 화4",
"화4 화5 화6 화7",
"수1 수2 수3 수4",
"수4 수5 수6 수7",
"목1 목2 목3 목4",
"목4 목5 목6 목7",
"금1 금2 금3 금4",
"금4 금5 금6 금7"]

def avoid_series_teachers(avoids = None) : 
    if avoids is None : 
        series = sem.avoid_series 
    else : 
        series = avoids
    t_names = [n for n in teachers if n.find("조교") < 0 ] 
    output = [ ] 
    for n in t_names : 
        slot_str = used_slot_str( used_slots_of(n))
        if list_util.some_true( 
                [ slot_str.find(c) >= 0 for c in series ] ) : 
            print(n, subjects_in_tabbing(n) )
            output.append(n)
    return output
        




def export_weekly_tables_of_rooms( filename, room_list = sem.all_classrooms ) : 
    print("Reading style from <weekly-table-template.xlsx>." )
    shutil.copyfile( "weekly-table-template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb["Template"]
    for room in room_list : 
        #print "Writing weekly table for %s." % room
        print("Writing weekly table for {}.".format(room))
        new_worksheet = wb.create_sheet(  room)
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        new_worksheet.page_setup.orientation='landscape'
        new_worksheet.cell( row = 1, column = 2 ).value = sem.semester_str
        new_worksheet.cell( row = 1, column = 3 ).value = room 
        for i in range(7) : 
            for j in range(5) : 
                if (i,j) in slot_names : 
                    curr_slot = slot_names[ (i,j) ] 
                    curr_part = room_slot_part(room, curr_slot)
                    if len(curr_part) > 0 : 
                        curr_clss = ' '.join( [para_classname(
                            class_name_of(pt)) for pt in curr_part.split()] )
                        t_names = [ n[:-1] for n in t_members_of(curr_part) 
                            if n.find("조교") < 0]
                        class_and_teachers = curr_clss + "\n" \
                                                    + ' '.join(t_names) 
                        new_worksheet.cell( row = i + 3, 
                column = j + 2 ).value = apply_shortening(class_and_teachers, 
                                                sem.weekly_table_shortening)
    wb.remove(template_ws)
    #print "Saving to <%s>."%filename
    print("Saving to <{}>.".format(filename))
    wb.save(filename)


def export_subject_intersections_xlsx(filename) : 
    output = { } 
    for subj_one in subjects : 
        for subj_two in subjects : 
            output[ (subj_one, subj_two) ] = len( list_util.intersect( 
                        members(subj_one), members(subj_two) ))
    xlsx_util.write_dict_into_xlsx(filename, output)


def export_subject_names_xlsx( output_filename ) : 
    output = [ ]
    global subjects
    print("updating subjects..")
    subjects = list(set( [c[1] for c in name_subjects]))
    subjects.sort()
    list_util.npr(subjects)
    for subj in subjects : 
        output.append( [ subj ] + sorted( members(subj) ) )
    xlsx_util.write_tuples_into_xlsx( output_filename, 
                list_util.transposed(output))

def export_name_subjects_xlsx(filename) : 
    output = [ ] 
    for st_name in students : 
        output.append( [st_name] + sorted( subjects_of(st_name)) )
    xlsx_util.write_tuples_into_xlsx( filename, output)


        

def export_compatible_subjects_xlsx(filename) : 
    output = [ ]
    for subj in subjects : 
        output.append( [ subj ] 
            +  [ ss if len( list_util.intersect( 
                        members(subj), members(ss) )) == 0 else "" 
                            for ss in subjects ]  )
    xlsx_util.write_tuples_into_xlsx( filename, output)
        









def read_enrollment_db(txt_filename) :
    global name_subjects 
    name_subjects = [ ]
    with open(txt_filename) as f :
        for line in f : 
            sr = re.search( r'(.*),\d{5},(\S+),(\d+),(\d+),(\d+)', line)
            if sr is not None : 
                subj = utf_util.convert_romans(sr.group(1))
                #st_name = ("%s%s%02d%s" % (sr.group(3), sr.group(4), int(sr.group(5)), sr.group(2) ) )
                st_name = ("{}{}{:02d}{}".format(sr.group(3), sr.group(4),
                            int(sr.group(5)), sr.group(2) ) )
                #print "%s -> %s" % (st_name, subj)
                print("{} -> {}".format(st_name, subj))
                name_subjects.append( (st_name, subj) )


            
def reset_classpart_slot() : 
    global classpart_slot
    classpart_slot = { }
    for c in classparts : 
        classpart_slot[c] = None 



def junior_skip_subject( st_name ) : 
    c_num = int(st_name[1]) 
    s_num = sem.semester_prefix[-2]
    spring_subj = "정치와법"
    fall_subj = "세계문화지리" 
    if s_num == "1" :
        if c_num < 5 : 
            return fall_subj
        else : 
            return spring_subj
    else : 
        if c_num < 5 : 
            return spring_subj
        else : 
            return fall_subj

    
def enroll_one_junior(st_name):
    for subj in sem.junior_subjects : 
        if subj == junior_skip_subject(st_name) : 
            continue
        else : 
            check_append( name_subjects, (st_name, subj), 
                    lambda x:' -> '.join(x) ) 



def enroll_juniors() :
    for st_name in sem.junior_names: 
        for subj in sem.junior_subjects : 
            if subj == junior_skip_subject(st_name) : 
                continue
            else : 
                check_append( name_subjects, (st_name, subj), 
                        lambda x:' -> '.join(x) ) 





def pair_str( pair ) : 
    #return ('(%s, %s)' % pair)
    return "({}, {})".format(* pair)


def check_remove( my_list, old_item, str_fn = str ) : 
    if old_item not in my_list : 
        #print "%s is not in the list. Doing nothing." % str_fn(old_item)
        print("{} is not in the list. Doing nothing.".format(str_fn(old_item)))
        return False
    else : 
        my_list.remove(old_item)
        #print "%s is removed." % str_fn(old_item)
        print("{} is removed.".format(str_fn(old_item)))
        return True

def clear_names_from( subj ) : 
    item_list = [c for c in name_subjects if c[1] == subj]
    for pair in item_list : 
        check_remove(name_subjects, pair, pair_str)


def delete_names_from_subject( subj, name_list ) : 
    for n in name_list : 
        check_remove( name_subjects, (n, subj) , pair_str)

def delete_students(name_list):
    for n in name_list : 
        remove_name_from_all_classes(n)
    for subj in subjects : 
        delete_names_from_subject(subj, name_list)



def add_names_to_subject( subj, name_list ) : 
    for n in name_list : 
        check_append( name_subjects, (n, subj) , pair_str)

def modify_for_philosophy() : 
    first_half_names = list_util.union( 
        [ c[0] for c in name_subjects if c[0].startswith("2") 
            and int(c[0][1]) < 5 ] )
    second_half_names = list_util.union( 
        [ c[0] for c in name_subjects if c[0].startswith("2") 
            and int(c[0][1]) >= 5 ] )
    first_half_subj = "한국사"
    second_half_subj = "철학" 
    delete_names_from_subject( first_half_subj, second_half_names )
    delete_names_from_subject( second_half_subj, first_half_names )
    add_names_to_subject( second_half_subj, second_half_names )
    add_names_to_subject( first_half_subj, first_half_names )


korean_digits = "공 일 이 삼 사 오 륙 칠 팔 구 열".split()
korean_gabja = "갑 을 병 정 무 기 경 신 임 계 은".split()
korean_surnames = "강 김 민 박 송 이 장 조 차 한 홍".split()

def num_to_name(four_digits) : 
    sur_name = sem.korean_surnames[ (int(four_digits[1:3]) * 7) % 11 ]
    first_letter = sem.korean_digits[ (int(four_digits[2:4]) * 3) % 11 ]
    second_letter = sem.korean_gabja[ (int(four_digits[1:4]) * 4) % 11 ]
    return (sur_name + first_letter + second_letter)
    

def generate_junior_numbers(n=128) : 
    first_letter = "0"
    second_letters = "1 2 3 4 5 6 7 8".split()
    nums = "01 02 03 04 05 06 07 08 09 10 11 12 13 14 15 16 17 18 19 20".split()
    output = [ ] 
    for num_str in nums : 
        for ban_str in second_letters : 
            output.append( first_letter + ban_str + num_str )
            if len(output) >= n : 
                return output
    return output
    


def generate_junior_names(n=128) : 
    return map( lambda x: x + num_to_name(x), generate_junior_numbers(n))







def hours( subj_clss ) : 
    if is_classname(subj_clss) : 
        subj = subject_name_of(subj_clss) 
    else : 
        subj = subj_clss
    return sem.hours_dict[subj]








def put_junior_names() : 
    if sem.semester_prefix[-2] == '2' : 
        first_subj = "세계문화지리"  
        second_subj = "정치와법"  
    else : 
        second_subj = "세계문화지리"  
        first_subj = "정치와법"  
    for subj in sem.junior_subjects : 
        for st_name in sem.junior_names : 
            if subj == second_subj and int(st_name[1])>4 : 
                c = classes_of(subj)[ int(st_name[1]) - 5 ]
            elif subj == first_subj  and int(st_name[1]) > 4 : 
                pass 
            else : 
                c = classes_of(subj)[ int(st_name[1]) - 1 ]
            put_name_into_class(st_name, c)


def put_pre_junior_names() : 
    if sem.semester_prefix[-2] == '2' : 
        first_subj = "세계문화지리"  
        second_subj = "정치와법"  
    else : 
        second_subj = "세계문화지리"  
        first_subj = "정치와법"
    for subj in sem.junior_subjects : 
        for st_name in sem.pre_junior_names : 
            if subj == second_subj and int(st_name[1])>4 : 
                c = classes_of(subj)[ int(st_name[1]) - 5 ]
            elif subj == first_subj  and int(st_name[1]) > 4 : 
                pass 
            else : 
                c = classes_of(subj)[ int(st_name[1]) - 1 ]
            put_name_into_class(st_name, c)


def assign_none_classrooms() : 
    print("Assigning None to all classrooms[pt]")
    for pt in classparts : 
        classrooms[pt] = None
    print("Done")



def read_year_advance(xl_filename) : 
    name_convs = xlsx_util.read_xlsx_sheet_into_list(xl_filename ,2,1)
    output = { } 
    for c in name_convs : 
        #if c[0].startswith("0") : 
            #output[ c[0] + num_to_name(c[0]) ] = c[1]
        #else : 
            #output[ c[0] ] = c[1]
        output[ c[0] ] = c[1]
    return output 

def apply_year_advance(advance_dict) : 
    first_names = [c for c in advance_dict if c.startswith("0")]
    second_names = [c for c in advance_dict if c.startswith("1")]
    third_names = [c for c in advance_dict 
        if c.startswith("2") or c.startswith("3")]
    for my_list in [third_names, second_names, first_names] : 
        for n in my_list : 
            change_name(n, advance_dict[n], name_classparts)
            change_name(n, advance_dict[n], name_subjects)
    

def assign_partially(st_name, count = 50000, verbose = False, includes=[], 
            randomize=False ) : 
    c_list = search_class_assignments(st_name, count = count, 
                input_includes  = includes,
                    verbose = verbose, yield_partial = True, 
                    randomize=randomize)
    remove_name_from_all_classes(st_name)
    put_name_into_classes(st_name, c_list[0])
    show_info(st_name)
    output = remaining_subjects(st_name)
    list_util.npr(output)
    return output


def current_slots_unsorted( classname ) : 
    my_slots = [ classpart_slot[pt] for pt in parts_of(classname) ]
    return my_slots

def sssc(subj, c_num, avoid_slot_list = [ ], disp_c_num = None, 
            put = 0, st_list = [ ], 
            sort_key = remaining_shorthands, show_count = False) : 
    if len(avoid_slot_list) > 0 : 
        my_slots = avoid_slot_list 
    else : 
        my_slots = list_util.complement(current_slots_unsorted( 
                    subj + "_" + str(c_num)), [None])
    if c_num == 0 : 
            return show_classes_of_subject(subj, my_slots, st_list=st_list, 
                        sort_key=sort_key, show_count = show_count )
    if put == 0: 
        if (disp_c_num is None or disp_c_num == 0) : 
            return show_classes_of_subject(subj, my_slots )
        else : 
            return show_classes_of_subject(subj, my_slots, disp_c_num )
    else: 
        if (disp_c_num == 0 or disp_c_num is None) : 
            if put > 0 : 
                put_names_into_class( 
                show_classes_of_subject(subj, my_slots )[:put], 
                    subj + "_" + str(c_num) )
            else : 
                put_names_into_class( 
                show_classes_of_subject(subj, my_slots )[put:], 
                    subj + "_" + str(c_num) )
        else : 
            put_names_into_class( 
                show_classes_of_subject(subj, 
                    my_slots, disp_c_num )[-put:], 
                    subj + "_" + str(c_num) )



def free_slots_of(st_name) : 
    return list_util.complement(slots, used_slots_of(st_name))


def show_classes_of_subject(subj, avoid_slot_list = [ ], c_num = None, 
        st_list = [ ], sort_key = remaining_shorthands, show_count=False ) : 
    for t_name in t_members_of(subj) : 
        print(korean_ljust(t_name, 24), ":", end=" " )
        print(teacher_used_slot_str( t_name))
    slot_list = [ ] 
    slot_counter = { } 
    double_slot_counter = { }
    for sl in slots : 
        slot_counter[sl] = 0
        double_slot_counter[sl] = 0
    if c_num is None : 
        if len(st_list) == 0 : 
            disp_names = remaining_names_of(subj) 
        else : 
            disp_names = list_util.intersect( members(subj), st_list )
    else : 
        disp_names = st_members_of(subj + "_" + str(c_num)) 
    for st_name in disp_names : 
        st_slots = used_slots_of(st_name)
        if len( list_util.intersect( avoid_slot_list, st_slots ) ) == 0  : 
            slot_list.append( (st_name, used_slot_str( st_slots ) ) )
            for sl in slot_counter : 
                if sl not in st_slots : slot_counter[sl] += 1
            for i, sl in enumerate(slots[:-1]) : 
                if (sl not in st_slots) and (slots[i+1] not in st_slots) : 
                    double_slot_counter[sl] += 1
    slot_list.sort(key = (lambda x:[ sort_key(x[0]),  
                            - len(used_slots_of(x[0])), x[1]]))
    for i, sl in enumerate(slot_list) : 
        print(str(i).rjust(2), korean_ljust(sl[0],20), ":", 
                subjects_in_tabbing(sl[0]))
    print("  ".rjust(2), korean_ljust(" ",20), ":", end=" " )
    print(' '.join( [str( slot_counter[sl] ).ljust(3) for sl in slots ]))
    print("  ".rjust(2), korean_ljust(" ",20), ":", end=" " )
    print(' '.join( [str( double_slot_counter[sl] ).ljust(3) for sl in slots ]))
    print()
    for c in parts_of(subj)  :
        print(korean_ljust(c + " " + str(len(st_members_of(c))), 20), end=" " )
        if classpart_slot[c] is not None : 
            print(":"+ classpart_slot[c] + ":", end=" ")
        else :
            print(":"+ "   " + ":", end=" ")
        if show_count : 
            print(slot_count_str(st_members_of(c),cl_slots=[classpart_slot[c]]))
        else : 
            print(slots_in_tabbing( available_slots_of(c), [classpart_slot[c]]))
    return [c[0] for c in slot_list]

scs = show_classes_of_subject



def show_classes_of_teachers_of(st_name, slot_list = None, 
        sort_key = remaining_shorthands, free_at = None, include = None, 
        c_num=True, indent=16 ) : 
    if slot_list is None : 
        constrain_slots = slots
    else : 
        constrain_slots = slot_list

    st_slots = list_util.intersect( constrain_slots, used_slots_of(st_name))
    st_parts = [ name_slot_part(st_name, sl) for sl in st_slots]
    t_list = list_util.union(list_util.flatten( 
                    [t_members_of(pt) for pt in st_parts]))
    if free_at is not None: 
        t_list = [ c for c in t_list if len(list_util.intersect(free_at , 
                    used_slots_of(c))) == 0 ] 
    t_list.sort()
    if include is not None: 
        t_list = include + t_list
    show_classes_of_teachers( [st_name] + t_list, c_num = c_num, indent=indent)

    #print  korean_ljust(' ',20), ":", slots_in_tabbing(slots)
    #print  korean_ljust(st_name,20), ":", subjects_in_tabbing(st_name)
    #print
    #for t_name in t_list : 
        #print  korean_ljust(t_name,20), ":", subjects_in_tabbing(t_name)


def show_classes_of_teachers(t_list, sort_key = remaining_shorthands, 
        c_num=False, indent=0) : 
    #print  korean_ljust(' ',20), ":", slots_in_tabbing(slots)
    if c_num : 
        tabs_num = 8
        tabbing = 5
    else : 
        tabs_num = 20
        tabbing = 4
    if indent > 0 : 
        tabs_num = indent
    print(korean_ljust(" ", tabs_num), end=" " )
    print(":", all_slots_in_tabbing(tabbing=tabbing))
    for t_name in t_list : 
        print( korean_ljust(t_name,tabs_num), ":", subjects_in_tabbing(t_name,
                    c_num=c_num))











def free_name_from_slots(st_name, in_slots) : 
    if type(in_slots) is list : 
        slot_list = in_slots
    else : 
        slot_list = [in_slots]
    my_slots = [ slot_dict[sl] if sl in slot_dict else sl for sl in slot_list]
    my_classes = list_util.complement(list_util.union( 
                    [ class_name_of( name_slot_part(st_name, sl)) 
                        for sl in my_slots]), [""])
    for cl in my_classes : 
        remove_name_from_class(st_name, cl)
    
def free_names_from_slots(st_names, in_slots) : 
    for st_n in st_names : 
        free_name_from_slots(st_n, in_slots)






s_subjs = [
"물리학III",
"화학III",
"정수론",
"자료구조",
"현대문학",
"고급커뮤니케이션",
"생활미술",
"수리정보탐구",
"중국어I",
"화학실험I",
"물리학실험I",
"생명과학실험I",
"철학",
"한국사",
"영어III",
"영어독해",
"기초통계학",
"생명과학III",
"수학III",
"지구과학I",
"여가와체육I",
"음악I",
"독서III" ] 

def do_cycle(my_list) : 
    my_list.append( my_list[0])
    my_list.pop(0)
    #print "Cycling.. 0:%s -> %d:%s " % ( my_list[-1], len(my_list)-1, my_list[-1])
    #print "Now 0:%s" % my_list[0]
    print("Cycling.. 0:{} -> {}:{} ".format( my_list[-1],
            len(my_list)-1, my_list[-1]))
    print("Now 0:{}".format(my_list[0]))
    return True

def do_uncycle(my_list) : 
    my_list.insert(0, my_list[-1])
    my_list.pop()
    #print "Cycling.. %d:%s -> 0:%s " % ( len(my_list)-1, my_list[0], my_list[0])
    print("Cycling.. {}:{} -> 0:{} ".format( len(my_list)-1, my_list[0],
             my_list[0]))
    return True

def fill_up_class( cl, add_num ) : 
    subj = subject_name_of(cl)
    c_num = int( cl[-1] )
    sssc(subj, c_num, put=add_num)

def fill_up_classes(subj, num_str, add_num) : 
    nums = [int(c) for c in num_str.split()] 
    for i in nums : 
        sssc(subj, i, put=add_num)
    
def ordered_work_names() : 
    output = [c for c in students if len(remaining_subjects(c))>0]
    output.sort( key = (lambda x: [ - total_hours(x), 
            sorted( [len(classes_of(s)) for s in subjects_of(x)]), 
            [sem.shorthands[s] for s in subjects_of(x)] ]))
    return output


def under_classes(st_name, num_bound) : 
    output = [c for c in current_classes(st_name) 
                if len(st_members_of(c))<num_bound] 
    output.sort()
    return output

def size_of(my_class) : 
    return len(st_members_of(my_class))


def empty_big_classes(subj, fix_names = [ ], size = 0) : 
    if size == 0 : 
        size_cut = average_size(subj)
    else : 
        size_cut = size
    my_classes = [ c for c in classes_of(subj) if size_of(c) > size_cut]
    for cl in my_classes : 
        remove_all_names_from_class(cl, fix_names = fix_names)

def fill_up_small_classes(subj, zero=False) : 
    av_size = int(math.ceil(average_size(subj)))
    if zero : 
        my_classes = [ c for c in classes_of(subj) if size_of(c) == 0 ]
    else : 
        my_classes = [ c for c in classes_of(subj) if size_of(c) > 0 
                    and size_of(c) < av_size ]
    my_classes.sort(key = size_of)
    for cl in my_classes : 
        fill_up_class(cl, av_size - size_of(cl))

def try_assign_remains(subj , includes = [ ], num = -1,count=1,
       excludes = [], fix_subjects = [ ], bound=0, across = False): 
    if num == -1 : 
        st_list = remaining_names(subj)
    elif num == 0 : 
        av_size = int(math.floor(average_size(subj)))
        if av_size > size_of( includes[0] ) : 
            my_num = av_size - size_of( includes[0] )
            st_list = remaining_names(subj)[-my_num:]
        else : 
            st_list = [ ]
    else : 
        st_list = remaining_names(subj)[-num:]
    for st in st_list : 
        try:
            choose_and_assign(st, count=count, includes=includes, 
                excludes=excludes, bound=bound, across = across)
        except (IndexError, KeyboardInterrupt):
            pass


def try_assign_names(st_list, across = False, sort_key = None, bound = 20, 
        fix_subjects = [ ] ) : 
    n = len(st_list)
    for i, st in enumerate(st_list) : 
        try:
            print(i+1, "/", n, ": ")
            choose_and_assign(st, count=1, bound=bound, across = across, 
                sort_key = sort_key, 
            fix_subjects = list_util.intersect(fix_subjects, subjects_of(st)) ) 
        except (IndexError, KeyboardInterrupt):
            pass





def num_to_students( num ) : 
    output = [c for c in names if c.startswith(str(num))]
    return output

def batch_put_sophomores(subj, philosophy=False) : 
    my_classes = classes_of(subj)
    if philosophy : 
        offset = 4 
    else : 
        offset = 0
    for i, cl in enumerate(my_classes):
        put_names_into_class( num_to_students( 21 + i + offset), cl)

def class_number_of(clss) : 
    sr = re.search( r'_(\d+)', clss)
    return sr.group(1)


def subject_str(st_name) : 
    return ''.join( [ sem.shorthands[ s ] 
                    for s in sorted(subjects_of(st_name))] )

def remaining_subject_str(st_name) : 
    return ''.join( [ sem.shorthands[ s ] 
                    for s in sorted(remaining_subjects(st_name))] )
                                
                                

def class_str(st_name) : 
    return ''.join( [ sem.shorthands[ subject_name_of(c) ] + class_number_of(c) 
                    for c in sorted(current_classes(st_name))] )


def show_electives(input_st_list, shorthands=True ) : 
    st_list = input_st_list[:]
    st_list.sort(key = subject_str)
    for i, st in enumerate(st_list) : 
        if shorthands : 
            subj_str = ' '.join( map( gen_util.dict_to_fn(sem.shorthands), 
                sorted( list_util.complement(subjects_of(st), sem.s_majors))))
        else:
            subj_str = ' '.join( sorted( list_util.complement( 
                            subjects_of(st), sem.s_majors)))
        print(i, st, korean_ljust(subj_str , 30  )  , end=" " )
        print(''.join( [ sem.shorthands[ subject_name_of(c) ] 
            + class_number_of(c) for c in sorted(current_classes(st))] ), end=" ")
        print(''.join([ sem.shorthands[ s ] for s in remaining_subjects(st) ]), end=" ")
        print(total_hours(st))
    return st_list


def put_group_into_major_classes(st_list, c_num ) : 
    for s in sem.s_majors : 
        clss = classes_of(s)[c_num - 1]
        remove_all_names_from_class(clss)
        put_names_into_class(st_list, clss )


def free_names_from_subject(st_names, subj) : 
    for st in st_names : 
        clss = name_subject_class(st, subj)
        if len(clss) == 0 : 
            #print "%s is not in a class of %s. Doing nothing.." % (st, subj)
            print("{} is not in a class of {}. Doing nothing..".format(st, 
                subj))
        else : 
            remove_name_from_class(st, clss)
        
    
def copy_members_from(subj, subj_source) : 
    cls_one = classes_of(subj)
    cls_two = classes_of(subj_source)
    for i in range(len(cls_one)) :
        put_names_into_class(  st_members_of(cls_two[i])  , cls_one[i])


def exchange_st_members(clss_one, clss_two) : 
    if subject_name_of(clss_one) != subject_name_of(clss_two) : 
        #print "%s, %s are not in the same subject" % (clss_one, clss_two)
        print("{}, {} are not in the same subject".format(clss_one, clss_two))
        return False
    else : 
        sg_one = st_members_of(clss_one)
        sg_two = st_members_of(clss_two)
        put_names_into_class( sg_one, clss_two ) 
        put_names_into_class( sg_two, clss_one )
        return True
    

def exchange_st_members_across(clss_one, clss_two) : 
    sg_one = st_members_of(clss_one)
    sg_two = st_members_of(clss_two)
    remove_names_from_class(sg_one, clss_one)
    put_names_into_class( sg_one, clss_two ) 
    remove_names_from_class(sg_two, clss_two)
    put_names_into_class( sg_two, clss_one )
    return True
    



def exchange_slots_of_classes(clss_one, clss_two) : 
    if subject_name_of(clss_one) != subject_name_of(clss_two) : 
        #print "%s, %s are not in the same subject" % (clss_one, clss_two)
        print("{}, {} are not in the same subject".format(clss_one, clss_two))
        return False
    else : 
        sg_one = slots_of(clss_one)
        sg_two = slots_of(clss_two)
        assign_class_slots( clss_two , sg_one ) 
        assign_class_slots( clss_one, sg_two)
        return True

def subject_slot_class(subj, slot): 
    if slot in slot_dict : 
        my_slot  = slot_dict[slot]
    else : 
        my_slot  = slot
    output = [c for c in classes_of(subj) if my_slot in slots_of(c)]
    if len(output) > 1 : 
        #print "More than one class of %s at %s" % (subj, my_slot)
        print("More than one class of {} at {}".format(subj, my_slot))
        return None
    elif len(output) == 0 : 
        #print "No class of %s at %s" % (subj, my_slot)
        print("No class of {} at {}".format(subj, my_slot))
        return None
    else : 
        return output[0]

def exchange_classes(clss_one, clss_two):
    exchange_slots_of_classes(clss_one, clss_two)
    exchange_st_members(clss_one, clss_two)
    exchange_classrooms(clss_one, clss_two) 
    
def exchange_classes_by_slots(subj, slot_one, slot_two): 
    clss_one = subject_slot_class(subj, slot_one)
    clss_two = subject_slot_class(subj, slot_two)
    exchange_slots_of_classes(clss_one, clss_two)
    exchange_st_members(clss_one, clss_two)



def exchange_slots_of_student(st_name, slot_one, slot_two) : 
    part_one = name_slot_part(st_name, slot_one)
    part_two = name_slot_part(st_name, slot_two)
    assign_part_slot(part_one, slot_two)  
    assign_part_slot(part_two, slot_one)  

def exchange_slots(t_name, in_slot_list) : 
    slot_list = [ slot_dict[sl] if sl in slot_dict else sl  
                    for sl in in_slot_list]
    n = len(slot_list)/2
    slots_one = slot_list[:n]
    slots_two = slot_list[n:2*n]
    for i in range(len(slots_one)) : 
        part_one = name_slot_part(t_name, slots_one[i])
        part_two = name_slot_part(t_name, slots_two[i])
        if len(part_one)>0 : 
            assign_part_slot(part_one, slots_two[i])  
        if len(part_two)>0 : 
            assign_part_slot(part_two, slots_one[i])  
    


def move_slot_of_student(st_name, slot_one, slot_two) : 
    part_one = name_slot_part(st_name, slot_one)
    assign_part_slot(part_one, slot_two)  

    
def delete_part( pt ) : 
    c_list = [c for c in name_classparts if c[1] == pt ]
    for c in c_list : 
            check_remove(name_classparts, c)
    if pt in classpart_slot : 
        #print "Deleting %s: %s" % (pt, str(classpart_slot[pt]))
        print("Deleting {}: {}".format(pt, str(classpart_slot[pt])))
        del classpart_slot[pt]
    if pt in classrooms : 
        #print "Deleting %s: %s" % (pt, str(classrooms[pt]))
        print("Deleting {}: {}".format(pt, str(classrooms[pt])))
        del classrooms[pt]



subjects_by_department = [ 
"융합과학탐구",
"국어I",
"독서I",
"독서III",
"매체언어비평",
"작문",
"현대문학",
"고급물리학I",
"고급물리학II",
"과학사",
"물리학I",
"물리학III",
"물리학실험I",
"보건/진로",
"세계문화지리",
"세계사",
"정치와법",
"철학",
"한국사",
"고급생명과학II",
"생명과학I",
"생명과학III",
"생명과학실험I",
"기초통계학",
"미적분학I",
"미적분학II",
"선형대수학",
"수학I",
"수학III",
"정수론",
"수리정보탐구",
"고급커뮤니케이션",
"영어I",
"영어III",
"영어독해",
"영어회화I",
"영작문",
"건강과체육I",
"미술I",
"생활미술",
"생활음악",
"생활체육",
"여가와체육I",
"음악I",
"중국어I",
"객체지향프로그래밍",
"자료구조",
"컴퓨터과학I",
"지구과학I",
"지구과학III",
"고급화학I",
"고급화학II",
"화학I",
"화학III",
"화학실험I"]



def classes_in_tabbing( room, t_name=False ) : 
    output = [ ] 
    for sl in slots : 
        cl = room_slot_class(room, sl)
        if len(cl)>0 : 
            if t_name is False : 
                output.append( sem.shorthands[ subject_name_of(cl)] 
                    + class_number_of(cl))
            else : 
                curr_name = short_tname(teachers_of(cl.split()[0])[0])
                output.append( curr_name + class_number_of(cl))
        else : 
            output.append( "----/") 
    return ''.join(output)




def all_slots_in_tabbing(tabbing=5):
    return ''.join( [korean_ljust(sl, tabbing) for sl in slots ] )
    
def classrooms_of_class(clss) : 
    room_list = [ ] 
    for pt in parts_of(clss) : 
        if classrooms[pt]:
            room_list.extend( classrooms[pt].split("/") )
    return list_util.union( room_list )

def short_tname(t_name):
    if len(t_name)>=9 : 
        return t_name[3:9]
    else : 
        return t_name[:-1]

def show_classrooms(subj_list,t_name=False) : 
    print(korean_ljust(" ", 15),  ":", all_slots_in_tabbing())
    c_list = list_util.union(* [classes_of(subj) for subj in subj_list] )
    c_list.sort()
    room_list = list_util.union( *[classrooms_of_class(cl) for cl in c_list] )
    room_list.sort()
    for r in room_list : 
        print(korean_ljust(r, 15),  ":", classes_in_tabbing(r,t_name=t_name))



def fcc(clss, room = None) : 
    if room is None : 
        return free_classrooms_for_class(clss)
    else : 
        return assign_classroom(clss, room)

def exchange_classrooms(clss_one, clss_two) : 
    room_one = classrooms[ parts_of(clss_one)[0] ]
    room_two = classrooms[ parts_of(clss_two)[0] ]
    assign_classroom(clss_one, room_two)
    assign_classroom(clss_two, room_one)



def export_class_list_xlsx(filename) : 
    c_list = list_util.union( * [classes_of(s) for s in subjects] )
    output = [ (sem.subject_group_dict[subject_name_of(a)], a, 
                subject_name_of(a), str(class_number_of(class_name_of(a))) ) 
                    for a in c_list ]
    output.sort( key = (lambda x: [ sem.ordered_group.index(x[0]), x[2], x[3] ]) )
    for c in output : 
        print(' '.join(c))
    xlsx_util.write_tuples_into_xlsx(filename, output)


def class_sizes_of(st_name) : 
    c_list = current_classes(st_name)
    c_list.sort(key=(lambda x: -size_of(x)))
    return [ map( lambda x: -size_of(x), c_list)   , c_list ]


def class_stat_str(st_name, smalls=[]) :
    c_list = current_classes(st_name)
    small_num = len(list_util.intersect(c_list, smalls))
    c_list.sort(key=size_of)
    output = [ ] 
    for c in c_list : 
        #output.append( "%s<%d>" % ( sem.shorthands[subject_name_of(c)] + class_number_of(c), size_of(c)) )
        output.append( "{}<{}>".format( sem.shorthands[subject_name_of(c)]
                + class_number_of(c), size_of(c)) )
    return st_name + " " + str(small_num) + " " + " ".join(output)


def export_team_weekly_table(filename) : 
    template_filename = "team_weekly_table_template.xlsx"
    #print "Reading style from <%s>" % template_filename
    print("Reading style from <{}>".format(template_filename))
    shutil.copyfile( template_filename, filename)
    #print "Copied %s to %s" %( template_filename, filename)
    print("Copied {} to {}".format( template_filename, filename))
    wb = openpyxl.load_workbook(filename)

    ws = wb[ "부서별명단" ]
    output = [ ]
    team_list = list_util.union( sem.name_team_dict.values() ) 
    team_list.sort()

    my_teachers = [c for c in teachers if not c.endswith("조교t") ]
    for team in team_list : 
        rec = [ team ] 
        for t_name in my_teachers : 
            if team == sem.name_team_dict[t_name] : 
                rec.append(t_name[:-1])
        output.append( rec )
    dept_list = list_util.union( sem.name_department_dict.values() ) 
    dept_list.sort()
    for dept in dept_list : 
        rec = [ dept ] 
        for t_name in my_teachers : 
            if dept == sem.name_department_dict[t_name] : 
                rec.append(t_name[:-1])
        output.append( rec )

    for i in range(1,4):
        rec = [ str(i) + "학년담임" ] 
        rec.extend( [ sem.bujang_dict[k] for k in sem.bujang_dict 
                        if k.startswith( str(i) ) ] )
        output.append(rec)

    xlsx_util.fill_in_sheet_by_tuples( ws, list_util.transposed(output))

    name_slots_dict = { }
    for t_name in my_teachers : 
        name_slots_dict[t_name] = used_slots_of(t_name, list_util.complement( 
            sem.skip_subject_list, "부장회의".split()) )
        if sem.is_in_sci_dept(t_name) : 
            name_slots_dict[t_name].extend( "수5 수6 수7".split() )
    my_slots = [ ]
    for d in days :
        my_slots.extend( [ d + str(i) for i in range(1,8) ] )
    slot_names_dict = { }
    for sl in my_slots : 
        slot_names_dict[sl] = sorted([ t_name[:-1] for t_name in my_teachers 
                                    if sl in name_slots_dict[t_name] ])
    output = [ ] 
    for sl in my_slots : 
        output.append(  [sl] + slot_names_dict[sl] )
    ws = wb[ "교시별교사명단" ]
    xlsx_util.fill_in_sheet_by_tuples( ws, list_util.transposed(output))

    #print "Saving to <%s>.." % filename
    print("Saving to <{}>..".format(filename))
    wb.save( filename )
    

    
def move_teacher_from_part_to_part( t_name, pt_one, pt_two ) :     
    if (t_name, pt_one) not in name_classparts : 
        #print "%s not in name_classparts" % pair_str( (t_name, pt_one) )
        print("{} not in name_classparts".format(pair_str( (t_name, pt_one))))
        return False 
    if (t_name, pt_two) in name_classparts : 
        #print "%s already in in name_classparts" % pair_str( (t_name, pt_two))
        print("{} already in in name_classparts".format(pair_str( 
            (t_name, pt_two))))
        return False 
    check_remove( name_classparts,  (t_name, pt_one))
    check_append( name_classparts,  (t_name, pt_two))
    return True


    
def replace_teacher_for_part( pt, name_one, name_two ) :     
    if (name_one, pt) not in name_classparts : 
        #print "%s not in name_classparts" % pair_str( (name_one, pt) )
        print("{} not in name_classparts".format(pair_str( (name_one, pt) )))
        return False 
    if (name_two, pt) in name_classparts : 
        #print "%s already in in name_classparts" % pair_str( (name_two, pt))
        print("{} already in in name_classparts".format(pair_str( 
            (name_two, pt))))
        return False 
    #check_remove( name_classparts,(name_one, pt), lambda x: ("(%s, %s)" % x) )
    check_remove( name_classparts,  (name_one, pt),
        lambda x: ("({}, {})".format(*x) ))
    check_append( name_classparts,  (name_two, pt), lambda x:' -> '.join(x))
    return True

def replace_teacher_for_class( clss, name_one, name_two ) :     
    for pt in parts_of(clss):
        replace_teacher_for_part(pt, name_one, name_two)




def batch_export() : 
    export_teacher_table_xlsx( sem.semester_prefix + "전체시간표.xlsx" )
    export_baedangpyo_xlsx(sem.semester_prefix + "수업배당표.xlsx" )
    export_weekly_tables_xlsx( sem.semester_prefix + "교사별_시간표.xlsx" )
    export_weekly_tables_of_rooms( sem.semester_prefix + "강의실별_시간표.xlsx" )
    export_weekly_tables_of_all_classes( sem.semester_prefix 
                + "반별_학생시간표.xlsx" )
    export_weekly_tables_of_all_classes( sem.semester_prefix 
                + "반별_학생시간표_담당교사명.xlsx", t_name=True )
    export_chulseokbu_of_all_subjects_xlsx(sem.semester_prefix +"분반별_명단.xlsx")
    #export_chulseokbu_of_all_departments() 
    export_bunban_table_xlsx(sem.semester_prefix + "분반배정표.xlsx" )   
    export_team_weekly_table(sem.semester_prefix + "team_weekly_table_data.xlsx" ) 
    export_teacher_schedule_txt( "t_schedule_" + sem.semester_prefix[:-1] + ".txt") 


def show_class_members_neis(clss, excludes = None ) : 
    if excludes is None : 
        ex_c_nums = range(1, int(class_number_of(clss)))
    else : 
        ex_c_nums = excludes 

    subj = subject_name_of(clss)
    ex_members = list_util.flatten( 
                [st_members_of( subj + "_" + str(i)) for i in ex_c_nums]) 
    base_list = list_util.complement( members(subj), ex_members )

    base_list.sort()
    target_list = st_members_of(clss)
    print("===\n", clss, len(target_list), "\n===")
    for st_name in base_list : 
        if st_name in target_list : 
            print("", end=" ")
        print(st_name, end=" ")
        if st_name in target_list : 
            print(st_name, end=" " )
        print
    print(len(target_list))


def export_bunban_neis_xlsx(filename) :

    wb = openpyxl.Workbook()
    ws = wb.active


    my_classes = list_util.flatten( [ classes_of(subj) for subj 
                        in list_util.complement(subjects, sem.junior_subjects)])
    my_classes.sort()
    output = [ ] 
    for clss in my_classes : 
        output.append( [para_classname(clss)] + sorted(st_members_of(clss)))
        
    xlsx_util.fill_in_sheet_by_tuples( ws, list_util.transposed(output))

    wb.create_sheet("Sheet2")
    ws = wb["Sheet2"]
    output=[]
    for clss in my_classes : 
        subj = subject_name_of(clss)
        ex_c_nums = range(1, int(class_number_of(clss)))
        ex_members = list_util.flatten( 
                [st_members_of( subj + "_" + str(i)) for i in ex_c_nums]) 
        base_list = list_util.complement( members(subj), ex_members )
        output.append( [para_classname(clss)] + sorted(base_list))

    xlsx_util.fill_in_sheet_by_tuples( ws, list_util.transposed(output))
    wb.save(filename)

def neis_name_fix(st_name) : 
    if st_name in names : 
        return st_name 
    else : 
        return name( st_name[1:])


def compare_with_neis_bunban(filename) : 
    neis_name_classes = [ ] 
    with open(filename) as f : 
        for line in f : 
            my_pair = tuple( line.strip().split() )
            neis_name_classes.append( (neis_name_fix(my_pair[0]), my_pair[1]))
    curr_name_classes = list_util.union( 
                [ (st_name, pt[:-2]) for st_name, pt in name_classparts])
    print("Wrong NEIS inputs" )
    for c in list_util.complement( neis_name_classes , curr_name_classes ):
        print(pair_str(c))
    print("Missed inputs" )
    for c in list_util.complement( curr_name_classes, neis_name_classes ):
        if c[0].endswith("t") or c[0].endswith("a")  or c[0].startswith("1") : 
            continue 
        else : 
            print(pair_str(c))

    




def export_picture_list_data(filename) : 
    col_names = [ "학번", "이름" ] + \
                    [ "과목" + str(i) + "(분반번호)"  for i in range(1,17)]
    sheetname = "학생별분반목록" 
    output = [ ] 
    
    for st_name in students : 
        output.append( [int(st_name[:4]), st_name[4:]] + map( 
            para_classname,  sorted( current_classes(st_name)) ))
    xlsx_util.write_tuples_into_xlsx( filename, output, 
        in_column_names = col_names, sheetname = sheetname)



def generate_password(salt, user_name ) : 
    m = hashlib.md5()
    m.update(salt + user_name)
    return m.hexdigest()[:6]




def export_sugang_checks_of_all_classes( filename, class_list = sem.ban_list, 
            date = None, salt = None) : 
    print("Reading style from <sugang-check-template.xlsx>." )
    shutil.copyfile( "sugang-check-template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb["Template"]

    if salt is None : 
        my_salt = datetime.datetime.today().strftime("%Y-%m-%d") 
    else : 
        my_salt = salt

    for clss in class_list : 
        #print "Writing subjects lists for %s." % clss
        print("Writing subjects lists for {}.".format(clss))
        new_worksheet = wb.create_sheet(  clss)
        new_worksheet.print_area = "A1:E400"
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        prefix = clss.replace("-", "")
        name_list = [nm for nm in students if nm.startswith(prefix)]
        name_list.sort()
        for i, st_name in enumerate(name_list) : 
            fill_in_sugang_checks(new_worksheet, i*20, st_name, date=date )
        for k in range( (i+1)*20 + 1, 401  ): 
            new_worksheet.row_dimensions[k].hidden = True
        if len(my_salt)>0: 
            new_worksheet.protection.password = generate_password(my_salt,clss) 
    wb.remove(template_ws)
    #print "Saving to <%s>."%filename
    print("Saving to <{}>.".format(filename))
    wb.save(filename)



def fill_in_sugang_checks(new_worksheet, offset, st_name, date=None ) : 
    if date is None : 
        date_str = datetime.datetime.today().strftime("%Y-%m-%d") 
    else : 
        date_str = date
    name_in_xlsx = convert_name_for_xlsx(st_name)
    #print "Writing sugang check for %s." % st_name
    print("Writing sugang check for {}.".format(st_name))
    new_worksheet.cell( row = 3 + offset, column = 5 ).value = sem.semester_str
    new_worksheet.cell( row = 2 + offset, column = 5 ).value = name_in_xlsx 
    new_worksheet.cell( row = 2 + offset, column = 2 ).value = date_str
    new_worksheet.cell( row = 19 + offset, column = 2 ).value = \
        sem.semester_str + " 수강신청을 위와 같이 하였음을 확인합니다." 
    temp_list = sorted( subjects_of(st_name) )
    st_hour_list = [  hours(subj)  for subj in temp_list]
    st_size_list  = [ ] 

    st_subj_list = [ ]
    for subj in temp_list : 
        if (subj in "철학 한국사".split()) and sem.semester_prefix[-2]=="1" : 
            st_subj_list.append( "한국사/철학") 
            if sem.is_spring_semester() : 
                #st_size_list.append( "(%d명)" % int(sem.num_juniors/2.0) )
                st_size_list.append( "({}명)".format(int(sem.num_juniors/2.0)))
            else : 
                #st_size_list.append( "(%d명)" % int(sem.num_sophomores/2.0) )
                st_size_list.append("({}명)".format(
                    int(sem.num_sophomores/2.0)))
        else : 
            st_subj_list.append( subj)
            #st_size_list.append( "(%d명)" % len(members(subj)) )
            st_size_list.append( "({}명)".format(len(members(subj)) ))
        

    
    xlsx_util.fill_in_sheet_by_list(new_worksheet, st_subj_list, 
                5 + offset, 2, direction="column")
    xlsx_util.fill_in_sheet_by_list(new_worksheet, st_size_list, 
                5 + offset, 5, direction="column")
    xlsx_util.fill_in_sheet_by_list(new_worksheet, st_hour_list, 
                5 + offset, 6, direction="column")
    #new_worksheet.cell( row = 19 + offset, column = 6 ).value = ("합계: %d 시간" % sum(st_hour_list) )
    new_worksheet.cell( row = 19 + offset, column = 6 ).value = \
            ("합계: {} 시간".format(sum(st_hour_list) ))




def generate_classparts(xl_filename) : 
    subj_info = xlsx_util.read_xlsx_sheet_into_dict(xl_filename, 
                    2,3,"수강신청현황값고정") 
    for c in subj_info.keys() : print(' '.join(c))
    output = [ ] 
    subjs_by_dept = [subj for subj in subjects]    
    subjs_by_dept.sort( key = (lambda x: sem.ordered_group.index(
                            sem.subject_group_dict[x])) ) 
    for subj in subjs_by_dept :
        num_of_classes = int( subj_info[ (subj, "분반수") ] )
        for c_num in  range(1,num_of_classes + 1) : 
            for p_num in range(1, hours(subj) + 1) : 
                output.append(subj + "_" + str(c_num) + "$" + str(p_num))
    return output


def remove_names_from_all(st_names) : 
    for st_name in st_names : 
        remove_name_from_all_classes(st_name) 
    n_s_list = [ c for c in name_subjects if c[0] in st_names ] 
    for pair in n_s_list : 
        check_remove(name_subjects, pair)


def clear_slots_classrooms(part_list = classparts ) : 
    for pt in part_list : 
        assign_part_slot(pt, None) 
        assign_classroom_to_part(pt, None )  


def print_student_slots(st_name) : 
    st_slots = used_slots_of(st_name)
    slot_list = [ (st_name, used_slot_str( st_slots ) ) ]
    for i, sl in enumerate(slot_list) : 
        print(str(i).rjust(2), korean_ljust(sl[0],20), ":", 
                subjects_in_tabbing(sl[0]))


def show_teacher_slots(st_name_or_classes) : 
    if type(st_name_or_classes) is list : 
        st_classes = st_name_or_classes 
    else : 
        st_classes = current_classes(st_name_or_classes)
    t_names = [ ] 
    for cl in st_classes : 
        t_names.extend( t_members_of(cl) )
    for t_name in t_names : 
        print(korean_ljust(t_name, 24), ":", end=" " )
        print(teacher_used_slot_str( t_name))
    if type(st_name_or_classes) is not list : 
        print_student_slots(st_name_or_classes)


def grades_of(subj) : 
    return sorted(list_util.union( [ c[0] for c in members(subj)]))



def include_exclude_of(st_name) : 
    incl = [ ] 
    for k in sem.includes_dict : 
        if st_name in k : 
            incl.append( sem.includes_dict[k] )
    excl = [ ] 
    for k in sem.excludes_dict : 
        if st_name not in k : 
            excl.append( sem.excludes_dict[k] )
    return [ incl, excl]

def remove_exceptions() : 
    for pair in sem.exception_list : 
        check_remove(name_classparts, pair, pair_str)



def include_list(st_name, bound_dict) : 
    output = [ ] 
    for clss in current_classes(st_name) : 
        if (size_of(clss) - 1) < bound_dict[ subject_name_of(clss) ][0] :
            output.append(clss)
    return output

def exclude_list(st_name, bound_dict) : 
    output = [ ] 
    for subj in subjects_of(st_name) : 
        for clss in classes_of(subj) : 
            if expected_class_size(st_name, clss) > bound_dict[subj][1] :
                output.append(clss)
    return output




def try_move_one_to( st_name, clss, n = 20, exp_n = 13, count = 1, 
        includes = [], excludes = [], bound_dict = None ):
    if bound_dict is None : 
        inc_classes = [clss] + list_util.intersect( 
                        current_classes(st_name), includes )
        big_cs = list_util.union(* [ [c for c in classes_of(subj) 
            if overfulled_with(c, st_name, n, exp_n) ] 
                for subj in subjects_of(st_name) ])
        ex_classes = list_util.union(big_cs , excludes)
    else : 
        inc_classes = [clss] + include_list(st_name, bound_dict) \
                  + list_util.intersect( current_classes(st_name), includes ) 
        ex_classes = list_util.union(exclude_list(st_name,bound_dict), excludes)
    try: 
        choose_and_assign(st_name, count, includes = inc_classes, 
                excludes = ex_classes)
    except IndexError : 
        return False
    return True


def try_move_one_from( st_name, clss, n = 20, exp_n = 13, count = 1, 
        includes = [], excludes = [], bound_dict = None ):
    subj = subject_name_of(clss)
    big_size = size_of(clss)
    from_classes = [c for c in classes_of(subj) if size_of(c) + 1 >= big_size]

    if bound_dict is None : 
        big_cs = list_util.union(* [ [c for c in classes_of(subj) 
            if overfulled_with(c, st_name, n, exp_n) ] 
                for subj in subjects_of(st_name) ])
        ex_classes = list_util.union( big_cs, from_classes, excludes)
        inc_classes = includes 
    else : 
        ex_classes = list_util.union([clss], from_classes,
                    exclude_list(st_name,bound_dict), excludes)
        inc_classes = include_list(st_name, bound_dict) \
                  + list_util.intersect( current_classes(st_name), includes ) 

    try: 
        choose_and_assign(st_name, count, excludes = ex_classes, 
            includes=inc_classes )
    except IndexError : 
        return False
    return True


def repeat_balancing(subj, iter_num = 5, from_top = False, n = 20, exp_n=13, 
        fix_classes = [ ], fix_update = None, bound_dict = None  ) : 
    clsses = classes_of(subj)
    if fix_update is None : 
        my_includes_fn = (lambda : list_util.intersect( clsses, fix_classes ))
    else : 
        my_includes_fn = (lambda : list_util.intersect( clsses, fix_update() ))
    if from_top : 
        for i in range(iter_num) : 
            clsses.sort(key = (lambda x: len(st_members_of(x))))
            max_clss = clsses[-1]
            max_size = len(st_members_of(max_clss))
            big_classes = [c for c in clsses 
                if max_size == len(st_members_of(c))]
            candis = list_util.flatten( map( st_members_of, big_classes))
            st_name = random.choice( candis )
            try_move_one_from(st_name, max_clss, n=n,exp_n=exp_n, 
                    includes = my_includes_fn())

    else : 
        for i in range(iter_num) : 
            clsses.sort(key = (lambda x: len(st_members_of(x))))
            min_clss = clsses[0]
            min_size = len(st_members_of(min_clss))
            bigger_classes = [c for c in clsses 
                if min_size < len(st_members_of(c)) - 1] 
            candis = list_util.flatten( map( st_members_of, bigger_classes))
            st_name = random.choice( candis )
            try_move_one_to(st_name, min_clss, n=n, exp_n=exp_n, 
                includes = my_includes_fn(), bound_dict = bound_dict )
    clsses.sort(key=size_of)
    #print ' '.join( ["%s<%d>" % ( sem.shorthands[subject_name_of(c)] + class_number_of(c), size_of(c)) for c in clsses ] ) 
    print(' '.join( ["{}<{}>".format( sem.shorthands[subject_name_of(c)]
                + class_number_of(c), size_of(c)) for c in clsses ] ))
    return clsses


def biggest_class_of(st_name) : 
    my_classes = current_classes(st_name)
    my_classes.sort(key=size_of)
    return my_classes[-1]



def balance_one_from(st_name, clss,  n=20, exp_n=13, includes = [ ] ) : 
    my_includes = list_util.intersect(current_classes(st_name), includes)
    my_excludes = [clss] 
    for clss in current_classes(st_name) : 
        if size_of(clss) < n :
            new_excl = [c for c in classes_of(subject_name_of(clss)) 
                         if expected_class_size(st_name, c) >= n]
            my_excludes.extend(new_excl) 
    try_move_one_from(st_name, clss, n=n, exp_n=exp_n, 
            includes = my_includes, excludes = my_excludes)



def balance_one(st_name, n=20, exp_n=13, includes = [ ] ) : 
    my_includes = list_util.intersect(current_classes(st_name), includes)
    my_excludes = [ ] 
    for clss in current_classes(st_name) : 
        if size_of(clss) < n :
            new_excl = [c for c in classes_of(subject_name_of(clss)) 
                         if expected_class_size(st_name, c) >= n]
            my_excludes.extend(new_excl) 
    try_move_one_from(st_name, biggest_class_of(st_name), n=n, exp_n=exp_n, 
            includes = my_includes, excludes = my_excludes)


def smallest_classes_of(subj, exact=False) : 
    my_classes = classes_of(subj)
    min_size = min( map(size_of, my_classes))
    if exact is False : 
        output = [c for c in my_classes if size_of(c) <= min_size + 1]
    else : 
        output = [c for c in my_classes if size_of(c) <= min_size ]
    return output

def smallest_classes() : 
    subjs = list_util.complement( subjects, sem.junior_subjects )
    output = [ ] 
    for s in subjs : 
        output.extend(smallest_classes_of(s))
    return output
    
def full_class_members(subj, n=20) : 
    output = [ ] 
    for c in classes_of(subj) : 
        if size_of(c) >= n : 
            output.extend(st_members_of(c))
    return output


def export_baedang_template_data(filename) : 
    classes = list_util.union( [ class_name_of(pt) for pt in classparts 
                        if subject_name_of(pt) not in sem.skip_subject_list])
    output = [ (sem.subject_group_dict[ subject_name_of(c) ], 
                    c, subject_name_of(c) , class_number_of(c)) 
                for c in classes]
    #output.sort(key = (lambda x: (sem.ordered_group.index(x[0]), x[2], "%02d"%int(x[3])  )))
    output.sort(key = (lambda x: (sem.ordered_group.index(x[0]),
            x[2], "{:02d}".format(int(x[3])  ))))
    xlsx_util.write_tuples_into_xlsx(filename, output)


table_slots = [ [d + str(i) for d in days] for i in range(1,8) ] 

table_slot_str = '\n'.join( [ ' '.join(row) for row in table_slots] )


def show_slots_on_table(in_slots) : 
    my_dict = { }
    for c in list_util.flatten(table_slots) : 
        if c not in in_slots : 
            my_dict[ c ] = "___"
        else : 
            my_dict[ c ] = c
    output = list_util.partition( 
                map( lambda x: my_dict[x], list_util.flatten(table_slots)), 5)
    print('\n'.join( [ ' '.join(row) for row in output] ))
    
    

def show_weekly_table(clss) : 
    show_slots_on_table( [ str(classpart_slot[pt]) for pt in parts_of(clss) ] ) 

def show_weekly_tables(subj) : 
    for c in classes_of(subj) : 
        print(c, ' '.join(list_util.intersect(teachers,t_members_of(c))))
        show_weekly_table(c)
        print("\n")

def is_exp_subject(subj) : 
    if subj.find("실험") >= 0 : 
        return True 
    else : 
        return False


def make_min_max_dict(min_dict = sem.min_subjects_dict, 
                max_dict = sem.max_subjects_dict, current = False, 
                max_num = 20, exp_max_num = 13 ) : 
    output = { }
    if current is True : 
        for subj in subjects : 
            size_list = [size_of(c) for c in classes_of(subj)]
            if is_exp_subject(subj) : 
                max_bound = exp_max_num
            else : 
                max_bound = max_num
            min_bound = min(size_list)
            output[subj] = (min_bound, max_bound)
    else : 
        for subj in subjects : 
            if is_exp_subject(subj) : 
                output[subj] = (7, 13)
            else : 
                output[subj] = (7, 20)

    subj_max_dict = { }
    for num in max_dict : 
        for subj in max_dict[num] : 
            subj_max_dict[ subj ] = num 
    subj_min_dict = { }
    for num in min_dict : 
        for subj in min_dict[num] : 
            subj_min_dict[ subj ] = num 
    
    for subj in subj_max_dict : 
        min_num, max_num = output[subj]
        output[subj] = (min_num, subj_max_dict[subj])
    for subj in subj_min_dict : 
        min_num, max_num = output[subj]
        output[subj] = ( subj_min_dict[subj], max_num)
    return output
    

def teacher_hours(t_name) :
    pts = [pt for pt in classparts_of(t_name) if subject_name_of(pt) 
            in subjects ]
    output = 0.0
    for pt in pts :
        t_count = len( [ c for c in t_members_of(pt) 
            if not is_assistant_name(c) ] )
        output += (1.0 / t_count)
    return output


def teacher_prepare_hours(t_name) :
    pts = [pt for pt in classparts_of(t_name) if subject_name_of(pt) 
             in subjects]
    pts_uniq = list_util.union([ re.sub(r'_\d+', "", pt) for pt in pts ] )
    
    output = 0.0
    for pt in pts_uniq :
        t_count = len( [ c for c in t_members_of(pts[0]) 
            if not is_assistant_name(c) ] )
        output += (1.0 / t_count)
    return output


all_double_slots = list_util.partition( re.split(r'\s+',
'''월1 월2 월2 월3 월3 월4 월5 월6 
화1 화2 화2 화3 화3 화4 화5 화6 화6 화7 
수1 수2 수2 수3 수3 수4 
목1 목2 목2 목3 목3 목4 목5 목6 
금1 금2 금2 금3 금3 금4 금5 금6'''), 2)

all_slots = list_util.union(list_util.flatten(all_double_slots))

all_triple_slots = [ (c[0], c[1], d) for c in all_double_slots 
                    for d in all_slots if c[1][:-1] != d[:-1] ]
all_quadruple_slots = list_util.union( 
        [ tuple(sorted((c[0], c[1], d[0], d[1]))) for c in all_double_slots 
                    for d in all_double_slots if c[1][:-1] != d[1][:-1] ])

all_chinese_pairs = [ (c, d) for c in re.split(r'\s+',
'''월1 월2 월3 월4 월5 월6 화1 화2 화3 화4 화5 화6 화7 
수1 수2 수3 수4''') for d in  re.split(r'\s+',
'''목1 목2 목3 목4 목5 목6 금1 금2 금3 금4 금5 금6''')]


def count_free_chinese_pairs(name_list) : 
    pair_slot_counter = {}
    for pair in all_chinese_pairs : 
        pair_slot_counter[pair] = 0
    for st_name in name_list : 
        st_slots = used_slots_of(st_name)
        for pair in pair_slot_counter : 
            if len(list_util.intersect(pair, st_slots)) == 0 :   
                pair_slot_counter[pair] += 1
    return pair_slot_counter




def count_free_triples(name_list, t_names=[]) : 
    triple_slot_counter = {}
    for triple in all_triple_slots : 
        triple_slot_counter[triple] = 0
    for st_name in name_list : 
        st_slots = used_slots_of(st_name)
        for triple in triple_slot_counter : 
            if len(list_util.intersect(triple, st_slots)) == 0 :   
                triple_slot_counter[triple] += 1
    for t_name in t_names : 
        t_slots = used_slots_of(t_name)
        for triple in triple_slot_counter.keys():
            if len(list_util.intersect(triple, t_slots)) > 0 :   
                triple_slot_counter[triple] = 0
    return triple_slot_counter


def count_free_quadruples(name_list) : 
    quadruple_slot_counter = {}
    for quadruple in all_quadruple_slots : 
        quadruple_slot_counter[quadruple] = 0
    for st_name in name_list : 
        st_slots = used_slots_of(st_name)
        for quadruple in quadruple_slot_counter : 
            if len(list_util.intersect(quadruple, st_slots)) == 0 :   
                quadruple_slot_counter[quadruple] += 1
    return quadruple_slot_counter

def subject_num_class(subj, cnum) : 
    return (subj + "_" + str(cnum))
        

def used_slots_of_class(clss) : 
    t_st_mems = list_util.union( list_util.flatten( 
                    [members(cp) for cp in parts_of(clss)]))
    tmp_slots = list_util.union( list_util.flatten( [ used_slots_of(name) 
                    for name in t_st_mems] ), [ classpart_slot[cp] ]  )
    return  tmp_slots 


def show_free_pairs(name_list, size=14, subj = None, cnum = None) : 
    if subj is None : 
        filter_fn = (lambda x: True)
    else : 
        clss = subject_num_class(subj,cnum)
        u_sl = used_slots_of_class(clss)
        filter_fn = (lambda x: len(list_util.intersect(x[:2], u_sl))==0)
    t_dict = count_free_chinese_pairs(name_list)
    output = filter(filter_fn,  
                [ c + (t_dict[c], ) for c in t_dict if t_dict[c]>= size])
    output.sort(key = (lambda x:(-x[-1], x)))
    list_util.join_npr( [ map(str,c) for c in output] )
    return [ list(c[:2]) for c in output]




def show_free_triples(name_list, size=14, subj = None, cnum = None, 
        t_names=[] ) : 
    if subj is None : 
        filter_fn = (lambda x: True)
    else : 
        clss = subject_num_class(subj,cnum)
        u_sl = used_slots_of_class(clss)
        filter_fn = (lambda x: len(list_util.intersect(x[:3], u_sl))==0)
    t_dict = count_free_triples(name_list, t_names=t_names)
    output = filter(filter_fn,  
                [ c + (t_dict[c], ) for c in t_dict if t_dict[c]>= size])
    output.sort(key = (lambda x:(-x[-1], x)))
    list_util.join_npr( [ map(str,c) for c in output] )
    return [ list(c[:3]) for c in output]

def show_free_quadruples(name_list, size=14, subj = None, cnum = None) : 
    if subj is None : 
        filter_fn = (lambda x: True)
    else : 
        clss = subject_num_class(subj,cnum)
        u_sl = used_slots_of_class(clss)
        filter_fn = (lambda x: len(list_util.intersect(x[:3], u_sl))==0)
    t_dict = count_free_quadruples(name_list)
    output = filter(filter_fn,  
                [ c + (t_dict[c], ) for c in t_dict if t_dict[c]>= size])
    output.sort(key = (lambda x:(-x[-1], x)))
    list_util.join_npr( [ map(str,c) for c in output] )
    return [ list(c[:3]) for c in output]


def is_soph_subject(subj, prefix = "1") : 
    if prefix in [c[0] for c in members(subj)] : 
        return True
    else : 
        return False


def slot_count_str(name_list, cl_slots = [ ]): 
    collected_slots = list_util.flatten( [ used_slots_of(name) 
                    for name in name_list] )
    tmp_slots = list_util.union( collected_slots   )
    my_slot_dict = {}
    total = len(name_list)
    for sl in tmp_slots : 
        tmp_count = collected_slots.count(sl)
        if total > tmp_count : 
            my_slot_dict[sl] = (total - collected_slots.count(sl))
    av_c = list_util.complement( slots, tmp_slots )
    return slots_in_tabbing( av_c, cl_slots = cl_slots, 
                count_dict = my_slot_dict )


def move_class_to_slots(clss, slot_str) : 
    remove_all_names_from_class(clss)
    clear_slot(clss)
    slot_list = slot_str.split()
    pt_list = parts_of(clss)
    for i,sl in enumerate(slot_list) : 
        assign_part_slot( pt_list[i], sl)
        
    

def divide_into_groups(subj, num) : 
    clsses = classes_of(subj)
    cl_groups = list_util.partition( clsses , num )
    num_per_clss = int( 1 + len(members(subj)) 
                    / float(len(list_util.flatten(cl_groups))))
    st_members = members(subj); 
    random.shuffle(st_members)
    tmp_clsses = list_util.partition( st_members, num_per_clss )
    st_groups = map(list_util.flatten, list_util.partition(tmp_clsses, num))
    output = [ ] 
    for i, cl_gp in enumerate(cl_groups) : 
        output.append(  [ list_util.complement(clsses, cl_gp), st_groups[i]])
    return output
    

def divide_into_groups_of_classes(subj, num_list) : 
    clsses = classes_of(subj)
    cl_groups = list_util.partition_by_list( clsses , num_list )
    num_per_clss = int( 1 + len(members(subj)) 
                    / float(len(list_util.flatten(cl_groups))))
    st_members = members(subj); 
    random.shuffle(st_members)
    tmp_clsses = list_util.partition( st_members, num_per_clss )
    st_groups = map(list_util.flatten, 
            list_util.partition_by_list(tmp_clsses, num_list))
    output = [ ] 
    for i, cl_gp in enumerate(cl_groups) : 
        output.append(  [ list_util.complement(clsses, cl_gp), st_groups[i]])
    return output
    
    
    

def save_exclusions(pairs, save_filename = "exclusions.semp" ) : 
    leave_a_backup( save_filename )
    with open( save_filename, "wb" ) as f :
        #print "Writing exclusions into <%s> " % save_filename
        print("Writing exclusions into <{}> ".format(save_filename))
        pickle.dump( pairs, f )


def count_slot_fits(clss, name_list) : 
    cl_slots = current_slots(clss)
    output = 0 
    for st_name in name_list : 
        st_slots = used_slots_of(st_name)
        if len(list_util.intersect(cl_slots, st_slots)) == 0 :   
            output += 1
    return output


def cc(subj, c_num, command = "show",  st_list = [ ], across = False) : 
    clss = subj + "_" + str(c_num)
    my_slots = list_util.complement(current_slots_unsorted(clss), [None])
    exc_names = sem.exclude_names_of(clss, across = across)

    if len(st_list) == 0 : 
        disp_st_list = list_util.complement( remaining_names_of(subj), 
           exc_names )
    else : 
        disp_st_list = list_util.complement( st_list, exc_names )
    if command == "show" : 
        return show_classes_of_subject(subj, my_slots, st_list=disp_st_list, 
                    sort_key=remaining_shorthands )
    if command == "show_c" : 
        return show_classes_of_subject(subj, my_slots, st_list=disp_st_list, 
                    sort_key=remaining_shorthands, show_count=True )
    elif command == "show_all":
        return show_classes_of_subject(subj, [ ] , st_list=members(subj), 
                    sort_key=remaining_shorthands )
    elif command == "show_class":
        return show_classes_of_subject(subj, [ ] , 
                st_list=st_members_of(subj + "_" + str(c_num)), 
                    sort_key=remaining_shorthands )
    elif command.startswith("show_p "): 
        my_num = int(command.split()[1])
        print(command, my_num)
        my_list = show_classes_of_subject(subj, my_slots, st_list=disp_st_list, 
                    sort_key=remaining_shorthands )
        print(my_list)
        return show_free_pairs(my_list, size = my_num)
    elif command.startswith("show_t "): 
        my_num = int(command.split()[1])
        my_list = show_classes_of_subject(subj, my_slots, st_list=disp_st_list, 
                    sort_key=remaining_shorthands )
        show_free_triples(my_list, size = my_num, 
            t_names = t_members_of(subj + "_" + str(c_num)))
    elif command.startswith("show_q "): 
        my_num = int(command.split()[1])
        my_list = show_classes_of_subject(subj, my_slots, st_list=disp_st_list, 
                    sort_key=remaining_shorthands )
        show_free_quadruples(my_list, size = my_num)
    elif command == "show_i": 
        tmp_list = list_util.intersect(disp_st_list, remaining_names_of(subj))
        for clss in classes_of(subj) : 
            print(clss, count_slot_fits(clss, tmp_list))
        return show_classes_of_subject(subj, my_slots, 
         st_list= tmp_list, 
                    sort_key=remaining_shorthands )
    elif command.startswith("show_from "): 
        my_num = int(command.split()[1])
        my_list = show_classes_of_subject(subj, my_slots, 
                    st_list = st_members_of(subj + "_" + str(my_num)))
        return my_list
    elif command.startswith("put_from "): 
        iter_num, my_num = map(int, command.split()[1:])
        my_list = show_classes_of_subject(subj, my_slots, 
           st_list = list_util.complement(
                    st_members_of(subj + "_" + str(my_num)), exc_names)  )
        cc(subj, c_num, "put_all" , my_list[:iter_num], across = across)
    elif command.startswith("show_compl"): 
        my_list = show_classes_of_subject(subj, my_slots, 
            st_list = list_util.complement(members(subj), 
                        st_members_of(clss), exc_names))
        return my_list
    elif command.startswith("put_compl "): 
        my_num = int(command.split()[1])
        my_list = show_classes_of_subject(subj, my_slots, 
            st_list = list_util.complement(members(subj), 
                        st_members_of(clss), exc_names))
        cc(subj, c_num, "put_all" , my_list[:my_num], across = across)
    elif command.startswith("random_from "): 
        my_num, iter_num  = map(int, command.split()[1:])
        try_move_random_ones_from_to( subj, my_num, c_num, 
            fix_names = [ ], excl_classes = [ ], n = 19, exp_n = 12, count = 1, 
        lower_bound = 9, iter_num = iter_num, stop_count=0, randomize=True )  
    elif command.startswith("put ") : 
        my_num = int(command.split()[1])
        st_names = show_classes_of_subject(subj, my_slots, 
                        st_list=disp_st_list, sort_key=remaining_shorthands ) 
        put_names_into_class(st_names[:my_num], clss ) 
    elif command.startswith("put_all") : 
        st_names = show_classes_of_subject(subj, my_slots, 
                        st_list=disp_st_list, sort_key=remaining_shorthands ) 
        put_names_into_class(st_names, clss ) 
    elif command.startswith("put_i ") : 
        my_num = int(command.split()[1])
        st_names = show_classes_of_subject(subj, my_slots, 
         st_list = list_util.intersect(disp_st_list, remaining_names_of(subj)), 
                    sort_key=remaining_shorthands )
        put_names_into_class(st_names[:my_num], clss ) 
    elif command.startswith("fill_up_by "): 
        tmp_list = command.split()
        size = int(tmp_list[1])
        nums = map(int, tmp_list[2:])
        for i in nums : 
            cc(subj, i, "put " + str(size))
        return True
    elif command.startswith("try_rem"): 
        try_assign_remains(subj,bound=0)
    elif command.startswith("try_c"): 
        try_assign_remains(subj, across = True)
    elif command == "remove_all_names" : 
        return remove_all_names_from_parts_of(subj)
    elif command == "remove_names" : 
        return remove_all_names_from_class(clss)
    elif command == "clear_class" : 
        remove_all_names_from_class(clss)
        return clear_slot(clss)
    elif command.startswith("slots ") : 
        tmp_list = command.split()
        my_slots = tmp_list[1:]
        return assign_c_slots(subj, c_num, my_slots)
    elif command.startswith("sl_ass "):
        cc(subj, c_num, command.replace("sl_ass", "slots"))
        cc(subj, c_num, "put_all", st_list, across = across)
        return cc(subj, 0, "show")
    elif command.startswith("sl_put "):
        tmp_list = command.split()
        my_slots = tmp_list[2:]
        assign_c_slots(subj, c_num, my_slots)
        cc(subj, c_num, "put " + tmp_list[1], st_list, across = across)
        return cc(subj, 0, "show")
    elif command.startswith("clear_slots") : 
        clear_slot(clss)
    elif command.startswith("clear_from_slots ") : 
        tmp_list = command.split()
        my_slots = tmp_list[1:]
        clear_class_from_slots(clss, my_slots)
    elif command.startswith("free_names ") : 
        my_slots = command.split()[1:]
        free_names_from_slots( 
            list_util.intersect(st_list, members(subj)), my_slots) 
    elif command == "rem_subj" : 
        for st_name in st_list : 
            list_util.jpr(remaining_subjects(st_name))
        return True
    else : 
        return None


def fill_and_try(subj ) : 
    cl_num = len(classes_of(subj))
    size = int( float(len(members(subj)))/cl_num )
    for i in range(1, len(classes_of(subj))+1) : 
        #command = ("put %d" % (size - size_of(subj + "_" +str(i))))
        command = ("put {}".format(size - size_of(subj + "_" +str(i))))
        cc(subj, i, command ) 
    try_assign_remains(subj, across = False)
    try_assign_remains(subj, across = True)


def reset_slots() : 
    for subj in subjects : 
        clear_slots_of(subj)
    for cp in sem.fixed_slots :
        assign_part_slot(cp, sem.fixed_slots[cp])

def assign_fixed_slots(subj) : 
    for cp in parts_of(subj) : 
        if cp in sem.fixed_slots : 
            assign_part_slot(cp, sem.fixed_slots[cp])


def collect_names_by_subjects(subj_list, st_list = students) : 
    output = [ n for n in st_list 
                if list_util.is_subset(subjects_of(n), subj_list) ] 
    return output

def collect_names_by_lists(subj_lists) : 
    output = list_util.union( list_util.flatten( 
                [collect_names_by_subjects(c) for c in subj_lists]))
    output.sort()
    return output

def maximal_subject_lists(st_list) : 
    subj_lists = list_util.union(
            [sorted(subjects_of(st_name)) for st_name in st_list])
    m_lists = [ c for c in subj_lists 
                    if list_util.is_maximal_in(c, subj_lists)]
    m_lists.sort()
    return m_lists

def take_a_group(subj_lists, st_list, size=12) : 
    output = [ ]
    for s_list in subj_lists : 
        new_candis = collect_names_by_subjects(s_list, st_list)
        for st_name in new_candis : 
            check_append(output, st_name)
        if len(output) >= size : 
            return output
    else : 
        return output

def group_from_remains(subj, size=15) :
    my_list = remaining_names_of(subj)
    subj_lists = maximal_subject_lists(my_list)
    output = take_a_group(subj_lists, my_list, size=size)
    return output

def exclude_classes_by_gap(st_name, gap=5) : 
    subj_list = [subj for subj in subjects_of(st_name) 
                    if gap_of_subject(subj) > gap - 2 ]
    output = [ ] 
    for subj in subj_list : 
        c_list = classes_of(subj)
        max_size = max( map(size_of, c_list))
        output.extend( [c for c in c_list if size_of(c) >=  max_size ])
    return output
        
def max_classes(st_name, probability = 0.3) :     
    subj_list = subjects_of(st_name)
    output = [ ] 
    for subj in subj_list : 
        if random.random() < probability : 
            continue
        max_size = max( map(size_of, classes_of(subj)))
        max_list = [c for c in classes_of(subj) if size_of(c) >= max_size - 1]
        if len(classes_of(subj)) - len(max_list) > 1 : 
            output.extend(max_list)
            continue
        max_list = [c for c in classes_of(subj) if size_of(c) >= max_size ]
        if len(classes_of(subj)) - len(max_list) > 1 : 
            output.extend(max_list)
    return output
        
        
        
    


def rehash(subj, sort_key = None, max_num = None, fix_max = True, gap=None, 
        filter_fn = (lambda x:True), avoid_max = True, probability=0.3) : 
    if fix_max is True : 
        max_size = max([size_of(c) for c in classes_of(subj)])
        clsses = [c for c in classes_of(subj) if size_of(c) == max_size] 
        name_list = list_util.flatten(
            [st_members_of(c) for c in classes_of(subj)])
    else : 
        name_list = members(subj)

    name_list = filter(filter_fn, name_list)[:]

    random.shuffle(name_list) 

    if max_num is None : 
        max_count = len(name_list) 
    else : 
        max_count = max_num 
    count = 0
    if gap is None : 
        my_gap = gap_of_subject(subj)
    else : 
        my_gap = gap

    for st_name in name_list : 
        try: 
            if avoid_max is True : 
                tmp_excls = max_classes(st_name, probability=probability)
            else : 
                tmp_excls = exclude_classes_by_gap(st_name, gap=my_gap)
            choose_and_assign(st_name, sort_key = sort_key, 
                excludes = tmp_excls)
            count += 1
            if count >= max_count : 
                break 
        except (IndexError, KeyboardInterrupt): 
            #print "No class-assignment found for <%s>" % st_name
            print("No class-assignment found for <{}>".format(st_name))
            
    
def subject_order_fn(gap_subjects):
    subj_list = [c for c in reversed(gap_subjects)] + \
           sorted(list_util.complement(subjects, gap_subjects), 
            key = (lambda x: len(classes_of(x))))
    return subj_list.index


def rank_in(my_num, num_list) : 
    return len( filter( lambda x: x >= my_num  , num_list))

def size_rank_dict_of(subj, weighted=True):
    cl_list = classes_of(subj)
    cl_sizes = [ size_of(c) for c in cl_list ] 
    ranks = [ rank_in(sz, cl_sizes) for sz in cl_sizes]
    if weighted : 
        ranks = [ (8 - len(cl_list))/2 + rk for rk in ranks]
    return dict( list_util.transposed( [cl_list, ranks] ))
    
def size_rank_dict() :     
    output = {}
    for subj in subjects : 
        output.update( size_rank_dict_of(subj))
    return output

def size_ranks_of(st_name, rank_dict) : 
    ranks = [ rank_dict[c] for c in current_classes(st_name) ]
    return sorted(ranks)
    
def str_sorted_classes_of(st_name, rank_dict = None, c_num=False):
    output = [ ] 
    output.append(st_name)
    if rank_dict : 
        sort_fn = (lambda x:rank_dict[x])
    else : 
        sort_fn = (lambda x: -size_of(x))
    for cl in sorted(current_classes(st_name), key = sort_fn ):
        if c_num :
            cl_str = class_shorthands(cl) 
        else : 
            cl_str = sem.shorthands[subject_name_of(cl)] 
        #output.append( "<%s %d>" % (cl_str, size_of(cl)))
        output.append( "<{} {}>".format(cl_str, size_of(cl)))
    return ' '.join(output)

def min_classes_of(st_name, rank_dict):
    output = [ ]
    for cl in current_classes(st_name) : 
        if max( [rank_dict[c] for c in classes_of(subject_name_of(cl))] ) \
            == rank_dict[cl] : 
            output.append(cl)
    return output

def overfulling_classes_of(st_name, subj) : 
    my_clss = name_subject_class(st_name, subj)
    if len(my_clss) == 0 : 
        return [ ]
    max_n = max( map(size_of, classes_of(subj)))
    maxed_classes = [ c for c in classes_of(subj) if size_of(c) == max_n ]
    if size_of(my_clss) < max_n : 
        return maxed_classes
    else : 
        output = list_util.complement( maxed_classes, [ my_clss ] )
        return output


def try_move(st_name, rank_dict, from_which = 1, bunch=True, sort_key = None, 
        target = None, include_class = None, fix_subjects = [ ], 
          max_n = 20, new_fix_subjects = [ ], skip_names = None, 
            comp_dict=None ):
    if include_class is not None : 
        bal_excludes = list_util.complement(
            classes_of(subject_name_of(include_class)), [include_class])
    elif target is None : 
        bal_excludes = [ ] 
    else : 
        max_cl = min( classes_of(target), key = lambda x: rank_dict[x])
        bal_excludes = [c for c in classes_of(target) 
                            if size_of(c) + 1 >= size_of(max_cl)]
    if max_n > 0 : 
        tmp_classes = list_util.flatten( 
            [classes_of(subj) for subj in subjects_of(st_name)])
        bal_excludes.extend( list_util.complement( 
            [ c for c in tmp_classes if size_of(c)>= max_n ], 
            current_classes(st_name)) )
        
    for subj in new_fix_subjects : 
        bal_excludes.extend( overfulling_classes_of(st_name, subj) )

        
    includes = min_classes_of(st_name, rank_dict)
    for subj in fix_subjects : 
        my_clss = name_subject_class(st_name, subj)
        if len(my_clss) > 0 : 
            includes.append( my_clss)

    includes = list_util.union(includes)

    if bunch : 
        excludes = sorted(current_classes(st_name), 
            key = lambda x:rank_dict[x])[:from_which]
    else : 
        excludes = [sorted(current_classes(st_name), 
            key = lambda x:rank_dict[x])[from_which]]
    excludes.extend(bal_excludes)
    excludes = list_util.union( excludes )
    try: 
        #print "excludes = %s, \nincludes = %s" % (' '.join(excludes), ' '.join(includes))
        print("excludes = {}, \nincludes = {}".format(' '.join(excludes), 
                    ' '.join(includes)))
        print(str_sorted_classes_of(st_name, rank_dict))
        if comp_dict is None : 
            choose_and_assign(st_name, includes = includes, 
                excludes = excludes, sort_key = sort_key )
        else : 
            result = choose_first_by_dict_and_assign(comp_dict, st_name, 
                   includes = includes, excludes = excludes, verbose=True )
            if result is False:
                raise IndexError
        return True
    except (IndexError, KeyboardInterrupt): 
        print("No assignment found, or interrupted")
        if type(skip_names) is list : 
            #print "Adding %s to skip_names of %d students" % (st_name, len(skip_names))
            print("Adding {} to skip_names of {} students".format(st_name,
                        len(skip_names)))
            skip_names.append(st_name)
        return False




def members_not_in_smallest_classes(subj, by_one=True) : 
    st_list = list_util.complement(members(subj), 
                members_in_smallest_classes(subj, by_one=by_one))
    return st_list

def copy_and_put(super_list, sub_list , verbose=True, partial=True):
    pairs = [ (a,b) for a in super_list for b in sub_list
                    if is_subject_subset(b,a)]
    uniq_pairs = [ ]
    sub_one_list = [ ] 
    for super_one, sub_one in pairs : 
        if sub_one not in sub_one_list : 
            sub_one_list.append(sub_one)
            uniq_pairs.append( (super_one, sub_one) )

    if len(pairs)>0:
        if verbose : 
            #print "Found %d many pairs (%s, %s), ..." % (len(uniq_pairs), uniq_pairs[0][0], uniq_pairs[0][1])
            print("Found {} many pairs ({}, {}), ...".format(len(uniq_pairs),
                    uniq_pairs[0][0], uniq_pairs[0][1]))

        for my_pair in uniq_pairs : 
            copy_from_and_put( * my_pair, partial=partial )
        return True
    else : 
        print("No pair found")
        return False


def try_move_by_bound_dict(st_name, in_bound_dict = None , target = None,
    increase_allows = [ ], decrease_allows = [ ], 
    include_class = None):
    bound_dict = { }

    if target is not None : 
        t_subject = target
    else : 
        t_subject = max( subjects_of(st_name), 
            key = lambda x: max( map(size_of, classes_of(x))) )

    print("target:", t_subject)

    for subj in subjects_of(st_name): 
        min_size = min(map(size_of, classes_of(subj)))
        max_size = max(map(size_of, classes_of(subj)))
        if subj == t_subject:
            bound_dict[subj] = (min_size, max_size - 1)
        elif subj in list_util.intersect(increase_allows, 
                        decrease_allows):
            bound_dict[subj] = (min_size-1, max_size+1)
        elif subj in increase_allows:
            bound_dict[subj] = (min_size, max_size+1)
        elif subj in decrease_allows:
            bound_dict[subj] = (min_size-1, max_size)
        else:
            bound_dict[subj] = (min_size, max_size)

    if in_bound_dict is not None : 
        bound_dict.update(in_bound_dict)


    includes = [ ] 
    excludes = [ ]
    for subj in subjects_of(st_name):
        for clss in classes_of(subj):
            if expected_class_size(st_name, clss) > bound_dict[subj][1] : 
                excludes.append(clss)

    for clss in current_classes(st_name):
        if expected_class_size(st_name, clss, removing=True) < \
                     bound_dict[subj][0] : 
            includes.append(clss)

    if include_class is not None : 
        tmp_includes = filter(
    lambda x: subject_name_of(include_class)!=subject_name_of(x), includes)
        includes = tmp_includes + [include_class]

    try: 
        #print "excludes = %s, \nincludes = %s" % (' '.join(map( class_shorthands,excludes)), ' '.join(includes))
        print("excludes = {}, \nincludes = {}".format(' '.join(map(
                class_shorthands,excludes)), ' '.join(includes)))
        print(str_sorted_classes_of(st_name))
        choose_and_assign(st_name, includes = includes, excludes = excludes)
        return True
    except (IndexError, KeyboardInterrupt): 
        print("No assignment found, or interrupted")
        return False

def try_move_random_one_by_bound_dict(clss , in_bound_dict = None , 
         increase_allows = [ ], decrease_allows=[]):
    random.seed()
    st_name = random.choice(st_members_of(clss))
    subj = subject_name_of(clss)
    try_move_by_bound_dict(st_name, target=subj, in_bound_dict=in_bound_dict,
        increase_allows=increase_allows, decrease_allows=decrease_allows)



def try_move_on_list(name_list, rank_dict, from_which = 1, 
     bunch=True, sort_key = None, 
        target = None, include_class = None, fix_subjects = [ ] , max_n = 20, 
        new_fix_subjects = [ ] , skip_names = None, comp_dict=None ):
    n = len(name_list)
    for i, st_name in enumerate(name_list) : 
        print(i+1, "/", n, ":")
        result = try_move(st_name, rank_dict, from_which = from_which, 
                    bunch = bunch, sort_key = sort_key, 
                    target = target, include_class = include_class, 
                    fix_subjects = fix_subjects, max_n = max_n, 
                    new_fix_subjects = new_fix_subjects,
                    skip_names = skip_names, comp_dict=comp_dict ) 
        if result : 
            return True
    else: 
        return False


def biggest_class(subj) : 
    return max( classes_of(subj), key = size_of )

def members_in_biggest_classes(subj, gap = 0 ) : 
    max_n = max( map(size_of, classes_of(subj)))
    max_classes = [c for c in classes_of(subj) if size_of(c) >= max_n - gap ]
    return list_util.union( * [st_members_of(c) for c in max_classes])




def one_way_excludes_of(st_name, target = None) : 
    output = [ ]
    for subj in subjects_of(st_name) : 
        clss = name_subject_class(st_name, subj)
        my_size = size_of(clss)
        tmp_list = [c for c in classes_of(subj) if size_of(c) >= my_size]
        output.extend( list_util.complement( tmp_list, [ clss ]))
    if target is not None : 
        clss = name_subject_class(st_name, target)
        my_size = size_of(clss)
        if len(clss)>0 : 
            tmp_list = [c for c in classes_of(target) 
                    if size_of(c)+1 >= my_size]
            output.extend( tmp_list )
    return output
        

def try_one_way(st_name, sort_key = None, target = None) : 
    my_excludes = one_way_excludes_of(st_name, target = target)
    try: 
        #print "excludes = %s" % (' '.join(my_excludes),)
        print("excludes = {}".format(' '.join(my_excludes)))
        print(str_sorted_classes_of(st_name))
        choose_and_assign(st_name,  excludes = my_excludes, 
            sort_key = sort_key )
        return True
    except (IndexError, KeyboardInterrupt): 
        print("No assignment found, or interrupted")
        return False

def try_one_way_on_list(name_list,  sort_key = None, target = None) : 
    for st_name in name_list : 
        result = try_one_way(st_name, sort_key = sort_key, target = target) 
        if result : 
            return True
    else: 
        return False

def read_name_subjects_from_table(filename, sheet_name = "Sheet2"):
    sub_namelists = list_util.transposed( 
        xlsx_util.read_xlsx_sheet_into_list(filename, sheet_name = sheet_name))
    output = [ ] 
    for line in sub_namelists : 
        subj = line[0]
        output.extend( [ (st_name, subj) for st_name in line[1:] 
            if len(st_name) > 0 ])
    return output
    
    

def read_assignment_data(my_dict) : 
    name_class_hours = {}
    for sheet in my_dict : 
        if sheet == "교과목개설현황" : 
            continue
        lines = my_dict[sheet]
        for i, line in enumerate(lines) : 
            if ''.join(line).find("23456") >= 0 :
                start_ind = i
                break
        for i, line in enumerate(lines) : 
            if i >=  start_ind and len(lines[i][3])==0 \
                and len(lines[i+1][3])==0 : 
                end_ind = i 
                break 
        data_lines = lines[start_ind:end_ind]
        ind_cnum_dict = { }
        for i,cnum in enumerate(data_lines[0]):
            if re.search(r'\d+', cnum ) and int(cnum)>0 : 
                ind_cnum_dict[i] = cnum
        curr_name = None
        for line in data_lines[1:]:
            if len(line[1])>0 and line[1]!= curr_name : 
                curr_name = line[1].strip()
            curr_subj = utf_util.convert_romans(line[3])
            if len(curr_subj) == 0 : 
                continue
            for ind in ind_cnum_dict : 
                if len(line[ind])>0 : 
                    clss = curr_subj + "_" + ind_cnum_dict[ind]
                    name_class_hours[ (curr_name, clss) ] = float(line[ind])
    return name_class_hours



def find_new_name(st_name, new_names) : 
    if st_name.startswith("0") : 
        candis = [c for c in new_names if c.startswith( "1" + st_name[1:4])]
    elif st_name.startswith("1") :
        candis = [c for c in new_names if c.startswith( "2" ) 
                        and c.endswith( st_name[4:] ) ]
    elif st_name.startswith("2") :
        candis = [c for c in new_names if c.startswith( "3" ) 
                        and c.endswith( st_name[4:] ) ]
    else : 
        return None
    if len(candis) == 1 : 
        return candis[0]
    else : 
        print(' '.join(candis))
        return None



    
def enroll_one_four_five_eight() : 
    soph_names = [c for c in students if c.startswith("2")]
    one_four_names = [c for c in soph_names if c[1] in "1 2 3 4".split() ]
    five_eight_names = list_util.complement(soph_names, one_four_names)
    for st_name in soph_names : 
        pair = (st_name,  sem.one_four_spring_soph) 
        check_remove(name_subjects, pair)
        pair = (st_name,  sem.five_eight_spring_soph) 
        check_remove(name_subjects, pair)

    for st_name in one_four_names : 
        pair = (st_name,  sem.one_four_spring_soph) 
        check_append(name_subjects, pair, lambda x:' -> '.join(x))
    for st_name in five_eight_names :
        pair = (st_name,  sem.five_eight_spring_soph) 
        check_append(name_subjects, pair, lambda x:' -> '.join(x))

def export_day_slot_names_xlsx(filename, name_list): 
    triples = [ ] 
    for t_name in name_list : 
        for pt in classparts_of( t_name ) : 
            if subject_name_of(pt) in "담임업무".split() : 
                continue
            else : 
                sl = classpart_slot[pt]
                my_day = sl[:-1]
                my_sl_str = sl[-1:] + "교시"
                triples.append( (my_day, my_sl_str, t_name.replace("t","")) )
    output = { }                
    for day in days : 
        for i in range(1,8) : 
            sl_str = str(i) + "교시"
            data_list = sorted( [ trip[2] for trip in triples 
                            if (day, sl_str) == (trip[0], trip[1]) ])
            output[ (day, sl_str) ] = ",".join(data_list)
    xlsx_util.write_dict_into_xlsx(filename, output, in_row_names = days) 
        


def read_neis_time_table(filename) : 
    ind_slot_dict = dict( list_util.partition(
re.split(r'\s+',
'''3 월1
4 월2
5 월3
6 월4
7 월5
8 월6
9 월7
13 화1
14 화2
15 화3
16 화4
17 화5
18 화6
19 화7
23 수1
24 수2
25 수3
27 수4
34 목1
35 목2
36 목3
38 목4
40 목5
41 목6
42 목7
46 금1
47 금2
48 금3
49 금4
50 금5
51 금6
52 금7'''), 2))
    data = xlsx_util.read_xlsx_sheet_into_list(filename)
    output = [ ] 
    for rec in data : 
        if rec[2].find("(")>=0 : 
            sr = re.search( r'\((\S+)\)', rec[2])
            t_name = sr.group(1) + "t"
            for k in ind_slot_dict : 
                if len( rec[ int(k) ] ) > 0 : 
                    sl = ind_slot_dict[k]
                    su = re.search(r'\((.+)\((\d+)\)\)', rec[int(k)] )
                    if su is None : 
                        print(rec[int(k)])
                        continue
                    clss = utf_util.convert_romans(su.group(1)) + "_" \
                            + su.group(2) 
                    output.append( (t_name, clss, sl) )
    return output
                
        
    
def diff_with_neis_time_table(filename) : 
    data = read_neis_time_table(filename) 
    tmp =  list_util.flatten( 
            [[ (t_name, class_name_of( name_slot_part(t_name, slot )), slot) 
          for slot in used_slots_of(t_name) ]  for t_name in teachers ])
    mod_fn = (lambda x: re.sub(r'(창의융합특강[^/]+)/[^_]+(_.*)', r'\1\2' ,x) )
    orig_data = [ c for c in tmp if subject_name_of(c[1]) in subjects ]
    mod_data = map(tuple, list_util.map_at(mod_fn, orig_data, level = 2))
    print("Missing items")
    list_util.join_npr( sorted( list_util.complement(mod_data, data)) )
    print("Wrong items")
    list_util.join_npr( sorted( list_util.complement(data,mod_data)) )
    

def grade_list_of(st_names) : 
    output = list_util.union( [c[0] for c in st_names])
    output.sort()
    return output

def students_taught_by(t_name) : 
    output = list_util.union( list_util.flatten( 
                [ st_members_of(pt) for pt in classparts_of(t_name)]))
    return output



def search_names( substr ) :
    pat = re.compile( str(substr) )
    output = [ ]
    for c in names :
        if pat.search(c) : 
            output.append(c)
    return ' '.join(output)


def fill_down_on_column( tuples, col_num ) : 
    output = [ ] 
    curr_val = 'BLANK'
    for rec in tuples : 
        if len(rec[ col_num ]) > 0 : 
            curr_val = rec[ col_num ]
        out_rec = rec[:]
        out_rec[col_num] = curr_val
        output.append(out_rec)
    return output




def distribute_into_classes(subj) : 
    c_list = classes_of(subj)
    my_size = int(average_size(subj))
    for i in range(len(c_list) - 1):
        put_names_into_class( remaining_names(subj)[:my_size], c_list[i])
    put_names_into_class( remaining_names(subj), c_list[-1])



def assistant_part_pairs(name_classroom_dict) : 
    output = [ ] 
    t_names = sorted(name_classroom_dict.keys(), 
        key = lambda x: (sem.ordered_departments.index(
                sem.name_department_dict[x]), x) )
    for t_name in t_names : 
        for room in name_classroom_dict[t_name].split('/') : 
            for cp in list_util.flatten( [ parts_of(c) for c 
                            in current_classes(t_name)]) : 
                list_util.check_append(output,  (room + "조교t", cp) )
    return output

def add_assistant_parts() : 
    for pair in sem.default_assistant_part_pairs : 
        list_util.check_append( name_classparts, pair)

def add_assistant_to_subject(a_name, subj) : 
    for pt in parts_of(subj) : 
        list_util.check_append( name_classparts, (a_name, pt), pair_str )


def second_year_subjects() : 
    output = [ ]
    for subj in subjects : 
        prefs = list_util.union([ st_name[0] for st_name in members(subj) ])
        if "2" in prefs : 
            output.append(subj)
    return output

def change_assistant(clss_or_part, from_str, to_str) : 
    if clss_or_part.find("$")>0 : 
        pt_list = [clss_or_part]
    else :
        pt_list = parts_of(clss_or_part)
    pairs = [ c for c in name_classparts 
        if c[1] in pt_list and is_assistant_name(c[0])]
    new_pairs = list_util.union([(c[0].replace(from_str, to_str), c[1]) 
                    for c in pairs])
    for pair in pairs : 
        check_remove(name_classparts, pair, pair_str)
    for pair in new_pairs : 
        check_append(name_classparts, pair, pair_str)



def order_with_priority_fn(pnum=1) : 
    if type(pnum) is int : 
        subj_list = subjects[:]
        subj_list.sort( key = lambda x : - gap_of_subject(x) )
        rest = subj_list[pnum:]
        rest.sort( key =  (lambda s: len(classes_of(s)) ) ) 
        output = subj_list[:pnum] + rest 
    else : 
        head_list = pnum
        rest = list_util.complement(subjects, head_list)
        rest.sort( key =  (lambda s: len(classes_of(s)) ) ) 
        output = head_list + rest
    return output.index



def members_in_smallest_classes(subj, by_one=False ) : 
    min_n = min( map(size_of, classes_of(subj)))
    if by_one : 
        min_n += 1
    smallest_classes = [c for c in classes_of(subj) if size_of(c) <=  min_n ]
    return list_util.union( * [st_members_of(c) for c in smallest_classes])


def is_in_smallest_class(st_name, subj) : 
    return (st_name in members_in_smallest_classes(subj))


def try_assign_one_name_from(st_list, across = False, sort_key = None, 
        fix_subjects = [], bound = 20, subj = None, small = False, 
        stay_in_small = True, small_except = [ ]  ) : 
    n = len(st_list)
    for i, st in enumerate(st_list) : 
        try:
            print(i+1, "/", n, ": ")
            if subj is not None : 
                target_subj = subj
                clss = name_subject_class(st, target_subj)
                big_classes = [c for c in classes_of(target_subj) 
                        if size_of(c) >= size_of(clss) - 1 ]
                if small is True : 
                    smallest_size = min([size_of(c) for c in classes_of(subj)])
                    big_classes = [c for c in classes_of(subj) 
                                     if size_of(c) > smallest_size ]
                #print "Excluding %s" % ' '.join(big_classes)
                print("Excluding {}" % ' '.join(big_classes))
            else : 
                big_classes = [ ] 

            includes = [ ] 
            if stay_in_small : 
                for st_subj in subjects_of(st) : 
                    if (st_subj not in small_except) and \
                            is_in_smallest_classes(st, st_subj) : 
                        includes.append( name_subject_class(st, st_subj))
                #print "Including %s" % ' '.join(includes)
                print("Including {}" % ' '.join(includes))
            
            choose_and_assign(st, count=1, bound=bound, across = across, 
                sort_key = sort_key, 
            fix_subjects = list_util.intersect(fix_subjects, subjects_of(st)), 
                excludes = big_classes, includes = includes) 
            return True
        except (IndexError, KeyboardInterrupt):
            pass
    else : 
        return False


def remove_assistant_from(as_name, clss_or_part) : 
    if clss_or_part.find("$")>0 : 
        pt_list = [clss_or_part]
    else :
        pt_list = parts_of(clss_or_part)
    pairs = [ (as_name, pt) for pt in pt_list ] 
    for pair in pairs : 
        check_remove(name_classparts, pair, pair_str)

def classroom_by_assistants(pt) : 
    as_names = filter( is_assistant_name, t_members_of(pt) )
    if len(as_names) == 0 : 
        return None
    room_names = [ c.replace("조교t", "") for c in as_names ]
    room_names.sort()
    return '/'.join(room_names)

def assign_classrooms_by_assistants() : 
    pt_list = list_util.flatten( [ parts_of(subj) for subj in subjects])
    for pt in pt_list : 
        assign_classroom_to_part(pt, classroom_by_assistants(pt))





def export_time_tables_of_all_departments() : 
    d_list = list_util.union( sem.name_department_dict.values() )
    dept_names = list_util.reversed_dict(sem.name_department_dict)
    for dept in d_list : 
        export_weekly_tables_xlsx( dept + "_" + sem.semester_prefix 
            + "교사별_시간표.xlsx",  in_name_list = sorted(dept_names[dept])) 
        
def export_slots_xlsx(filename, c_parts, t_name = False) : 
    if t_name is False : 
        output = [  (pt, classpart_slot[pt]) for pt in c_parts]
    else : 
        output = [  (pt, classpart_slot[pt], 
            ' '.join(filter(lambda x: not is_assistant_name(x), 
                    t_members_of(pt)  )) ) for pt in c_parts]
    
    xlsx_util.write_tuples_into_xlsx(filename, output)


def read_slot_dict_from_xlsx(filename) : 
    data = xlsx_util.read_xlsx_sheet_into_list(filename)
    return dict( [ (c[0], c[1]) for c in data if c[0] in classparts])

def update_avoid_slots_from(filename): 
    data = xlsx_util.read_xlsx_sheet_into_list(filename)[1:]
    my_dict = dict(list_util.transposed(["1 2 3 4 5 6 7".split(), 
        "일 이 삼 사 오 륙 칠".split()]))
    output = [ ] 
    for rec in data : 
        t_name = rec[0] + "t"
        my_slots = list_util.intersect(rec[1:], slots)
        for sl in my_slots : 
            cp = gen_util.replace_by_dict(my_dict, sl) + "_1$1"
            output.append( (t_name, cp) )
            add_teacher_to_part( t_name, cp )
            check_append(classparts, cp)
            assign_part_slot(cp, sl) 
    for cp in list_util.union( [c[1] for c in output]): 
        assign_classroom_to_part(cp, None) 


def clear_slots_of_teacher(t_name) : 
    my_parts = [ cp for cp in classparts_of(t_name) 
                    if  subject_name_of(cp) in subjects]
    for cp in my_parts : 
        assign_part_slot(cp, None) 

def clear_slots_of_subject(subj) :     
    for cp in parts_of(subj) : 
        assign_part_slot(cp, None) 

    
def filter_by_class_slots(clss, st_list):
    class_slots = list_util.complement(current_slots_unsorted(clss ), [None])
    output = [ ] 
    for st_name in st_list : 
        st_slots = used_slots_of(st_name)
        if len( list_util.intersect( class_slots, st_slots )) == 0 : 
            output.append(st_name)
    return output

def classes_at_slot(st_list, sl):
    my_parts = [ name_slot_part(c, sl) for c in st_list] 
    output = list_util.union( [class_name_of(c) for c 
                in  list_util.complement(my_parts, ['']) ] )
    return output

def count_classes(st_list) : 
    nums = [len(classes_at_slot(st_list, c)) for c in slots]
    print("  ".rjust(2), korean_ljust(" ",20), ":", end=" " )
    print(' '.join( [str( sl ).ljust(3) for sl in slots ]))
    print("  ".rjust(2), korean_ljust(" ",20), ":", end=" " )
    print(' '.join( [str( num ).ljust(3) for num in nums ]))

def are_slots_complete(subj):
    my_slots = list_util.union( 
        [ classpart_slot[cp] for cp in parts_of(subj) ] )
    if None not in my_slots : 
        return True
    else : 
        return False

def apply_slot_dict(slot_dict, subj) : 
    my_parts = list_util.intersect( parts_of(subj), slot_dict.keys())
    for cp in my_parts : 
        assign_part_slot(cp, slot_dict[cp])



def classrooms_in_tabbing( st_name , cl_slots = [ ] ) : 
    output = [ ] 
    st_slots = used_slots_of(st_name)
    for sl in slots : 
        if sl in st_slots : 
            pt = name_slot_part(st_name, sl)
            if subject_name_of(pt) in subjects : 
                if classrooms[pt] is not None : 
                    output.append( sem.classroom_shorthand[ classrooms[pt]]) 
                else : 
                    output.append( "None/") 
            else : 
                output.append( "----/") 
        else : 
            output.append( "----/") 
    return ''.join(output)

def show_classes_of_names(t_list, classrooms=False): 
    print(korean_ljust(' ', 16), ":", end=" " )
    if classrooms : 
        print('  '.join(slots))
        str_fn = classrooms_in_tabbing
    else : 
        print('  '.join(slots))
        str_fn = (lambda x: subjects_in_tabbing(x,c_num=True))
    for t_name in t_list : 
        print(korean_ljust(t_name, 16), ":", end=" " )
        print(str_fn(t_name))



def assign_classroom_to_classes(room, the_classes):
    for clss in the_classes : 
        assign_classroom(clss, room)

def subject_classes_of(t_name) : 
    output = [c for c in current_classes( t_name ) 
        if subject_name_of(c) in subjects]
    output.sort()
    return output

def assign_classroom_to_name(room, t_name):
    the_classes = [c for c in current_classes( t_name ) 
        if subject_name_of(c) in subjects]
    for clss in the_classes : 
        assign_classroom(clss, room)

def expand_to_full_classname(short_c_name): 
    c_num = re.sub( r'\D+', '',  short_c_name)
    short_subj = re.sub( r'_{0,1}\d+', '',  short_c_name)
    subj = list_util.reversed_dict(sem.shorthands)[short_subj][0]
    return subj + "_" + c_num

def short_classname(c_name) : 
    output =  sem.shorthands[ subject_name_of(c_name)] + \
                    class_number_of(c_name)
    return output


def export_tables_of_rooms_by_department() : 
    dn_dict = list_util.reversed_dict(sem.name_department_dict)
    depts = dn_dict.keys()
    depts.sort()
    for dt in depts : 
        t_names = dn_dict[dt]
        #filename = ("%s%s과_강의실시간표.xlsx" % (sem.semester_prefix, dt))
        filename = ("{}{}과_강의실시간표.xlsx".format(sem.semester_prefix, dt))
        c_list = list_util.flatten([subject_classes_of(c) for c in t_names])
        if len(c_list)==0: 
            continue
        r_list = list_util.union( list_util.flatten( 
                    [classrooms_of_class(c) for c in c_list]))
        r_list.sort()
        export_weekly_tables_of_rooms( filename, room_list = r_list )


def export_class_sizes_xlsx(filename):
    dn_dict = list_util.reversed_dict(sem.name_department_dict)
    depts = dn_dict.keys()
    depts.sort(key = sem.ordered_departments.index)
    output = [ ]
    for dt in depts :
        t_names = dn_dict[dt]
        c_list = list_util.flatten([subject_classes_of(c) for c in t_names])
        subjs = list_util.union([subject_name_of(c) for c in c_list])
        subjs.sort()
        for subj in subjs :
            size_list = [str(size_of(c)) for c in classes_of(subj)]
            output.append( [dt, subj] + size_list )
    xlsx_util.write_tuples_into_xlsx(filename, output)


def count_free_names(slot, name_list = students) : 
    return len( [ c for c in name_list if slot not in used_slots_of(c) ] )
def count_free_slots(st_name, slot_list = slots) : 
    return len(list_util.complement(slots, used_slots_of(st_name)))



def teachers_of(clss) : 
    return filter( lambda x : not is_assistant_name(x), t_members_of(clss) )


def export_name_class_teachers_xlsx(filename): 
    output = [ ]
    for st_name in students : 
        recs = [st_name] + [ para_classname(c) + '\n' + 
            ' '.join(map(convert_name_for_xlsx, teachers_of(c))) 
            for c in sorted(current_classes(st_name))]
        output.append(recs)
    xlsx_util.write_tuples_into_xlsx(filename, output)



def export_weekly_table_of_each_student(freshmen = False): 
    if freshmen is False : 
        st_list = [c for c in students if not c.startswith("1") ]
        for st_name in st_list : 
            phone_num = sem.phone_dict[st_name].replace("-", "")
            export_weekly_tables_of_all_classes( 
                "수업시간표_" + st_name + "_" + phone_num + ".xlsx", 
                t_name = True, st_list = [st_name]) 
    else : 
        st_list = [c for c in students if c.startswith("1") ]
        for ban in range(1,9): 
            prefix = "1" + str(ban)
            ban_st_list = [st for st in st_list if st.startswith(prefix)]
            export_weekly_tables_of_all_classes( 
                "수업시간표-1-" + str(ban) + ".xlsx", 
                t_name = True, st_list = ban_st_list ) 

def divide_and_put(subj):
    st_list = members(subj)
    random.shuffle(st_list)
    c_list = classes_of(subj)
    n = len(c_list)
    for i, st_name in enumerate(st_list) : 
        put_name_into_class(st_name, c_list[i%n])

def related_class_pairs():
    adv_subjs = list_util.complement(subjects, subjects_of(name(1101)), 
                    subjects_of(name(1801))) 
    #print ' '.join(adv_subjs)
    related_subj_pairs = [tuple(sorted(c)) 
        for c in list_util.distinct_pairs(adv_subjs) 
                if [] != list_util.intersect(members(c[0]), members(c[1]))]
    c_list = list_util.flatten([classes_of(subj) for subj in adv_subjs])
    output = [tuple(sorted(c)) for c in list_util.distinct_pairs(c_list) 
       if tuple(sorted([subject_name_of(c[0]), subject_name_of(c[1])])) 
            in related_subj_pairs]
    return output

def build_related_class_names():
    adv_subjs = list_util.complement(subjects, subjects_of(name(1101)), 
                    subjects_of(name(1801))) 
    c_list = list_util.flatten([classes_of(subj) for subj in adv_subjs])
    output = { }
    for clss in c_list : 
        output[clss] = st_members_of(clss)
    return output

def build_common_member_counts(class_members):
    pairs = [ tuple(sorted( [c, d] )) for c in class_members.keys() 
        for d in class_members.keys() 
            if subject_name_of(c)!=subject_name_of(d) ]
    output = { }
    for p in pairs : 
        output[p] = len(list_util.intersect(class_members[p[0]], 
                        class_members[p[1]]))
    return output
        

    


def build_class_common_members( class_pairs ) : 
    output = { }
    i = 0
    for cl_one, cl_two in class_pairs : 
        i += 1
        st_list = list_util.intersect(members(cl_one), members(cl_two))
        output[ (cl_one, cl_two) ] = st_list
        if i%50 == 0 : 
            print(i, cl_one, cl_two, len(st_list), ' '.join(st_list))
    return output

def name_class_inters(triples, name, clss):
    output = [c[2] for c in triples if c[0] == name and c[1]==clss]
    return output

        
def most_intersecting_class(common_member_counts, clss, c_list):
    candi_classes = list_util.complement(classes_of(subject_name_of(clss)), 
                        [clss])
    max_n = 0
    output = None
    for cl in candi_classes : 
        new_count = len( [ c for c in c_list 
            if  common_member_counts[tuple(sorted([cl, c]))]>0 ] )
        if new_count >= max_n : 
            max_n = new_count 
            output = cl
    return output
    
def put_name_and_update_common(st_name, clss, class_members, 
        common_member_counts, verbose=False):
    subj = subject_name_of(clss)
    old_class = name_subject_class(st_name, subj)
    c_list = list_util.complement(current_classes(st_name), [old_class])
    put_name_into_class(st_name, clss)

    class_members[old_class] = members(old_class)
    class_members[clss] = members(clss)

    my_pairs = [ tuple( sorted([old_class, c])) for c in c_list ] + [ 
                    tuple( sorted([clss, c])) for c in c_list ]
    my_pairs.sort(key = lambda x: (old_class in x, 
        len(list_util.intersect(members(x[0]), members(x[1]))) ) )

    for cl_one, cl_two in my_pairs:
        common_member_counts[ (cl_one, cl_two) ] = len( list_util.intersect( 
            class_members[cl_one], class_members[cl_two]))
        if verbose:
            print(sem.shorthands[
                subject_name_of(cl_one)]+class_number_of(cl_one), end=" ")
            print(sem.shorthands[
                subject_name_of(cl_two)]+class_number_of(cl_two), end=" ")
            print( common_member_counts[ (cl_one, cl_two) ])

def intersection_weight(common_member_counts, class_one, class_two):
    return common_member_counts[ tuple( sorted( [class_one, class_two]))]
    
def evaluate_moving(common_member_counts, st_name, curr_class, new_class):
    if curr_class == new_class : 
        return 0
    other_classes = current_classes(st_name)
    other_classes.remove(curr_class)
    count_ones = len( [c for c in other_classes 
        if 1==intersection_weight(common_member_counts, curr_class, c)])
    count_twos = len( [c for c in other_classes 
        if 2==intersection_weight(common_member_counts, curr_class, c)])
    count_loss_zeros = len( [c for c in other_classes 
        if 0==intersection_weight(common_member_counts, new_class, c)])
    count_loss_ones = len( [c for c in other_classes 
        if 1==intersection_weight(common_member_counts, new_class, c)])
    return (count_loss_zeros + count_loss_ones - count_ones - count_twos)


def update_upon_moving(class_members, clpair_num_dict, 
        st_name, from_clss, to_clss):
    for pair in clpair_num_dict : 
        if from_clss in pair : 
            other_clss = list_util.complement(list(pair),[from_clss])[0]
            if st_name in class_members[other_clss]:
                clpair_num_dict[pair] -= 1
        if to_clss in pair : 
            other_clss = list_util.complement(list(pair),[to_clss])[0]
            if st_name in class_members[other_clss]:
                clpair_num_dict[pair] += 1


def evaluate_cycling(class_members, common_member_counts, name_class_pairs):
    c_list = [c[1] for c in name_class_pairs]
    if len( list_util.union( [ subject_name_of(clss) for clss in c_list]))>1:
        #print "subject name mismatches %s" % ' '.join( list_util.union( [ subject_name_of(clss) for clss in c_list]))
        print("subject name mismatches {}" % ' '.join(
            list_util.union( [ subject_name_of(clss) for clss in c_list])))
        return None
    if len(list_util.union(c_list)) != len(c_list):
        #print "not simple cycle %s" % ' '.join(c_list)
        print("not simple cycle {}" % ' '.join(c_list))
        return None
    subj = subject_name_of(c_list[0])
    cycled_pairs = name_class_pairs + [ name_class_pairs[0] ]
    val = 0
    n = len(name_class_pairs)
    involving_classes = list_util.union(list_util.flatten( 
        [current_classes(st_name) for st_name, clss in name_class_pairs]))
    subj_classes = [clss for clss in involving_classes if belongs_to(clss,subj)]
    outer_classes = list_util.complement(involving_classes, subj_classes)
    ccm_keys = [ tuple( sorted( [subj_clss, outer_clss] )) 
        for subj_clss in subj_classes 
        for outer_clss in outer_classes ]
    #for c in ccm_keys : 
        #print ' '.join(c), len(class_common_members[c])
    pair_num_dict = { }
    for pair in ccm_keys : 
        pair_num_dict[pair] = common_member_counts[pair]
    inter_counts = pair_num_dict.values()
    count_list = [ inter_counts.count(i) for i in range(10) ]  
    #print count_list
    new_pn_dict = dict(pair_num_dict)
        
    for i, pair in enumerate(cycled_pairs):
        if i >= n : 
            break
        st_name = pair[0]
        curr_class = pair[1]
        new_class = cycled_pairs[i+1][1]
        #print st_name, curr_class, new_class
        update_upon_moving(class_members, new_pn_dict, 
                st_name, curr_class, new_class)
    new_inter_counts = new_pn_dict.values()
    new_count_list = [ new_inter_counts.count(i) for i in range(10) ]  
    #print new_count_list
    if new_count_list > count_list : 
        return -1 
    else : 
        return 1


def get_name_class_pairs(subj, name_list):
    output = [ ]
    for st_name in name_list: 
        clss = name_subject_class(st_name, subj)
        if clss is None: 
            return None
        output.append( (st_name, clss) )
    return output

def recommend_best_transfer(common_member_counts, class_one, class_two):
    if subject_name_of(class_one) != subject_name_of(class_two):
        return None
    curr_min = 0
    output = None
    for st_name in st_members_of(class_one):
        new_val = evaluate_moving(common_member_counts, st_name, 
                class_one, class_two)
        if new_val < curr_min : 
            curr_min = new_val
            output = st_name
    #print "Found a %d transfer.." %  curr_min
    print("Found a {} transfer..".format(curr_min))
    return output

def choose_first_negative_triple(class_common_members, 
        class_one, class_two, class_three):
    return None

        
def exchange_best_transfers(class_members, common_member_counts, 
        class_one, class_two): 
    st_one = recommend_best_transfer(common_member_counts, class_one, class_two)
    if st_one is None : 
        return None
    st_two = recommend_best_transfer(common_member_counts, class_two, class_one)
    if st_two is None : 
        return None
    put_name_and_update_common(st_one, class_two, class_members, 
        common_member_counts)
    put_name_and_update_common(st_two, class_one, class_members, 
        common_member_counts)

    print(subject_name_of(class_two), end=" ")
    for num in map(size_of, classes_of(subject_name_of(class_two))):
        #print "<%d>" % num,
        print("<{}>".format(num), end=' ')
    print

    weight_ones =  [c for c in common_member_counts
        if common_member_counts[c] == 1 ]
    #print "There are %d many weight one intersections" % len(weight_ones)
    print("There are {} many weight one intersections".format(len(weight_ones)))
    weight_zeros =  [c for c in common_member_counts
        if common_member_counts[c] == 0 ]
    #print "There are %d many weight zero intersections" % len(weight_zeros)
    print("There are {} many weight zero intersections".format(
        len(weight_zeros)))
    return len(weight_zeros)

def repeat_exchange(class_common_members, class_one, class_two):
    zero_count = 0 
    for i in range(30):
        new_val = exchange_best_transfers(class_common_members, 
            class_one, class_two)
        if new_val is None : 
            return None
        elif new_val <= zero_count : 
            print("No progress")
            return None
        else : 
            zero_count = new_val
        

def common_members(class_members, class_pair):
    return list_util.intersect( class_members[class_pair[0]], 
        class_members[class_pair[1]] )
    
def work_on_weight_one(class_members,common_member_counts, nth=0, clss = None):
    weight_ones =  [c for c in common_member_counts
        if common_member_counts[c] == 1 ]
    weight_zeros =  [c for c in common_member_counts
        if common_member_counts[c] == 0 ]
    #print "There are %d many weight one intersections" % len(weight_ones)
    #print "There are %d many weight zero intersections" % len(weight_zeros)
    print("There are {} many weight one intersections".format(len(weight_ones)))
    print("There are {} many weight zero intersections".format(
        len(weight_zeros)))
    triples =  list_util.flatten( [[ ( ' '.join(common_members(class_members, 
            c ) ), c[0], c[1]) , 
        (' '.join(common_members(class_members, c ) ), c[1], c[0] )] 
            for c in weight_ones] )
    if clss is not None:
        triples = [c for c in triples if clss in c]
    st_cl = list_util.mode( [ (c[0], c[1]) for c in triples 
        if len(classes_of(subject_name_of(c[1]))) > 1 
            and (not c[0].endswith("t"))  ], nth=nth ) 
    print(' '.join(st_cl))
    c_list = name_class_inters(triples, * st_cl )
    new_class = most_intersecting_class(common_member_counts, st_cl[1], 
                    c_list )
    put_name_and_update_common(st_cl[0], new_class, 
        class_members, common_member_counts)
    print(subject_name_of(new_class), end=" ")
    for num in map(size_of, classes_of(subject_name_of(new_class))):
        #print "<%d>" % num,
        print("<{}>".format(num), end=' ')
    print
    return new_class

def choose_random_name_classes(subj, n):
    c_list = classes_of(subj)
    if n <= 0 : 
        return [ ] 
    elif n > len(c_list): 
        #print "%d > number of classes, doing nothing.."
        print("{} > number of classes, doing nothing..")
        return None
    distinct_classes = [ ]
    for i in range(n):
        new_clss = random.choice( 
            list_util.complement(c_list, distinct_classes) ) 
        distinct_classes.append(new_clss)
    output = [ ] 
    for clss in distinct_classes : 
        output.append( (random.choice(st_members_of(clss)), clss) )
    return output
    
def show_slot_counts():
    #show_slots = slots[:18] + "수5 수6 수7".split() + slots[18:]+"금7".split()
    show_slots = slots[:18] + "수5 수6 수7".split() + slots[18:].split()
    output = [sl + " " + str( 
        list_util.count(classparts, lambda x:classpart_slot[x]==sl))
                for sl in show_slots]
    gen_util.print_table( list_util.transposed(list_util.partition(output, 7)))
        
        
def put_subject_number_index(subj, cnum, st_ind):
    st_list = remaining_names(subj)
    clss = subj + "_" + str(cnum)
    put_name_into_class( st_list[st_ind], clss )


def show_intersection_stats(common_member_counts):
    vals = common_member_counts.values()
    for i in range(10):
        #print "%d :" % i, vals.count(i)
        print("{} :".format(i, vals.count(i)))
    print(">=10 :", list_util.count(vals, lambda x: x>=10))

def show_small_intersections(common_member_counts, subj, n=1):
    c_list = classes_of(subj)
    pairs = [c for c in common_member_counts 
                if common_member_counts[c]<=n and common_member_counts[c]>0 
                and subj in map(subject_name_of,c)]
    #pairs.sort( key = lambda x: (common_member_counts[x], x[0], x[1])  )
    for i in range(1,n+1):
        for clss in c_list: 
            print(i, clss, " ", end=" " )
            for p in pairs: 
                if common_member_counts[p] == i and clss in p : 
                    print(list_util.complement(p, [clss])[0], end=" ")
            print(" ")
    
def choose_random_cycle(class_members, common_member_counts, subj, 
        n=1, length=2 ):
    c_list = classes_of(subj)
    pairs = [c for c in common_member_counts if common_member_counts[c]==n 
                and subj in map(subject_name_of,c)]
    #pairs.sort( key = lambda x: (common_member_counts[x], x[0], x[1])  )
    subj_pairs = [ ] 
    c_collection = [ ] 
    for i in range(20):
        p = random.choice(pairs)
        s_cl = filter(lambda x: belongs_to(x, subj), p)[0]
        if s_cl not in c_collection : 
            subj_pairs.append(p)
            c_collection.append(s_cl)
            if len(subj_pairs)>=length : 
                break 
    output = [ ] 
    for i, clss in enumerate(c_collection):
        output.append( (list_util.intersect(class_members[subj_pairs[i][0]], 
            class_members[subj_pairs[i][1]])[0], clss)  )
    return output
     
def choose_best_target(class_members, common_member_counts, subj):
    c_list = classes_of(subj)
    pairs = [c for c in common_member_counts if common_member_counts[c]==1]
    st_list = [ list_util.intersect(* 
        map(gen_util.dict_to_fn(class_members), p))[0] for p in pairs]
    subj_members = members(subj)
    restricted = [st_name for st_name in st_list  if st_name in subj_members]
    candi_st = list_util.mode(restricted)
                
    #pairs.sort( key = lambda x: (common_member_counts[x], x[0], x[1])  )
    subj_pairs = [ ] 
    c_collection = [ ] 
    for i in range(20):
        p = random.choice(pairs)
        s_cl = filter(lambda x: belongs_to(x, subj), p)[0]
        if s_cl not in c_collection : 
            subj_pairs.append(p)
            c_collection.append(s_cl)
            if len(subj_pairs)>=length : 
                break 
    output = [ ] 
    for i, clss in enumerate(c_collection):
        output.append( (list_util.intersect(class_members[subj_pairs[i][0]], 
            class_members[subj_pairs[i][1]])[0], clss)  )
    return output
               


    
def exchange_on_weight_one(class_members,common_member_counts, 
        nth=0, clss = None):
    weight_ones =  [c for c in common_member_counts
        if common_member_counts[c] == 1 ]
    weight_zeros =  [c for c in common_member_counts
        if common_member_counts[c] == 0 ]
    #print "There are %d many weight one intersections" % len(weight_ones)
    #print "There are %d many weight zero intersections" % len(weight_zeros)
    print("There are {} many weight one intersections".format(len(weight_ones)))
    print("There are {} many weight zero intersections".format(
        len(weight_zeros)))
    triples =  list_util.flatten( [[ ( ' '.join(common_members(class_members, 
            c ) ), c[0], c[1]) , 
        (' '.join(common_members(class_members, c ) ), c[1], c[0] )] 
            for c in weight_ones] )
    if clss is not None:
        triples = [c for c in triples if clss in c[1:]]
    st_cl = list_util.mode( [ (c[0], c[1]) for c in triples 
        if len(classes_of(subject_name_of(c[1]))) > 1 
            and (not c[0].endswith("t"))  ], nth=nth ) 
    print(' '.join(st_cl))
    c_list = name_class_inters(triples, * st_cl )

    old_class = st_cl[1]
    moving_st = st_cl[0]
    new_class = most_intersecting_class(common_member_counts, old_class, 
                    c_list )
    for new_name in st_members_of(new_class):
        if 0 > evaluate_cycling(class_members, common_member_counts, 
                [ (moving_st, old_class), (new_name, new_class) ] ):
            coming_st = new_name 
            break 
    else : 
        print("no coming_st found, quitting..")
        return False

    put_name_and_update_common(moving_st, new_class, 
        class_members, common_member_counts)
    put_name_and_update_common(coming_st, old_class, 
        class_members, common_member_counts)
    print(subject_name_of(new_class), end=" ")
    for num in map(size_of, classes_of(subject_name_of(new_class))):
        #print "<%d>" % num,
        print("<{}>".format(num), end=' ')
    print
    weight_ones =  [c for c in common_member_counts
        if common_member_counts[c] == 1 ]
    weight_zeros =  [c for c in common_member_counts
        if common_member_counts[c] == 0 ]
    #print "There are %d many weight one intersections" % len(weight_ones)
    #print "There are %d many weight zero intersections" % len(weight_zeros)
    print("There are {} many weight one intersections".format(
        len(weight_ones)))
    print("There are {} many weight zero intersections".format(
        len(weight_zeros)))
    return True

def work_on_subject(class_members,common_member_counts, subj, nth=0):
    c_list = classes_of(subj)
    for clss in c_list : 
        for i in range(20):
            result = exchange_on_weight_one(class_members,common_member_counts, 
                        nth=nth, clss = clss)
            if result is False : 
                break 
    
def work_on_subjects(class_members,common_member_counts, subj_list, nth=0):
    for subj in subj_list : 
        try: 
            work_on_subject(class_members,common_member_counts, subj,nth=nth)
        except IndexError : 
            continue
    show_intersection_stats(common_member_counts)



def clear_by_name_slot(st_name, in_slot):
    pt = name_slot_part(st_name, in_slot)
    if len(pt) == 0 : 
        #print "Wrong input %s %s" % (st_name, in_slot), 
        print("Wrong input {} {}".format(st_name, in_slot), end=' ')
        print("Doing nothing..")
        return False
    else : 
        clear_slot(pt)

def transferable_classes(subj, st_name):
    c_list = classes_of(subj)
    output = [ ] 
    for clss in c_list : 
        if len( filter_by_class_slots(clss, [st_name] )) > 0 : 
            output.append(clss)
    print(list_util.jpr(output))
    return output

def clear_class_from_slots(clss, slot_list):
    if type(slot_list) is str : 
        loc_slot_list = slot_list.split()
    else : 
        loc_slot_list = slot_list
    my_slots = list_util.intersect(slots, 
        map(gen_util.dict_to_fn(slot_dict),loc_slot_list))
    
    st_list = [st_name for st_name in st_members_of(clss) 
              if [] != list_util.intersect(my_slots, used_slots_of(st_name)) ] 
    remove_names_from_class(st_list, clss)

def reversed_shorthands(short_clss):
    r_dict = list_util.reversed_dict(sem.shorthands)
    sr = re.search(r'^(\D+)(\d+)$', short_clss)
    if sr : 
        s_subj = sr.group(1)
        num = sr.group(2)
        if s_subj in r_dict : 
            #print s_subj, num
            return (r_dict[ s_subj ][0] + "_" + num)
    else : 
        s_subj = short_clss
        if s_subj in r_dict : 
            return r_dict[ s_subj ][0]
        else : 
            return None
    
def subjects_of_names(st_list, ordered=True, shorthands=False ):
    output = list_util.union(list_util.flatten([subjects_of(st_name) 
                for st_name in st_list]))
    if ordered : 
        output.sort(key= sem.ordered_subjects.index)
    else : 
        output.sort()
    if shorthands : 
        output = map(gen_util.dict_to_fn(sem.shorthands), output)
    return output

def filter_by_slots(clss, st_list): 
    if is_classname(clss):
        in_slots = list_util.complement(current_slots_unsorted(clss), [None])
    elif type(clss) is str:
        in_slots = clss.split()
    else:
        in_slots = clss
    my_slots = map(gen_util.dict_to_fn(slot_dict), in_slots)
    output = filter( lambda x: []== list_util.intersect(my_slots, 
                        used_slots_of(x))  ,st_list)
    return output


def teacher_size_class_hour_list(t_name):
    c_list = [ c for c in current_classes(t_name) 
        if subject_name_of(c) in subjects]
    c_list.sort()
    #output = [ "%02d명#%s(%.2f)" % (size_of(c), c,  float(len(parts_of(c))) / len(teachers_of(c)) ) for c in c_list ]
    output = [ "{:02d}명#{}({:.2f})".format(size_of(c), c,
            float(len(parts_of(c))) / len(teachers_of(c)) ) for c in c_list ]
    return [ convert_name_for_xlsx(t_name) ] + output

def has_two_days(clss):
    my_slots = slots_of(clss)
    if None in my_slots : 
        print(' '.join(map(str,my_slots)))
        return False
    if len(my_slots) < 3:
        return True
    if len(list_util.union( [c[:-1] for c in my_slots])) < 2 :
        print(' '.join(map(str,my_slots)))
        return False
    else : 
        return True



def copy_from_and_put(source_name, target_name, partial=False  ):
    subjs = subjects_of(target_name)
    target_classes = [clss for clss in current_classes(source_name) 
                        if subject_name_of(clss) in subjs ] 
    remains = list_util.complement(subjs, map(subject_name_of, target_classes))
    if partial is False and len(remains)>0 : 
        #print "Subjects %s of %s missing.." % (target_name, ','.join(remains))
        print("Subjects {} of {} missing..".format(
            target_name, ','.join(remains)))
        return False
    put_name_into_classes(target_name, target_classes)




def put_random_one_into_class( clss, candidates = None, n=19, exp_n=16, 
        count=1, randomize=False, lower_bound = 13, fix_names = []) : 
    subj = subject_name_of(clss)
    if candidates is None : 
        candis = members(subj)
    else : 
        candis = list_util.intersect(candidates, members(subj))
    st_name = random.choice( list_util.complement(candis, 
                                st_members_of(clss), fix_names))
    subjs = [s for s in subjects_of(st_name) if len(classes_of(s)) > 2 ]
    small_cs = list_util.union(* [ [c for c in classes_of(subj) 
         if expected_class_size(st_name, c) < lower_bound] for subj in subjs 
            if has_class_with_assigned_slots(subj) ])
    big_cs = list_util.union(* [ [c for c in classes_of(subj) 
         if overfulled_with(c, st_name, n, exp_n) ] 
            for subj in subjects_of(st_name) ])

def is_subject_subset( sub_name, super_name):
    return list_util.is_subset( subjects_of(sub_name), subjects_of(super_name))






def measure_badness(st_name, in_class_list, verbose=False, check_full=True, 
        min_max_info = None): 
    subjs = subjects_of(st_name)
    class_list = [clss for clss in in_class_list 
                    if subject_name_of(clss) in subjs ] 
    badness = 0 

    if len(class_list) < len(subjs) : 
        badness += 1000

    if min_max_info:
        min_classes, max_classes = min_max_info

    for clss in class_list : 
        subj = subject_name_of(clss)
        curr_clss = name_subject_class(st_name,subj)
        if min_max_info : 
            curr_min = size_of(list_util.intersect(
                        classes_of(subj), min_classes)[0])
            curr_max = size_of(list_util.intersect(
                        classes_of(subj), max_classes)[0])
        else : 
            curr_class_sizes =  map(size_of, classes_of(subj))
            curr_max = max(curr_class_sizes)
            curr_min = min(curr_class_sizes)
        
        if size_of(curr_clss) == curr_max  and \
                expected_class_size(st_name, clss) < curr_max : 
            badness -= 1
        if curr_clss != clss and size_of(clss) == curr_min : 
            badness -= 1
        if expected_class_size(st_name, clss) > curr_max : 
            badness += 1
        if curr_clss != clss and size_of(curr_clss) == curr_min :
            badness += 3
    if verbose: 
        print(st_name, badness, end=" " )
        print(''.join( [class_shorthands(cl) for cl in class_list]))
    return badness


def show_best_copy_transfer(super_list, sub_list, min_max_info=None):
    uniq_sub_list = [ ]
    temp_clists = [ ] 
    for st_name in sub_list : 
        c_list = sorted(current_classes(st_name))
        if c_list not in temp_clists : 
            temp_clists.append(c_list)
            uniq_sub_list.append(st_name)

    uniq_super_list = [ ]
    temp_clists = [ ] 
    for st_name in super_list : 
        c_list = sorted(current_classes(st_name))
        if c_list not in temp_clists : 
            temp_clists.append(c_list)
            uniq_super_list.append(st_name)


    pairs = [ (a,b) for a in uniq_super_list for b in uniq_sub_list
                    if is_subject_subset(b,a)]
    name_clists = [ ] 
    for super_one, sub_one in pairs : 
        subj_list = subjects_of(sub_one)
        c_list = sorted([c for c in current_classes(super_one) 
                    if subject_name_of(c) in subj_list])
        if [ sub_one, c_list] not in name_clists:
            name_clists.append( [sub_one, c_list])
    with_badness = [ [ measure_badness(st_name, c_list, 
            min_max_info=min_max_info), st_name, 
        sorted(c_list, key=lambda x:expected_class_size(st_name,x)) ] 
                        for st_name, c_list in name_clists ]
    with_badness.sort()
    for i, item in enumerate(with_badness):
        st_name = item[1]
        c_list = current_classes(st_name)

        temp_slist = map(subject_name_of,item[2])
        c_list.sort(key=lambda x:temp_slist.index(subject_name_of(x)))

        print(i, item[0], st_name, ','.join([ class_shorthands(c) + " " + 
         str(expected_class_size(st_name, c,removing=True)) for c in c_list]))
        print(i, item[0], item[1], ','.join([ class_shorthands(c) + " " + 
            str(expected_class_size(item[1], c)) for c in item[2]] ))
        print()
        if i >= 10 : 
            break
    return [ [item[1], item[2]] for item in with_badness ][:5]

        
    


def subset_trans(subj, i, j, randomize=False, verbose=True, only_full=False, 
        missing_sub=False ):
    from_class = subj + "_" + str(i)
    to_class = subj + "_" + str(j)
    pairs = [ (a,b) for a in st_members_of(to_class) 
                    for b in st_members_of(from_class) 
                    if is_subject_subset(b,a)]

    if only_full : 
        full_pairs = [ c for c in pairs 
            if list_util.is_subset(subjects_of(c[1]), 
                map(subject_name_of,current_classes(c[0])))]
        pairs = full_pairs
    if missing_sub : 
        missing_pairs = [ c for c in pairs if remaining_subjects(c[1])!=[] ]
        pairs = missing_pairs

    uniq_pairs = [ ]
    c_lists = [ ] 
    for super_one, sub_one in pairs : 
        new_list = sorted(current_classes(super_one))
        if new_list not in c_lists : 
            c_lists.append(new_list)
            uniq_pairs.append( (super_one, sub_one) )

    if len(pairs)>0:
        if verbose : 
            #print "Found %d many pairs (%s, %s), ..." % (len(uniq_pairs), uniq_pairs[0][0], uniq_pairs[0][1])
            print("Found {} many pairs ({}, {}), ...".format(len(uniq_pairs),
                    uniq_pairs[0][0], uniq_pairs[0][1]))
        if randomize : 
            random.seed()
            my_pair = random.choice(uniq_pairs)
        else : 
            my_pair = min( uniq_pairs, key = lambda x: measure_badness(x[1], 
                current_classes(x[0]), verbose=verbose))
        copy_from_and_put( * my_pair, partial=missing_sub )
        return True
    else : 
        print("No pair found")
        return False




def subset_auto_trans(subj, big_size = None, small_size = None, 
        missing_sub = False):
    if big_size is None : 
        big_size = max( [size_of(clss) for clss in classes_of(subj)])
    if small_size is None : 
        small_size = min( [size_of(clss) for clss in classes_of(subj)])
    big_class = random.choice( [clss for clss in classes_of(subj) 
                    if size_of(clss) == big_size])
    small_class = random.choice( [clss for clss in classes_of(subj) 
                    if size_of(clss) == small_size])
    subset_trans(subj, class_number_of(big_class),class_number_of(small_class),
            missing_sub = missing_sub)


def put_name_into_single_classes(st_name):
    class_list = [ classes_of(subj)[0] for subj in subjects_of(st_name) 
            if len(classes_of(subj))==1 ]
    put_name_into_classes(st_name, class_list)


def name_subject_available_classes(st_name, subj):
    st_slots = used_slots_of(st_name)
    output = [ ] 
    for clss in classes_of(subj):
        c_slots = slots_of(clss)
        if len(c_slots)<hours(subj) : 
            #print "%s has None slots skipping it" % clss 
            print("{} has None slots skipping it".format(clss))
            continue
        collides = list_util.intersect(c_slots, st_slots)
        if len(collides) == 0:
            output.append(clss)
            print(clss, ' '.join(c_slots))
        else :
            print(" "*5, end=" "  )
            print(clss, ' '.join( list_util.complement(c_slots,collides) + 
                    [" "*5] + collides   ))
    return output



def put_name_into_inevitables(st_name, go_forward=False, skip_subjects=[]  ):
    #list_util.jpr(skip_subjects)
    remains = list_util.complement(remaining_subjects(st_name), skip_subjects)
    curr_remains = remains[:]
    #list_util.jpr(remains)
    while len(curr_remains) > 0 : 
        for subj in remains : 
            if subj not in curr_remains : 
                continue
            c_list = name_subject_available_classes(st_name, subj) 
            if len(c_list) == 0 : 
                #print "No available classes in %s" % subj
                print("No available classes in {}".format(subj))
                return False
            if len(c_list) == 1 : 
                #print "Found inevitable %s" % c_list[0]
                print("Found inevitable {}".format(c_list[0]))
                put_name_into_class(st_name, c_list[0])
                #remains = remaining_subjects(st_name)
                curr_remains =list_util.intersect(remains, 
                                remaining_subjects(st_name))
                break
            if go_forward : 
                print("Found more than one " + ' '.join( c_list ))
                print("Proceeding with a random")
                put_name_into_class(st_name, random.choice(c_list))
                #remains = remaining_subjects(st_name)
                curr_remains =list_util.intersect(remains, 
                                remaining_subjects(st_name))
                break
        else :
            print("Found no inevitables")
            return False
    print("All were inevitables" )
    return True
    
        
def part_size_slots_str(in_part):
    head = korean_ljust(in_part + " " + str(len(st_members_of(in_part))), 20)
    if classpart_slot[in_part] is not None : 
        middle = ":"+ classpart_slot[in_part] + ":"
    else :
        middle = ":"+ "   " + ":"
    tail = slots_in_tabbing( available_slots_of(in_part), 
            [classpart_slot[in_part]] )
    #print head, middle, tail
    return ' '.join( [ head, middle, tail] )


def show_classes_of_name_subjects(st_name, subj_list):
    st_list = [ ] 
    for subj in subj_list : 
        clss_str = name_subject_class(st_name, subj)
        if len(clss_str) == 0:
            st_list.extend( t_members_of(subj) )
            st_list.extend( remaining_names(subj) )
        else : 
            st_list.extend( members(clss_str) )
    show_classes_of_teachers(st_list)
        
def put_name_by_subject_slot(st_name, subj, in_slot, index=0) :
    if in_slot in slots : 
        slot = in_slot
    else : 
        slot = slot_dict[in_slot]
    c_list = [clss for clss in classes_of(subj) if slot in slots_of(clss)]
    if len(c_list) > 0 : 
        put_name_into_class(st_name, c_list[index])
        return True
    else : 
        #print "No class of %s at %s" % (subj, slot)
        print("No class of {} at {}".format(subj, slot))
        return False

def show_classes_by_subject_slots( in_subject_slot_list ):
    st_list = [ ] 
    if type(in_subject_slot_list) is str : 
        pairs = list_util.partition( in_subject_slot_list.split(), 2)
    else : 
        pairs = in_subject_slot_list
    subject_slot_list = [ ] 
    for subj, slot in pairs : 
        if subj in subjects : 
            new_subj = subj
        else : 
            new_subj = reversed_shorthands(subj)
        if slot in slots : 
            new_slot = slot
        else : 
            new_slot = slot_dict[slot]
        subject_slot_list.append( (new_subj, new_slot) )

    for subj, slot in subject_slot_list:
        c_list = [clss for clss in classes_of(subj) if slot in slots_of(clss)]
        if len(c_list)==0:
            #print "No classes of %s at %s Skipping" % (subj, slot)
            print("No classes of {} at {} Skipping".format(subj, slot))
            continue
        for clss in c_list : 
            st_list.extend( teachers_of(clss))
            st_list.extend( st_members_of(clss))
    show_classes_of_teachers(st_list)



def copy_and_put_one(super_list, sub_list , verbose=True, partial=True, 
        lower_bound_fixes = [ ] ):
    smallest_classes = [ ] 
    for subj in lower_bound_fixes: 
        min_size = min(map(size_of, classes_of(subj)))
        c_list = [clss for clss in classes_of(subj) if size_of(clss)==min_size]
        smallest_classes.extend(c_list)
    sc_members = list_util.union(list_util.flatten(
                        map(st_members_of,smallest_classes)))

    compl_sub_list = list_util.complement(sub_list, sc_members)
    
    pairs = [ (a,b) for a in super_list for b in compl_sub_list
                    if is_subject_subset(b,a)]
    uniq_pairs = [ ]
    sub_one_list = [ ] 
    for super_one, sub_one in pairs : 
        if sub_one not in sub_one_list : 
            sub_one_list.append(sub_one)
            uniq_pairs.append( (super_one, sub_one) )

    if len(pairs)>0:
        if verbose : 
            #print "Found %d many pairs (%s, %s), ..." % (len(uniq_pairs), uniq_pairs[0][0], uniq_pairs[0][1])
            print("Found {} many pairs ({}, {}), ...".format(len(uniq_pairs),
                    uniq_pairs[0][0], uniq_pairs[0][1]))

        my_pair  = uniq_pairs[0]
        copy_from_and_put( * my_pair, partial=partial )
        return True
    else : 
        print("No pair found")
        return False




def detect_consecutives():
    t_list = [ ] 
    for t_name in teachers : 
        s_slots = list_util.union(
            [classpart_slot[cp] for cp in classparts_of(t_name) 
                if subject_name_of(cp) in subjects])
        slot_str = ''.join(sorted(list_util.complement(s_slots, [None])))
        for day in days:
            #if slot_str.find("%s1%s2%s3" % (day,day,day))>=0 or slot_str.find("%s2%s3%s4" % (day,day,day))>=0 or slot_str.find("%s5%s6%s7" % (day,day,day))>=0 : 
            if slot_str.find("{}1{}2{}3".format(day,day,day))>=0 or \
                  slot_str.find("{}2{}3{}4".format(day,day,day))>=0 or \
                   slot_str.find("{}5{}6{}7".format(day,day,day))>=0 :
                t_list.append(t_name)
                print(t_name, ''.join(sorted([sl for sl in s_slots 
                                        if str(sl).startswith(day)])))
    return t_list
            


def try_move_by_bounds_with_allows(st_name, subj, bound_dict=None):
    c_list = [c for c in current_classes(st_name) if not belongs_to(c,subj)]
    random.seed()
    random.shuffle(c_list)
    i_allows = []
    d_allows = []

    result = try_move_by_bound_dict(st_name, target=subj, 
                increase_allows = i_allows, decrease_allows = d_allows , 
                in_bound_dict=bound_dict)
    if result : 
        return True
    for clss in c_list : 
        if size_of(clss) == min(
            map(size_of,classes_of(subject_name_of(clss)))):
            d_allows.append(subject_name_of(clss))
        else:
            temp_subj = subject_name_of(clss)
            if temp_subj in bound_dict  and \
        bound_dict[temp_subj][1] <= max(map(size_of,classes_of(temp_subj))):
                #print "not allowing %s for inc" % temp_subj
                print("not allowing {} for inc".format(temp_subj))
                continue
                
            i_allows.append(subject_name_of(clss))
        #print "Trying %s with inc %s dec %s" % (st_name, ','.join([sem.shorthands[c] for c in i_allows]), ','.join([sem.shorthands[c] for c in d_allows]) )
        print("Trying {} with inc {} dec {}".format(st_name,
                ','.join([sem.shorthands[c] for c in i_allows]),
                ','.join([sem.shorthands[c] for c in d_allows]) ))
        result = try_move_by_bound_dict(st_name, target=subj, 
                    increase_allows = i_allows, decrease_allows = d_allows, 
                in_bound_dict=bound_dict)
        if result : 
            return True
    else : 
        return False



def try_copy_from_and_put(source_name, target_name ):
    subjs = subjects_of(target_name)
    target_classes = [clss for clss in current_classes(source_name) 
                        if subject_name_of(clss) in subjs ] 
    try: 
        choose_and_assign(target_name, includes = target_classes)
        return True
    except (IndexError, KeyboardInterrupt): 
        print("No assignment found, or interrupted")
        return False

def try_classes_and_put(class_list, st_name, bound=16, exp_bound=13 ):
    subjs = subjects_of(st_name)
    target_classes = [clss for clss in class_list 
                        if subject_name_of(clss) in subjs ] 
    excludes = [clss for clss in list_util.flatten(map(classes_of, subjs))
                    if ( expected_class_size(st_name, clss) > bound
             or (is_exp_subject(clss) and 
                expected_class_size(st_name, clss) > exp_bound) ) ]

    try: 
        #print "Trying %s with incl %s excl %s" % (st_name, ','.join([class_shorthands(c) for c in target_classes]), ','.join([class_shorthands(c) for c in excludes]) )
        print("Trying {} with incl {} excl {}".format(st_name,
                ','.join([class_shorthands(c) for c in target_classes]),
                ','.join([class_shorthands(c) for c in excludes]) ))
        choose_and_assign(st_name, includes = target_classes, 
            excludes = excludes)
        return True
    except (IndexError, KeyboardInterrupt): 
        print("No assignment found, or interrupted")
        return False

def run_fn_until_count(  fn, arg_one, arg_two_list, count=1 ):
    my_count = 0
    n = len(arg_two_list)
    for i, item in enumerate(arg_two_list) : 
        print(i+1, "/", n, ":", "("+str(my_count)+"/"+str(count)+")")
        result = fn(arg_one, item)
        if result is True : 
            my_count += 1
            if my_count >= count : 
                return True
    else : 
        return False


def has_unassigned(st_name):
    all_empty_subjs = [subj for subj in subjects 
        if [None] == list_util.union(
            [classpart_slot[pt] for pt in  parts_of(subj)] ) ]
    if len(list_util.complement(
            remaining_subjects(st_name),all_empty_subjs))>0:
        return True
    else : 
        return False
    
def subject_nums_members(subj, num_str): 
    c_list = [ subj + "_" + c for c in num_str.split()]
    output = [ ] 
    for clss in c_list:
        output.extend(st_members_of(clss))
    return output
    
def subject_bound_members(subj, lower_bound ): 
    c_list = [ c for c in classes_of(subj) if size_of(c) >= lower_bound ]
    output = [ ] 
    for clss in c_list:
        output.extend(st_members_of(clss))
    return output

def subject_bound_random_class(subj, upper_bound):
    c_list = [ c for c in classes_of(subj) if size_of(c) <= upper_bound ]
    if len(c_list)==0:
        return None
    random.seed()
    return random.choice(c_list)


def build_compatibility_dict(st_list):
    all_subjs = sorted(list_util.union(list_util.flatten([subjects_of(st_name) 
                    for st_name in st_list] )))
    #print "Building comp-pair-dict for %d many subjects %s .. %s" % (len(all_subjs), all_subjs[0], all_subjs[-1])
    print("Building comp-pair-dict for {} many subjects {} .. {}".format(
                len(all_subjs), all_subjs[0], all_subjs[-1]))
    all_subj_pairs = list_util.union([tuple(sorted((a, b))) for a in all_subjs 
                                        for b in all_subjs if a != b])
    output = { }
    for subj_pair in all_subj_pairs:
        subj_one, subj_two = subj_pair
        print("Working on %s %s" % (subj_one, subj_two))
        one_list = [c for c in classes_of(subj_one) if not has_none_slot(c)]
        two_list = [c for c in classes_of(subj_two) if not has_none_slot(c)]
        comp_pairs = [ (a, b) for a in one_list for b in two_list 
                        if []==list_util.intersect(slots_of(a), slots_of(b))]
        #print "Found %d comp-pairs" % len(comp_pairs)
        print("Found {} comp-pairs".format(len(comp_pairs)))
        sub_dict = { }
        for clss in one_list : 
            sub_dict[clss] = set([ b for (a,b) in comp_pairs if a == clss])
        for clss in two_list : 
            sub_dict[clss] = set([ a for (a,b) in comp_pairs if b == clss])
        output[ subj_pair ] = sub_dict
    return output


def is_compatible_pair(comp_dict, clss_one, clss_two):
    if clss_two in comp_dict[ tuple(sorted( 
        (subject_name_of(clss_one), subject_name_of(clss_two))))][clss_one]:
        return True
    else: 
        return False


def number_of_compatible_classes(comp_dict, class_list, subj):    
    count = 0;
    for clss in classes_of(subj) : 
        if has_none_slot(clss) : 
            continue
        elif list_util.all_true(map(
                lambda x: is_compatible_pair(comp_dict,x,clss),class_list)):
            count += 1
    return count

def enumerate_assignments_by_compatibility_dict( comp_dict,
    class_list, subject_list, size_bound_pair = (0, None), 
        excludes = [ ], includes = [ ],  randomize = False ): 
    if len( subject_list ) == 0 : 
        yield class_list 
    else : 
        candi_classes = [c for c in  classes_of( subject_list[0] ) 
                    if (c not in excludes) and (not has_none_slot(c)) 
                    and not ( size_bound_pair[0] > 0 and 
        size_bound_pair[0] < expected_class_size(size_bound_pair[1], c)) ] 
        
        
        classes = [ ] 
        for clss in candi_classes : 
            if list_util.all_true(map(
                lambda x: is_compatible_pair(comp_dict,x,clss),class_list)):
                classes.append(clss)
            else:
                continue
                

        curr_includes = [c for c in includes 
                            if subject_list[0] == subject_name_of(c) ]
        if len(curr_includes) > 0 : 
            intersected_classes = list_util.intersect(classes, 
                curr_includes )
        else : 
            intersected_classes = classes
        if len(intersected_classes ) > 0 : 
            classes = intersected_classes 
            if randomize is True : 
                random.shuffle(classes)
            else : 
                classes.sort( key = (lambda x: len(st_members_of(x))))
            for c in classes : 
                new_list = class_list[:] + [c]
                new_subj_list = subject_list[1:]
                if len(new_subj_list)>0:
                    next_subj = min(new_subj_list, key=lambda x:
                        number_of_compatible_classes(comp_dict, class_list, x))

                    new_subj_list.remove(next_subj)
                    next_subj_list = [next_subj] + new_subj_list
                else : 
                    next_subj_list = [ ] 

                for c_list in enumerate_assignments_by_compatibility_dict(
                    comp_dict, new_list, next_subj_list, size_bound_pair, 
                    excludes, includes, randomize ) : 
                    yield c_list
        else : 
            yield class_list 



def print_current_class_sizes(st_name) : 
    c_list = current_classes(st_name)
    c_list.sort(key= size_of)
    for c in c_list : 
        #st_num = len( list_util.complement( st_members_of(c), [ name ] )) + 1 
        #print "%s %d," % (class_shorthands(c), size_of(c)), 
        print("{} {},".format(class_shorthands(c), size_of(c)), end=' ')
    print


def choose_first_by_dict_and_assign(comp_dict, st_name, includes = [ ], 
        excludes=[], bound = 0, randomize=False, verbose = False ) : 
    subj_list = filter( has_class_with_assigned_slots, subjects_of(st_name))

    #random.seed()
    #random.shuffle(subj_list)
    subj_list.sort(key = lambda s: num_of_assignable_classes(name, s, excludes, 
                        bound, includes=includes ) )
    subj_num = len(subj_list)
    for i, c_list in enumerate(enumerate_assignments_by_compatibility_dict(
            comp_dict, [], subj_list, size_bound_pair = (bound, st_name),
        includes = includes, excludes = excludes,  randomize = randomize )): 
        if verbose is True and i%500 == 0:
            print(i, ' '.join(map(class_shorthands,c_list)))
        if len(c_list) == subj_num : 
            result_list = c_list 
            break 
    else : 
        print("No full assignment found" )
        return False
    put_name_into_classes(st_name, result_list)
    if verbose:
        print_current_class_sizes(st_name)
    return True


    
def try_assign_remains_by_dict(comp_dict, subj, verbose=False):
    st_list = remaining_names(subj)
    n = len(st_list)
    output = [ ]
    for i, st_name in enumerate(st_list):
        #print "%d/%d" %(i,n), 
        print("{}/{}".format(i,n), end=' ')
        print(st_name)
        result = choose_first_by_dict_and_assign(comp_dict, st_name, 
                    verbose=verbose )
        if not result : 
            output.append(st_name)
    return output
        


def try_move_by_compatibility_and_bounds(comp_dict, 
    st_name, in_bound_dict = None , target = None,
    increase_allows = [ ], decrease_allows = [ ], 
    include_class = None):
    bound_dict = { }

    if target is not None : 
        t_subject = target
    else : 
        t_subject = max( subjects_of(st_name), 
            key = lambda x: max( map(size_of, classes_of(x))) )

    print("target:", t_subject)

    for subj in subjects_of(st_name): 
        min_size = min(map(size_of, classes_of(subj)))
        max_size = max(map(size_of, classes_of(subj)))
        if subj == t_subject:
            bound_dict[subj] = (min_size, max_size - 1)
        elif subj in list_util.intersect(increase_allows, 
                        decrease_allows):
            bound_dict[subj] = (min_size-1, max_size+1)
        elif subj in increase_allows:
            bound_dict[subj] = (min_size, max_size+1)
        elif subj in decrease_allows:
            bound_dict[subj] = (min_size-1, max_size)
        else:
            bound_dict[subj] = (min_size, max_size)

    if in_bound_dict is not None : 
        bound_dict.update(in_bound_dict)

    includes = [ ] 
    excludes = [ ]
    for subj in subjects_of(st_name):
        for clss in classes_of(subj):
            if expected_class_size(st_name, clss) > bound_dict[subj][1] : 
                excludes.append(clss)
            if expected_class_size(st_name, clss, removing=True) < \
                         bound_dict[subj][0] : 
                includes.append(clss)
    if include_class is not None : 
        tmp_includes = filter(
    lambda x: subject_name_of(include_class)!=subject_name_of(x), includes)
        includes = tmp_includes + [include_class]

    try: 
        #print "excludes = %s, \nincludes = %s" % (' '.join(map( class_shorthands,excludes)), ' '.join(includes))
        print("excludes = {}, \nincludes = {}".format(' '.join(map(
                class_shorthands,excludes)), ' '.join(includes)))
        print(str_sorted_classes_of(st_name))

        result =  choose_first_by_dict_and_assign(comp_dict, st_name, 
            includes = includes, excludes=excludes, verbose = True ) 

        return result
    except (IndexError, KeyboardInterrupt): 
        print("No assignment found, or interrupted")
        return False


def put_one_into_class_by_compatibility(comp_dict, st_name, clss):
    print(st_name, end=" " )
    print_current_class_sizes(st_name)
    result = choose_first_by_dict_and_assign(comp_dict, st_name, 
            includes = [clss], verbose = True ) 
    return result

def add_to_class(comp_dict, count, clss, from_list = None):
    curr_count = 0 
    if from_list is None : 
        st_list = list_util.complement(members(subject_name_of(clss)), 
                    st_members_of(clss))
    else : 
        st_list = from_list
    random.seed()
    random.shuffle(st_list)
    n = len(st_list)
    for i, st_name in enumerate(st_list):
        #print "%d/%d (%d/%d)" % (i+1, n, curr_count, count)
        print("{}/{} ({}/{})".format(i+1, n, curr_count, count))
        result = put_one_into_class_by_compatibility(comp_dict, st_name, clss)
        if result : 
            curr_count += 1
        if curr_count >= count:
            return True
    else:
        print("Name list run out")
        return False
        
def name_bound_includes_excludes(st_name, in_bound_dict, max_size=25, 
        min_size=11):
    bound_dict = { }
    if max_size > 0 : 
        for subj in subjects_of(st_name): 
            if is_exp_subject(subj):
                my_min = 6 
                my_max = 12
            else : 
                my_min = min_size 
                my_max = max_size
            bound_dict[subj] = (my_min, my_max)
    else : 
        for subj in subjects_of(st_name): 
            my_sizes = map(size_of,classes_of(subj))
            bound_dict[subj] = (min(my_sizes), max(my_sizes))
        

    if in_bound_dict is not None:
        bound_dict.update(in_bound_dict)

    includes = [ ] 
    excludes = [ ]
    for subj in subjects_of(st_name):
        for clss in classes_of(subj):
            if expected_class_size(st_name, clss) > bound_dict[subj][1] : 
                excludes.append(clss)
    for clss in current_classes(st_name):
        if expected_class_size(st_name, clss, removing=True) < \
                     bound_dict[subject_name_of(clss)][0] : 
            includes.append(clss)
    return [includes, excludes]



def try_resize_with_bounds(comp_dict, subj, in_bound_dict=None, upper_bound=25,
                lower_bound=11,
                deal_small=False ,st_list=None   ):
    bound_dict = { } 
    if in_bound_dict is not None:
        bound_dict.update(in_bound_dict)


    min_size = min(map(size_of, classes_of(subj)))
    max_size = max(map(size_of, classes_of(subj)))
    bound_dict[subj] = (min_size, max_size - 1)
    
    random.seed()
    if st_list is None : 
        candis = members(subj)
    else : 
        candis = st_list
    if deal_small is not True : 
        st_name = random.choice(list_util.intersect(
            members_in_biggest_classes(subj), candis))
    else : 
        min_clss = filter(lambda x: size_of(x)==min_size, classes_of(subj))[0]
        st_name = random.choice(list_util.intersect(
            members_not_in_smallest_classes(subj),candis))

    includes, excludes = name_bound_includes_excludes(st_name, 
                bound_dict, max_size=upper_bound, min_size=lower_bound)

    if deal_small is True:
        tmp_includes = filter(
    lambda x: subj !=subject_name_of(x), includes)
        includes = tmp_includes + [min_clss]


    print(st_name, end=" " )
    print_current_class_sizes(st_name)
    #print "excludes = %s, \nincludes = %s" % (' '.join(map( class_shorthands,excludes)), ' '.join(includes))
    print("excludes = {}, \nincludes = {}".format(' '.join(map(
                class_shorthands,excludes)), ' '.join(includes)))
    result = choose_first_by_dict_and_assign(comp_dict, st_name, 
            includes = includes, excludes=excludes, verbose = True ) 
    return result





def try_assign_one_with_bounds(comp_dict, st_name, subj, in_bound_dict=None, 
        upper_bound=25, deal_small=False , keep_current = False , 
            includes=[], no_bound = False ):
    bound_dict = { } 
    if in_bound_dict is not None:
        bound_dict.update(in_bound_dict)

    if not no_bound : 
        min_size = min(map(size_of, classes_of(subj)))
        max_size = max(map(size_of, classes_of(subj)))
        bound_dict[subj] = (min_size, max_size )
    else : 
        bound_dict[subj] = (0, upper_bound)

    new_includes, excludes = name_bound_includes_excludes(st_name, 
                bound_dict, max_size=upper_bound)

    new_includes = list_util.union(new_includes, includes)

    if deal_small is True:
        tmp_includes = filter(
    lambda x: subj !=subject_name_of(x), new_includes)
        min_clss = smallest_classes_of(subj,exact=True)[0]
        new_includes = tmp_includes + [min_clss]
    

    if keep_current is True:
        curr_classes = current_classes(st_name)
        s_list = map(subject_name_of, curr_classes)
        tmp_includes=[c for c in new_includes 
            if subject_name_of(c) not in s_list]
        new_includes = curr_classes + tmp_includes


    print(st_name, end=" " )
    print_current_class_sizes(st_name)
    #print "excludes = %s, \nincludes = %s" % (' '.join(map( class_shorthands,excludes)), ' '.join(new_includes))
    print("excludes = {}, \nincludes = {}".format(' '.join(map(
                class_shorthands,excludes)), ' '.join(new_includes)))
    result = choose_first_by_dict_and_assign(comp_dict, st_name, 
            includes = new_includes, excludes=excludes, verbose = True ) 
    return result

def build_min_max_classes(freshmen=False):
    min_classes = [ ] 
    max_classes = [ ] 
    if freshmen : 
        subj_list = subjects
    else : 
        subj_list = list_util.complement(subjects, sem.junior_subjects)
    for subj in subj_list:
        c_list = classes_of(subj)
        size_list = map(size_of, c_list)
        max_size = max(size_list)
        min_size = min(size_list)
        min_classes.extend([c for c in c_list if size_of(c) == min_size])
        max_classes.extend([c for c in c_list if size_of(c) == max_size])
    return [ set(min_classes), set(max_classes)]

def min_max_ordered(st_list, mm_pair):
    min_classes, max_classes = mm_pair
    output = [ ] 
    for st_name in st_list : 
        c_list = current_classes(st_name)
        min_num = len(list_util.intersect(c_list, min_classes))
        max_num = len(list_util.intersect(c_list, max_classes))
        output.append( ( (-min_num, max_num), st_name) )
    output.sort()
    return output
    




def try_put_into_class_with_bounds(comp_dict, st_name, clss , 
        in_bound_dict=None, upper_bound=25 ):
    bound_dict = { } 
    if in_bound_dict is not None:
        bound_dict.update(in_bound_dict)

    includes, excludes = name_bound_includes_excludes(st_name, 
                bound_dict, max_size=upper_bound)

    subj = subject_name_of(clss)

    tmp_includes = filter(
            lambda x: subj !=subject_name_of(x), includes)
    includes = tmp_includes + [clss]

    print(st_name, end=" " )
    print_current_class_sizes(st_name)
    #print "excludes = %s, \nincludes = %s" % (' '.join(map( class_shorthands,excludes)), ' '.join(includes))
    print("excludes = {}, \nincludes = {}".format(' '.join(map(
                class_shorthands,excludes)), ' '.join(includes)))
    result = choose_first_by_dict_and_assign(comp_dict, st_name, 
            includes = includes, excludes=excludes, verbose = True ) 
    return result

def arg_to_slots(arg):
    if type(arg) is str : 
        list_arg = arg.split()
    else : 
        list_arg = arg
    output = list_util.intersect(slots, 
        map(gen_util.dict_to_fn(slot_dict), list_arg))
    return output

def filter_by_free_slots( free_slots, name_list ):
    my_slots = arg_to_slots(free_slots)
    output = [c for c in name_list if list_util.intersect(my_slots, 
                used_slots_of(c)) == [ ] ]
    return output
    
def allow_bound(subj, delta=1 ):
    subj_size = len(members(subj))
    num_cls = len(classes_of(subj))
    if 0 == subj_size % num_cls :
        average = subj_size / num_cls 
        return ( average - 1 - delta, average + 1 + delta)
    else : 
        average = int(float(subj_size) / num_cls )
        return ( average - delta, average + 1 + delta)


