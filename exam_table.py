# -*- coding: utf-8 -*-

import re, os, shutil
import datetime

import list_util, xlsx_util, utf_util, gen_util 

import pickle, random, hashlib

import openpyxl


from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.worksheet.pagebreak import Break


import semester as sem
import time_table as t

import term

try : 
    slot_minutes_dict 
except NameError : 
    slot_minutes_dict = { }
    exam_subjects = [ ] 
    subject_minutes_dict = { }
    slot_subject_list_dict = { }
    stackable_subjects_dict = { }
    subject_clnum_dict = { }
    classroom_dict = { }
    supervisor_dict = { }
    

save_filename = "exam_info.p"
def save(save_filename = save_filename) :
    gen_util.leave_a_backup( save_filename )
    with open( save_filename, "wb" ) as f :
        #print("Writing table to <%s> " % save_filename)
        print("Writing table to <{}> ".format(save_filename))
        pickle.dump( [slot_minutes_dict, exam_subjects ,
                    subject_minutes_dict , slot_subject_list_dict ,
                    stackable_subjects_dict, subject_clnum_dict, 
                    classroom_dict, supervisor_dict ], 
                f )

def load( filename = save_filename  ) :
    global slot_minutes_dict, exam_subjects  
    global subject_minutes_dict , slot_subject_list_dict 
    global stackable_subjects_dict, subject_clnum_dict 
    global classroom_dict, supervisor_dict

    with open( filename ) as f :
        print("Loading " + filename)
        slot_minutes_dict, exam_subjects ,\
            subject_minutes_dict , slot_subject_list_dict, \
                    stackable_subjects_dict, subject_clnum_dict, \
                    classroom_dict,  supervisor_dict \
            = pickle.load( open(filename, "rb") )

        #print("%s has %d items" % ("slot_minutes_dict", len(slot_minutes_dict)) )

        #print("%s has %d items" % ("exam_subjects", len(exam_subjects ))  )
        #print("%s has %d items" % ("subject_minutes_dict", len(subject_minutes_dict ))  )
        #print("%s has %d items" % ("slot_subject_list_dict", len(slot_subject_list_dict)) )
        #print("%s has %d items" % ("stackable_subjects_dict", len(stackable_subjects_dict)) ) print("%s has %d items" % ("subject_clnum_dict", len(subject_clnum_dict)) ) print("%s has %d items" % ("classroom_dict", len(classroom_dict)) ) print("%s has %d items" % ("supervisor_dict", len(supervisor_dict)) )


        print("{} has {} items".format("slot_minutes_dict", len(slot_minutes_dict)))

        print("{} has {} items".format("exam_subjects", len(exam_subjects )))
        print("{} has {} items".format("subject_minutes_dict", 
            len(subject_minutes_dict )))
        print("{} has {} items".format("slot_subject_list_dict", 
            len(slot_subject_list_dict)) )
        print("{} has {} items".format("stackable_subjects_dict", len(stackable_subjects_dict)) ) 
        print("{} has {} items".format("subject_clnum_dict", len(subject_clnum_dict)) ) 
        print("{} has {} items".format("classroom_dict", len(classroom_dict)) )
        print("{} has {} items".format("supervisor_dict", 
            len(supervisor_dict)) )





def read_subject_minutes(filename) : 
    data = xlsx_util.read_xlsx_sheet_into_dict(filename, 2,3)
    output = {}
    for k in data : 
        if re.sub(r'\s+', '', k[1]) == "시간(분)" : 
            output[ utf_util.convert_romans(k[0]) ] = int(data[k])
    return output



def stackable_subjects(subj ) : 
    st_list = t.members(subj)
    output = [ s for s in exam_subjects
                if len( list_util.intersect(st_list, t.members(s))) == 0 ]
    output.sort()
    return output

def make_stackable_subjects_dict() : 
    global stackable_subjects_dict 
    stackable_subjects_dict  = { }
    for subj in exam_subjects : 
        stackable_subjects_dict[subj] = stackable_subjects(subj)
    for subj in exam_subjects : 
        print(subj, "->", ','.join(stackable_subjects_dict[subj]))
    


def make_minutes_dict(num_days) : 
    global slot_minutes_dict 
    slot_minutes_dict = { }
    for d in range(1, num_days + 1) : 
        for k, length in enumerate(term.period_lengths):
            slot_minutes_dict[ (d,k+1) ] = length
    print("slot_minutes_dict")
    for k in sorted(slot_minutes_dict.keys()) : 
        #print("%s -> %d" % (str(k), slot_minutes_dict[k]))
        print("{} -> {}".format(str(k), slot_minutes_dict[k]))


def make_exam_subjects(filename) : 
    global subject_minutes_dict 
    global exam_subjects
    subject_minutes_dict = read_subject_minutes(filename) 
    exam_subjects = sorted(subject_minutes_dict.keys())
    for subj in exam_subjects : 
        print(subj, subject_minutes_dict[subj])


def initialize_slot_subjects() : 
    print("initializing slot_subject_list_dict ..")
    global slot_subject_list_dict 
    for k in slot_minutes_dict : 
        slot_subject_list_dict[k] = [ ]
    for k in sorted(slot_minutes_dict.keys()) :
        print(k, "->", slot_subject_list_dict[k], ";", end=" " )

def minutes_fit_subjects(slot) : 
    output = [subj for subj in exam_subjects 
                if subject_minutes_dict[subj] <= slot_minutes_dict[slot]]
    return output
    
def stackable_subjects_of_slot(slot) : 
    if len(slot_subject_list_dict[slot]) == 0 : 
        return exam_subjects 
    else : 
        return list_util.intersect( * [ stackable_subjects_dict[subj] 
                for subj in slot_subject_list_dict[slot] ] )

def make_subject_clnum_dict() : 
    global subject_clnum_dict 
    subject_clnum_dict = { }
    for subj in exam_subjects : 
        subject_clnum_dict[subj] = len(t.classes_of(subj))
    for subj in exam_subjects : 
        print(subj, "->", subject_clnum_dict[subj] )
    
        

def count_classes(slot) : 
    output = sum( [ subject_clnum_dict[subj] for subj in 
                      slot_subject_list_dict[ slot ] ])
    return output


def compatible_subjects(slot, max_num = term.max_simul_num, only_new = True ) : 
    candis = minutes_fit_subjects(slot) 
    if  only_new : 
        candis = list_util.complement( candis, 
                    list_util.flatten( slot_subject_list_dict.values() ) )
    output =  list_util.intersect( candis, 
                  stackable_subjects_of_slot(slot) ) 
    bound = max(0, max_num - count_classes(slot))
    
    classes_counted = [ subj for subj in output 
            if subject_clnum_dict[subj] <= bound ]
    return classes_counted


def assign_slot_subject(slot, subj) : 
    for sl in slot_subject_list_dict : 
        if subj in slot_subject_list_dict[sl] : 
            slot_subject_list_dict[sl].remove(subj)
            print("%s is removed from %s" % (subj, str(sl)))
            print("{} is removed from {}".format(subj, str(sl)))
    #print("Adding %s to %s" % ( subj, str(slot) ))
    print("Adding {} to {}".format( subj, str(slot) ))
    return list_util.check_append(slot_subject_list_dict[slot], subj) 

def remove_slot_subject(sl, subj) : 
    if subj in slot_subject_list_dict[sl] : 
        slot_subject_list_dict[sl].remove(subj)
        #print("%s is removed from %s" % (subj, str(sl)))
        print("{} is removed from {}".format(subj, str(sl)))
    else : 
        #print("%s is not assigned at %s. Doing nothing.." % (subj, str(sl)))
        print("{} is not assigned at {}. Doing nothing..".format(subj, 
            str(sl)))



def move_subject_to_slot( subj, slot ) : 
    for sl in slot_subject_list_dict : 
        if subj in slot_subject_list_dict[sl] :
            remove_slot_subject(sl, subj)
    assign_slot_subject(slot, subj)



def remove_all_from(slot):
    subj_list = slot_subject_list_dict[slot][:]
    for subj in subj_list : 
        remove_slot_subject(slot,subj)


    

def assign_first_one(slot, only_new = True) : 
    subj_list = compatible_subjects(slot, only_new = only_new)
    if len(subj_list) == 0 : 
        return False
    else :
        return assign_slot_subject(slot, subj_list[0])

def remaining_subjects() : 
    assigned = list_util.flatten( slot_subject_list_dict.values() )
    output = sorted( list_util.complement(exam_subjects, assigned))
    list_util.npr(output)
    return output

def pad_str(my_str, length, pad_with = " ", on="left" ) : 
    given_length = utf_util.korean_len(my_str)
    if  given_length >= length : 
        return my_str 
    else : 
        if on == "right" : 
            return (my_str + pad_with*(length - given_length )  )
        else:
            return (pad_with*(length - given_length ) + my_str)

def print_table( str_tuples, align="right" ) : 
    if align == "right" : 
        pad_on = "left"
    else : 
        pad_on = "right"
    n = max( map(len, str_tuples))
    padded_list = [ list_util.pad(tpl,  n) for tpl in str_tuples]
    column_widths  = { }
    for i in range(n) : 
        column_widths[i] = max( map(
            lambda x: utf_util.korean_len(x[i]) + 1, padded_list))
    #print(column_widths)
    lines = [ ] 
    for tpl in str_tuples : 
        lines.append( ' '.join( [pad_str(c, column_widths[i], " ", on=pad_on) 
            for i,c in enumerate(tpl)]   ))
    for line in lines : 
        print(line)
    



def show_assigned_subjects(slots_per_day=4) : 
    ordered_slots = sorted( slot_subject_list_dict.keys())
    #print_table( list_util.transposed(list_util.partition(
            #map(str, ordered_slots ),4)))
    n = max( map(len, slot_subject_list_dict.values()))
    #print(n)
    output = [ ] 
    for slot_list in list_util.partition(ordered_slots,slots_per_day) : 
        output.append( [ ] )
        for slot in slot_list :
            output[-1].append( str(slot))
            #print(slot_subject_list_dict[slot])
            output[-1].extend( list_util.pad( list(
                map( lambda x: x + " " + str(subject_minutes_dict[x]), 
                    slot_subject_list_dict[slot])), n))
    print_table( list_util.transposed(output))


def grade_str(subj) : 
    return ''.join( map(str,  t.grade_list_of( t.members(subj))))


def export_exam_table_xlsx(filename, version_str = "", n=5) : 
    ordered_slots = sorted( slot_subject_list_dict.keys())
    #n = max( map(len, slot_subject_list_dict.values()))
    max_slot_num = max( [c[1] for c in ordered_slots])
    output = [ ] 
    for slot_list in list_util.partition(ordered_slots,max_slot_num) : 
        output.append( [ ] )
        for slot in slot_list :
            output[-1].extend( list_util.pad(
                list(map( lambda x: (x,  str(subject_minutes_dict[x]), 
                    grade_str(x)), 
                    sorted(slot_subject_list_dict[slot]))), n, ('', "",'') ))


    print("Reading style from <exam-table-template.xlsx>.")
    shutil.copyfile( "exam-table-template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb["Template"]

    print("Writing exam table..")
    new_worksheet = wb.create_sheet(  'Sheet1' )
    instance = WorksheetCopy(template_ws, new_worksheet)
    WorksheetCopy.copy_worksheet(instance)
    new_worksheet.page_setup.orientation='landscape'
    xlsx_util.fill_in_sheet_by_list( new_worksheet, 
            [ term.term_str + " 시간표" + version_str ], 1,1,direction="row" )
    date_line = list_util.flatten( 
        [ [c, "시간", "학년"] for c in term.date_day_list] )
    xlsx_util.fill_in_sheet_by_list( new_worksheet, date_line , 2, 2, 
            direction="row" )

    print(list_util.transposed(output))
    ex_table = map(list_util.flatten, list_util.transposed( output))
    xlsx_util.fill_in_sheet_by_tuples( new_worksheet, ex_table ,  3, 2)

    wb.remove(template_ws)
    #print("Saving to <%s>."%filename)
    print("Saving to <{}>.".format(filename))
    wb.save(filename)





def exam_table_of(st_name, n=4) : 
    ordered_slots = sorted( slot_subject_list_dict.keys())
    #n = max( map(len, slot_subject_list_dict.values()))
    output = [ ] 
    for slot_list in list_util.partition(ordered_slots,4) : 
        output.append( [ ] )
        for slot in slot_list :
            my_tmp =  list_util.pad(
                map( lambda x: (x,  str(subject_minutes_dict[x]), 
                    grade_str(x)), 
                    sorted(slot_subject_list_dict[slot])), n, ('', "",'') )
            filtered_list = list(map( lambda x: x if len(x)>0 
                    and st_name in t.members(x[0]) else ("","",""), my_tmp))
            output[-1].extend(filtered_list)
    return map(list_util.flatten, list_util.transposed( output))





def show_assigned_minutes(slots_per_day=4) : 
    ordered_slots = sorted( slot_subject_list_dict.keys())
    n = max( map(len, slot_subject_list_dict.values()))
    output = [ ] 
    for slot_list in list_util.partition(ordered_slots,slots_per_day) : 
        output.append( [ ] )
        for slot in slot_list :
            output[-1].append( str(slot))
            minutes = 0 
            clnum = 0 
            for subj in slot_subject_list_dict[slot]:
                minutes += subject_minutes_dict[subj]*subject_clnum_dict[subj]
                clnum += subject_clnum_dict[subj]
            output[-1].append(str(clnum)) 
            output[-1].append(str(minutes)) 
    print_table( list_util.transposed(output))



def available_slots(subj, verbose=False) : 
    output = []
    for slot in sorted(slot_subject_list_dict.keys()) : 
        if subj in compatible_subjects( slot, only_new = False ):
            output.append(slot)
        if verbose : 
            conflicts = list_util.complement(slot_subject_list_dict[slot], 
                            stackable_subjects_dict[subj])
            if len(conflicts) > 0 : 
                print(slot, ' '.join(conflicts))
    output.sort()
    pairs = [ (sl, subj) for sl in output]
    list_util.npr( [ str(c[0]) + " " + c[1] for c in pairs ])
    return pairs

slots_per_day = range(1,1+ len(term.period_lengths))

def exam_subjects_on_day( day_num, st_name ) : 
    day_subjs = list_util.flatten(
        [ slot_subject_list_dict[ (day_num, i)] for i in slots_per_day])
    return  list_util.intersect( day_subjs, t.subjects_of(st_name) )

def count_exam_subjects_on_day( day_num, st_name, subject=None ) : 
    day_subjs = list_util.flatten(
        [ slot_subject_list_dict[ (day_num, i)] for i in slots_per_day])
    if subject is not None and subject in exam_subjects : 
        if subject not in day_subjs : 
            day_subjs.append(subject)
    return len( list_util.intersect( day_subjs, t.subjects_of(st_name) ))

def show_moved_overloading(subj, n, day): 
    st_list = [st_name for st_name in t.members(subj) 
                if count_exam_subjects_on_day(day, st_name, subject=subj)>=n]
    day_subjs = list_util.flatten( [ slot_subject_list_dict[ 
                        (day, i)] for i in slots_per_day])
    if subj not in day_subjs : 
        day_subjs.append(subj)
    count_overloading = { }
    for st_name in st_list : 
        subj_tuple = tuple(sorted(list_util.intersect(day_subjs, 
                        t.subjects_of(st_name) )))
        if subj_tuple not in count_overloading : 
            count_overloading[subj_tuple] = 0
    for tp_k in count_overloading.keys() : 
        count_overloading[tp_k] = len( 
            list_util.intersect( * [t.members(subj) for subj in tp_k]))
        
    print_output = [ ] 
    for k in sorted( count_overloading.keys() ) : 
        print_output.append( ( str(count_overloading[k]) , ) + k)
    print_output.sort(key = lambda x: (-int(x[0]), x[1]))
    list_util.join_npr(print_output)
    return st_list
    #return count_overloading
        



def overloading_subjects(day_num) : 
    n = max( [ count_exam_subjects_on_day(day_num, st_name) 
            for st_name in t.students])
    day_subjs = list_util.flatten(
            [ slot_subject_list_dict[ (day_num, i)] for i in slots_per_day])
    output = [ ]
    for st_name in t.students : 
        subj_tuple = tuple(sorted(list_util.intersect(day_subjs, 
                        t.subjects_of(st_name) )))
        if len(subj_tuple) == n and subj_tuple not in output :
            output.append(subj_tuple)
    return list_util.flatten( [ c + ( "", ) for c in output] )

def show_overloading() : 
    day_nums = range(1, 1 + max( [c[0] for c in slot_minutes_dict]))
    output = [ ]
    for day_n in day_nums : 
        output.append( [ ] )
        #output[-1].append( "(%d, )" % day_n )
        output[-1].append( "({}, )".format(day_n ))
        output[-1].extend( overloading_subjects(day_n) )
    print_table( list_util.transposed(output))

def overloading_students(n):
    day_nums = range(1, 1 + max( [c[0] for c in slot_minutes_dict]))
    st_list = [ ] 
    for st_name in t.students : 
        max_n = max( [ count_exam_subjects_on_day(day_num, st_name) 
                for day_num in day_nums ])
        if max_n >= n : 
            st_list.append(st_name)
    count_overloading = { }
    for day_n in day_nums : 
        day_subjs = list_util.flatten( [ slot_subject_list_dict[ 
                        (day_n, i)] for i in slots_per_day])
        for st_name in st_list : 
            subj_tuple = tuple(sorted(list_util.intersect(day_subjs, 
                        t.subjects_of(st_name) )))
            if len(subj_tuple) >= n : 
                if subj_tuple not in count_overloading : 
                    count_overloading[subj_tuple] = 0
    for tp_k in count_overloading.keys() : 
        count_overloading[tp_k] = len( 
            list_util.intersect( * [t.members(subj) for subj in tp_k]))

    print_output = [ ] 
    for k in sorted( count_overloading.keys() ) : 
        print_output.append( ( str(count_overloading[k]) , ) + k)
    print_output.sort(key = lambda x: (-int(x[0]), x[1]))
    list_util.join_npr(print_output)
    return st_list
        



def classes_at( slot ):
    return list(sorted( [c[1] for c in classroom_dict if c[0] == slot]))


def exchange_between_slots( slot_one, slot_two ) : 
    subj_list_one = slot_subject_list_dict[slot_one][:]
    subj_list_two = slot_subject_list_dict[slot_two][:]
    remove_all_from(slot_one)
    remove_all_from(slot_two)
    for subj in subj_list_one : 
        assign_slot_subject(slot_two, subj)
    for subj in subj_list_two : 
        assign_slot_subject(slot_one, subj)
    old_keys = [ ] 
    new_dict = { }
    for clss in classes_at(slot_one) : 
        old_keys.append( (slot_one, clss) )
        new_dict[ (slot_two, clss) ] = classroom_dict[ (slot_one, clss) ]
    for clss in classes_at(slot_two) : 
        old_keys.append( (slot_two, clss) )
        new_dict[ (slot_one, clss) ] = classroom_dict[ (slot_two, clss) ]
    print("Deleting ..")
    for pair in old_keys:
        print(str(pair[0]), pair[1], "  ", end=" ")
        del classroom_dict[pair]
    print("Updating ..")
    for pair in sorted(new_dict.keys()) : 
        print(str(pair[0]), pair[1], "->", new_dict[pair])
    classroom_dict.update(new_dict)
        
def exchange_between_subjects(subj_one, subj_two):
    slot_one = subject_slot(subj_one)
    slot_two = subject_slot(subj_two)
    if None in [slot_one, slot_two] :
        print("None slot found doing nothing")
        return None
    #print("Exchanging betwwen %s %s, %s %s.." % (str(slot_one), subj_one, str(slot_two), subj_two))
    print("Exchanging betwwen {} {}, {} {}..".format(str(slot_one), 
        subj_one, str(slot_two), subj_two))
    assign_slot_subject(slot_two, subj_one)
    assign_slot_subject(slot_one, subj_two)


def generate_classes(subj, clnum) : 
    return [ subj + "_" + str(k) for k in range(1,clnum + 1) ]

def initialize_classroom_dict(): 
    global classroom_dict 
    classroom_dict = { }
    for slot in slot_minutes_dict : 
        for subj in slot_subject_list_dict[slot]:
            for clss in generate_classes(subj, subject_clnum_dict[subj]) : 
                classroom_dict[ (slot, clss) ] = None
    print("classroom_dict intialized.." )
    for k in sorted(classroom_dict.keys()) : 
        print(str(k[0]), k[1], "->", classroom_dict[k], end=" ")


                
                
def available_classrooms( sl_clss ) : 
    in_slot, in_clss = sl_clss 
    used_rooms = [ classroom_dict[ (slot, clss) ] 
            for slot, clss in classroom_dict if slot == in_slot 
                and classroom_dict[ (slot, clss) ] is not None ]
    return list_util.complement( term.exam_classrooms, used_rooms )

def assign_classroom( sl_clss, classroom ) : 
    slot, clss = sl_clss
    for bundle in term.bundle_classes : 
        if clss in bundle : 
            c_list = bundle[:]
            break 
    else : 
        c_list = [clss]
        
    for cl in c_list : 
        assign_dict_value( classroom_dict, (slot, cl) , classroom,
            str_fn = lambda x: "{} {}".format(str(x[0]), x[1] ) )

    #classroom_dict[ (slot, clss) ] = classroom
    #print(str(slot), clss, "->", classroom_dict[(slot, clss)])


def assign_first_available_classroom(sl_clss) : 
    rooms = available_classrooms(sl_clss)
    if len(rooms)>0:
        assign_classroom(sl_clss, rooms[0])
        return True
    else : 
        return False



def supervisor_pair( sl_clss ) : 
    first_one = str( supervisor_dict[ (sl_clss[0], sl_clss[1], svA) ] )
    if (sl_clss[0], sl_clss[1], svB) in supervisor_dict and \
         (sl_clss[0], sl_clss[1], svC) in supervisor_dict : 
        second_one = str( supervisor_dict[ (sl_clss[0], sl_clss[1], svB) ] ) +\
                 " " + str( supervisor_dict[ (sl_clss[0], sl_clss[1], svC) ] )
    elif (sl_clss[0], sl_clss[1], svB) in supervisor_dict : 
        second_one = str( supervisor_dict[ (sl_clss[0], sl_clss[1], svB) ] )
    elif (sl_clss[0], sl_clss[1], svC) in supervisor_dict : 
        second_one = str( supervisor_dict[ (sl_clss[0], sl_clss[1], svC) ] )
    else : 
        second_one = "" 
    return [ first_one, second_one ]
    

def supervisor_triple( sl_clss ) : 
    first_one = str( supervisor_dict[ (sl_clss[0], sl_clss[1], svA) ] )
    if (sl_clss[0], sl_clss[1], svB) in supervisor_dict : 
        second_one = str( supervisor_dict[ (sl_clss[0], sl_clss[1], svB) ] )
    else : 
        second_one = "" 
    if (sl_clss[0], sl_clss[1], svC) in supervisor_dict : 
        third_one = str( supervisor_dict[ (sl_clss[0], sl_clss[1], svC) ] )
    else : 
        third_one = "" 

    return [ first_one, second_one, third_one ]
    




def class_sort_fn(clss ) : 
    return (t.subject_name_of(clss), int(t.class_number_of(clss)))

def class_shorthands(x):
    return  (sem.shorthands[t.subject_name_of(x)]+ "_" + t.class_number_of(x) )

def show_classroom_assignment() : 
    max_slot_num = max( [c[0][1] for c in supervisor_dict])
    ordered_slots = sorted( [c for c in slot_minutes_dict.keys() 
                        if c[1]<= max_slot_num ])
    n = 0 
    for slot in ordered_slots : 
        candi_n = len( [ c for c in classroom_dict if c[0] == slot])
        n = max( n, candi_n)

    output = [ ] 
    for slot_list in list_util.partition(ordered_slots,max_slot_num) : 
        output.append( [ ] )
        for slot in slot_list :
            output[-1].append( str(slot))
            my_keys = [ c for c in classroom_dict if c[0] == slot]
            my_keys.sort(key = lambda x: (x[0], class_sort_fn(x[1])))
            output[-1].extend( list_util.pad(
                map( lambda x: sem.shorthands[t.subject_name_of(x[1])] 
                        + "_" + t.class_number_of(x[1]) + " " + 
                str(subject_minutes_dict[t.subject_name_of(x[1])]) + 
                " " + sem.classroom_shorthand[classroom_dict[x]] +
                " " + ' '.join( supervisor_pair(x) )  , 
                    my_keys ), n))
    print_table( list_util.transposed(output), align="left")


def classroom_list_at( slot ) : 
    class_list = classes_at(slot)
    class_list.sort( key = class_sort_fn )
    output = [ classroom_dict[ (slot, cl)] for cl in class_list  ]
    return output


def assign_slot_first_classrooms(slot):
    my_keys = [ c for c in classroom_dict if c[0] == slot]
    my_keys.sort(key = lambda x: (x[0], class_sort_fn(x[1])))
    for slot, clss in my_keys : 
        assign_first_available_classroom( (slot, clss) )
    

def assign_slot_classrooms( slot, room_list ) : 
    class_list = classes_at(slot)
    class_list.sort( key = class_sort_fn )
    n = min(len(class_list), len(room_list)) 
    for i in range(n) : 
        assign_classroom( (slot, class_list[i]), room_list[i] )


def exchange_classrooms_at_slot( slot, room_one, room_two) : 
    first_key = list(filter( lambda x: x[0] == slot and classroom_dict[x]==room_one,
                    classroom_dict.keys()))[0]
    second_key = list(filter( lambda x: x[0] == slot 
                    and classroom_dict[x]==room_two, classroom_dict.keys()))[0]
    #print("Exchanging classrooms of %s %s" % ( first_key[1], second_key[1]))
    print("Exchanging classrooms of {} {}".format( first_key[1], 
        second_key[1]))
    assign_classroom( first_key, room_two )
    assign_classroom( second_key, room_one )
    
def change_classroom_at_slot( slot, room_old, room_new) : 
    first_key = list(filter( lambda x: x[0] == slot and classroom_dict[x]==room_old,
                    classroom_dict.keys()))[0]
    assign_classroom( first_key, room_new )
    
    



svA = "정감독" 
svB = "부감독"
svC = "복도감독" 

def initialize_supervisor_dict(special_room=False, autonomous=True) : 
    global supervisor_dict 
    supervisor_dict = { }
    "Initializing supervisor_dict.."
    for slot, clss in classroom_dict : 
        subj = t.subject_name_of(clss)
        if clss.find("특별고사") >= 0 : 
            continue

        supervisor_dict[ (slot, clss, svA) ] = None
        max_minutes = max( [ subject_minutes_dict[t.subject_name_of(c)] 
                for c in classes_at(slot) if c.find("특별고사")<0 ])
        if clss in term.svB_classes : 
            supervisor_dict[ (slot, clss, svB) ] = None
        elif clss in term.svC_classes : 
            supervisor_dict[ (slot, clss, svC) ] = None
        elif subj in term.auto_subjects and autonomous is True : 
            supervisor_dict[ (slot, clss, svC) ] = None
            #if subject_minutes_dict[subj] > 50 : 
                #supervisor_dict[ (slot, clss, svB) ] = None
        elif subject_minutes_dict[subj] <= 50 : 
            supervisor_dict[ (slot, clss, svC) ] = None
        else :
            supervisor_dict[ (slot, clss, svB) ] = None
    if special_room is True : 
        all_slots = sorted( list_util.union( 
            [slot for slot, clss in classroom_dict]))
        for i, slot in enumerate(all_slots) : 
            max_minutes = max( [ subject_minutes_dict[t.subject_name_of(c)] 
                for c in classes_at(slot) if c.find("특별고사")<0 ])
            new_item = ( slot, "특별고사" + str(i+1) + "_1" )
            subject_minutes_dict[ t.subject_name_of(new_item[1]) ] = \
                     max_minutes
            classroom_dict[new_item] = "외국어강의실1" 
            supervisor_dict[ tuple( new_item + (svA, )) ] = None
        
    for slot, clss, sv in sorted(supervisor_dict.keys()) : 
        print(slot, clss, sv, "->", supervisor_dict[ (slot,clss, sv)])
        
prep_minutes = 15
auto_minutes = 25

def minutes_of(sl_clss_sv, raw = False, autonomous=True) : 
    slot, clss, sv  = sl_clss_sv
    subj_min = subject_minutes_dict[ t.subject_name_of(clss) ]
    if t.subject_name_of(clss) in term.auto_subjects and sv in [svA, svB] : 
        return auto_minutes 
    
    if raw : 
        prep_m = 0
    else : 
        prep_m = prep_minutes
    if sv in [svA ] : 
        return (prep_m + subj_min)
    else :
        return subj_min
    
        
def count_minutes(t_name, moderation = False, raw=False, 
        accumulation=False, day= None, autonomous=True , temp_added = False ) : 
    triples = [ ] 
    slot_list = [ ] 
    for slot, clss, sv in supervisor_dict : 
        if t_name == supervisor_dict[ (slot, clss, sv) ] : 
            if slot not in slot_list : 
                triples.append( (slot, clss, sv) )
                slot_list.append(slot)
    if day is not None:
        slot_list = [ sl for sl in slot_list if sl[0] == day]
        triples = [trp for trp in triples if trp[0][0] == day]
    if moderation : 
        if t_name in term.moderation_dict : 
            weight = term.moderation_dict[t_name]
        else : 
            weight = 0 
    else : 
            weight = 0 
    if accumulation : 
        if t_name in term.accumulated_time_dict : 
            weight += term.accumulated_time_dict[t_name]
    if temp_added and (t_name in term.temp_added_time_dict) : 
        weight += term.temp_added_time_dict[t_name]

    return (weight + sum( [ minutes_of( triple, raw=raw, autonomous=autonomous) 
                for triple in triples ]))

def count_unavailable_slots(t_name) : 
    if t_name in term.unavailable_teacher_slots_dict : 
        return(-len( list_util.union( 
            term.unavailable_teacher_slots_dict[t_name])))
    else : 
        return 0

def next_slot( slot ) : 
    return ( slot[0], slot[1] + 1)
def prev_slot(slot):
    return ( slot[0], slot[1] - 1)

def is_slot_teacher_available(slot, t_name): 
    if t_name in term.unavailable_teacher_slots_dict and  \
        slot in term.unavailable_teacher_slots_dict[t_name] : 
        return False
    my_slots = slots_of(t_name)
    if slot in my_slots : 
        return False 
    if len( [sl for sl in my_slots if sl[0] == slot[0]] ) >= 3 : 
        return False
    if True:
        if (slot[1] in [2]  and next_slot(slot) in my_slots ) or \
            (slot[1] in [3] and prev_slot(slot) in my_slots) : 
            return False
    if t_name in term.consec_avoid_teachers: 
        if (slot[1] in [1,2,3,4]  and next_slot(slot) in my_slots ) or \
            (slot[1] in [2,3,4,5] and prev_slot(slot) in my_slots) : 
            return False
    if t_name in term.first_third_avoid_teachers: 
        if (slot[1] == 1 and next_slot(next_slot(slot)) in my_slots ) or \
            (slot[1] == 3 and prev_slot(prev_slot(slot)) in my_slots ):
            return False
    if t_name in term.second_year_advisors + term.third_year_advisors and \
                    slot[1]==2 : 
            return False
    return True

    
def is_available_at( t_name, sl_clss_sv ) :  
    slot, clss, sv = sl_clss_sv 
    if not is_slot_teacher_available(slot, t_name):
        return False
    subj_conflicts = [c for c in term.shuffled_teachers 
                if len( list_util.intersect( t.subjects_of(t_name), 
                        slot_subject_list_dict[slot])) > 0 ] 
    if t_name in subj_conflicts : 
        return False
    if slot in slots_of(t_name):
        return False
    if slot[1] == 1 and sv == svA and t_name in term.aa_list :
        return False
    if t_name in term.unavailable_sv_teachers_dict[sv] : 
        return False
    else : 
        return True

def assign_name_slot_any(t_name, sl, enforce=False):
    if enforce is False : 
        triples = [trp for trp in supervisor_dict if trp[0]==sl and 
            supervisor_dict[trp] is None and is_available_at(t_name,trp) ]
    else: 
        triples = [trp for trp in supervisor_dict if trp[0]==sl and 
            supervisor_dict[trp] is None  ]
    if len(triples) == 0 : 
        #print("All at %s are assigned or unassignable, doing nothing" % str(sl))
        print("All at {} are assigned or unassignable, doing nothing".format(
            str(sl)))

    else : 
        random.shuffle(triples)
        assign_teacher_at(t_name, triples[0])

def assign_name_day_any(t_name, day):
    triples = [trp for trp in supervisor_dict if trp[0][0]==day and 
        supervisor_dict[trp] is None and is_available_at(t_name,trp) ]
    if len(triples) == 0 : 
        #print("All at %s are assigned or unassignable, doing nothing" % str(day))
        print("All at {} are assigned or unassignable, doing nothing".format(
            str(day)))
    else : 
        random.shuffle(triples)
        assign_teacher_at(t_name, triples[0])

def assign_name_for_all_days(t_name) : 
    day_list = sorted(term.day_str_dict.keys())
    #print("Working on %s" % t_name)
    print("Working on {}".format(t_name))
    for day in day_list : 
        #print("Looking at day %d" % day )
        print("Looking at day {}".format(day))
        assign_name_day_any(t_name, day)
 
def assign_teachers_for_all_days():
    for t_name in term.shuffled_teachers : 
        assign_name_for_all_days(t_name)

def has_more_slots_than(t_name, n):
    if len(slots_of(t_name)) > n : 
        return True
    else : 
        return False


def first_available_teacher(sl_clss_sv):
    slot, clss, sv = sl_clss_sv 
    minute_names = [ ( has_more_slots_than(x,6), has_more_slots_than(x,2), 
        count_minutes(x, moderation = True, accumulation = True, 
            temp_added = True), x ) for x in term.shuffled_teachers]
    minute_names.sort()
    for three_min, six_max, t_min, t_name in minute_names : 
        if is_available_at(t_name, sl_clss_sv):
            return t_name
    else:
        return None



def available_teachers( sl_clss_sv, top=3, avoid_serial = True  ) : 
    slot, clss, sv = sl_clss_sv 

    unavailables = [t_name for t_name in term.shuffled_teachers 
                        if not is_slot_teacher_available(slot, t_name)]
    #print("slot_unavail", ' '.join(unavailables))
    subj_conflicts = [t_name for t_name in term.shuffled_teachers 
                if len( list_util.intersect( t.subjects_of(t_name), 
                        slot_subject_list_dict[slot])) > 0 ] 
    #print("subj_confl", ' '.join(subj_conflicts))
    slot_occupieds = [ supervisor_dict[c] for c in supervisor_dict 
                        if c[0] == slot ]

    #if avoid_serial is True : 
        #serial_occupieds = [ supervisor_dict[c] for c in supervisor_dict 
           #if c[0] == (slot[0], slot[1]-1) or c[0] == (slot[0], slot[1]+1)]
        #candis = list_util.complement( term.shuffled_teachers , unavailables, 
                    #subj_conflicts, slot_occupieds, serial_occupieds )
    #else : 

    candis = list_util.complement( term.shuffled_teachers , unavailables, 
                    subj_conflicts, slot_occupieds )
    #candis = [c for c in term.shuffled_teachers if 
                    #is_available_at(c, sl_clss_sv)]

    #candis.sort( key = lambda x: ( count_minutes(x, moderation = True, 
        #accumulation = True, temp_added = True), 
                #term.shuffled_teachers.index(x) )  )
    output = candis[:top]
    #output = [ min( candis, key = 
        #lambda x: ( count_minutes(x, moderation = True, 
        #accumulation = True, temp_added = True), 
                #term.shuffled_teachers.index(x) )  )]
    random.shuffle(output)
    return output




def remaining_positions( svs = [svA, svB, svC]  ) : 
    return sorted([ k for k in supervisor_dict if supervisor_dict[k] is None  
                and k[2] in svs ], 
        key = lambda x: (150 - slot_minutes_dict[x[0]], x))

def assign_dict_value( my_dict, in_key, val, str_fn = str ) : 
    my_dict[in_key] = val 
    print(str_fn(in_key), "->", val)

def triple_str(x):
    #return "%s %s %s"% (str(x[0]), x[1], x[2]) 
    return "{} {} {}".format(str(x[0]), x[1], x[2])



def are_on_same_floor(room_one, room_two): 
    math_rooms = ["수학강의실" + str(i) for i in range(1,8)]
    fourth_floor = ["국어강의실1",  "국어강의실2",  "국어강의실3",  
           "우암공통강의실3", "사회강의실1", "사회강의실2", "사회강의실3" ]
    second_floor = ["우암공통강의실1", "우암공통강의실2"]
    eng_floor = ["외국어강의실1", "외국어강의실2", "외국어강의실3", 
                    "외국어강의실4", "외국어강의실5"] 
    if list_util.is_subset([room_one, room_two], math_rooms) : 
        return True 
    elif list_util.is_subset([room_one, room_two], fourth_floor) : 
        return True 
    elif list_util.is_subset([room_one, room_two], second_floor) : 
        return True 
    elif list_util.is_subset([room_one, room_two], eng_floor) : 
        return True 
    else:
        return False


def bundle_positions( sl_clss_sv ) : 
    slot, clss, sv = sl_clss_sv
    the_room = classroom_dict[ (slot, clss) ]
    svC_rooms = sorted([ classroom_dict[ (slot, c) ] for c in classes_at(slot) 
                    if sv == svC ])
    same_floor_rooms = tuple(sorted([room for room in svC_rooms 
        if are_on_same_floor(the_room, room)]))
    if t.subject_name_of(clss) in term.auto_subjects  and \
          same_floor_rooms in term.auto_bundle_classroom_dict :
        my_bundle_classrooms = term.auto_bundle_classroom_dict[same_floor_rooms]
    elif same_floor_rooms in term.bundle_classroom_dict : 
        my_bundle_classrooms = term.bundle_classroom_dict[same_floor_rooms]
    elif len(same_floor_rooms) <= 4 : 
        my_bundle_classrooms = term.special_bundle_classrooms 
    else : 
        my_bundle_classrooms = term.bundle_classrooms 

    rooms = list_util.flatten( [ c for c in my_bundle_classrooms 
                if classroom_dict[ (slot, clss) ] in c ] )
    c_list = list_util.flatten( [ c for c in term.bundle_classes if clss in c])
    print(' '.join(c_list))
    output = [ ] 
    for c in classes_at(slot) : 
        if sv == svC and classroom_dict[ (slot, c) ] in rooms : 
            if (slot, c, sv) in supervisor_dict : 
                output.append( (slot, c, sv) )
        if c in c_list : 
            if (slot, c, sv) in supervisor_dict : 
                output.append( (slot, c, sv) )
    return output



def assign_teacher_at(t_name, sl_clss_sv, bundle=True ) : 
    sl, clss, sv = sl_clss_sv
    b_positions = bundle_positions( sl_clss_sv ) 
    if len(b_positions) == 0 or bundle is False : 
        assign_dict_value( supervisor_dict, sl_clss_sv, t_name, 
            str_fn = triple_str  )
    else : 
        for position in b_positions : 
            assign_dict_value( supervisor_dict, position , t_name, 
                str_fn = triple_str )

def assign_first_teacher( sl_clss_sv, top=1, avoid_serial=True ): 
    #t_name = available_teachers( sl_clss_sv, top=top, 
                #avoid_serial=avoid_serial  )[0]
    t_name = first_available_teacher(sl_clss_sv)
    assign_teacher_at(t_name, sl_clss_sv) 

def assign_preassigned() : 
    for sl, clss, sv in supervisor_dict.keys() : 
        if (clss, sv) in term.preassigned_pairs : 
            t_name = term.preassigned_pairs[ (clss, sv) ]
            assign_teacher_at( t_name, (sl, clss, sv) )

def exchange_between_teachers( t_one, t_two ) : 
    rev_dict = list_util.reversed_dict( supervisor_dict )
    one_triples = rev_dict[ t_one ]
    two_triples = rev_dict[ t_two ]
    for triple in one_triples : 
        assign_teacher_at( t_two, triple) 
    for triple in two_triples : 
        assign_teacher_at( t_one, triple) 


def remove_subject( subj ) : 
    for sl in slot_subject_list_dict : 
        if subj in slot_subject_list_dict[sl] : 
            print("Removing {} from  {} -> {}".format( subj, str(sl), 
                ' '.join(slot_subject_list_dict[sl])) )
            slot_subject_list_dict[sl].remove(subj)
            print("Now {} -> {}".format( str(sl), 
                ' '.join(slot_subject_list_dict[sl])) )
    key_list = classroom_dict.keys() 

    #print("Removing %s from exam_subjects" % subj)
    print("Removing {} from exam_subjects".format(subj))
    exam_subjects.remove(subj)
    #print("Deleting subject_minutes_dict[%s]" % subj)
    print("Deleting subject_minutes_dict[{}]".format(subj))

    del subject_minutes_dict[subj]

    for sl, clss in  key_list :
        if subj == t.subject_name_of(clss) : 
            #print("Deleting (%s, %s) -> %s from classroom_dict " % (str(sl), clss, classroom_dict[ (sl, clss) ] ) )
            print("Deleting ({}, {}) -> {} from classroom_dict ".format(
                str(sl), clss, classroom_dict[ (sl, clss) ] ))
            del classroom_dict[ (sl, clss) ] 
            print("Deleted..")

    key_list = supervisor_dict.keys() 
    for sl, clss, sv in  key_list :
        if subj == t.subject_name_of(clss) : 
            #print("Deleting (%s, %s, %s) -> %s from classroom_dict " % (str(sl), clss, sv, supervisor_dict[ (sl, clss, sv) ] ) )
            print("Deleting ({}, {}, {}) -> {} from classroom_dict ".format(
                str(sl), clss, sv, supervisor_dict[ (sl, clss, sv) ] ))
            del supervisor_dict[ (sl, clss, sv) ] 
            print("Deleted..")
            

def show_teacher_minutes() : 
    t_list = term.shuffled_teachers[:]
    output = [ ( t_name, count_minutes(t_name) ) for t_name in t_list] 
    output.sort( key = lambda x: (x[1], x[0]) )
    for c in output : 
        print(c[0], c[1])




def check_compatibility() : 
    output = True 
    for sl in sorted(slot_subject_list_dict.keys()) : 
        print(sl, end=" ")
        for subj in slot_subject_list_dict[sl] : 
            if subject_minutes_dict[subj] <= slot_minutes_dict[sl] : 
                print(subj, subject_minutes_dict[subj], "<=" , end=" " )
                print(slot_minutes_dict[sl], '   ', end=" " )
            else : 
                output = False
                print()
                print(subj, subject_minutes_dict[subj], ">" , end=" " )
                print(slot_minutes_dict[sl] )
        print()
    for sl in sorted(slot_subject_list_dict.keys()) : 
        print(sl, end=" " )
        for pair in list_util.distinct_pairs(slot_subject_list_dict[sl]) : 
            my_tmp = list_util.intersect( t.members(pair[0]), 
                            t.members(pair[1]))
            if len(my_tmp) != 0 : 
                output = False 
                print()
                print(pair[0], pair[1], len(my_tmp) )
            else : 
                print(pair[0], pair[1], "OK", end=" ")
        print()
    return output

def slots_of(t_name) : 
    rev_dict =   list_util.reversed_dict(supervisor_dict)
    if t_name in rev_dict : 
        return sorted( list_util.union( [ c[0] for c in rev_dict[t_name] ]))
    else : 
        return [ ] 

def supervisor_triple_partners(t_name):
    t_triples = list_util.filter_dict( lambda x: x==t_name, 
                    supervisor_dict).keys() 
    t_triples = [c for c in t_triples if c[2] != svC ]
    output = [ ] 
    for tp in t_triples : 
        if tp[2]==svA and (tp[0],tp[1], svB) in supervisor_dict : 
            output.append( (tp[0],tp[1],svB, 
                    supervisor_dict[(tp[0],tp[1],svB)] ) )
        elif tp[2]==svB and (tp[0],tp[1], svA) in supervisor_dict : 
            output.append( (tp[0],tp[1],svA, 
                    supervisor_dict[(tp[0],tp[1],svA)] ) )
    return output


def check_supervisor_compatibility():
    output = True
    for t_name_one, t_name_two in term.avoid_teacher_pairs : 
        tp_names = supervisor_triple_partners(t_name_one)
        for quadruple in tp_names:
            if quadruple[3] == t_name_two : 
                print(t_name_one, quadruple[0], quadruple[1], end=" " )
                print(quadruple[2], quadruple[3] )
                output = False
        
    for triple in supervisor_dict : 
        if supervisor_dict[triple] is None : 
            print("{} -> None".format( triple_str(triple)))
            output = False
    rev_dict = list_util.reversed_dict(supervisor_dict)
    for t_name in term.shuffled_teachers : 
        if t_name in term.unavailable_teacher_slots_dict : 
            tmp_list = list_util.intersect( slots_of(t_name), 
                        term.unavailable_teacher_slots_dict[t_name]) 
            if len(tmp_list) > 0 : 
                #print("slot conflict %s %s" % (t_name, ' '.join(map(str,tmp_list))) )
                print("slot conflict {} {}".format(t_name, 
                    ' '.join(map(str,tmp_list))) )
                output = False

    for t_name in term.shuffled_teachers : 
        if t_name in rev_dict : 
            #triples = [ c for c in rev_dict[t_name] 
                #if (c[1],c[2]) not in term.preassigned_pairs]
            triples = [ c for c in rev_dict[t_name] ]
            check_slot_duplicates = list_util.union([ 
                (c[0], c[2]) for c in triples if c[2] == svC ])
            check_slot_duplicates.extend( [ 
                (c[0], c[2]) for c in triples if c[2] != svC ])
            for sl in list_util.union(list_util.firsts(check_slot_duplicates)):
                if list_util.count(check_slot_duplicates, 
                    lambda x: x[0] == sl)>1:
                        print("slot conflict {} {}".format(t_name, 
                            ' '.join(map(triple_str,
                         [c for c in triples if c[0] == sl]) )))
                        output = False

            for slot, clss, sv in triples : 
                tmp_list = list_util.intersect( t.subjects_of(t_name), 
                        slot_subject_list_dict[slot]) 
                if len(tmp_list) > 0 : 
                    #print("subject conflict %s %s" % (t_name, ' '.join(tmp_list)))
                    print("subject conflict {} {}".format(t_name, 
                        ' '.join(tmp_list)))
                    output = False
                
                if t_name in term.avoid_classes_dict and  \
                    clss in term.avoid_classes_dict[t_name] : 
                    print("AA conflict {} {}".format(t_name, clss))
                    output = False
                if t_name in (term.first_year_advisors + term.second_year_advisors + term.third_year_advisors )and \
                    slot[1]==1 and sv == svA : 
                    print("AA conflict {} {} {}".format(t_name, str(slot),sv))
                    output = False
                #if t_name in term.second_year_advisors + \
                        #term.third_year_advisors and \
                    #slot[1]==2 : 
                    #print("AA conflict {} {}".format(t_name, str(slot) ))
                    #output = False
    return output




def export_exam_tables_of_all_classes( filename, class_list = sem.ban_list,
                   t_name = False, st_list = None, version_str = "" ) :
    print("Reading style from <exam-table-template-for-class.xlsx>.")
    shutil.copyfile( "exam-table-template-for-class.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    template_ws = wb["Template"]
    for clss in class_list :
        prefix = clss.replace("-", "")
        name_list = [nm for nm in t.names if nm.startswith(prefix)]
        if st_list is not None :
            name_list = list_util.intersect(name_list, st_list)
        name_list.sort()
        if(len(name_list)==0) :
            continue

        #print("Writing exam table for %s." % clss)
        print("Writing exam table for {}.".format(clss))
        new_worksheet = wb.create_sheet(  clss )
        instance = WorksheetCopy(template_ws, new_worksheet)
        WorksheetCopy.copy_worksheet(instance)
        new_worksheet.page_setup.orientation='landscape'

        for k in range( 19, 401, 20  ):
            new_worksheet.row_dimensions[k+1].hidden = True
            new_worksheet.row_dimensions[k].hidden = True
            new_worksheet.row_dimensions[k-1].hidden = True
            #new_worksheet.row_dimensions[k-2].hidden = True
        for k in range(20,401,20):
            #new_worksheet.page_breaks.append( Break(id=k) )
            new_worksheet.row_breaks.append( Break(id=k) )

        date_line = list_util.flatten(
            [ [c, "시간", "학년"] for c in term.date_day_list] )
        for i, st_name in enumerate(name_list) :
            xlsx_util.fill_in_sheet_by_list( new_worksheet, 
                [t.convert_name_for_xlsx(st_name), 
                    '', '', '', term.term_str + " 시간표" + version_str  ], 
                        i*20 + 1, 1, direction="row" )
            xlsx_util.fill_in_sheet_by_list( new_worksheet, 
                 date_line , i*20 + 2, 2, direction="row" )

            xlsx_util.fill_in_sheet_by_tuples( new_worksheet, 
                exam_table_of(st_name), i*20 + 3, 2)
        for k in range( (i+1)*20 + 1, 401  ):
            new_worksheet.row_dimensions[k].hidden = True
    wb.remove(template_ws)
    #print("Saving to <%s>."%filename)
    print("Saving to <{}>.".format(filename))
    wb.save(filename)

def password_of(st_name): 
    m = hashlib.md5()
    m.update(term.salt + st_name)
    return  m.hexdigest()[:6]


def name_password_list(st_list = None) : 
    output = [ ] 
    if st_list is None : 
        f_st_list = t.students
    else : 
        f_st_list = st_list
     
    for st_name in f_st_list:
        m = hashlib.md5()
        m.update(term.salt + st_name)
        output.append( (st_name, m.hexdigest()[:6]))
    return  output

def num_name_split(st_name) : 
    sr = re.search( r'^(\d{4})(\S+)' , st_name)
    if sr : 
        return (sr.group(1), sr.group(2))
    else : 
        return None

#def export_redirect_pairs_xlsx(filename, st_list = None):
    #if st_list is None : 
        #f_st_list = [c for c in  t.students if c[0] in "2 3".split() ]
    #else : 
        #f_st_list = st_list
    #output = [ ] 
    #for st_name in f_st_list:
        #pw = password_of(st_name) 
        #st_num = num_name_split(st_name)[0]
        #long_url = term.prefilled_link.replace("NUMBER", 
                    #st_num).replace("CODE", pw)
        #output.append( (st_num + "_" + pw, long_url ) )
    #xlsx_util.write_tuples_into_xlsx(filename, output, sheetname="regist") 


def export_redirect_pairs_txt(filename, st_list = None, prefix = "notice"):
    if st_list is None : 
        f_st_list = [c for c in  t.students if c[0] in "2 3".split() ]
    else : 
        f_st_list = st_list
    output = [ ] 
    for st_name in f_st_list:
        pw = password_of(st_name) 
        st_num = num_name_split(st_name)[0]
        long_url = term.prefilled_link.replace("NUMBER", 
                    st_num).replace("CODE", pw)
        if prefix is None : 
            short_key = st_num + "_" + pw
        else : 
            short_key = prefix + "/" + st_num + "_" + pw
        output.append( (short_key, long_url ) )
    gen_util.write_tuples_into_txt(filename, output ) 

    
def write_messages_csv(filename, st_list = None) : 
    if st_list is None : 
        f_st_list = [c for c in  t.students if c[0] in "2 3".split() ]
    else : 
        f_st_list = st_list

    output = [ ] 
    for st_name in f_st_list:
        pw = password_of(st_name) 
        st_num = num_name_split(st_name)[0]
        message = term.message_template.replace("NUMBER", 
                    st_num).replace("CODE", pw)
        phone = term.phone_dict[ st_name ].replace("-", "")
        output.append( phone + "," +  message + "," + st_name)
    with open(filename, "wb") as f : 
        #print("Writing <%s>.." % filename )
        print("Writing <{}>..".format(filename ))
        f.write( '\n'.join(output) )


def schedule_str(t_name, moderation = True, raw=False) : 
    rev_dict = list_util.reversed_dict( supervisor_dict )
    output = [t_name, str(count_minutes(t_name, moderation=moderation,
        raw=raw, accumulation=True)), str(count_minutes(t_name, moderation)) ]
    if t_name not in rev_dict : 
        return ' '.join(output)
    else : 
        triples = rev_dict[t_name]
        slot_list = [ ] 
        str_list = [ ] 
        for slot, clss, sv in triples : 
            if slot not in slot_list : 
                str_list.append(str(slot).replace(" ","") + 
             class_shorthands(clss).replace("_","") + sv.replace("감독","") 
                        + str( minutes_of((slot,clss,sv), raw=raw)))
                slot_list.append(slot)
        str_list.sort()
        output.extend(str_list)
        return ' '.join(output)


    
def period_strings(day_num, max_slot_num=4):
    period_strs = [ ] 
    for i in range(1,max_slot_num + 1): 
        subj_list = slot_subject_list_dict[ (day_num, i) ]
        if len(subj_list) == 0:
            period_strs.append('')
        else : 
            max_minutes = max([subject_minutes_dict[c] for c in subj_list])
            end_datetime = term.period_starts[i-1] \
                                + datetime.timedelta(0,60*max_minutes)
            period_strs.append( term.period_starts[i-1].strftime("%H:%M") 
                    + "~" + end_datetime.strftime("%H:%M") )
    return period_strs
        



def export_classroom_assignment_xlsx(filename, max_slot_num=4) : 
    print("Reading style from <exam-table-notice-templates.xlsx>.")
    shutil.copyfile( "exam-table-notice-templates.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    target_sheet_name =  "고사실" 
    sheet_names = list_util.complement( wb.sheetnames, 
            [ target_sheet_name])
    for sn in sheet_names : 
        ws = wb[ sn ] 
        wb.remove(ws)
    ws = wb[ target_sheet_name ]

    #ordered_slots = sorted( slot_minutes_dict.keys())
    #max_slot_num = max( [c[0][1] for c in supervisor_dict])
    ordered_slots = sorted( [c for c in slot_minutes_dict.keys() 
                        if c[1]<= max_slot_num ])
    n = 0 
    for slot in ordered_slots : 
        candi_n = len( [ c for c in classroom_dict if c[0] == slot])
        n = max( n, candi_n)

    output = [ ] 
    for slot in ordered_slots :
        output.append( [ ] )
        my_keys = [ c for c in classroom_dict if c[0] == slot]
        my_keys.sort(key = lambda x: (x[0], class_sort_fn(x[1])))
        output[-1].extend( list_util.pad(
                map( lambda x: ( 
                utf_util.revert_romans( term.shorten_by_dict(
                    t.subject_name_of(x[1]))),  
                         t.class_number_of(x[1]) + "분반", 
                sem.classroom_shorthand[classroom_dict[x]]) , 
                    my_keys ), n, ( '','','') ))
    sec_output = list_util.transposed(output)  
    third_output = [ list_util.flatten(c) for c in sec_output]
    print_table(third_output)
    xlsx_util.fill_in_sheet_by_tuples( ws, third_output ,  11, 2)

    output = [ ] 
    n = max( map(len, slot_subject_list_dict.values()))
    for slot in ordered_slots :
        output.append( [ ] )
        output[-1].extend(  list_util.pad( map( lambda x: (x 
            + " (" + str(subject_minutes_dict[x]) + "')", '', '' ) , 
                    sorted(slot_subject_list_dict[slot])), n, ( '','','') ))
    third_output = [ list_util.flatten(c) for c in list_util.transposed(output)]
    print_table(third_output)
    xlsx_util.fill_in_sheet_by_tuples( ws, third_output ,  5, 2)


    xlsx_util.fill_in_sheet_by_tuples( ws,  [ list_util.flatten( [ [c,'',''] 
        for c in  list_util.flatten( 
     [period_strings(i, max_slot_num=max_slot_num) 
            for i in sorted(term.day_str_dict.keys()) ]) ]) ], 4, 2)


    fourth_output = list_util.flatten( [  [term.day_str_dict[k]] + 
            ([""]* (3*max_slot_num - 1) )   
            for k in sorted(term.day_str_dict.keys())] ) 
    xlsx_util.fill_in_sheet_by_tuples( ws, [fourth_output] ,  2, 2)
    xlsx_util.fill_in_sheet_by_tuples( ws, [(term.notice_str,) ] ,  1, 1)
    

    #print("Saving to <%s>."%filename)
    print("Saving to <{}>.".format(filename))
    wb.save(filename)






def export_assignment_by_day_xlsx( day_num, output_filename = None, 
     max_slot_num=4) : 
    if output_filename is None : 
        #filename = (term.term_str.replace(" ", "_") + "_%d일차_시험감독.xlsx" % day_num)
        filename = (term.term_str.replace(" ", "_") + 
            "_{}일차_시험감독.xlsx".format(day_num))

    else : 
        filename = output_filename
    print("Reading style from <exam-table-notice-templates.xlsx>.")
    shutil.copyfile( "exam-table-notice-templates.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    target_sheet_name =  "1일차" 
    sheet_names = list_util.complement( wb.sheetnames, 
            [ target_sheet_name])
    for sn in sheet_names : 
        ws = wb[ sn ] 
        wb.remove(ws)
    ws = wb[ target_sheet_name ]

    #ws.title = unicode( "%d일차" % day_num , 'utf-8' )
    ws.title = "{}일차".format(day_num)


    xlsx_util.fill_in_sheet_by_list(ws, [ term.term_str + " 감독관" ] , 1,1)
    xlsx_util.fill_in_sheet_by_list(ws, [ term.day_str_dict[day_num] ], 2,2)
    my_table = [ ] 
    for i in range(2) : 
        my_table.append( ['']*30 )

    period_strs = [ ] 
    for i in slots_per_day: 
        subj_list = slot_subject_list_dict[ (day_num, i) ]
        if len(subj_list) == 0:
            period_strs.append('')
        else : 
            max_minutes = max([subject_minutes_dict[c] for c in subj_list])
            end_datetime = term.period_starts[i-1] \
                                + datetime.timedelta(0,60*max_minutes)
            period_strs.append( term.period_starts[i-1].strftime("%H:%M") 
                    + "~" + end_datetime.strftime("%H:%M") )
        

    for i in range(1) : 
        for j in range(max_slot_num) : 
            my_table[i][ 6*j] = term.time_periods[i][j]
    for i in range(1,2) : 
        for j in range(max_slot_num) : 
            my_table[i][ 6*j] = period_strs[j] 

    xlsx_util.fill_in_sheet_by_tuples(ws,  my_table, 3,2)


    output = [ ] 
    n = max( map(len, slot_subject_list_dict.values()))

    ordered_slots = sorted( slot_minutes_dict.keys())
    day_slots = [ c for c in  ordered_slots if c[0] == day_num]
    for slot in day_slots  :
        output.append( [ ] )
        output[-1].extend(  list_util.pad( map( lambda x: (x 
            + " (" + str(subject_minutes_dict[x]) + "')",'','', '', '','' ) , 
                    sorted(slot_subject_list_dict[slot])), n, 
                        ('','', '','','','') ))
    third_output = [ list_util.flatten(c) for c in list_util.transposed(output)]
    print_table(third_output)
    xlsx_util.fill_in_sheet_by_tuples( ws, third_output ,  5, 2)

    output = []
    n = 0 
    for slot in day_slots : 
        candi_n = len( [ c for c in classroom_dict if c[0] == slot])
        n = max( n, candi_n)
    for slot in day_slots :
        output.append( [ ] )
        my_keys = [ c for c in classroom_dict if c[0] == slot]
        my_keys.sort(key = lambda x: (x[0], class_sort_fn(x[1])))
        output[-1].extend( list_util.pad(
             map(lambda x:( utf_util.revert_romans(
                term.shorten_by_dict(t.subject_name_of(x[1]))) 
                    if t.class_number_of(x[1]) == "1" else '' ,  
                         t.class_number_of(x[1]) + "분반", 
                sem.classroom_shorthand[classroom_dict[x]]) + 
      tuple(map( gen_util.convert_name_for_xlsx, supervisor_triple(x))), 
                    my_keys ), n, ('','', '','','','') ))
    print(output)
    for c in output : 
        for tp in c : 
            print(' '.join(tp), end=" " )
        print()
    sec_output = list_util.transposed(output)  
    third_output = [ list_util.flatten(c) for c in sec_output]
    print(third_output)
    xlsx_util.fill_in_sheet_by_tuples( ws, third_output ,  11, 2)


    #print("Saving to <%s>."%filename)
    print("Saving to <{}>.".format(filename))
    wb.save(filename)

def modify_shorthand(my_str) : 
    return my_str.replace("수강", "수학").replace("사강", "사회")

def export_classroom_assignment_tuples_xlsx(filename) : 
    pairs = list(classroom_dict.keys())
    pairs.sort( key = lambda x: (x[0], class_sort_fn( x[1] )) )
    output = [ ] 
    for slot, clss in pairs : 
        sr = re.search( r'\((\d+)\D+(\d+)\D+\s+(\S+)\)', 
                term.day_str_dict[ slot[0] ])
        month = int(sr.group(1))
        date = int(sr.group(2))
        day = sr.group(3)
        output.append(  [ t.class_number_of(clss) + "분반", 
            modify_shorthand(
            sem.classroom_shorthand[classroom_dict[ (slot, clss) ]]), 
            month, date, day, ("{}교시".format(slot[1])), 
            t.subject_name_of(clss), term.day_str_dict[ slot[0] ],
            ("{}교시({}분)".format(slot[1], 
                subject_minutes_dict[t.subject_name_of(clss)]  ))]  )
            #("%d교시(%d분)" % (slot[1], subject_minutes_dict[t.subject_name_of(clss)]  ))]  )

    xlsx_util.write_tuples_into_xlsx(filename, output)
        
    

def export_big_supervisor_table(filename, svD = False, minutes=True, 
        max_slot_num=4 ) : 
    print("Reading style from <supervisors-template.xlsx>.")
    shutil.copyfile( "supervisors-template.xlsx", filename)
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ordered_slots = sorted( slot_subject_list_dict.keys())
    n = max( map(len, slot_subject_list_dict.values()))
    output = [ ] 
    day_strings = [ ] 

    subj_minutes_list = [ ]

    for slot in ordered_slots : 
        output.append( map( utf_util.revert_romans, 
            sorted(slot_subject_list_dict[slot])))
        subj_list = slot_subject_list_dict[slot]
        if len(subj_list) == 0:
            subj_minutes_list.append('')
        else : 
            subj_minutes_list.append('/'.join(sorted(list_util.union(
                [str(subject_minutes_dict[c])+"'" for c in subj_list]))))
        
    ex_table =  list_util.transposed( output)

    svD_str = "시험감독" 
    svD_dict = { svA: svD_str, svB: svD_str, svC:svD_str, '':''}
    
    name_triples_dict = list_util.reversed_dict( supervisor_dict )

    t_list = sorted( term.shuffled_teachers, 
                key = (lambda t_name: ( sem.ordered_departments.index( 
                        sem.name_department_dict[t_name]), t_name) ))
    sup_table = [ ] 
    mod_acc_list =  [ ] 
    unavailable_list = [ ]
    role_list = [ ] 
    
    for t_name in t_list : 
        mod_acc_list.append( (term.moderation_dict[t_name] 
            if t_name in term.moderation_dict else 0, 
                term.accumulated_time_dict[t_name]) )  
        tmp_name = gen_util.convert_name_for_xlsx(t_name)
        subj_str = ','.join( [ utf_util.revert_romans(subj)
            for subj in sorted(t.subjects_of(t_name )) if subj in t.subjects ])

        if t_name in name_triples_dict : 
            triples = name_triples_dict[t_name]
        else : 
            triples = [ ] 
        tmp_dict = {}
        ordered_slots = sorted( slot_subject_list_dict.keys())
        for slot, clss, sv in triples : 
            if slot not in tmp_dict : 
                if minutes is False:
                    tmp_dict[slot] = sv
                else:
                    #tmp_dict[slot] = "%s%3d" % (sv, minutes_of((slot,clss,sv), raw=False))
                    tmp_dict[slot] = "{}{:3d}".format(sv, 
                        minutes_of((slot,clss,sv), raw=False))
        for slot in ordered_slots : 
            if slot not in tmp_dict : 
                tmp_dict[slot] = ''
        if svD is False : 
            sv_list = tuple([tmp_dict[slot] for slot in ordered_slots])
        else : 
            sv_list = tuple([svD_dict[tmp_dict[slot]]  
                for slot in ordered_slots])

        sup_table.append( ( tmp_name , 
            t.bujang_str(tmp_name), subj_str ) + sv_list )

        unavail_rec = [ ]
        for slot in ordered_slots: 
            rec_tmp = ''
            if (t_name in term.unavailable_teacher_slots_dict) and \
                (slot in term.unavailable_teacher_slots_dict[t_name]) :  
                rec_tmp = '1'
            if (t_name in term.hide_unavailable_slots ) and \
                (slot in term.hide_unavailable_slots[t_name]) :  
                rec_tmp = ''
            unavail_rec.append(rec_tmp)
        unavailable_list.append(unavail_rec)
        if t_name in term.role_dict : 
            role_list.append( term.role_dict[t_name] )
        else : 
            role_list.append( '' )
        

    day_str_list = [ term.day_str_dict[c] for c 
            in sorted( list_util.union( [k[0] for k in ordered_slots]))]
    #print(' '.join(day_str_list))

    print("Writing exam table..")
    xlsx_util.fill_in_sheet_by_list( ws, subj_minutes_list, 
        4,6,direction="row" )
    #xlsx_util.fill_in_sheet_by_list( ws, minute_list, 10,34,direction="column" )
    print(day_str_list)
    xlsx_util.fill_in_sheet_by_list( ws, day_str_list,2,
        40 +(max_slot_num-4)*6,direction="column" )
    xlsx_util.fill_in_sheet_by_list( ws, 
            [ term.term_str + " 감독관" ], 1,1,direction="row" )
    xlsx_util.fill_in_sheet_by_tuples( ws, ex_table ,  5, 6)
    xlsx_util.fill_in_sheet_by_tuples( ws, sup_table,  10, 3)
    xlsx_util.fill_in_sheet_by_tuples( ws, mod_acc_list,  
        10, 32+(max_slot_num-4)*6)
    xlsx_util.fill_in_sheet_by_tuples( ws, unavailable_list , 
        10, 41 +(max_slot_num-4)*6 )
    xlsx_util.fill_in_sheet_by_list( ws, role_list, 
        10,40+(max_slot_num-4)*6,direction="column" )
    #print("Saving to <%s>."%filename)
    print("Saving to <{}>.".format(filename))
    wb.save(filename)



def give_slot_name_to_name( slot, b_name, s_name) : 
    triple = list( filter( 
                lambda x: x[0] == slot and  supervisor_dict[x]==b_name, 
                supervisor_dict.keys() ))[0]
    #print("Found %s at %s,%s,%s" % (b_name, str(triple[0]),triple[1],triple[2]))
    print("Found {} at {},{},{}".format(b_name, 
            str(triple[0]),triple[1],triple[2]))
    #print("Giving it to %s" % s_name)
    print("Giving it to {}".format(s_name))
    assign_teacher_at(s_name, triple)


def name_slot_triple(b_name, slot_one) : 
    tr_one = list(filter( lambda x: x[0] == slot_one and  
            supervisor_dict[x]==b_name, supervisor_dict.keys() ))[0]
    return tr_one

def name_day_triples(t_name, day_num) : 
    return list(filter( lambda x: x[0][0] == day_num and  
            supervisor_dict[x]==t_name, supervisor_dict.keys() ))

def name_slot_triple(t_name, slot) : 
    output = list(filter( lambda x: x[0] == slot and  
            supervisor_dict[x]==t_name, supervisor_dict.keys() ))
    if output == [ ] : 
        return None
    else : 
        return output[0]

    
def exchange_name_slots( b_name, slot_one,  s_name, slot_two ) : 
    tr_one = list(filter( lambda x: x[0] == slot_one and  
            supervisor_dict[x]==b_name, supervisor_dict.keys() ))[0]
    print("Found {} at {},{},{}".format(b_name, 
                str(tr_one[0]),tr_one[1],tr_one[2]))
    tr_two = list(filter( lambda x: x[0] == slot_two and  
            supervisor_dict[x]==s_name, supervisor_dict.keys() ))[0]
    print("Found {} at {},{},{}".format(s_name, 
        str(tr_two[0]),tr_two[1],tr_two[2]))
    print("Exchanging them")

    assign_teacher_at(b_name, tr_two)
    assign_teacher_at(s_name, tr_one)

    
def relieve_name_slot( b_name, slot_one) : 
    tr_one = list(filter( lambda x: x[0] == slot_one and  
            supervisor_dict[x]==b_name, supervisor_dict.keys() ))[0]
    #print("Found %s at %s,%s,%s" % (b_name, str(tr_one[0]),tr_one[1],tr_one[2]))
    print("Found {} at {},{},{}".format(b_name, 
        str(tr_one[0]),tr_one[1],tr_one[2]))
    print("Relieving .. ")
    assign_teacher_at(None, tr_one)

def add_flu_classes() : 
    classroom_dict.update( term.flu_classroom_dict )
    supervisor_dict.update( term.flu_supervisor_dict )


def delete_slot_classroom( slot, classroom) : 
    clss = list(filter( lambda x: x[0] == slot and classroom_dict[x] == classroom, 
                classroom_dict.keys()))[0][1]
    #print("Found %s %s -> %s" % ( slot, clss, classroom_dict[ (slot,clss)]))
    print("Found {} {} -> {}".format( slot, clss, 
        classroom_dict[ (slot,clss)]))
    triples = [c for c in supervisor_dict.keys() 
        if c[0] == slot and c[1] == clss]
    for tr in triples : 
        #print("Found %s %s %s -> %s" % ( tr[0],tr[1], tr[2],supervisor_dict[tr]))
        print("Found {} {} {} -> {}".format( tr[0],tr[1], tr[2],
            supervisor_dict[tr]))
    print("Deleting them")
    del classroom_dict[ (slot,clss)]
    for tr in triples : 
        del supervisor_dict[ tr ]
    
    
def change_supervisor_dict_keys( triples, delete=True ) : 
    for triple in triples : 
        if triple[2] == svA : 
            #print("Cannot delete %s %s %s"%(str(triple[0]),triple[1],triple[2]) )
            print("Cannot delete {} {} {}".format(str(triple[0]),
                triple[1],triple[2]) )
            return False
    new_triples = [ (c[0], c[1], svC) for c in triples  ]
    for triple in triples : 
        #print("Deleting %s %s %s"%(str(triple[0]), triple[1], triple[2]) )
        print("Deleting {} {} {}".format(str(triple[0]), 
            triple[1], triple[2]) )
        del supervisor_dict[triple]
    if not delete :
        for triple in new_triples : 
            #print("Writing %s %s %s -> None"%(str(triple[0]),triple[1], triple[2])  )
            print("Writing {} {} {} -> None".format(str(triple[0]),
                triple[1], triple[2]))
            supervisor_dict[ triple ] = None

def subject_slot(subj) : 
    for sl in slot_subject_list_dict : 
        if subj in slot_subject_list_dict[sl] : 
            return sl 
    else : 
        return None

    
def show_pair_weights(subj) : 
    subj_pairs = [c for c in term.pair_weights.keys() if subj in c]
    subj_pairs.sort(key = lambda x: -term.pair_weights[x])
    for pair in subj_pairs : 
        t_subj = list_util.complement(pair, [subj])[0]
        my_slot = subject_slot(t_subj)
        slot_str = '' if my_slot is None else str(my_slot)
        print(t_subj + slot_str, subj, ":", term.pair_weights[pair])

def subj_slot_str(subj):
    my_slot = subject_slot(subj)
    slot_str = '' if my_slot is None else str(my_slot)
    return slot_str

def compute_pair_weights() : 
    output_pairs = [ ] 
    days = sorted( list_util.union( [c[0] for c in slot_minutes_dict.keys()]))
    pair_keys = term.pair_weights.keys()
    for day_num in days : 
        day_subjs = list_util.flatten(
            [ slot_subject_list_dict[ (day_num, i)] for i in slots_per_day])
        for pair in list_util.distinct_pairs(day_subjs) : 
            if pair in pair_keys : 
                output_pairs.append(pair)
            elif (pair[1],pair[0]) in pair_keys : 
                output_pairs.append((pair[1],pair[0]))
    output_pairs.sort( key = gen_util.dict_to_fn( term.pair_weights ))
    for pair in output_pairs : 
        print(pair[0], subj_slot_str(pair[0]), end=" " )
        print(pair[1], subj_slot_str(pair[1]), end=" " )
        print(":", term.pair_weights[pair])
    return output_pairs 
    
    

def count_intersections( slot_one, slot_two ) : 
    first_subjs = slot_subject_list_dict[slot_one]
    second_subjs = slot_subject_list_dict[slot_two]
    for subj_one in first_subjs : 
        for subj_two in second_subjs : 
            print(subj_one, subj_two, ":", end=" " )
            print(len(list_util.intersect(t.members(subj_one), 
                                        t.members(subj_two))) )

def available_slots_for_teacher(t_name): 
    #output = [ ] 
    my_slots = list_util.union( [trp[0] for trp in supervisor_dict])
    #for slot in my_slots : 
        ##print(slot)
        #triple = ([ trp for trp in supervisor_dict if trp[0] == slot])[0]
        #if t_name in  available_teachers( triple, top=70,avoid_serial=True ) :
            #output.append(slot)
    avail_triples = [ tp for tp in supervisor_dict 
        if is_available_at( t_name, tp ) ]
    return sorted( list_util.union( [c[0] for c in avail_triples]))

    #return sorted(output)


def overworked_teachers(max_minutes=180): 
    t_list = [ ] 
    for t_name in term.shuffled_teachers : 
        for i in range(1,7) : 
            c_minutes = count_minutes(t_name, day = i)
            if c_minutes >= max_minutes : 
                list_util.check_append(t_list, t_name)
    return t_list


def teacher_slot_triple(t_name, slot) : 
    for triple in supervisor_dict : 
        if triple[0] == slot and supervisor_dict[triple]==t_name : 
            return triple
    else : 
        #print("%s is not assigned at %s" % (t_name, str(slot)))
        print("{} is not assigned at {}".format(t_name, str(slot)))
        return None


def subject_slot_cnums() : 
    output = [ ] 
    rev_dict = list_util.reversed_list_dict( slot_subject_list_dict )
    for subj in exam_subjects : 
        slot = rev_dict[subj]
        #day_str = "제%d일 %d교시" % slot 
        day_str = "제{}일 {}교시".format(slot)
        output.append( (subj, day_str, len(t.classes_of(subj))))
    return output

def paracname_subjects():
    e_classes = list_util.flatten([t.classes_of(subj) 
        for subj in exam_subjects])
    e_classes.sort()
    output = [(t.para_classname(c), t.subject_name_of(c)) for c in e_classes]
    return output
    
def count_students_per_teacher(subj) : 
    st_count = len(t.members(subj))
    t_count = len( [c for c in t.t_members_of(subj) 
        if not t.is_assistant_name(c)] )
    return (float(st_count)/t_count)



def first_available_triple(t_name, avoid_serial=True) : 
    try: 
        for sl_clss_sv in remaining_positions(): 
            avails = available_teachers( sl_clss_sv, top=80, 
                avoid_serial = avoid_serial) 
            if t_name in avails : 
                return sl_clss_sv
        else : 
            return None
    except (KeyboardInterrupt, IndexError) as e : 
        print(e)
        return None

def assign_teacher_at_slot(t_name, in_slot, idx=0):
    pos_candis = [c for c in remaining_positions() if c[0] == in_slot]
    if len(pos_candis) < idx + 1 : 
        #print("Positions at slot %s are less than %d." % (str(in_slot), idx + 1))
        print("Positions at slot {} are less than {}.".format(str(in_slot), 
                idx + 1))
        return None
    else : 
        assign_teacher_at(t_name, pos_candis[idx] ) 



def assign_first_triple_first_teacher(idx=0) : 
    t_list = [c for c in term.shuffled_teachers 
        if c not in term.unavailable_teacher_slots_dict or \
        len(term.unavailable_teacher_slots_dict[c]) < len(slot_minutes_dict)]
    t_list.sort(key = lambda x: count_minutes(x, moderation = True,  
            accumulation=True)) 
    t_name = t_list[idx]
    trp = first_available_triple(t_name)
    if trp is not None : 
        assign_teacher_at(t_name, trp) 
        print(schedule_str(t_name, moderation=True, raw=False))
        return True
    else : 
        return False

def assign_firsts(idx=0) : 
    toggle = True
    for i in range(1000) : 
        if toggle is False : 
            break 
        else : 
            toggle = assign_first_triple_first_teacher(idx = idx)
wah_nums = "2 4 5 9".split()
gwah_nums = "1 3 6 7 8 0".split()

def and_josa_after(my_string):
    if my_string[-1] in wah_nums : 
        return "와"
    else:
        return "과"

def teacher_schedule_str(t_name):
    my_slots = sorted(slots_of(t_name))
    slot_strings = [ ] 
    for sl in my_slots : 
        slot_strings.append( term.day_day_dict[sl[0]] + str(sl[1]) )
    output_string = ''.join(slot_strings)
    return output_string

def sms_message_for(t_name): 
    my_slots = sorted(slots_of(t_name))
    slot_strings = [ ] 
    for sl in my_slots : 
        slot_strings.append( term.day_day_dict[sl[0]] + str(sl[1]) )
    output_string = ''.join(slot_strings)
    msg = "안녕하세요. " + t_name.replace("t", '') + \
        " 부득이하게 시감을 다시 짰습니다. 선생님의 기말고사 시험감독 일정은 "  + output_string  +\
      and_josa_after(output_string)+ " 같습니다." +\
        " 최원근 올림" 
    return msg

def has_double_day(t_name, day=None):
    slot_list = slots_of(t_name)
    firsts = [c[0] for c in slot_list]
    if day is None : 
        indices = sorted(term.day_str_dict.keys())
    else : 
        indices = [day]
    for i in indices : 
        if firsts.count(i)>1 : 
            return True
    else:
        return False
    

def has_free_day(t_name, day=None):
    slot_list = slots_of(t_name)
    firsts = [c[0] for c in slot_list]
    if day is None : 
        indices = sorted(term.day_str_dict.keys())
    else : 
        indices = [day]
    for i in indices : 
        if firsts.count(i)==0 : 
            return True
    else:
        return False
    
def exchange_candidates(t_list, day_one, day_two):
    group_one = [t_name for t_name in t_list if has_free_day(t_name,day_one) 
                    and has_double_day(t_name, day_two)]
    group_two = [t_name for t_name in t_list if has_double_day(t_name,day_one) 
                    and has_free_day(t_name, day_two)]
    if len(group_one)>0 and len(group_two)>0 : 
        return (group_one + group_two)
    else :
        return []
    

def name_slot_subject(st_name, slot): 
    for subj in slot_subject_list_dict[slot]:
        if st_name in t.members(subj):
            return subj
    else : 
        return None

def name_slot_class(st_name, slot):
    subj = name_slot_subject(st_name, slot)
    if subj is None : 
        return None
    else : 
        return t.name_subject_class(st_name, subj)
    
def slot_str(slot):
    day_str = term.date_day_list[ slot[0] - 1 ]
    p_num = str(slot[1])
    #return "%s %s교시" % (day_str, p_num) 
    return "{} {}교시".format(day_str, p_num)


def export_name_slot_classes(filename):
    output = { }
    for slot in sorted(slot_minutes_dict.keys()):
        for st_name in t.students : 
            key_str = slot_str(slot)
            output[ (st_name, key_str) ] = ''
            clss = name_slot_class(st_name, slot)
            if clss is not None:
                output[ (st_name, key_str) ] = t.para_classname(clss)
    xlsx_util.write_dict_into_xlsx(filename, output)
            

    return None



def count_minutes_tuple(t_name, moderation = True, raw=False) : 
    rev_dict = list_util.reversed_dict( supervisor_dict )
    #output = [t_name, str(count_minutes(t_name, moderation=moderation,
        #raw=raw, accumulation=True)), str(count_minutes(t_name, moderation)) ]

    if moderation : 
        if t_name in term.moderation_dict : 
            weight = term.moderation_dict[t_name]
        else : 
            weight = 0 
    else : 
        weight = 0 

    output = [t_name.replace("t",""), weight  ] 
    if t_name not in rev_dict : 
        return output
    else : 
        triples = rev_dict[t_name]
        slot_list = [ ] 
        period_list = [ ] 
        
        for slot, clss, sv in triples : 
            if slot not in slot_list : 
                period_list.append( minutes_of((slot,clss,sv), raw=raw))
                slot_list.append(slot)
        output.extend(period_list)
        return output

def delete_fifths():
    my_slots = [c for c in slot_subject_list_dict.keys() if c[1]==5]
    for sl in my_slots : 
        #print("Deleting %s from slot_subject_list_dict" % str(sl))
        print("Deleting {} from slot_subject_list_dict".format(str(sl)))
        del slot_subject_list_dict[sl]

    my_slots = [c for c in slot_minutes_dict.keys() if c[1]==5]
    for sl in my_slots : 
        #print("Deleting %s from slot_minutes_dict" % str(sl))
        print("Deleting {} from slot_minutes_dict".format(str(sl)))
        del slot_minutes_dict[sl]



def consecutive_slots(): 
    output = [ ] 
    for t_name in term.shuffled_teachers : 
        my_slots = slots_of(t_name)
        for day in range(1, len(term.date_day_list)+1):
            #if list_util.is_subset( [ (day,1), (day,2) ] , my_slots ): 
                #output.append( [t_name, (day,1), (day,2) ])
            if list_util.is_subset( [ (day,2), (day,3) ] , my_slots ): 
                output.append( [t_name, (day,2), (day,3) ])
    output.sort()
    return output


def exchange_candidates(t_name, slot):
    candi_slots = available_slots_for_teacher(t_name)
    tmp_names = available_teachers(name_slot_triple(t_name, slot), top=70)
    candi_names = [c for c in tmp_names 
        if list_util.intersect(candi_slots, slots_of(c)) != [] ]
    return candi_names

def subject_slot_clnums():
    output = [ ] 
    for subj in exam_subjects : 
        #output.append( (subj, "제%d일 %d교시" % subject_slot(subj), subject_clnum_dict[subj]))
        output.append( (subj, "제{}일 {}교시".format(subject_slot(subj), 
            subject_clnum_dict[subj])))
    output.sort()
    return output

def class_classroom(in_class):
    for k in classroom_dict : 
        if k[1] == in_class : 
            return classroom_dict[k]
    else:
        return None

def class_subject_classrooms():
    my_classes = list_util.flatten( 
        [t.classes_of(subj) for subj in exam_subjects] )
    my_classes.sort( key = lambda x: (t.subject_name_of(x), 
                                    int(t.class_number_of(x))))
    output = [ ] 
    for clss in my_classes : 
        output.append( ( t.para_classname(clss), t.subject_name_of(clss), 
            class_classroom(clss) ) )
    return output


    
def export_supervisor_tuples_xlsx(filename):
    quints = [ ] 
    for sl, clss, sv in supervisor_dict : 
        t_name = supervisor_dict[ (sl, clss, sv) ].replace("t", "")
        subj = t.subject_name_of(clss)
        period = subject_minutes_dict[subj]
        if subj in term.auto_subjects : 
            out_sv = "자율감독" 
        else : 
            out_sv = sv
        quints.append( (t_name, sl, subj, period, out_sv) )
    quints.sort()

    name_slots = [ ] 
    output = [ ]
    for rec in quints : 
        if (rec[0], rec[1]) in name_slots : 
            continue
        else:
            name_slots.append((rec[0], rec[1]))
            output.append( (rec[0], rec[2],rec[3],rec[4]) )
        
    output.sort()

    xlsx_util.write_tuples_into_xlsx(filename, output)
        

